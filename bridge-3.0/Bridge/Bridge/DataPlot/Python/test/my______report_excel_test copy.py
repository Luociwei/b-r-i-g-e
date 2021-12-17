#! /usr/bin/env python3
# --*-- coding: utf-8 ---*---

import sys,os,time,math,re
import time
import threading
import datetime
from pytz import timezone
import pytz



BASE_DIR=os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0,BASE_DIR+'/site-packages/')

try:
    import csv
except Exception as e:
    print('e---->',e)

# print('python import ---->matplotlib')
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except Exception as e:
    print('e---->',e)

# print('python import ----> matplotlib.colors')
try:
    import matplotlib.colors as colors
except Exception as e:
    print('e---->',e)

# print('python import ----> FontProperties')
try:
    from matplotlib.font_manager import FontProperties
except Exception as e:
    print('e---->',e)


# print('python import ----> numpy')
try:
    import numpy as np
except Exception as e:
    print('e--->',e)

# print('python import ----> pandas')
try:
    import pandas as pd
except Exception as e:
    print('e--->',e)

# print('python import ----> openpyxl')
try:
    import openpyxl
except Exception as e:
    print('import openpyxl error:',e)
# print('python import ----> xlsxwriter')
try:
    import xlsxwriter
except Exception as e:
    print('import xlsxwriter error:',e)

# print('python import ----> diptest')
try:
    import diptest
except Exception as e:
    print('import diptest error:',e)


try:
    import zmq
except Exception as e:
    print('import zmq error:',e)
# print('python import ----> zmg')

try:
    import redis
except Exception as e:
    print('import redis error:',e)
# print('python import ----> redis')

print(sys.getdefaultencoding())

redisClient = redis.Redis(host='localhost', port=6379, db=0)

context = zmq.Context()
socket = context.socket(zmq.REP)
socket.setsockopt(zmq.LINGER,0)
socket.bind("tcp://127.0.0.1:3130")

filelogname = '/tmp/CPK_Log/temp/.excel.txt'
filelognamehash = '/tmp/CPK_Log/temp/.excel_hash.txt'
p_val_checked = ''

# def correlation(message):
#     print("this function is generate correlation plot......")
#     val = r.get(message)
#     # time.sleep(5)  #测试python 执行时间 5s
#     if val:
#         return val
#     else:
#         return b'None'

def get_redis_data(zmqMsg):
    tb = redisClient.get(zmqMsg)
    tb_data=[]
    if tb:
        tb=tb.decode('utf-8')
        tb=tb.split("\n")
        tb=(tb[1:-1])   #去掉数据库首尾元素
        for i in tb:
            k=re.sub('\"','',i)  #去掉数据库引号
            h=re.sub(',','',k)   #去掉数据库逗号
            m=h.strip()          #去掉数据库首尾空白
            if is_number(m):
                tb_data.append(eval(m))   #去掉数字的引号
            else:
                tb_data.append(m)
    else:
        tb_data.append('')
    return tb_data

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
 
    return False


def open_one_item_csv(event,all_csv_path,data_select,remove_fail):

    tmp_lst = []
    with open(all_csv_path, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('fct') != -1:
                # print("FW version---->")
                pass
            elif row[0].lower().find('display name') != -1:
                pass
            elif row[0].lower().find('pdca priority') != -1:
                pass
            elif row[0].lower().find('upper limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('lower limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('measurement unit') != -1:  # "Measurement Unit ----->" in row:
                pass
            elif row[0].lower().find('site') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # print("index---->", tmp_lst[0])
    header_list = tmp_lst[0][:]#
    temp_header_list = tmp_lst[0]

    if str(header_list[12]).lower() == 'fixture channel id' or str(header_list[13]).lower() == 'fixture channel id':
        temp_header_list[14]= 'slot_id' #rename fixture channel id

    if header_list[12] == header_list[13]:
        temp_header_list[13] = temp_header_list[13]+'_2'
        temp_df = pd.DataFrame(tmp_lst[1:], columns=temp_header_list)
        correlation_header_df =temp_df[0:2]




    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    header_df =df[0:2]
    # print('header_df before--->', header_df)
    data_df = df[2:]
    correlation_data_df = data_df
    # print('data_df before--->', data_df)
    
    # print('99999888--->',type(correlation_data_df.columns),correlation_data_df.columns[12],correlation_data_df.columns[13])
    if correlation_data_df.columns[12] != correlation_data_df.columns[13]:
        print(' one_item_plot before remove empty--->', len(correlation_data_df.values.tolist()))
        correlation_data_df=correlation_data_df[~correlation_data_df[correlation_data_df.columns[12]].isin([''])]#Remove SN Empty
        correlation_data_df=correlation_data_df[~correlation_data_df[correlation_data_df.columns[13]].isin([''])]#Remove SN Empty
        print(' one_item_plot after remove empty--->', len(correlation_data_df.values.tolist()))


    print('csv data row number before remove SN empty--->', len(correlation_data_df.values.tolist()))
    correlation_data_df=correlation_data_df[~correlation_data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(correlation_data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(correlation_data_df.values.tolist()))
    if remove_fail== 'yes':
        correlation_data_df=correlation_data_df[correlation_data_df['Test Pass/Fail Status'].isin(['PASS'])]
        # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(correlation_data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(correlation_data_df.values.tolist()))
    # print('correlation_data_df--->',type(correlation_data_df),len(correlation_data_df.values.tolist()))
    if data_select == 'first':
        correlation_data_df = correlation_data_df.sort_values(axis=0,by=['StartTime'],ascending='True')
        correlation_data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        correlation_data_df = correlation_data_df.sort_values(axis=0,by=['StartTime'],ascending='True')
        correlation_data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        correlation_data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(correlation_data_df.values.tolist()))
    print("==========================<<2222>>>>>>here")
    correlation_start_time_first,correlation_start_time_last = '',''
    print("==========================<<>>>>>>here")
    if len(correlation_data_df.values.tolist()) >2:
        correlation_start_time_l  = correlation_data_df['StartTime'].values.tolist() #StartTime
        correlation_start_time_first = min(correlation_start_time_l)
        correlation_start_time_last  = max(correlation_start_time_l)
        print('<correlation first time -- last time>',correlation_start_time_first,correlation_start_time_last)

    if header_list[12] == header_list[13]:
        df_correlation = correlation_header_df.append(correlation_data_df)
    else:
        df_correlation = header_df.append(correlation_data_df)






    print('csv data row number before remove SN empty--->', len(data_df.values.tolist()))
    data_df=data_df[~data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(data_df.values.tolist()))
    if remove_fail== 'yes':
        data_df=data_df[data_df['Test Pass/Fail Status'].isin(['PASS'])]
        # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(data_df.values.tolist()))
    if data_select == 'first':
        data_df = data_df.sort_values(axis=0,by=['StartTime'],ascending='True')
        data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        data_df = data_df.sort_values(axis=0,by=['StartTime'],ascending='True')
        data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(data_df.values.tolist()))

    start_time_l  = data_df['StartTime'].values.tolist() #StartTime
    # print('start_time_l:',start_time_l)
    if len(start_time_l)>0:
        start_time_first = min(start_time_l)
        start_time_last  = max(start_time_l)
    else:
        start_time_first = ''
        start_time_last  = ''
    # print('<first time -- last time>',start_time_first,start_time_last)

    if header_list[12] == header_list[13]:
        df = correlation_header_df.append(data_df)
    else:
        df = header_df.append(data_df)

    # print("index1---->", header_list)

    # print('df after--->',df_correlation.columns.values.tolist(), df._stat_axis.values.tolist(),df)
    # print('df.values ---->', df.values)#array([[ ]])
    return header_list,df,df_correlation,start_time_first,start_time_last,correlation_start_time_first,correlation_start_time_last

def clear_files(path):
    for root,dirs,files in os.walk(path):
        for file in files:
            os.remove(path+'/'+file)

def creat_excel_report_file(path,file,cpk_lsl,cpk_usl,event,fail_plot_to_excel):
    if event == 'keynote-report':
        excel_file_path = '/tmp/CPK_Log/temp/cpk.xlsx'
    else:
        excel_file_path = path + file
        global excel_report_name 
        excel_report_name = excel_file_path
        global hash_csv_path
        hash_csv_path = path

    book = xlsxwriter.Workbook(excel_file_path)#'cpk.xlsx'
    report_sheet = book.add_worksheet('report')#'report'
    ssh_sheet = book.add_worksheet('ssh')#'hash number '
    if event == 'excel-report' and fail_plot_to_excel == 'yes':
        plot_sheet = book.add_worksheet('fail plot')#'plot'
    else:
        plot_sheet = None
    

    report_sheet.set_column("B:B", 80)  # 设定A列列宽为40
    report_sheet.set_column("M:M", 10)  # 
    report_sheet.set_column("AB:AB",40)  # 

    report_sheet.set_row(0, 30)    #设置行高度

    cpk_result ='CPK_Result(cpk_lsl:'+str(cpk_lsl)+';cpk_usl:'+str(cpk_usl)+')'
    if p_val_checked == '1':
        report_sheet_title = [u'No', u'Item_name',u'BC',u'P_Val',u'a_Q',u'a_irr',u'3CV',
                                u'ORIG LSL',u'Target', u'ORIG USL', u'Min', u'Mean',u'Max', 
                                u'Std',u'CPL', u'CPU', u'CPK',cpk_result, u'New LSL',u'New Target',
                                u'New USL',u'New CPL',u'New CPU',u'New CPK','New '+cpk_result,
                                u'Reviewer Name',u'Review Date',u'Comment']
    else:
        report_sheet_title = [u'No', u'Item_name',u'BC',u'a_Q',u'a_irr',u'3CV',
                                u'ORIG LSL',u'Target', u'ORIG USL', u'Min', u'Mean',
                                u'Max', u'Std',u'CPL', u'CPU', u'CPK',cpk_result, 
                                u'New LSL',u'New Target',u'New USL',u'New CPL',
                                u'New CPU',u'New CPK','New '+cpk_result,
                                u'Reviewer Name',u'Review Date',u'Comment']

    ssh_sheet.set_column("B:B", 20)  
    ssh_sheet.set_column("C:C", 50)  
    ssh_sheet.set_row(0, 30)    #设置行高度
    ssh_sheet_title = [u'Index', u'Item',u'ssh code']

    format_normal = book.add_format()  # 定义format格式对象
    format_normal.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_normal.set_valign('center') # 定义format_titile对象单元格对齐方式

    format_normal.set_border(1)  # 定义format对象单元格边框加粗的格式
    # format_normal.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    format_normal.set_text_wrap()  # 内容换行



    new_format_pass = book.add_format()  # 定义format格式对象
    new_format_pass.set_align('center')  # 定义format_titile对象单元格对齐方式
    new_format_pass.set_valign('center') # 定义format_titile对象单元格对齐方式
    new_format_pass.set_bg_color('#00FF00')  # 定义format_titile对象单元格背景颜色

    new_format_pass.set_border(1)  # 定义format对象单元格边框加粗的格式
    # new_format_pass.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    new_format_pass.set_text_wrap()  # 内容换行


    format_highlight = book.add_format()  # 定义format_title 格式对象
    format_highlight.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    format_highlight.set_bg_color('yellow')  # 定义format_titile对象单元格背景颜色
    format_highlight.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_highlight.set_valign('center') # 定义format_titile对象单元格对齐方式

    # format_highlight.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    format_highlight.set_text_wrap()  # 内容换行



    new_format_fail = book.add_format()  # 定义format_title 格式对象
    new_format_fail.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    new_format_fail.set_bg_color('red')  # 定义format_titile对象单元格背景颜色
    new_format_fail.set_align('center')  # 定义format_titile对象单元格对齐方式
    new_format_fail.set_valign('center') # 定义format_titile对象单元格对齐方式

    # format_fail.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    new_format_fail.set_text_wrap()  # 内容换行


    format_titile = book.add_format()  # 定义format_title 格式对象
    format_titile.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    format_titile.set_bg_color('#cddccc')  # 定义format_titile对象单元格背景颜色
    format_titile.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_titile.set_valign('center')  # 定义format_titile对象单元格对齐方式
    format_titile.set_bold()  # 定义format_titile对象内容加粗
    format_titile.set_text_wrap()  # 内容换行


    format_ave = book.add_format()  # 定义format_ave格式对象
    format_ave.set_border(1)  # 边框加粗的格式
    format_ave.set_num_format('0.00')  # 定义format_ave对象单元格数字类别显示格式
    report_sheet.write_row('A1', report_sheet_title, format_titile)
    ssh_sheet.write_row('A1', ssh_sheet_title, format_titile)

    print('Excel report head created finished!')

    return book,report_sheet,plot_sheet,ssh_sheet,format_highlight,format_normal,format_titile,new_format_fail,new_format_pass



def valid_column(test_item_name,column_list):

    # column_list = df[test_item_name].tolist()
    # print('valid_column item_name-->',test_item_name)
    # print('valid_column--->',len(column_list),len(set(column_list)),column_list)

    # print('column_value_list--->',column_list)
    if test_item_name.lower().find('fixture vendor_id') == -1 and test_item_name.lower().find('unit number') == -1:
        pass
    else:
        return 'not_cpk state1'
    if len(column_list)< 3:
        # column_list[0] == '0' and column_list[1] == '0' or column_list[0] == '1' and column_list[1] == '1' or column_list[0] == column_list[1] and  column_list[0] !='NA' or 
        # print('====>',column_list[0],column_list[1])
        return "not_cpk state2"
    else:
        # pattern = re.compile(r'^[+-]?[0-9]*\.?[0-9]+$')
        # print('valid_column len----->',test_item_name,len(column_list))
        # print('v====>',column_list[0],column_list[1],column_list,len(column_list))
        j=0
        for i in range(2,len(column_list),1):
            # print(str(i),'valid_column----->',test_item_name,pattern.match(column_list[i]))
            if is_number(column_list[i]):
                j=j+1
                # print(str(j)+',len(set(column_list))-->',j,len(set(column_list)))
                if j>0 and len(set(column_list))>=1:#
                    # print('need_cpk')
                    return "need_cpk"
 
        return "not_cpk state3"


def test_value_to_numeric(test_data_list):
    column_list = []
    i = 0
    for x in test_data_list:
        if (i ==0 and (x == 'NA' or x == '')) or (i ==1 and (x == 'NA' or x == '')):
            column_list.append(x)
        else:
            try:
              x = float(x)
              column_list.append(x)
            except Exception as e:
                pass
            # print('-------------------- it is not number --------------')
        i = i + 1

    # print('column_list--->',column_list)
    return column_list

def is_empty_list(l):
    temp_l = []
    for value in l:
        if value != '':
            temp_l.append(value)
    n = len(temp_l)
    if n == 0:
        # print('it is empty list!')
        return True
    else:
        return False




def parse_all_csv_local(header_list,df,color_by1,selected_category1,event,color_by2,selected_category2,param_item_start_index):

    color_l = ['#0000FF','#FF0000','#008000','#00FFFF','#000000',
                '#8B008B','#B8860B','#FF6347','#A9A9A9','#FFFF00',
                '#A52A2A','#7FFF00','#D2691E','#6495ED','#FF00FF']
    table_data = []#[[]]
    table_category_data = []#[[[]]]
    column_list = []
    n=0
    no_valid_column_name_l = []
    for item_name in header_list:
        # if event == 'one_item_plot1':
        #     # print('one column--->',type(df.iloc[:, [n]].values),df.iloc[:, [n]].values.tolist())
        #     temp_l=[]
        #     for i in df.iloc[:, [n]].values.tolist():
        #         temp_l.append(i[0])
        #     column_list = temp_l  # 读取指定键值列的所有行
        #     print('column_list--->',column_list)
        #     n=n+1
        # else:
        # print('item_name:',item_name)
        # print('header_list in parse_all_csv:',header_list)
        # print('df:',df)
        try:
            column_list = df[item_name].tolist()
        except Exception as e:
            if event == 'excel-report' and header_list.index(item_name) >= param_item_start_index:
                no_valid_column_name_l.append(item_name)
            print(item_name,'is duplicate ? pls check!',e)
            continue
        # print('column_list--->',type(df[item_name]),column_list)

        need_cpk = valid_column(item_name,column_list)
        # print('need_cpk:---->',need_cpk)
        column_num_list = []
        if need_cpk == 'need_cpk':
            column_list = test_value_to_numeric(column_list)
            column_list.insert(0, item_name)  # item name
            usl = column_list[1]
            lsl = column_list[2]
            # print('color_by1:',color_by1)
            # 'Off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'/'Product'/'StartTime'/'Special Build Description'
            if color_by1 == 'SerialNumber' or color_by1 == 'Version' or color_by1 == 'Station ID' or color_by1 == 'Special Build Name' or color_by1 == 'Product' or color_by1 == 'StartTime' or color_by1 =='Special Build Description' or color_by1 =='Fixture Channel ID' or color_by1 =='Diags_Version':
                column_temp = []
                first_filter_category_data_len=0
                second_filter_category_data_len=0
                i =0
                for x in selected_category1:
                    # print('x:',x,color_by2)
                    if color_by2 == 'SerialNumber' or color_by2 == 'Version' or color_by2 == 'Station ID' or color_by2 == 'Special Build Name' or color_by2 == 'Product' or color_by2 == 'StartTime' or color_by2 =='Special Build Description' or color_by2 =='Fixture Channel ID' or color_by2 =='Diags_Version':
                        
                        for xx in selected_category2:

                            # print('xxxxx:', item_name,color_by1,x[0],x[1],color_by2,xx[0],xx[1])
                            if color_by2 == 'Fixture Channel ID':
                                # print('the same!',df.columns.values.tolist()[14])
                                index2 = df.columns.values.tolist()[14]
                            index2 = color_by2
                            one_category_list = df.loc[(df[color_by1] == x[0]) & (df[index2] == xx[0]), item_name].tolist()  # 
                            if is_empty_list(one_category_list) != True:
                                # print('one category_data with second filter-->', len(one_category_list),one_category_list)
                                second_filter_category_data_len = second_filter_category_data_len +len(one_category_list)
                                one_category_list = test_value_to_numeric(one_category_list)
                                column_temp = column_temp +one_category_list

  
                                # print('usl--->', usl)
                                # print('lsl--->', lsl)
                                category_value = x[0]+'&'+xx[0]
                                # print('category_value--->',category_value)
                                one_category_list.insert(0, category_value)  # insert category
                                # print('i==>',i)
                                if i >14:
                                    i = 0 
                                one_category_list.insert(1, color_l[i])  # insert color
                                one_category_list.insert(2,item_name)  # item_name
                                if lsl != 'NA':
                                    lsl = float(lsl)
                                if usl != 'NA':
                                    usl = float(usl)

                                one_category_list.insert(3, usl)  # usl
                                one_category_list.insert(4, lsl)  # lsl
                                # print('one_category_list-->', one_category_list)
                                column_num_list.append(one_category_list)
                                i=i+1

                    elif color_by2 == 'Off':

                        # print('color_by2= Off xxxxx:', x[0],x[1])
                        one_category_list = df.loc[df[color_by1] == x[0], item_name].tolist()  # 
                        if is_empty_list(one_category_list) != True:
                            # print('one category_data with first filter-->', one_category_list)
                            first_filter_category_data_len = first_filter_category_data_len +len(one_category_list)
                            one_category_list = test_value_to_numeric(one_category_list)
                            column_temp = column_temp +one_category_list

                            # usl = column_list[1]
                            # lsl = column_list[2]
                            # print('usl--->', usl)
                            # print('lsl--->', lsl)
                            one_category_list.insert(0, x[0])  # insert category
                            one_category_list.insert(1, x[1])  # insert color
                            one_category_list.insert(2,item_name)  # item_name
                            if lsl != 'NA':
                                lsl = float(lsl)

                            if usl != 'NA':
                                usl = float(usl)

                            one_category_list.insert(3, usl)  # usl
                            one_category_list.insert(4, lsl)  # lsl
                            # print('one_category_list-->', one_category_list)
                            column_num_list.append(one_category_list)
                # print('one category_data total with first filter-->', first_filter_category_data_len)
                # print('one category_data total with second filter-->', second_filter_category_data_len)
                
                column_temp.insert(0,item_name)  # item_name
                if lsl != 'NA':
                    lsl = float(lsl)
                if usl != 'NA':
                    usl = float(usl)

                column_temp.insert(1, usl)  # usl
                column_temp.insert(2, lsl)  # lsl
                # print('column_temp-->', column_temp)
                table_data.append(column_temp)# one column's all category data []
                # print('column_num_list-->', column_num_list)
                table_category_data.append(column_num_list)# category [[],[],...]]
            elif color_by1 == 'Off':
                if event == 'excel-report':
                    if header_list.index(item_name) >= param_item_start_index:
                        table_data.append(column_list)  # item_name,usl,lsl,data
                else:
                    table_data.append(column_list)  # item_name,usl,lsl,data
        else:
            # print('no cpk item:',event,header_list.index(item_name),param_item_start_index)
            if event == 'excel-report' and header_list.index(item_name) >= param_item_start_index:
                no_valid_column_name_l.append(item_name)
    # print('table_data-->',len(table_data[0]),len(table_data),len(table_category_data),table_data,table_category_data)
    if event == 'excel-report':
        pass
    else:
        no_valid_column_name_l = []
    # print('no_valid_column_name_l-->',no_valid_column_name_l)
    return table_data,table_category_data,no_valid_column_name_l #[[[ ]]]




def parse_all_csv(header_list,df,color_by1,selected_category1,event,color_by2,selected_category2,param_item_start_index):

    color_l = ['#0000FF','#FF0000','#008000','#00FFFF','#000000',
                '#8B008B','#B8860B','#FF6347','#A9A9A9','#FFFF00',
                '#A52A2A','#7FFF00','#D2691E','#6495ED','#FF00FF']
    table_data = []#[[]]
    table_category_data = []#[[[]]]
    column_list = []
    n=0
    no_valid_column_name_l = []
    for item_name in header_list:
        # if event == 'one_item_plot1':
        #     # print('one column--->',type(df.iloc[:, [n]].values),df.iloc[:, [n]].values.tolist())
        #     temp_l=[]
        #     for i in df.iloc[:, [n]].values.tolist():
        #         temp_l.append(i[0])
        #     column_list = temp_l  # 读取指定键值列的所有行
        #     print('column_list--->',column_list)
        #     n=n+1
        # else:
        # print('item_name:',item_name)
        # print('header_list in parse_all_csv:',header_list)
        # print('df:',df)
        try:
            column_list = df[item_name].tolist()
        except Exception as e:
            if event == 'excel-report' and header_list.index(item_name) >= param_item_start_index:
                no_valid_column_name_l.append(item_name)
            print(item_name,'is duplicate ? pls check!',e)
            continue
        # print('column_list--->',type(df[item_name]),column_list)

        need_cpk = valid_column(item_name,column_list)
        # print('need_cpk:---->',need_cpk)
        column_num_list = []
        if need_cpk == 'need_cpk':
            column_list = test_value_to_numeric(column_list)
            column_list.insert(0, item_name)  # item name
            usl = column_list[1]
            lsl = column_list[2]
            # print('color_by1:',color_by1)
            # 'Off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'/'Product'/'StartTime'/'Special Build Description'
            if color_by1 == 'SerialNumber' or color_by1 == 'Version' or color_by1 == 'Station ID' or color_by1 == 'Special Build Name' or color_by1 == 'Product' or color_by1 == 'StartTime' or color_by1 =='Special Build Description' or color_by1 =='Fixture Channel ID' or color_by1 =='Diags_Version':
                column_temp = []
                first_filter_category_data_len=0
                second_filter_category_data_len=0
                i =0
                for x in selected_category1:
                    # print('x:',x,color_by2)
                    if color_by2 == 'SerialNumber' or color_by2 == 'Version' or color_by2 == 'Station ID' or color_by2 == 'Special Build Name' or color_by2 == 'Product' or color_by2 == 'StartTime' or color_by2 =='Special Build Description' or color_by2 =='Fixture Channel ID' or color_by2 =='Diags_Version':
                        
                        for xx in selected_category2:

                            # print('xxxxx:', item_name,color_by1,x[0],x[1],color_by2,xx[0],xx[1])
                            if color_by2 == 'Fixture Channel ID':
                                # print('the same!',df.columns.values.tolist()[14])
                                index2 = df.columns.values.tolist()[14]
                            index2 = color_by2
                            one_category_list = df.loc[(df[color_by1] == x[0]) & (df[index2] == xx[0]), item_name].tolist()  # 
                            if is_empty_list(one_category_list) != True:
                                # print('one category_data with second filter-->', len(one_category_list),one_category_list)
                                second_filter_category_data_len = second_filter_category_data_len +len(one_category_list)
                                one_category_list = test_value_to_numeric(one_category_list)
                                column_temp = column_temp +one_category_list

  
                                # print('usl--->', usl)
                                # print('lsl--->', lsl)
                                category_value = x[0]+'&'+xx[0]
                                # print('category_value--->',category_value)
                                one_category_list.insert(0, category_value)  # insert category
                                # print('i==>',i)
                                if i >14:
                                    i = 0 
                                one_category_list.insert(1, color_l[i])  # insert color
                                one_category_list.insert(2,item_name)  # item_name
                                if lsl != 'NA':
                                    lsl = float(lsl)
                                if usl != 'NA':
                                    usl = float(usl)

                                one_category_list.insert(3, usl)  # usl
                                one_category_list.insert(4, lsl)  # lsl
                                # print('one_category_list-->', one_category_list)
                                column_num_list.append(one_category_list)
                                i=i+1

                    elif color_by2 == 'Off':

                        # print('color_by2= Off xxxxx:', x[0],x[1])
                        one_category_list = df.loc[df[color_by1] == x[0], item_name].tolist()  # 
                        if is_empty_list(one_category_list) != True:
                            # print('one category_data with first filter-->', one_category_list)
                            first_filter_category_data_len = first_filter_category_data_len +len(one_category_list)
                            one_category_list = test_value_to_numeric(one_category_list)
                            column_temp = column_temp +one_category_list

                            # usl = column_list[1]
                            # lsl = column_list[2]
                            # print('usl--->', usl)
                            # print('lsl--->', lsl)
                            one_category_list.insert(0, x[0])  # insert category
                            one_category_list.insert(1, x[1])  # insert color
                            one_category_list.insert(2,item_name)  # item_name
                            if lsl != 'NA':
                                lsl = float(lsl)
                            if usl != 'NA':
                                usl = float(usl)

                            one_category_list.insert(3, usl)  # usl
                            one_category_list.insert(4, lsl)  # lsl
                            # print('one_category_list-->', one_category_list)
                            column_num_list.append(one_category_list)
                # print('one category_data total with first filter-->', first_filter_category_data_len)
                # print('one category_data total with second filter-->', second_filter_category_data_len)
                
                column_temp.insert(0,item_name)  # item_name
                if lsl != 'NA':
                    lsl = float(lsl)

                if usl != 'NA':
                    usl = float(usl)

                column_temp.insert(1, usl)  # usl
                column_temp.insert(2, lsl)  # lsl
                # print('column_temp-->', column_temp)
                table_data.append(column_temp)# one column's all category data []
                # print('column_num_list-->', column_num_list)
                table_category_data.append(column_num_list)# category [[],[],...]]
            elif color_by1 == 'Off':
                if event == 'excel-report':
                    if header_list.index(item_name) >= param_item_start_index:
                        table_data.append(column_list)  # item_name,usl,lsl,data
                else:
                    table_data.append(column_list)  # item_name,usl,lsl,data
        else:
            # print('no cpk item:',event,header_list.index(item_name),param_item_start_index)
            if event == 'excel-report' and header_list.index(item_name) >= param_item_start_index:
                no_valid_column_name_l.append(item_name)
    # print('table_data-->',len(table_data[0]),len(table_data),len(table_category_data),table_data,table_category_data)
    if event == 'excel-report':
        pass
    else:
        no_valid_column_name_l = []
    # print('no_valid_column_name_l-->',no_valid_column_name_l)
    return table_data,table_category_data,no_valid_column_name_l #[[[ ]]]



def get_coefficients(value_l):
    '''
    param value_l: need float list
    return: bc,p_val,a_Q,a_irr,three_σ_x100_divide_mean
    1σ＝690000／1000000 #fault rate
    2σ＝308000／1000000
    3σ＝66800／1000000
    4σ＝6210／1000000
    5σ＝230／1000000
    6σ＝3.4／1000000 
    7σ＝0／1000000
    '''
    if len(value_l) <= 3:
        return '','','','',''
    
    column_stdev = np.std(value_l,ddof=1,axis=0)
    three_sigma= 3*column_stdev
    # print('three_sigma:',column_stdev,three_sigma)
    temp_l= value_l
    if len(value_l) < 10:
        temp_l = value_l + value_l
        if len(temp_l)<10:
            temp_l = value_l + value_l + value_l + value_l+value_l + value_l + value_l + value_l+value_l + value_l

    # print('------>>>temp_l:',temp_l)
    data0 = np.array(temp_l)
    # print('------>>>data0:',data0)
    try:
        dip, p_val = diptest.diptest(data0)
        p_val = '%f'%p_val
        # print('------>>>p_val :',p_val)
    except RuntimeWarning as w:
        print('calculate dip,p_val RuntimeWarning:',w)
        return '','','','',''        
    except Exception as e:
        print('calculate dip,p_val error:',e)
        return '','','','',''
    # print('dip,p_val:',dip,p_val)
    item_name ='value1'
    data = pd.DataFrame({item_name:value_l})
    # print('data--->',type(data),data)
    u1 = data[item_name].mean() # 计算均值
    # std1 = data[item_name].std() # 计算标准差
    # t,pval=stats.kstest(data[item_name], 'norm', (u1, std1))
    # print('normality test t,pval:',str(t),str(pval))

    # print('------')
    # 正态性检验 → pvalue >0.05

    n= float(len(value_l))
    #Item (xi-ẍ)^2
    item_l_1 =[]
    item_l_2 = []
    item_l_3 = []
    for i in value_l:
        temp1 = (i-u1)**2
        temp2 = (i-u1)**3
        temp3 = (i-u1)**4
        item_l_1.append(temp1)
        item_l_2.append(temp2)
        item_l_3.append(temp3)
    # print('item_l_1--->',item_l_1)
    # print('item_l_2--->',item_l_2)
    # print('item_l_3--->',item_l_3)
    sum_item_l_1 = sum(item_l_1)
    sum_item_l_2 = sum(item_l_2)
    sum_item_l_3 = sum(item_l_3)
    # print('sum_item_l_1',sum_item_l_1)
    # print('sum_item_l_2',sum_item_l_2)
    # print('sum_item_l_3',sum_item_l_3)
    if n<=3 or sum_item_l_1==0 or sum_item_l_2==0 or sum_item_l_3==0:
        # print('len < 3--->')
        if abs(u1) == 0:
            return 'Nan',str(p_val),'Nan','Nan','Nan'
        else:
            three_CV = three_sigma*100/abs(u1)
            # print('three_CV:',three_CV)
            return 'Nan',str(p_val),'Nan','Nan',str(round(three_CV,6))
    else:
        try:
            m3 = np.sqrt(n*(n-1))/(n-2)*((1/n*sum_item_l_2)/np.sqrt(1/n*sum_item_l_1)**3)

            # print('m3=',m3)
            # print('d8:',n+1)
            # print('d15:',1/n*sum_item_l_3/(1/n*sum_item_l_1)**2)
            # print('d16:',(n+1)*1/n*sum_item_l_3/(1/n*sum_item_l_1)**2-3*(n-1))
            # print('d6/d7',(n-1)/((n-2)*(n-3)))
            m4 = ((n-1)/((n-2)*(n-3)))*((n+1)*1/n*sum_item_l_3/(1/n*sum_item_l_1)**2-3*(n-1))#(d6/d7)*d16
            # print('m4=',m4)
            #(d14**2+1)/(d17+3*(d10/d7))
            bc =(m3**2+1)/(m4+3*((n-1)**2/((n-2)*(n-3))))
            # print('bc:',bc)
            a_L=0.05
            a_M=0.1
            a_U=0.32
            a_Q = (a_U-a_L)*bc**2+a_L
            # print ('a_Q:',a_Q)
            a_irr = np.sqrt((a_U-a_L)**2*bc)+a_L
            # print('a_irr:',a_irr)
        except Exception as e:
            # print('calculate error',e)
            if abs(u1) == 0:
                return 'Nan',str(p_val),'Nan','Nan','Nan'
            else:
                three_CV = three_sigma*100/abs(u1)
                # print('three_CV:',three_CV)
                return 'Nan',str(p_val),'Nan','Nan',str(round(three_CV,6))


    if abs(u1) == 0:
        return str(round(bc,6)),str(p_val),str(round(a_Q,6)),str(round(a_irr,6)),'Nan'
    else:
        three_CV = three_sigma*100/abs(u1)
        # print('three_CV:',three_CV)
        return str(round(bc,6)),str(p_val),str(round(a_Q,6)),str(round(a_irr,6)),str(round(three_CV,6))



def cpk_calc(df_data,lsl,usl):
    """
    :param df_data: list
    :param usl: 数据指标上限
    :param lsl: 数据指标下限
    :return:
    """

    sigma = 3
    # print('limit--->',lsl,usl)
    # 数据平均值
    # print('df_data in cpk_calc:',df_data)
    mean = np.mean(df_data)#
    # print('mean ---->',mean)

    # 数据max值
    max_num = max(df_data)
    # print('max_num ---->',max_num)

    # 数据min值
    min_num = min(df_data)
    # print('min_num ---->',min_num)

    # a = np.array([[1, 2], [3, 4]])
    # print('a--->', type(a))
    # print('gobal std:',np.std(a))#全局标准差
    # print('each line std:',np.std(a, axis=0,ddof=1))
    # print("each row std:",np.std(a, axis=1,ddof=1))

    # 数据标准差
    if len(df_data)==1:
        stdev =0.00
        return (mean,max_num,min_num,stdev,None,None,None,None,None)
    else:
        try:
            stdev = np.std(df_data,ddof=1,axis=0)
        except Exception as e:
            return (mean,max_num,min_num,stdev,None,None,None,None,None)


    # print('stdev ---->',stdev)
    if stdev == 0:#stop count cpk
        return (mean,max_num,min_num,stdev,None,None,None,None,None)
    # 生成横轴数据平均分布
    # x1 = np.linspace(mean - sigma * stdev, mean + sigma * stdev, 1000)
    # print('x1 ---->',x1)
    # 计算正态分布曲线
    # y1 = np.exp(-(x1 - mean) ** 2 / (2 * stdev ** 2)) / (math.sqrt(2 * math.pi) * stdev)
    # print('y1 ---->',y1)
    x1,y1 = None,None

    if (lsl != 'NA' and  lsl != '') and (usl == 'NA' or usl == ''):
        cpl = (mean - lsl) / (sigma * stdev)
        # print('====>>>>>>cpl',cpl)
        return (mean,max_num,min_num,stdev,None,None,None,cpl,None)

    if (usl != 'NA' and usl != '') and (lsl == 'NA' or  lsl == ''):
        # print('====>>>>>>=====cpu')
        cpu = (usl - mean) / (sigma * stdev)
        # print('====>>>>>>cpu',cpu)
        return (mean,max_num,min_num,stdev,None,None,cpu,None,None)

    if lsl == 'NA' or usl == 'NA' or lsl == '' or usl == '':
        return (mean,max_num,min_num,stdev,None,None,None,None,None)
    cpu = (usl - mean) / (sigma * stdev)
    cpl = (mean - lsl) / (sigma * stdev)


    # print('cpu ---->',cpu)
    # print('cpl ---->',cpl)

    # 得出cpk
    cpk = min(cpu, cpl)
    return (mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk)



def get_target_value(lsl,usl):
    if lsl != 'NA' and usl != 'NA' and lsl != '' and usl != '':
        target_value = round(((lsl + usl) / 2.0), 5)
    else:
        target_value = 'Nan'
    return target_value


def probability_distribution_extend(data,bins,margin,item_name,lsl,usl,mean,
                                    max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,
                                    pic_path,bins_l,bins_h,start_time_first,
                                    start_time_last,bmc,zoom_type='limit'):
    bins = sorted(bins)
    length = len(bins)
    intervals = np.zeros(length+1)
    for value in data:
        i = 0
        while i < length and value >= bins[i]:
            i += 1
        intervals[i] += 1
    intervals = intervals / float(len(data))
    plt.ion()  # 开启interactive mode

    plt.figure(1)  # 创建图表1
    plt.xlim(min(bins) - margin, max(bins) + margin)
    bins.insert(0, -999)
    # plt.title("probability-distribution",size=8,verticalalignment='bottom')
    plt.bar(bins, intervals,color=['r'], label='')#频率分布
    x_ticks,labels = plt.xticks()
    #x_ticks_start=round(x_ticks[0],2)
    #x_ticks_end = round(x_ticks[len(x_ticks) - 1],2)
    y_ticks, labels = plt.yticks()
    #print('=====y_ticks, labels:',y_ticks, labels)
    # print('y_ticks--->',y_ticks,y_ticks[len(y_ticks) - 1],y_ticks[0])
    y_ticks=round(y_ticks[len(y_ticks) - 1],5)
    # print("x_ticks_start,x_ticks_end--->",x_ticks_start,x_ticks_end)
    # print('y_ticks_end--->',y_ticks)
    # plt.show()
    plt.close(1)



    plt.figure(2,dpi=150)  # 创建图表2
    fig, axes = plt.subplots(1, 0, figsize=(6, 5), facecolor='#ccddef')
    plt.axes([0.1, 0.17, 0.85, 0.7])  # [左, 下, 宽, 高] 规定的矩形区域 （全部是0~1之间的数，表示比例）

    if stdev =='nan':
        pass  
    else:
        if stdev > 999999:
            stdev = str("%.3e" % stdev)
        else:
            stdev = str("%.3f" % stdev)

    if cpl ==None:

        cpl_value =''
    else:
        if cpl > 999999:
            cpl_value = str("%.3e" % cpl)
        else:
            cpl_value = str("%.3f" % cpl)

    if cpu ==None:
        cpu_value =''
    else:
        if cpu > 999999:
            cpu_value = str("%.3e" % cpu)
        else:
            cpu_value = str("%.3f" % cpu)


    if cpk ==None:
        cpk_value =''
    else:
        if cpk > 999999:
            cpk_value = str("%.3e" % cpk)
        else:
            cpk_value = str("%.3f" % cpk)


    if max_num > 999999:
        max_num = str("%.3e" % max_num)
    else:
        max_num = str("%.3f" % max_num)

    if mean > 999999:
        mean = str("%.3e" % mean)
    else:
        mean = str("%.3f" % mean)

    if min_num > 999999:
        min_num = str("%.3e" % min_num)
    else:
        min_num = str("%.3f" % min_num)


    if len(data) > 1000000000000:
        sample_n = str("%.e" % len(data))
    else:
        sample_n = str("%.f" % len(data))

    info = "Samples:" + sample_n + '  ' +"Max:" + max_num + '  ' + "Mean:" + mean + '  ' + "Min:" + min_num + '\n' + "Std:" + stdev + '  ' + "Cpl:" + cpl_value + '  ' + "Cpu:" + cpu_value + '  ' + "Cpk:" + cpk_value + '\n' + "Bimodal:" + bmc

    if len(item_name) > 60:
        item_name = item_name[0:60] + '\n' + item_name[60:]
    # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
    # plt.title(item_name,FontProperties=font)
    plt.title(item_name,size=11,verticalalignment='bottom')
    # plt.xlabel(str(start_time_first)+' -- '+str(start_time_last))
    plt.ylabel('Count')
    bins = [round(x,5) for x in bins]
    bins=sorted(bins)

    plt.hist(data, bins=bins, label=info, histtype='stepfilled',color = 'blue', edgecolor='blue', linewidth=1.0,align='mid',density=False) #time分布
    
    range_value = get_limit_range(bins_l, bins_h)
    range_value =round(range_value/5,5)
    # print('plot bar--->',bins_l,range)

    if zoom_type =='data':
        plt.xlim(float(min_num)*0.999, float(max_num)*1.001)

    else:
        if (lsl =='' or lsl =='NA') and (usl!='NA'and usl!=''):
            plt.xlim(float(min_num)*0.999, usl+range_value)
        elif (usl =='' or usl =='NA') and (lsl!='NA'and lsl!=''):
            plt.xlim(lsl-range_value, float(max_num)*1.001)
        elif (usl =='' or usl =='NA') and (lsl=='NA'or lsl==''):
            plt.xlim(float(min_num)*0.999, float(max_num)*1.001)
        else:
            plt.xlim(bins_l-range_value, bins_h+range_value)

    y_ticks = len(data) * y_ticks
    plt.ylim((0, y_ticks))  # 设置y轴scopex

    ax=plt.gca()
    ax.spines['bottom'].set_linewidth(1)
    ax.spines['left'].set_linewidth(1)
    ax.spines['right'].set_linewidth(1)
    ax.spines['top'].set_linewidth(1)


    # if lsl !='NA' and usl != 'NA' and zoom_type =='limit':
    #     plt.plot([lsl, lsl, ], [0, y_ticks, ], 'k--', linewidth=1.0, color='red')  # 画lower limit线，
    #     plt.plot([usl, usl, ], [0, y_ticks, ], 'k--', linewidth=1.0, color='red')  # 画upper limit线，

    #     plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 8, 'color': 'r'})
    #     plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 8, 'color': 'r'})

    if lsl !='' and lsl !='NA' and zoom_type =='limit':
        plt.plot([lsl, lsl, ], [0, y_ticks, ], 'k--', linewidth=1.0, color='red')  # 画lower limit线，
        # plt.plot([usl, usl, ], [0, y_ticks, ], 'k--', linewidth=1.0, color='red')  # 画upper limit线，
        plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 8, 'color': 'r'})
        # plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 8, 'color': 'r'})

    if usl != '' and usl != 'NA' and zoom_type =='limit':
        # plt.plot([lsl, lsl, ], [0, y_ticks, ], 'k--', linewidth=1.0, color='red')  # 画lower limit线，
        plt.plot([usl, usl, ], [0, y_ticks, ], 'k--', linewidth=1.0, color='red')  # 画upper limit线，
        # plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 8, 'color': 'r'})
        plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 8, 'color': 'r'})

    plt.legend(bbox_to_anchor=(0.9,-0.09),loc="best",fontsize=10,framealpha=0,edgecolor='royalblue',borderaxespad=0.1)
    plt.grid(linestyle=':',c='gray')  # 生成网格

    plt.savefig(pic_path,dpi=200)
    plt.draw()

    plt.close('all')

    plt.ioff()  




def get_bins(min_num,max_num,lsl,usl,set_bins,zoom_type='limit'):
    # print('min_num,max_num,lsl,usl,set_bins=====>',min_num,max_num,lsl,usl,set_bins)

    bins_l = 0
    bins_h = 0
    if lsl == 'NA' or usl == 'NA' or lsl == '' or usl == '' or zoom_type == 'data':
        bins_l,bins_h = min_num,max_num

    else:
        if min_num < lsl and max_num < lsl:
            bins_l = min_num
            bins_h = usl
        elif min_num < lsl and max_num > lsl and max_num <= usl:
            bins_l = min_num
            bins_h = usl
        elif min_num < lsl and max_num > usl:
            bins_l = min_num
            bins_h = max_num
        elif lsl <= min_num and min_num <= usl and max_num <= usl:
            bins_l = lsl
            bins_h = usl
        elif lsl <= min_num and min_num <= usl and max_num > usl:
            bins_l = lsl
            bins_h = max_num
        elif min_num > usl:
            bins_l = lsl
            bins_h = max_num
    range_value = get_limit_range(bins_l,bins_h)
    if lsl == 'NA' or usl == 'NA' or  lsl == '' or usl == '' or zoom_type == 'data':
        if range_value ==0  and min_num > 0:
            range_value = min_num*0.2
            bins_l = (bins_l - min_num*0.1)
            bins_h = (bins_h + min_num*0.1)
 
        elif range_value ==0 and min_num == 0:
            range_value = 6
            bins_l =  - 3
            bins_h = 3
        elif range_value ==0 and min_num <0:
            range_value = 6
            bins_l =  min_num - 3
            bins_h =  min_num + 3
            # print('range_value0-->',range_value,min_num,bins_l,bins_h)

        else:
            if min_num > 1 and range_value < 1 and range_value !=0:
                range_value = min_num*0.05
                bins_l = (bins_l - min_num*0.025)
                bins_h = (bins_h + min_num*0.025)
            elif min_num > 0.001 and min_num < 1 and range_value < 1 and range_value !=0:
                range_value = min_num*0.2
                bins_l = (bins_l - min_num*0.1)
                bins_h = (bins_h + min_num*0.1)

            else:
                range_value = range_value*0.4
                bins_l = (bins_l - range_value*0.2)
                bins_h = (bins_h + range_value*0.2)
          

    else:
        if range_value == 0 and lsl !=0:
            range_value = lsl*0.2
            bins_l = bins_l - lsl*0.1
            bins_h = bins_h + usl*0.1
        elif range_value == 0 and lsl ==0:
            range_value = 6
            bins_l =  - 3
            bins_h = 3


    # print('range_value1-->',range_value)
    range_value = round((range_value/set_bins),12)
    # print('range_value2-->',range_value)
    # print('=====>',bins_l,bins_h,range_value)
    # bins = np.arange(bins_l, bins_h, range_value)#必须是单调递增的

    bins = np.arange(bins_l, bins_h, range_value)#必须是单调递增的
    if bins_l<0:
        bins = np.arange(bins_h, bins_l, -range_value)
        
    # print('lsl,usl,min_num,max_num,bins_l,bins_h in get_bins=====>',lsl,usl,min_num,max_num,bins_l,bins_h)
    return bins,bins_l,bins_h

def get_limit_range(lsl,usl):
    # print('lsl,usl----->', lsl, usl)
    range_value = 0
    if lsl < 0 and usl <= 0:
        range_value = abs(lsl) - abs(usl)
    elif lsl < 0 and usl >= 0:
        range_value = abs(lsl) + usl
    elif lsl >= 0 and (usl > 0):
        range_value = usl - lsl
    else:
        print('get_limit_range 00000')
    range_value = round(range_value, 5)
    # print('range in get_limit_range----->', range_value)
    return range_value





def probability_distribution_extend_by_color(column_category_data_list,data,bins,margin,item_name,
                                            lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,
                                            pic_path,bins_l,bins_h,start_time_first,start_time_last,zoom_type):
    # print('one column data len:',len(data),item_name,data,column_category_data_list)
    bins = sorted(bins)
    length = len(bins)
    intervals = np.zeros(length+1)
    for value in data:
        i = 0
        while i < length and value >= bins[i]:
            i += 1
        intervals[i] += 1
    intervals = intervals / float(len(data))
    plt.ion()  # 开启interactive mode

    plt.figure(1)  # 创建图表1
    plt.xlim(min(bins) - margin, max(bins) + margin)
    bins.insert(0, -999)
    plt.title("probability-distribution")
    plt.bar(bins, intervals,color=['r'], label='')#频率分布
    x_ticks,labels = plt.xticks()
    x_ticks_start=round(x_ticks[0],2)
    x_ticks_end = round(x_ticks[len(x_ticks) - 1],2)
    y_ticks, labels = plt.yticks()
    # print('y_ticks--->',y_ticks,y_ticks[len(y_ticks) - 1],y_ticks[0])
    y_ticks=round(y_ticks[len(y_ticks) - 1],5)
    # print("x_ticks_start,x_ticks_end--->",x_ticks_start,x_ticks_end)
    # print('y_ticks_end--->',y_ticks)
    # plt.show()
    plt.close(1)


    plt.figure(2,dpi=150)  # 创建图表2,facecolor='blue',edgecolor='black'

    if stdev =='nan':
        pass  
    else:
        if stdev > 999999:
            stdev = str("%.3e" % stdev)
        else:
            stdev = str("%.3f" % stdev)

    if cpl ==None:
        cpl_value =''
    else:
        if cpl > 999999:
            cpl_value = str("%.3e" % cpl)
        else:
            cpl_value = str("%.3f" % cpl)

    if cpu ==None:
        cpu_value =''
    else:
        if cpu > 999999:
            cpu_value = str("%.3e" % cpu)
        else:
            cpu_value = str("%.3f" % cpu)


    if cpk ==None:
        cpk_value =''
    else:
        if cpk > 999999:
            cpk_value = str("%.3e" % cpk)
        else:
            cpk_value = str("%.3f" % cpk)


    if max_num > 999999:
        max_num = str("%.3e" % max_num)
    else:
        max_num = str("%.3f" % max_num)

    if mean > 999999:
        mean = str("%.3e" % mean)
    else:
        mean = str("%.3f" % mean)

    if min_num > 999999:
        min_num = str("%.3e" % min_num)
    else:
        min_num = str("%.3f" % min_num)


    if len(data) > 1000000000000:
        sample_n = str("%.e" % len(data))
    else:
        sample_n = str("%.f" % len(data))

    info = "Sample: " + sample_n + '\n' +"Max: " + max_num + '\n' + "Mean: " + mean + '\n' + "Min: " + min_num + '\n' + "Std: " + stdev + '\n' + "Cpl: " + cpl_value + '\n' + "Cpu: " + cpu_value + '\n' + "Cpk: " + cpk_value
    # info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
    #     "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + str(
    #     "%.3f" % cpl) + '\n' + "Cpu: " + str("%.3f" % cpu) + '\n' + "Cpk: " + str("%.3f" % cpk)
    if len(item_name) > 55:
        item_name = item_name[0:55] + '\n' + item_name[55:]
    # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
    plt.title(item_name,size=10)
    plt.xlabel(str(start_time_first)+' -- '+str(start_time_last))
    plt.ylabel('Count')

    # plt.title(item_name+"\nCpk={0}".format(str("%.6f" % cpk)))
    # plt.hist(x=data, bins=bins, density=False, histtype='bar', color=['r'])
    bins = [round(x,5) for x in bins]
    bins=sorted(bins)
    # print('----->bins-->',bins)
    # print(' more draw category data:',len(column_category_data_list),column_category_data_list)
    l=0
    l_len=[]
    for category_data in column_category_data_list:
        # print('category_test_data--->', category_data[5:])
        # print('category_name,category_color--->', category_data[0],str(category_data[1]))
        category_name,category_color = category_data[0],str(category_data[1])
        n=len(category_data[5:])
        l=l+n
        # print('category len:',n)
        l_len.append(n)
        if len(column_category_data_list) == 1:
            plt.hist(category_data[5:], bins=bins, label=category_data[0], color=category_color ,histtype='stepfilled',edgecolor=category_color,linewidth=1.5,align='mid',density=False) #time分布
        else:
            plt.hist(category_data[5:], bins=bins, label=category_data[0], color='white' ,histtype='step',edgecolor=category_color,linewidth=1.5,align='mid',density=False) #time分布
    # print('----->a column len,max in one category:-->',l,max(l_len))

   
    y_ticks = max(l_len) * (y_ticks+0.04)
    range_value = get_limit_range(bins_l, bins_h)
    range_value =round(range_value/5,5)
    # print('plot bar--->',bins_l,range)
    plt.xlim(bins_l-range_value, bins_h+range_value)
    plt.ylim((0, y_ticks))  # 设置y轴scopex

    ax=plt.gca()
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)

    # plt.plot(x1, y1, 'k--', label="", linewidth=1.0, color='lime')  # 画正态分布曲线
    if lsl !='NA' and usl != 'NA' and zoom_type=='limit':
        plt.plot([lsl, lsl, ], [0, y_ticks, ], 'k--', linewidth=3.0, color='red')  # 画lower limit线，
        plt.plot([usl, usl, ], [0, y_ticks, ], 'k--', linewidth=3.0, color='red')  # 画upper limit线，
        # 添加文字
        # plt.text(0.1,20, r'$\mu=100,\ \sigma=15$')
        plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 10, 'color': 'r'})
        plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 10, 'color': 'r'})
        # plt.axis([-0.2, 0.2,0, 100])#x 轴，y 轴

    plt.text(bins_l+range_value/3, y_ticks*0.78, info, size=10, rotation=0.0, alpha=0.85,fontsize=8,ha="left",
             va="center",bbox=dict(boxstyle="round", ec=('royalblue'),linestyle='-.',lw=1, fc=('white'), ))



    #os.system('mkdir fail')
    if len(column_category_data_list) < 30:    
        plt.legend(loc="upper right",framealpha=1,edgecolor='royalblue',borderaxespad=0.3,fontsize=6)#facecolor ='None',

    # plt.legend(bbox_to_anchor=(1.01, 1), loc=2, borderaxespad=0)

    plt.grid(linestyle=':',c='gray')  # 生成网格
    # path="/Users/rex/PycharmProjects/my/fail/"

    plt.savefig(pic_path,dpi=200)
    plt.draw()

    # plt.show()
    plt.close('all')

    plt.ioff()  

def save_image_to_excel(location,pic_path,plot_sheet):
    #save picture to excel
    plot_sheet.insert_image(location,pic_path,{'x_scale': 1.2, 'y_scale': 1.2})
    return True




def get_one_item_new_limit_from_csv(new_limit_path,item_name):
    new_limit_file = '/tmp/CPK_Log/temp/item_limit.csv'
    if not os.path.exists(new_limit_file):
        return None,None

    tmp_lst = []
    # try:
    with open(new_limit_file, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('item') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # except IOError as exc:
    #   print('ERROR:',exc)
    header_list = tmp_lst[0]
    # print('header_list--->',header_list)
    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    # print(df)
    # print(item_name)
    try:
        new_lsl=df.loc[df['item'] == item_name, 'new_lsl'].tolist()[0]
    except IndexError as e:
        # print('lsl-->',e)
        new_lsl = None
    try:
        new_usl=df.loc[df['item'] == item_name, 'new_usl'].tolist()[0]
    except IndexError as e:
         # print('usl-->',e)
         new_usl = None
    
    try:
        reviewer = df.loc[df['item'] == item_name, 'reviewer'].tolist()[0]
    except IndexError as e:
        # print('lsl-->',e)
        reviewer = ''
    try:
        review_date = df.loc[df['item'] == item_name, 'date'].tolist()[0]
    except IndexError as e:
        # print('lsl-->',e)
        review_date = ''

    try:
        comment = df.loc[df['item'] == item_name, 'comment'].tolist()[0]
    except IndexError as e:
        # print('lsl-->',e)
        comment = ''

    return new_lsl,new_usl,reviewer,review_date,comment



def write_hash_to_excel(file_path,sheet_name,cell_format):
    hash_csv = '/tmp/CPK_Log/temp/data_hash.csv'
    if not os.path.exists(hash_csv):
        return None,None

    tmp_lst = []
    with open(hash_csv, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if i>1:
                location ='A'+str(i)
                sheet_name.write_row(location,row,cell_format)
            i = i + 1
    return 'hash save finish! ok'


def get_file_pic_name(file_dir):
  pic_file_l = []
  for root,dirs,files in os.walk(file_dir):
      # print(root)
      # print('fail picture path:',dirs)
      # print(files)
      for file in files:
          if os.path.splitext(file)[1] == '.png':
              # print('suffix name---->',os.path.splitext(file)[1])
              pic_file_l.insert(0,os.path.splitext(file)[0])
  return pic_file_l


def append_hash_to_excel(table_data,excel_file_path,hashfile_path):

    workbook = openpyxl.load_workbook(excel_file_path)
    wb_sheet = workbook.sheetnames
    ssh_sheet = workbook['ssh']

    hash_csv = '/tmp/CPK_Log/temp/data_hash.csv'
    if not os.path.exists(hash_csv):
        return None,None

    tmp_lst = []
    print('hashfile_path:',hash_csv)
    with open(hash_csv, 'r') as f:
        csv_reader = csv.reader(f)
        i = 1
        for row in csv_reader:
            print('row:',row)
            if i>1:
                for col,v in enumerate(row):
                    ssh_sheet.cell(i,col+1).value = v.encode('utf-8')
            i = i + 1
    workbook.save(excel_file_path)

    if 'fail plot' in wb_sheet:
        workbook = openpyxl.load_workbook(excel_file_path)
        # fail_plot_sheet = workbook['fail plot']
        # fail_pic_path = hashfile_path+'fail_plot'

        # file_l=[]
        # file_l = get_file_pic_name(fail_pic_path)

        # j=0
        # for f_name in file_l:
        #     if j==0:
        #         location='A1'
        #     else:
        #         location = 'A'+str(30*j)
        #     one_fail_pic_path = fail_pic_path +'/'+f_name+'.png'
        #     print('===one_fail_pic_path :',one_fail_pic_path)
        #     # fail_plot_sheet.insert_image(location,one_fail_pic_path,{'x_scale': 1.2, 'y_scale': 1.2})
        #     j=j+1
        workbook.save(excel_file_path)

    
    push2git_checkBox = table_data[2]
    if push2git_checkBox == 1:
        update_limit_path = table_data[1]
        gitAddr = table_data[3]
        gitComment = table_data[4]
        print('cd ' + str(update_limit_path))
        os.chdir(str(update_limit_path))
        file_name = os.path.split(excel_file_path)[1]
        print('git add ' + str(file_name))
        os.system('git add ' + str(file_name))
        print('git commit -m "'+ str(gitComment) + '"')
        os.system('git commit -m "'+ str(gitComment) + '"')
        print('git remote add origin "'+ str(gitAddr) + '"')
        os.system('git remote add origin "'+ str(gitAddr) + '"')
        print('git push -u origin master')
        os.system('git push -u origin master')

    with open(filelognamehash, 'w') as file_object:
        file_object.write("Finished,create excel report finish")

    print('hash save finish! done')





def verify_limit(lsl,usl):
    if lsl != None:
        lsl.replace(' ','')
    if usl != None:
        usl.replace(' ','')

    if lsl =='NA' or lsl =='':
        lsl = None
    else:
        try:
            lsl = float(eval(lsl))
        except:
            lsl = None

    if usl =='NA' or usl == '':
        usl = None
    else:    
        try:
            usl = float(eval(usl))
        except:
            usl = None
    # if type(lsl) == float and type(usl) == float:
    # print('after verify lsl===>',lsl)
    # print('after verify usl===>',usl)
    return lsl,usl

def write_invalid_item_to_excel(invalid_item_name_l,sheet_name,row,cell_format,limit_csv_path):
    for i,v in enumerate(invalid_item_name_l):
        sheet_name.write(row+1+i,0,row+1+i,cell_format)
        sheet_name.write(row+1+i,1,v,cell_format)
        new_lsl,new_usl,reviewer,review_date,comment = get_one_item_new_limit_from_csv(limit_csv_path,v)
        if p_val_checked == '1':
            for n in range(2,28,1):
                if new_lsl != None and n == 18:
                    sheet_name.write(row+1+i,n,new_lsl,cell_format)
                elif new_usl !=None and n == 20:
                    sheet_name.write(row+1+i,n,new_usl,cell_format)
                elif reviewer != '' and n == 25:
                    sheet_name.write(row+1+i,n,reviewer,cell_format)
                elif review_date != '' and n == 26:
                    sheet_name.write(row+1+i,n,review_date,cell_format)
                elif comment != '' and n == 27:
                    sheet_name.write(row+1+i,n,comment,cell_format)
                else:
                    sheet_name.write(row+1+i,n,'',cell_format)

        else:
            for n in range(2,27,1):
                if new_lsl != None and n == 17:
                    sheet_name.write(row+1+i,n,new_lsl,cell_format)
                elif new_usl !=None and n == 19:
                    sheet_name.write(row+1+i,n,new_usl,cell_format)
                elif reviewer != '' and n == 24:
                    sheet_name.write(row+1+i,n,reviewer,cell_format)
                elif review_date != '' and n == 25:
                    sheet_name.write(row+1+i,n,review_date,cell_format)
                elif comment != '' and n == 26:
                    sheet_name.write(row+1+i,n,comment,cell_format)
                else:
                    sheet_name.write(row+1+i,n,'',cell_format)



def draw_histogram(column_data,item_name,lsl,usl,mean,max_num,min_num,stdev,
                    x1,y1,cpu,cpl,cpk,pic_path,set_bins,start_time_first,
                    start_time_last,bmc,zoom_type):


    bins,bins_l,bins_h = get_bins(min_num,max_num,lsl,usl,set_bins,zoom_type)
    # print(len(column_data),'---->',min(column_data),max(column_data),bins)
    probability_distribution_extend(column_data,bins,0,item_name,lsl,usl,
                                    mean,max_num,min_num,stdev,x1,y1,cpu,cpl,
                                    cpk,pic_path,bins_l,bins_h,start_time_first,
                                    start_time_last,bmc,zoom_type)

    return True


def draw_more_histogram(column_category_data_list,column_data, item_name, lsl, usl, mean, 
                        max_num, min_num, stdev, x1, y1,cpu, cpl, cpk, pic_path,set_bins,
                        start_time_first,start_time_last,zoom_type):
    """

    """
    # print('9999 column_category_data_list--->',column_category_data_list)
    # print('9999 column_data--->',len(column_data),column_data)

    # range = get_limit_range(lsl, usl)
    # range = round((range /set_bins), 5)
    # bins = np.arange(lsl, usl, range)  # 必须是单调递增的
    bins,bins_l,bins_h = get_bins(min_num,max_num,lsl,usl,set_bins,zoom_type)
    # print('9999 bins len--->',len(bins))


    # print(len(column_data),'---->',min(column_data),max(column_data),bins)
    probability_distribution_extend_by_color(column_category_data_list,column_data,bins, 0, item_name, lsl, usl, 
                                            mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,bins_l,
                                            bins_h,start_time_first,start_time_last,zoom_type)

    return True


def get_project_info(data_df):
    # print('one row data:--->',data_df[1:2].values[0].tolist(),len(data_df[1:2].values[0].tolist()))
    # project_code = data_df[1:2].values[0].tolist()[1] #Product first cell
    # print("=======get_project_info======",data_df)
    project_code = get_project_code(data_df)
    # print('--------------------project_code:'+project_code+'------------------------')

    build_stage = get_build_stage(data_df)
    # print('--------------------build_stage:'+build_stage+'-------------------------')

    station_name = get_station_name(data_df)
    # print('--------------------station_name:'+station_name+'------------------------')
    return project_code,build_stage,station_name

def get_build_stage(data_df):
    special_build_name  = data_df['Special Build Name'].values.tolist() #Special Build Name
    # pattern =re.compile(r'.*\-(.+)\-')
    # result=pattern.match(build_stage)
    # build_stage = result.group(1)
    build_stage =''
    build_stage_l = list(set(special_build_name))
    print('build_stage_l-->',build_stage_l)
    
    temp_l=[]
    for i in build_stage_l:
        pattern = re.compile(r'.*\-(.+)')
        result = pattern.match(i)
        if result:
            temp_l.append(result.group(1))
    n=1 
    for j in list(set(temp_l)):
        if n==1:
            build_stage = j
        else:
            build_stage = build_stage +'&'+j
        n=n+1
    print("======>>>>>>>build_stage",build_stage)
    return build_stage

def get_station_name(data_df):
    # station_name = data_df[1:2].values[0].tolist()[6] #Station ID first cell
    station_name_l = data_df['Station ID'].values.tolist() #Station ID first cell
    station_name_l = list(set(station_name_l))
    station_name =''
    if len(station_name_l) == 0:
        print('data is empty!')
        station_name == 'xxx'
    else:
        for i in station_name_l:
            if i != '':
                pattern =re.compile(r'.*\_([0-9]+)\_(\D+)')
                result=pattern.match(i)
                try:
                    station_name = result.group(2)
                    return station_name
                except Exception as e:
                    print('match station_name error')

    return station_name


def get_project_code(data_df):
    product_l = list(set(data_df['Product'].values.tolist())) #Product first cell
    print("=============>>>>>>product_l",product_l)
    project_code =''
    n=1
    for i in product_l:
        if n == 1:
            project_code = i
        else:
            project_code = project_code+'&'+i
        n=n+1

    return project_code

def get_pst_time():
    date_format=  '%Y-%m-%d' #'%m-%d%Y_%H_%M_%S_%Z'
    date = datetime.datetime.now(tz=pytz.utc)
    date = date.astimezone(timezone('US/Pacific'))
    pstDateTime=date.strftime(date_format)
    return pstDateTime

def open_all_csv_local(event,all_csv_path,data_select,remove_fail):
    all_csv_path = os.path.join(all_csv_path+ '')
    tmp_lst = []
    with open(all_csv_path, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            # if row[0].lower().find('fct') != -1:
            #     # print("FW version---->")
            #     pass

            if row[0].lower().find('upper limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('lower limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('measurement unit') != -1:  # "Measurement Unit ----->" in row:
                pass
            elif row[0].lower().find('site') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # print("index---->", tmp_lst[0])
    param_item_start_index = tmp_lst[0].index('Test Stop Time')+1
    # print('location parametric:',param_item_start_index)
    header_list = tmp_lst[0]
  

    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    try:
        pd.to_datetime(df['Test Start Time'])
    except Exception as e:
        print('check Test Start Time,csv format wrong!')
        return e
    header_df =df[0:2]
    data_df = df[2:]

 

    print('csv data row number before remove SN empty--->', len(data_df.values.tolist()))
    data_df=data_df[~data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(data_df.values.tolist()))
    if remove_fail== 'yes':
        data_df=data_df[data_df['PASS/FAIL'].isin(['PASS'])]
            # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(data_df.values.tolist()))
    if data_select == 'first':
        data_df = data_df.sort_values(axis=0,by=['Test Start Time'],ascending='True')
        data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        data_df = data_df.sort_values(axis=0,by=['Test Start Time'],ascending='True')
        data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(data_df.values.tolist()))

    project_code,build_stage,station_name = '','',''

    start_time_l  = data_df['Test Start Time'].values.tolist() #StartTime
    if len(start_time_l)>0:
        start_time_first = min(start_time_l)
        start_time_last  = max(start_time_l)
    else:
        start_time_first = ''
        start_time_last  = ''
    print('<first time -- last time>',start_time_first,start_time_last)

    df = header_df.append(data_df)



    # print('df after--->', df)
    # print('df.values ---->', df.values)#array([[ ]])
    return header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last,param_item_start_index








def open_all_csv(event,all_csv_path,data_select,remove_fail):
    all_csv_path = os.path.join(all_csv_path+ '')
    tmp_lst = []
    with open(all_csv_path, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            # if row[0].lower().find('fct') != -1:
            #     # print("FW version---->")
            #     pass
            if row[0].lower().find('display name') != -1:
                pass
            elif row[0].lower().find('pdca priority') != -1:
                pass
            elif row[0].lower().find('upper limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('lower limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('measurement unit') != -1:  # "Measurement Unit ----->" in row:
                pass
            elif row[0].lower().find('site') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # print("index---->", tmp_lst[0])
    param_item_start_index = tmp_lst[0].index('Parametric')
    # print('location parametric:',param_item_start_index)
    header_list = tmp_lst[1]
  



    df = pd.DataFrame(tmp_lst[2:], columns=tmp_lst[1])
    try:
        pd.to_datetime(df['StartTime'])
    except Exception as e:
        print('check StartTime,csv format wrong!')
        return e
    header_df =df[0:2]
    # print('header_df before--->', header_df)
    data_df = df[2:]
    # print('data_df before--->', data_df)
 

    print('csv data row number before remove SN empty--->', len(data_df.values.tolist()))
    data_df=data_df[~data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(data_df.values.tolist()))
    if remove_fail== 'yes':
        data_df=data_df[data_df['Test Pass/Fail Status'].isin(['PASS'])]
            # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(data_df.values.tolist()))
    if data_select == 'first':
        data_df = data_df.sort_values(axis=0,by=['StartTime'],ascending='True')
        data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        data_df = data_df.sort_values(axis=0,by=['StartTime'],ascending='True')
        data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(data_df.values.tolist()))
    print("----------------<<<1>>>>>>>>>>>>>>")
    if event == 'keynote-report' or event == 'excel-report':
        project_code,build_stage,station_name = get_project_info(data_df)
    else:
        project_code,build_stage,station_name = '','',''

    start_time_l  = data_df['StartTime'].values.tolist() #StartTime
    if len(start_time_l)>0:
        start_time_first = min(start_time_l)
        start_time_last  = max(start_time_l)
    else:
        start_time_first = ''
        start_time_last  = ''
    print('<first time -- last time>',start_time_first,start_time_last)

    df = header_df.append(data_df)

 
    # if event != 'one_item_plot':
    #     station_id_l = df['Station ID'].values.tolist()
    #     fixture_channel_id = df['Fixture Channel ID'].values.tolist()
    #     # print('station_id_l:',station_id_l)
    #     # print('fixture_channel_id:',fixture_channel_id)
    #     temp_l = []
    #     for i in range(0,len(station_id_l[2:])):
    #         temp_l.append(station_id_l[i+2]+'_'+fixture_channel_id[i+2])
    #     temp_l.insert(0,'')
    #     temp_l.insert(0,'')
    #     fixture_channel_id = temp_l
    #     # print(fixture_channel_id)
    #     df['Fixture Channel ID'] = pd.DataFrame({'Fixture Channel ID':fixture_channel_id})
    #     # print('fixture channel id:',df['Fixture Channel ID'].values.tolist())

    #     print('*******************************')


    # print('df after--->', df)
    # print('df.values ---->', df.values)#array([[ ]])
    return header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last,param_item_start_index



def create_report(event,header_list,df,color_by1,pic_path,select_category1,cpk_lsl,cpk_usl,
                    save_all_cpk_path,set_bins,excel_name,project_code,build_stage,station_name,
                    start_time_first,start_time_last,color_by2,select_category_l2,excel_report_item,
                    fail_plot_to_excel,zoom_type,param_item_start_index):
    clear_files(pic_path)
    book,report_sheet,plot_sheet,ssh_sheet,format_highlight,format_normal,format_titile,new_format_fail,new_format_pass = creat_excel_report_file(save_all_cpk_path,excel_name,cpk_lsl,cpk_usl,event,fail_plot_to_excel)
    # excel_report_data,excel_report_table,max_row,max_col = load_excel_table(save_all_cpk_path+excel_name)
    table_data,table_category_data,no_valid_column_name_l = parse_all_csv_local(header_list,df,color_by1,select_category1,event,color_by2,select_category_l2,param_item_start_index)#
    i,j,n,t=0,0,0,0
    path=save_all_cpk_path
    result='pass'

    for column_data in table_data:
        item_name=column_data[0]
        wr_excel = 'No'
        # if  str(item_name).lower() != 'head id':
        usl = column_data[1]
        lsl = column_data[2]
        column_data = column_data[3:]
        # print('item_name:',item_name)
        # print('---->>usl:',usl)
        # print('---->>lsl:',lsl)
        # print(str(t)+' column test value:',len(column_data),column_data)
        if len(column_data) >0:   
            bc,p_val,a_Q,a_irr,three_CV = get_coefficients(column_data)

            BMC = ''

            # if p_val_checked == '1':
            if bc != '' and bc != 'Nan' and p_val != '' and a_Q != '' and p_val != 'Nan' and a_Q != 'Nan':
                if float(p_val) <= float(a_Q) and float(bc)>0.555:
                    BMC = 'YES'
                elif float(p_val) <= float(a_Q) and float(bc)<0.555:
                    BMC = 'YES'
                elif float(p_val) > float(a_Q) and float(bc)>0.555:
                    BMC = 'YES'
                elif float(p_val) > float(a_Q) and float(bc)<0.555 and (float(bc)-float(p_val))>=-0.1:
                    BMC = 'YES'
                elif float(p_val) > float(a_Q) and float(bc)<0.555 and (float(bc)-float(p_val))<-0.1:
                    BMC = 'NO'
                else:
                    BMC = ''
            else:
                BMC = ''

            row_data = []
            target_value = 9999999999
            mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)

            if cpk:
                if is_number(str(cpk)):
                    if float(cpk)>10:
                        BMC = ''

            result=''

            if stdev == 0 or stdev == 'nan':
                n=n+1
                # print('old limit--->')
                # print(excel_report_item,item_name+': do not draw plot due to stdev == 0!')#stdev ==0
                location = 'A' + str(j+2)

                result='Nan'
                target_value = get_target_value(lsl,usl)
                # print('target_value--->',lsl,usl,target_value)

                if p_val_checked == '1':
                    if event=='excel-report' and excel_report_item == 'all':#'Fail'
                        row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,
                                    lsl,target_value, usl, min_num, mean, max_num, 
                                    stdev, cpl, cpu, cpk, result,'','','','','','','']
                        report_sheet.write_row(location,row_data,format_normal)
                        wr_excel = 'Yes'
                else:
                    if event=='excel-report' and excel_report_item == 'all':#'Fail'
                        row_data = [j + 1, item_name,bc,a_Q,a_irr,three_CV,lsl,
                                    target_value, usl, min_num, mean, max_num, 
                                    stdev, cpl, cpu, cpk, result,'','','','','','','']
                        report_sheet.write_row(location,row_data,format_normal)
                        wr_excel = 'Yes'
            else:

                location = 'A' + str(j+2)
                # print("location,j--->",location,j)

                # print('cpk:',cpk)
                if cpk == None:
                    result='Nan'
                    target_value = get_target_value(lsl,usl)
                    # print('target_value--->',lsl,usl,target_value)
                    if p_val_checked == '1':
                        if event=='excel-report' and excel_report_item == 'all':#'Fail'
                            row_data = [str(j + 1), item_name,bc,p_val,a_Q,a_irr,
                                        three_CV,lsl,target_value, usl, min_num, 
                                        mean, max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                            report_sheet.write_row(location,row_data,format_normal)
                            wr_excel = 'Yes'
                    else:
                        if event=='excel-report' and excel_report_item == 'all':#'Fail'
                            row_data = [str(j + 1), item_name,bc,a_Q,a_irr,three_CV,lsl,
                                        target_value, usl, min_num, mean, max_num, stdev, 
                                        cpl, cpu, cpk, result,'','','','','','','']
                            report_sheet.write_row(location,row_data,format_normal)
                            wr_excel = 'Yes'

                else:
                    if cpk < cpk_lsl or cpk > cpk_usl:
                        # print('----------------------cpk value fail--------------------------')
                        # print('cpk_lsl:',cpk_lsl)
                        # print('cpk_usl:',cpk_usl)
                        result='Fail'
                        target_value = get_target_value(lsl,usl)
                        # print('target_value--->',lsl,usl,target_value)

                        if p_val_checked == '1':
                            row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,
                                        lsl,target_value, usl, min_num, mean, max_num, 
                                        stdev, cpl, cpu, cpk, result,'','','','','','','']
                            report_sheet.write_row(location,row_data,format_normal)
                            report_sheet.write(j+1,17,result,new_format_fail)
                        else:
                            row_data = [j + 1, item_name,bc,a_Q,a_irr,three_CV,
                                        lsl,target_value, usl, min_num, mean, 
                                        max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                            report_sheet.write_row(location,row_data,format_normal)
                            report_sheet.write(j+1,16,result,new_format_fail)



                        wr_excel = 'Yes'
                        if not os.path.exists(path):
                            os.makedirs(path)

                        if fail_plot_to_excel == 'yes' and event =='excel-report':
                            image_name = item_name.replace('/','_')+".png"
                            pic_path = '/tmp/CPK_Log/fail_plot/' + image_name
                            # if fail_plot_to_excel == 'yes' and event =='excel-report':
                            if len(table_category_data) == 0:
                                # print("============pic_path:",pic_path,image_name)
                                draw_histogram(column_data,item_name,lsl, usl, 
                                                mean, max_num, min_num, stdev, x1, y1, 
                                                cpu, cpl, cpk, pic_path,set_bins,
                                                start_time_first,start_time_last,BMC,zoom_type)
                            else:
                                draw_more_histogram(table_category_data[t],column_data,
                                                    item_name,lsl, usl, mean, max_num, min_num, 
                                                    stdev, x1, y1, cpu, cpl, cpk, pic_path,
                                                    set_bins,start_time_first,start_time_last,zoom_type)
                            
                            if i==0:
                                location='A1'
                            else:
                                location = 'A'+str(30*i)
                            if fail_plot_to_excel == 'yes' and event =='excel-report':
                                print('=====save_image_to_excel pic_path:',pic_path)
                                save_image_to_excel(location,pic_path,plot_sheet)
                                i = i + 1
                    else:
                        if excel_report_item == 'all' or event == 'keynote-report':#'Fail'
                            result='Pass'
                            target_value = get_target_value(lsl,usl)
                            # print('target_value--->',lsl,usl,target_value)

                            if p_val_checked == '1':
                                row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,
                                            three_CV,lsl,target_value,usl, min_num, 
                                            mean, max_num, stdev, cpl, cpu, cpk, 
                                            result,'','','','','','','']
                                report_sheet.write_row(location,row_data,format_normal)
                                report_sheet.write(j+1,17,result,new_format_pass)
                                wr_excel = 'Yes'
                            else:
                                row_data = [j + 1, item_name,bc,a_Q,a_irr,three_CV,
                                            lsl,target_value,usl, min_num, mean, 
                                            max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                                report_sheet.write_row(location,row_data,format_normal)
                                report_sheet.write(j+1,16,result,new_format_pass)
                                wr_excel = 'Yes'


                        else:
                            pass


            #--------------------------------get new limit from csv start------------------------
            new_lsl,new_usl,reviewer,review_date,comment = get_one_item_new_limit_from_csv(save_all_cpk_path,item_name)
            new_lsl_backup,new_usl_backup = new_lsl,new_usl
            # print('=========new----',new_lsl_backup,new_usl_backup)
            new_lsl,new_usl = verify_limit(new_lsl,new_usl)

            old_mean, old_max_num,old_target_value,old_min_num, old_stdev,old_cpu, old_cpl, old_cpk,old_lsl,old_usl,old_result = mean, max_num,target_value,min_num, stdev, cpu, cpl, cpk,lsl,usl,result
            if new_lsl != None :
                lsl = new_lsl
            if new_usl != None :
                usl = new_usl

            if new_lsl == None and new_usl != None:
                lsl = ''
            if new_lsl != None and new_usl == None:
                usl = ''

            if new_lsl != None or new_usl != None:
            #--------------------------------get new limit from csv end------------------------
                mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)
                if stdev == 0 or stdev == 'nan':
                    n=n+1
                    result = 'Nan'
                    target_value = get_target_value(lsl,usl)
                    print(item_name+':  stdev == 0!(new limit)')#stdev ==0
                    if event=='excel-report' and excel_report_item == 'all':#'Fail'
                        if p_val_checked == '1':
                            report_sheet.write(j+1,18,new_lsl_backup,format_highlight)
                            report_sheet.write(j+1,19,target_value,format_highlight)
                            report_sheet.write(j+1,20,new_usl_backup,format_highlight)
                            report_sheet.write(j+1,21,cpl,format_highlight)
                            report_sheet.write(j+1,22,cpu,format_highlight)
                            report_sheet.write(j+1,23,cpk,format_highlight)
                            report_sheet.write(j+1,24,result,format_highlight)
                        else:
                            report_sheet.write(j+1,17,new_lsl_backup,format_highlight)
                            report_sheet.write(j+1,18,target_value,format_highlight)
                            report_sheet.write(j+1,19,new_usl_backup,format_highlight)
                            report_sheet.write(j+1,20,cpl,format_highlight)
                            report_sheet.write(j+1,21,cpu,format_highlight)
                            report_sheet.write(j+1,22,cpk,format_highlight)
                            report_sheet.write(j+1,23,result,format_highlight)
                        wr_excel = 'Yes'


                else:

                    location = 'A' + str(j+2)
                    # print("location,j--->",location,j)

                    # print('cpk:',cpk)

                    if cpk == None:
                        result='Nan'
                        target_value = get_target_value(lsl,usl)
                        # print('target_value--->',lsl,usl,target_value)
                        # append_new_limit_data_to_excel_report(excel_report_data,excel_report_table,save_all_cpk_path+excel_name,j+2,lsl,usl,cpl,cpu,cpk)
                        
                        if p_val_checked == '1':
                            row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,
                                        old_lsl,old_target_value, old_usl, old_min_num, 
                                        old_mean, old_max_num, old_stdev, old_cpl, 
                                        old_cpu, old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                            report_sheet.write_row(location,row_data,format_normal)
                            if old_result == 'Fail':
                                report_sheet.write(j+1,17,old_result,new_format_fail)
                            elif old_result == 'Pass':
                                report_sheet.write(j+1,17,old_result,new_format_pass)
                            report_sheet.write(j+1,18,new_lsl_backup,format_highlight)
                            report_sheet.write(j+1,19,target_value,format_highlight)
                            report_sheet.write(j+1,20,new_usl_backup,format_highlight)
                            report_sheet.write(j+1,21,cpl,format_highlight)
                            report_sheet.write(j+1,22,cpu,format_highlight)
                            report_sheet.write(j+1,23,cpk,format_highlight)
                            report_sheet.write(j+1,24,result,new_format_fail)
                        else:
                            row_data = [j + 1, item_name,bc,a_Q,a_irr,three_CV,
                                        old_lsl,old_target_value, old_usl, 
                                        old_min_num, old_mean, old_max_num, 
                                        old_stdev, old_cpl, old_cpu, old_cpk, 
                                        old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                            report_sheet.write_row(location,row_data,format_normal)
                            if old_result == 'Fail':
                                report_sheet.write(j+1,16,old_result,new_format_fail)
                            elif old_result == 'Pass':
                                report_sheet.write(j+1,16,old_result,new_format_pass)
                            report_sheet.write(j+1,17,new_lsl_backup,format_highlight)
                            report_sheet.write(j+1,18,target_value,format_highlight)
                            report_sheet.write(j+1,19,new_usl_backup,format_highlight)
                            report_sheet.write(j+1,20,cpl,format_highlight)
                            report_sheet.write(j+1,21,cpu,format_highlight)
                            report_sheet.write(j+1,22,cpk,format_highlight)
                            report_sheet.write(j+1,23,result,new_format_fail)


                        wr_excel = 'Yes'
                        picFail_path = '/tmp/CPK_Log/fail_plot/'
                        if not os.path.exists(picFail_path):
                            os.makedirs(picFail_path)

                        if fail_plot_to_excel == 'yes' and event =='excel-report':
                            image_name = item_name.replace('/','_')+" new_limit.png"
                            pic_path = picFail_path + image_name
                            if len(table_category_data) == 0:
                                draw_histogram(column_data,item_name,lsl, usl, mean, 
                                                max_num, min_num, stdev, x1, y1, cpu, 
                                                cpl, cpk, pic_path,set_bins,start_time_first,
                                                start_time_last,BMC,zoom_type)
                            else:
                                #[[],[],[]...],only value category lists
                                draw_more_histogram(table_category_data[t],column_data,
                                                    item_name,lsl, usl, mean, max_num, 
                                                    min_num, stdev, x1, y1, cpu, cpl, 
                                                    cpk, pic_path,set_bins,start_time_first,
                                                    start_time_last,zoom_type)
                            if i==0:
                                location='A1'
                            else:
                                location = 'A'+str(29*i)
                            if fail_plot_to_excel == 'yes' and event =='excel-report':

                                save_image_to_excel(location,pic_path,plot_sheet)
                                i = i + 1


                    else:


                        if cpk < cpk_lsl or cpk > cpk_usl:
                            # print('----------------------cpk value result  fail--------------------------')
                            # print('cpk_lsl:',cpk_lsl)
                            # print('cpk_usl:',cpk_usl)
                            result='Fail'
                            target_value = get_target_value(lsl,usl)
                            # print('target_value--->',lsl,usl,target_value)
                            # append_new_limit_data_to_excel_report(excel_report_data,excel_report_table,save_all_cpk_path+excel_name,j+2,lsl,usl,cpl,cpu,cpk)
                            
                            if p_val_checked == '1':
                                row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,
                                            old_lsl,old_target_value, old_usl, old_min_num, 
                                            old_mean, old_max_num, old_stdev, old_cpl, 
                                            old_cpu, old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                                report_sheet.write_row(location,row_data,format_normal)
                                if old_result == 'Fail':
                                    report_sheet.write(j+1,17,old_result,new_format_fail)
                                elif old_result == 'Pass':
                                    report_sheet.write(j+1,17,old_result,new_format_pass)
                                report_sheet.write(j+1,18,new_lsl_backup,format_highlight)
                                report_sheet.write(j+1,19,target_value,format_highlight)
                                report_sheet.write(j+1,20,new_usl_backup,format_highlight)
                                report_sheet.write(j+1,21,cpl,format_highlight)
                                report_sheet.write(j+1,22,cpu,format_highlight)
                                report_sheet.write(j+1,23,cpk,format_highlight)
                                report_sheet.write(j+1,24,result,new_format_fail)
                            else:
                                row_data = [j + 1, item_name,bc,a_Q,a_irr,three_CV,
                                            old_lsl,old_target_value, old_usl, old_min_num, 
                                            old_mean, old_max_num, old_stdev, old_cpl, old_cpu, 
                                            old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                                report_sheet.write_row(location,row_data,format_normal)
                                if old_result == 'Fail':
                                    report_sheet.write(j+1,16,old_result,new_format_fail)
                                elif old_result == 'Pass':
                                    report_sheet.write(j+1,16,old_result,new_format_pass)
                                report_sheet.write(j+1,17,new_lsl_backup,format_highlight)
                                report_sheet.write(j+1,18,target_value,format_highlight)
                                report_sheet.write(j+1,19,new_usl_backup,format_highlight)
                                report_sheet.write(j+1,20,cpl,format_highlight)
                                report_sheet.write(j+1,21,cpu,format_highlight)
                                report_sheet.write(j+1,22,cpk,format_highlight)
                                report_sheet.write(j+1,23,result,new_format_fail)


                            wr_excel = 'Yes'
                            picFail_path = '/tmp/CPK_Log/fail_plot/'
                            if not os.path.exists(picFail_path):
                                os.makedirs(picFail_path)

                            if fail_plot_to_excel == 'yes' and event =='excel-report':
                                image_name = item_name.replace('/','_')+" new_limit.png"
                                pic_path = picFail_path + image_name
                                if len(table_category_data) == 0:
                                    draw_histogram(column_data,item_name,lsl, usl, mean, 
                                                    max_num, min_num, stdev, x1, y1, cpu, cpl, 
                                                    cpk, pic_path,set_bins,start_time_first,
                                                    start_time_last,BMC,zoom_type)
                                else:

                                    #[[],[],[]...],only value category lists
                                    draw_more_histogram(table_category_data[t],column_data,
                                                        item_name,lsl, usl, mean, max_num, 
                                                        min_num, stdev, x1, y1, cpu, cpl, cpk, 
                                                        pic_path,set_bins,start_time_first,start_time_last,zoom_type)
                                if i==0:
                                    location='A1'
                                else:
                                    location = 'A'+str(29*i)
                                if fail_plot_to_excel == 'yes' and event =='excel-report':
                                    save_image_to_excel(location,pic_path,plot_sheet)
                                    i = i + 1
                        else:
                            if excel_report_item == 'all' or event == 'keynote-report':#'Fail'
                                result='Pass'
                                target_value = get_target_value(lsl,usl)
                                # print('target_value--->',lsl,usl,target_value)
                                # append_new_limit_data_to_excel_report(excel_report_data,excel_report_table,save_all_cpk_path+excel_name,j+2,lsl,usl,cpl,cpu,cpk)
                                if p_val_checked == '1':
                                    row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,
                                                three_CV,old_lsl,old_target_value, 
                                                old_usl, old_min_num, old_mean, old_max_num, 
                                                old_stdev, old_cpl, old_cpu, old_cpk, old_result,
                                                lsl,target_value,usl,cpl,cpu,cpk,result]
                                    report_sheet.write_row(location,row_data,format_normal)
                                    if old_result == 'Fail':
                                        report_sheet.write(j+1,17,old_result,new_format_fail)
                                    elif old_result == 'Pass':
                                        report_sheet.write(j+1,17,old_result,new_format_pass)
                                    report_sheet.write(j+1,18,new_lsl_backup,format_highlight)
                                    report_sheet.write(j+1,19,target_value,format_highlight)
                                    report_sheet.write(j+1,20,new_usl_backup,format_highlight)
                                    report_sheet.write(j+1,21,cpl,format_highlight)
                                    report_sheet.write(j+1,22,cpu,format_highlight)
                                    report_sheet.write(j+1,23,cpk,format_highlight)
                                    report_sheet.write(j+1,24,result,new_format_pass)

                                else:
                                    row_data = [j + 1, item_name,bc,a_Q,a_irr,three_CV,
                                                old_lsl,old_target_value, old_usl, old_min_num, 
                                                old_mean, old_max_num, old_stdev, old_cpl, 
                                                old_cpu, old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                                    report_sheet.write_row(location,row_data,format_normal)
                                    if old_result == 'Fail':
                                        report_sheet.write(j+1,16,old_result,new_format_fail)
                                    elif old_result == 'Pass':
                                        report_sheet.write(j+1,16,old_result,new_format_pass)
                                    report_sheet.write(j+1,17,new_lsl_backup,format_highlight)
                                    report_sheet.write(j+1,18,target_value,format_highlight)
                                    report_sheet.write(j+1,19,new_usl_backup,format_highlight)
                                    report_sheet.write(j+1,20,cpl,format_highlight)
                                    report_sheet.write(j+1,21,cpu,format_highlight)
                                    report_sheet.write(j+1,22,cpk,format_highlight)
                                    report_sheet.write(j+1,23,result,new_format_pass)


                                wr_excel = 'Yes'
            # print('reviewer,review_date----->',reviewer,review_date)
            elif new_lsl_backup == 'NA' or new_usl_backup == 'NA':
                result='Nan'
                target_value = 'NA'
                cpl = 'NA'
                cpu = 'NA'
                cpk = 'NA'
                
                if p_val_checked == '1':      
                    report_sheet.write(j+1,18,new_lsl_backup,format_highlight)
                    report_sheet.write(j+1,19,target_value,format_highlight)
                    report_sheet.write(j+1,20,new_usl_backup,format_highlight)
                    report_sheet.write(j+1,21,cpl,format_highlight)
                    report_sheet.write(j+1,22,cpu,format_highlight)
                    report_sheet.write(j+1,23,cpk,format_highlight)
                    report_sheet.write(j+1,24,result,new_format_fail)
                    wr_excel = 'Yes'
                    
                else:
                    report_sheet.write(j+1,17,new_lsl_backup,format_highlight)
                    report_sheet.write(j+1,18,target_value,format_highlight)
                    report_sheet.write(j+1,19,new_usl_backup,format_highlight)
                    report_sheet.write(j+1,20,cpl,format_highlight)
                    report_sheet.write(j+1,21,cpu,format_highlight)
                    report_sheet.write(j+1,22,cpk,format_highlight)
                    report_sheet.write(j+1,23,result,new_format_fail)
                    wr_excel = 'Yes'

            if p_val_checked == '1':
                report_sheet.write(j+1,25,reviewer,format_normal)
                report_sheet.write(j+1,26,review_date,format_normal)
                report_sheet.write(j+1,27,comment,format_normal)
            else:

                report_sheet.write(j+1,24,reviewer,format_normal)
                report_sheet.write(j+1,25,review_date,format_normal)
                report_sheet.write(j+1,26,comment,format_normal)
                
            if event == 'excel-report' and excel_report_item == 'all' or event == 'excel-report' and excel_report_item == 'fail' and old_result =='Fail' or event == 'excel-report' and excel_report_item == 'fail' and result =='Fail' or  stdev != 0 and stdev != 'nan' and old_cpk != None and event == 'keynote-report':
                # print(stdev,event,'add j:')
                j=j+1
        else:
            print('>data empty ,no record in excel!')
        t = t + 1
    # print('Total {} items not calulate cpk'.format('%.0f' % n))
    if event == 'excel-report' and excel_report_item == 'all':
        write_invalid_item_to_excel(no_valid_column_name_l,report_sheet,j,format_normal,save_all_cpk_path)

    book.close()
    print('All items excel report finished!')


def local_csv_generate_excel(table_data):
    print('---local csv generate excel')
    # exportAllItems,exportPassItems,cpkLow,cpkHigh,populate,userName,projectName,targetBuild,cpk_path,set_bin,csv_data_Path
    global filelogname
    global filelognamehash
    cpk_lsl = table_data[2]
    # cpk_usl = table_data[3]
    cpk_usl = float("inf")
    excel_report_user = table_data[5]
    project_name = table_data[6]
    excel_report_stage = table_data[7]
    
    cpk_path = table_data[8]#"/Users/RyanGao/Desktop/CPK_Log/"
    set_bins = table_data[9] #250
    filelogname = '/tmp/CPK_Log/temp/.excel.txt'
    filelognamehash = '/tmp/CPK_Log/temp/.excel_hash.txt'

    all_csv_path = table_data[10] #'/tmp/CPK_Log/Temp/keynote_data_temp.csv'
    data_select = 'all'
    remove_fail = 'yes'
    event = 'excel-report'

    start_time = get_pst_time()

    excel_report_file_name ='Limit_Update_'+str(excel_report_user)+'_'+str(project_name)+'_For_'+str(excel_report_stage)+'_'+str(start_time)+'.xlsx'#%Y-%m-%d_%H-%M-%S
    with open('/tmp/CPK_Log/temp/.excelreportname.txt', 'w') as file_object:
        file_object.write(excel_report_file_name)

    zoom_type = 'limit'
    color_by1 = 'Off'
    fail_pic_path =cpk_path+'fail_plot/'
    select_category_l1 =[]

    color_by2 = 'Off'
    select_category_l2 =[]


    
    exportAllItems = table_data[0]
    exportPassItems = table_data[1]

    excel_report_item = 'all'  #fail   all
    if exportAllItems == 1:
        excel_report_item = 'all'  
    if exportPassItems == 1:
        excel_report_item = 'fail' 


    populate = table_data[4]
    fail_plot_to_excel = 'no'#'no'   # yes
    if populate==1:
        fail_plot_to_excel = 'yes'
    

    header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last,param_item_start_index =  open_all_csv_local(event,all_csv_path,data_select,remove_fail)
    create_report(event,header_list,df,color_by1,fail_pic_path,select_category_l1,cpk_lsl,
                    cpk_usl,cpk_path,set_bins,excel_report_file_name,project_code,build_stage,station_name,
                    start_time_first,start_time_last,color_by2,select_category_l2,excel_report_item,
                    fail_plot_to_excel,zoom_type,param_item_start_index)
    print('['+str(start_time)+' '+str(datetime.datetime.now())+']','create excel report finished!')

    with open(filelogname, 'w') as file_object:
        file_object.write("Finished," + excel_report_name)


def generate_excel(table_data):
    # exportAllItems,exportPassItems,cpkLow,cpkHigh,populate,userName,projectName,targetBuild,cpk_path,set_bin,csv_data_Path
    global filelogname
    global filelognamehash
    global p_val_checked

    cpk_lsl = table_data[2]
    # cpk_usl = table_data[3]
    cpk_usl = float("inf")
    excel_report_user = table_data[5]
    project_name = table_data[6]
    excel_report_stage = table_data[7]
    
    cpk_path = table_data[8] #"[NSString stringWithFormat:@"%@/CPK_Log/",desktopPath]"
    set_bins = table_data[9] #250

    filelogname = '/tmp/CPK_Log/temp/.excel.txt'
    filelognamehash = '/tmp/CPK_Log/temp/.excel_hash.txt'

    all_csv_path = table_data[10] #'[m_configDictionary valueForKey:Load_Csv_Path]  raw data for load path'
    p_val_checked = str(table_data[14]) #choose p_val

    data_select = 'all'
    remove_fail = 'yes'
    event = 'excel-report'

    start_time = get_pst_time() #datetime.datetime.now() 
    excel_report_file_name ='Limit_Update_'+str(excel_report_user)+'_'+str(project_name)+'_For_'+str(excel_report_stage)+'_'+str(start_time)+'.xlsx'#%Y-%m-%d_%H-%M-%S
    with open('/tmp/CPK_Log/temp/.excelreportname.txt', 'w') as file_object:
        file_object.write(excel_report_file_name)
    zoom_type = 'limit'
    color_by1 = 'Off'
    fail_pic_path ='/tmp/CPK_Log/fail_plot/'
    select_category_l1 =[]

    color_by2 = 'Off'
    select_category_l2 =[]


    
    exportAllItems = table_data[0]
    exportPassItems = table_data[1]

    excel_report_item = 'all'  #fail   all
    if exportAllItems == 1:
        excel_report_item = 'all'  
    if exportPassItems == 1:
        excel_report_item = 'fail' 


    populate = table_data[4]
    fail_plot_to_excel = 'no'#'no'   # yes
    if populate==1:
        fail_plot_to_excel = 'yes'
    

    header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last,param_item_start_index =  open_all_csv(event,all_csv_path,data_select,remove_fail)
    create_report(event,header_list,df,color_by1,fail_pic_path,select_category_l1,cpk_lsl,cpk_usl,
                    cpk_path,set_bins,excel_report_file_name,project_code,build_stage,station_name,
                    start_time_first,start_time_last,color_by2,select_category_l2,excel_report_item,
                    fail_plot_to_excel,zoom_type,param_item_start_index)
    
    print('==[='+str(start_time)+' '+str(datetime.datetime.now())+']','create excel report finished!')

    with open(filelogname, 'w') as file_object:
        file_object.write("Finished," + excel_report_name)

def excel_hash_to_csv(excel_path,csv_hash_path):
    data = pd.read_excel(excel_path,'ssh',index_col=0,keep_default_na=False)
    data.to_csv(csv_hash_path,encoding='utf-8')

def excel_limitupdate_to_csv(excel_path,csv_limit_path):
    data = pd.read_excel(excel_path,'report',index_col=0,keep_default_na=False)
    data.to_csv(csv_limit_path,encoding='utf-8')


def run(n):

    while True:
        try:
            print("wait for excel client ...")
            zmqMsg = socket.recv()
            socket.send(b'excel.csv')  # socket.send(ret.decode('utf-8').encode('ascii'))
            if len(zmqMsg)>0:
                key = zmqMsg.decode('utf-8')
                print("message from excel client:", key)

                if key == 'generate_excel_sheet1_hash':
                    table_data = get_redis_data(key)
                    if len(table_data)>0:
                        append_hash_to_excel(table_data,excel_report_name,hash_csv_path)
                    else:
                        print("---get generate_excel_sheet1_hash error")
                        with open(filelognamehash, 'w') as file_object:
                            file_object.write("Finished,create excel report error")

                elif key == 'generate_excel':
                    table_data = get_redis_data(key)
                    if len(table_data)>0:
                        generate_excel(table_data)
                    else:
                        print("---get generate_excel error")
                        with open(filelogname, 'w') as file_object:
                            file_object.write("Finished," + str(excel_report_name))



                elif key == 'local_csv_generate_excel':
                    table_data = get_redis_data(key)
                    if len(table_data)>0:
                        local_csv_generate_excel(table_data)
                    else:
                        print("---get local_csv_generate_excel error")
                        with open(filelogname, 'w') as file_object:
                            file_object.write("Finished," + str(excel_report_name))

                elif key == 'excel_hash_to_csv':
                    table_data = get_redis_data(key)
                    if len(table_data)>0:
                        excel_path = table_data[0]
                        csv_hash_path = table_data[1]
                        excel_hash_to_csv(excel_path,csv_hash_path)
                    else:
                        print("---get excel_hash_to_csv error")
                        with open(filelogname, 'w') as file_object:
                            file_object.write("Finished," + str(excel_report_name))

                elif key == 'excel_limit_update_to_csv':
                    table_data = get_redis_data(key)
                    if len(table_data)>0:
                        excel_path = table_data[0]
                        csv_limit_path = table_data[1]
                        excel_limitupdate_to_csv(excel_path,csv_limit_path)
                    else:
                        print("---get excel_limit_update_to_csv error")
                        
                
            else:
                time.sleep(0.05)

        except Exception as e:
            print('error excel:',e)
            with open(filelogname, 'w') as file_object:
                file_object.write("Finished,create excel report error: " + str(e))
            with open(filelognamehash, 'w') as file_object:
                file_object.write("Finished,create excel report error: " + str(e))

if __name__ == '__main__':
    # t1 = threading.Thread(target=run, args=("<<correlation>>",))
    # t1.start()

    run(0)
    # l=['All','yes',1.67,float("inf"),'','tom','j171a','pvt','/Users/rex/Desktop/CPK_Log/',250,'/Users/rex/Desktop/CPK_Log/loadFile/data/222---.csv']
    # generate_excel(l)
    
