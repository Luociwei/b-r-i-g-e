#! /usr/bin/env python3
# --*-- coding: utf-8 ---*---


import sys,os,time,math,re
import datetime
start_time = datetime.datetime.now()  #
print('-------------- script update on 20200705 ---------------')
show_log_flag = True

BASE_DIR=os.path.dirname(os.path.abspath(__file__))
print('BASE_DIR--->',BASE_DIR)
sys.path.insert(0,BASE_DIR+'/site-packages/')

print('python import ----> csv')
try:
    import csv
except Exception as e:
    print('e---->',e)

print('python import ---->matplotlib')
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
except Exception as e:
    print('e---->',e)

print('python import ----> matplotlib.colors')
try:
    import matplotlib.colors as colors
except Exception as e:
    print('e---->',e)

print('python import ----> FontProperties')
try:
    from matplotlib.font_manager import FontProperties
except Exception as e:
    print('e---->',e)


print('python import ----> numpy')
try:
    import numpy as np
except Exception as e:
    print('e--->',e)

print('python import ----> pandas')
try:
    import pandas as pd
except Exception as e:
    print('e--->',e)

print('python import ----> openpyxl')
try:
    import openpyxl
except Exception as e:
  print('import openpyxl error:',e)
print('python import ----> xlsxwriter')
try:
    import xlsxwriter
except Exception as e:
  print('import xlsxwriter error:',e)


# print('python import ----> scipy')
# try:
#     from scipy import stats
# except Exception as e:
#     print('import scipy error:',e)


print('python import ----> diptest')
try:
    import diptest
except Exception as e:
    print('import diptest error:',e)

print(sys.getdefaultencoding())



current_dir = os.path.dirname(os.path.realpath(__file__))
keynote_lib_path = current_dir+'/python_keynote'
print('cpk file--->',keynote_lib_path)
# sys.path.append(keynote_lib_path)
# sys.path.insert(0,keynote_lib_path)
cmd = 'cd '+current_dir+'/python_keynote;export PYTHONPATH='+keynote_lib_path
# print('cpk file execute cmd--->',cmd)
print(os.system(cmd))
print('python import ----> python_keynote')
try:
    from python_keynote import generate_keynote
except Exception as e:
  print('e---->',e)



# vals =np.random.normal(5,3,30)#分布的均值,分布的标准差,个数
# print(vals)
over_time = datetime.datetime.now()   #
total_time = (over_time-start_time).total_seconds()
print('import expend total time is: %s s' % total_time)


def count_time(func):
    def int_time(*args, **kwargs):
        start_time = datetime.datetime.now()  #
        func(*args, **kwargs)
        over_time = datetime.datetime.now()   #
        total_time = (over_time-start_time).total_seconds()
        print('Function < '+func.__name__+' > expend total time is: %s s' % total_time)
    return int_time


def list_to_df(column_item_name,column_value_list):
    a = []
    for i in column_value_list:
        b = []
        b.append(i)
        a.append(b)
    #a-->[['a'], ['b'], ['x']]
    df = pd.DataFrame(columns=[column_item_name], data=a)
    print('df--->',df)
    return df
    

def show_log(message):
    if show_log_flag == True:
        print(message)
    return 
    


def cpk(vals,lower,upper=None):
    mean = np.mean(vals)
    std = np.std(vals)
    if upper:
        cpl = (mean -lower)/(3*std)
        cpu = (upper -mean)/(3*std)
        cpk = min(cpl,cpu)
        # print('Mean: {},Standard Deviation:{}'.format(mean,std))
        # print('Cpl:{},Cpu: {}'.format(cpl,cpu))
        # print('Cpk:{}'.format(cpk))
    else:
        cpk = abs(mean-lower)/(3*std)
        # print('Mean:{},Standard Deviation:{}'.format(mean,std))
        # print('Cpk:{}'.format(cpk))
# cpk(vals,3)




def cpk_calc(df_data,lsl,usl):
    """
    :param df_data: list
    :param usl: 数据指标上限
    :param lsl: 数据指标下限
    :return:
    """

    sigma = 3
    # print('limit--->',lsl,usl)
    # 数据平均值
    mean = np.mean(df_data)#
    # print('mean ---->',mean)

    # 数据max值
    max_num = max(df_data)
    # print('max_num ---->',max_num)

    # 数据min值
    min_num = min(df_data)
    # print('min_num ---->',min_num)

    # a = np.array([[1, 2], [3, 4]])
    # print('a--->', type(a))
    # print('gobal std:',np.std(a))#全局标准差
    # print('each line std:',np.std(a, axis=0,ddof=1))
    # print("each row std:",np.std(a, axis=1,ddof=1))

    # 数据标准差
    stdev = np.std(df_data,ddof=1,axis=0)
    # print('stdev ---->',stdev)
    if stdev == 0:#stop count cpk
        return (mean,max_num,min_num,stdev,None,None,None,None,None)
    # 生成横轴数据平均分布
    # x1 = np.linspace(mean - sigma * stdev, mean + sigma * stdev, 1000)
    # print('x1 ---->',x1)
    # 计算正态分布曲线
    # y1 = np.exp(-(x1 - mean) ** 2 / (2 * stdev ** 2)) / (math.sqrt(2 * math.pi) * stdev)
    # print('y1 ---->',y1)
    x1,y1 = None,None
    if lsl == 'NA' or usl == 'NA':
        return (mean,max_num,min_num,stdev,None,None,None,None,None)
    cpu = (usl - mean) / (sigma * stdev)
    cpl = (mean - lsl) / (sigma * stdev)
    # print('cpu ---->',cpu)
    # print('cpl ---->',cpl)

    # 得出cpk
    cpk = min(cpu, cpl)
    return (mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk)

def correlation_coefficient_calc(data_l1,data_l2,item_name1,item_name2):

    '''
    item_name1 : string
    item_name2 : string
    data_l2 :[]
    data_l1 :[]
    '''
    if item_name1 == item_name2:# can not be the same name.
       item_name2 = item_name2+'_2'
    data = pd.DataFrame({item_name1:data_l1,
                       item_name2:data_l2})
    # print('correlation data:',data)
    pearson = data.corr(method='pearson').values[0].tolist()[1]
    spearman = data.corr(method='spearman').values[0].tolist()[1]
    pearson = round(pearson,5)
    spearman = round(spearman,5)
    # print('pearson:',pearson)
    # print('spearman:',spearman)
    return pearson,spearman




def get_coefficients(value_l):
    '''
    param value_l: need float list
    return: bc,p_val,a_Q,a_irr,three_σ_x100_divide_mean
    1σ＝690000／1000000 #fault rate
    2σ＝308000／1000000
    3σ＝66800／1000000
    4σ＝6210／1000000
    5σ＝230／1000000
    6σ＝3.4／1000000 
    7σ＝0／1000000
    '''
    
    column_stdev = np.std(value_l,ddof=1,axis=0)
    three_sigma= 3*column_stdev
    # print('three_sigma:',column_stdev,three_sigma)
    data = np.array(value_l)
    if len(data) <= 3:
        return '','','','',''
    try:
        dip, p_val = diptest.diptest(data)
    except Exception as e:
        print('calculate dip,p_val error:',e)
        return '','','','',''
    # print('dip,p_val:',dip,p_val)
    item_name ='value1'
    data = pd.DataFrame({item_name:value_l})
    # print('data--->',type(data),data)
    u1 = data[item_name].mean() # 计算均值
    # std1 = data[item_name].std() # 计算标准差
    # t,pval=stats.kstest(data[item_name], 'norm', (u1, std1))
    # print('normality test t,pval:',str(t),str(pval))

    # print('------')
    # 正态性检验 → pvalue >0.05

    n= float(len(value_l))
    #Item (xi-ẍ)^2
    item_l_1 =[]
    item_l_2 = []
    item_l_3 = []
    for i in value_l:
        temp1 = (i-u1)**2
        temp2 = (i-u1)**3
        temp3 = (i-u1)**4
        item_l_1.append(temp1)
        item_l_2.append(temp2)
        item_l_3.append(temp3)
    # print('item_l_1--->',item_l_1)
    # print('item_l_2--->',item_l_2)
    # print('item_l_3--->',item_l_3)
    sum_item_l_1 = sum(item_l_1)
    sum_item_l_2 = sum(item_l_2)
    sum_item_l_3 = sum(item_l_3)
    # print('sum_item_l_1',sum_item_l_1)
    # print('sum_item_l_2',sum_item_l_2)
    # print('sum_item_l_3',sum_item_l_3)
    if n<=3 or sum_item_l_1==0 or sum_item_l_2==0 or sum_item_l_3==0:
        # print('len < 3--->')
        if abs(u1) == 0:
            return 'Nan',round(p_val,6),'Nan','Nan','Nan'
        else:
            three_CV = three_sigma*100/abs(u1)
            # print('three_CV:',three_CV)
            return 'Nan',round(p_val,6),'Nan','Nan',round(three_CV,6)
    else:
        try:
            m3 = np.sqrt(n*(n-1))/(n-2)*((1/n*sum_item_l_2)/np.sqrt(1/n*sum_item_l_1)**3)

            # print('m3=',m3)
            # print('d8:',n+1)
            # print('d15:',1/n*sum_item_l_3/(1/n*sum_item_l_1)**2)
            # print('d16:',(n+1)*1/n*sum_item_l_3/(1/n*sum_item_l_1)**2-3*(n-1))
            # print('d6/d7',(n-1)/((n-2)*(n-3)))
            m4 = ((n-1)/((n-2)*(n-3)))*((n+1)*1/n*sum_item_l_3/(1/n*sum_item_l_1)**2-3*(n-1))#(d6/d7)*d16
            # print('m4=',m4)
            #(d14**2+1)/(d17+3*(d10/d7))
            bc =(m3**2+1)/(m4+3*((n-1)**2/((n-2)*(n-3))))
            # print('bc:',bc)
            a_L=0.05
            a_M=0.1
            a_U=0.32
            a_Q = (a_U-a_L)*bc**2+a_L
            # print ('a_Q:',a_Q)
            a_irr = np.sqrt((a_U-a_L)**2*bc)+a_L
            # print('a_irr:',a_irr)
        except Exception as e:
            # print('calculate error',e)
            if abs(u1) == 0:
                return 'Nan',round(p_val,6),'Nan','Nan','Nan'
            else:
                three_CV = three_sigma*100/abs(u1)
                # print('three_CV:',three_CV)
                return 'Nan',round(p_val,6),'Nan','Nan',round(three_CV,6)


    if abs(u1) == 0:
        return round(bc,6),round(p_val,6),round(a_Q,6),round(a_irr,6),'Nan'
    else:
        three_CV = three_sigma*100/abs(u1)
        # print('three_CV:',three_CV)
        return round(bc,6),round(p_val,6),round(a_Q,6),round(a_irr,6),round(three_CV,6)





def write_row_data_to_excel(location,row_data,format,sheet_name):
    sheet_name.write_row(location, row_data, format)
    # report_sheet.write_row('A1', title, format_titile)

def write_column_data_to_excel(location, column_data, format,sheet_name):
    sheet_name.write_column(location, column_data, format)




# 按照固定区间长度绘制频率分布直方图
# bins_interval 区间的长度
# margin        设定的左边和右边空留的大小

def probability_distribution(data, bins_interval=1, margin=1):
    bins = range(int(min(data)), int(max(data)) + bins_interval - 1, bins_interval)
    print(len(bins))
    for i in range(0, len(bins)):
        print(bins[i])
    plt.xlim(min(data) - margin, max(data) + margin)
    plt.title("Probability-distribution")
    plt.xlabel('Interval')
    plt.ylabel('Probability')
    # 频率分布normed=True，频次分布normed=False
    prob,left,rectangle = plt.hist(x=data, bins=bins, density=False, histtype='bar', color=['r'])
    for x, y in zip(left, prob):
        # 字体上边文字
        # 频率分布数据 normed=True
        # plt.text(x + bins_interval / 2, y + 0.003, '%.2f' % y, ha='center', va='top')
        # 频次分布数据 normed=False
        plt.text(x + bins_interval / 2, y + 0.25, '%.2f' % y, ha='center', va='top')
    plt.show()







def probability_distribution_extend(data,bins,margin,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,bins_l,bins_h,start_time_first,start_time_last):
  bins = sorted(bins)
  length = len(bins)
  intervals = np.zeros(length+1)
  for value in data:
    i = 0
    while i < length and value >= bins[i]:
      i += 1
    intervals[i] += 1
  intervals = intervals / float(len(data))
  plt.ion()  # 开启interactive mode

  plt.figure(1)  # 创建图表1
  plt.xlim(min(bins) - margin, max(bins) + margin)
  bins.insert(0, -999)
  plt.title("probability-distribution")
  plt.bar(bins, intervals,color=['r'], label='')#频率分布
  x_ticks,labels = plt.xticks()
  x_ticks_start=round(x_ticks[0],2)
  x_ticks_end = round(x_ticks[len(x_ticks) - 1],2)
  y_ticks, labels = plt.yticks()
  # print('y_ticks--->',y_ticks,y_ticks[len(y_ticks) - 1],y_ticks[0])
  y_ticks=round(y_ticks[len(y_ticks) - 1],5)
  # print("x_ticks_start,x_ticks_end--->",x_ticks_start,x_ticks_end)
  # print('y_ticks_end--->',y_ticks)
  # plt.show()
  plt.close(1)


  plt.figure(2,dpi=150)  # 创建图表2

  if cpl ==None:
    cpl_value =''
  else:
    cpl_value = str("%.3f" % cpl)

  if cpu ==None:
    cpu_value =''
  else:
    cpu_value = str("%.3f" % cpu)

  if cpk ==None:
    cpk_value =''
  else:
    cpk_value = str("%.3f" % cpk)

  info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
      "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + cpl_value + '\n' + "Cpu: " + cpu_value + '\n' + "Cpk: " + cpk_value
  # info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
  #     "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + str(
  #     "%.3f" % cpl) + '\n' + "Cpu: " + str("%.3f" % cpu) + '\n' + "Cpk: " + str("%.3f" % cpk)
  if len(item_name) > 55:
      item_name = item_name[0:55] + '\n' + item_name[55:]
  # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
  # plt.title(item_name,FontProperties=font)
  plt.title(item_name,size=10)
  plt.xlabel(str(start_time_first)+' -- '+str(start_time_last))
  plt.ylabel('Count')
  # plt.title(item_name+"\nCpk={0}".format(str("%.6f" % cpk)))
  # plt.hist(x=data, bins=bins, density=False, histtype='bar', color=['r'])
  bins = [round(x,5) for x in bins]
  bins=sorted(bins)
  # print('----->bins-->',bins)
  plt.hist(data, bins=bins, label=info, histtype='stepfilled',color = 'blue', edgecolor='blue', linewidth=1.5,align='mid',density=False) #time分布
  
  range = get_limit_range(bins_l, bins_h)
  range =round(range/5,5)
  # print('plot bar--->',bins_l,range)
  plt.xlim(bins_l-range, bins_h+range)

  y_ticks = len(data) * y_ticks
  plt.ylim((0, y_ticks))  # 设置y轴scopex
  ax=plt.gca()
  ax.spines['bottom'].set_linewidth(1.5)
  ax.spines['left'].set_linewidth(1.5)
  ax.spines['right'].set_linewidth(1.5)
  ax.spines['top'].set_linewidth(1.5)
  # plt.plot(x1, y1, 'k--', label="", linewidth=1.0, color='lime')  # 画正态分布曲线
  if lsl !='NA' and usl != 'NA':
      plt.plot([lsl, lsl, ], [0, 100000, ], 'k--', linewidth=3.0, color='red')  # 画lower limit线，
      plt.plot([usl, usl, ], [0, 100000, ], 'k--', linewidth=3.0, color='red')  # 画upper limit线，

      # 添加文字
      # plt.text(0.1,20, r'$\mu=100,\ \sigma=15$')
      plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 10, 'color': 'r'})
      plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 10, 'color': 'r'})
      # plt.axis([-0.2, 0.2,0, 100])#x 轴，y 轴

  #os.system('mkdir fail')
  plt.legend(loc="upper right",framealpha=1,edgecolor='royalblue',borderaxespad=0.3)
  plt.grid(linestyle=':',c='gray')  # 生成网格
  # path="/Users/rex/PycharmProjects/my/fail/"

  plt.savefig(pic_path,dpi=200)
  plt.draw()

  # plt.show()
  plt.close('all')

  plt.ioff()  






def probability_distribution_extend_by_color(column_category_data_list,data,bins,margin,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,bins_l,bins_h,start_time_first,start_time_last):
  # print('one column data len:',len(data),item_name,data,column_category_data_list)
  bins = sorted(bins)
  length = len(bins)
  intervals = np.zeros(length+1)
  for value in data:
    i = 0
    while i < length and value >= bins[i]:
      i += 1
    intervals[i] += 1
  intervals = intervals / float(len(data))
  plt.ion()  # 开启interactive mode

  plt.figure(1)  # 创建图表1
  plt.xlim(min(bins) - margin, max(bins) + margin)
  bins.insert(0, -999)
  plt.title("probability-distribution")
  plt.bar(bins, intervals,color=['r'], label='')#频率分布
  x_ticks,labels = plt.xticks()
  x_ticks_start=round(x_ticks[0],2)
  x_ticks_end = round(x_ticks[len(x_ticks) - 1],2)
  y_ticks, labels = plt.yticks()
  # print('y_ticks--->',y_ticks,y_ticks[len(y_ticks) - 1],y_ticks[0])
  y_ticks=round(y_ticks[len(y_ticks) - 1],5)
  # print("x_ticks_start,x_ticks_end--->",x_ticks_start,x_ticks_end)
  # print('y_ticks_end--->',y_ticks)
  # plt.show()
  plt.close(1)


  plt.figure(2,dpi=150)  # 创建图表2,facecolor='blue',edgecolor='black'

  if cpl ==None:
    cpl_value =''
  else:
    cpl_value = str("%.3f" % cpl)

  if cpu ==None:
    cpu_value =''
  else:
    cpu_value = str("%.3f" % cpu)

  if cpk ==None:
    cpk_value =''
  else:
    cpk_value = str("%.3f" % cpk)

  info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
      "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + cpl_value + '\n' + "Cpu: " + cpu_value + '\n' + "Cpk: " + cpk_value
  


  # info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
  #     "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + str(
  #     "%.3f" % cpl) + '\n' + "Cpu: " + str("%.3f" % cpu) + '\n' + "Cpk: " + str("%.3f" % cpk)
  if len(item_name) > 55:
      item_name = item_name[0:55] + '\n' + item_name[55:]
  # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
  plt.title(item_name,size=10)
  plt.xlabel(str(start_time_first)+' -- '+str(start_time_last))
  plt.ylabel('Count')

  # plt.title(item_name+"\nCpk={0}".format(str("%.6f" % cpk)))
  # plt.hist(x=data, bins=bins, density=False, histtype='bar', color=['r'])
  bins = [round(x,5) for x in bins]
  bins=sorted(bins)
  # print('----->bins-->',bins)
  # print(' more draw category data:',len(column_category_data_list),column_category_data_list)
  l=0
  l_len=[]
  for category_data in column_category_data_list:
      # print('category_test_data--->', category_data[5:])
      # print('category_name,category_color--->', category_data[0],str(category_data[1]))
      category_name,category_color = category_data[0],str(category_data[1])
      n=len(category_data[5:])
      l=l+n
      # print('category len:',n)
      l_len.append(n)
      if len(column_category_data_list) == 1:
          plt.hist(category_data[5:], bins=bins, label=category_data[0], color=category_color ,histtype='stepfilled',edgecolor=category_color,linewidth=1.5,align='mid',density=False) #time分布
      else:
          plt.hist(category_data[5:], bins=bins, label=category_data[0], color='white' ,histtype='step',edgecolor=category_color,linewidth=1.5,align='mid',density=False) #time分布
  # print('----->a column len,max in one category:-->',l,max(l_len))

 
  y_ticks = max(l_len) * (y_ticks+0.04)
  range = get_limit_range(bins_l, bins_h)
  range =round(range/5,5)
  # print('plot bar--->',bins_l,range)
  plt.xlim(bins_l-range, bins_h+range)
  plt.ylim((0, y_ticks))  # 设置y轴scopex

  ax=plt.gca()
  ax.spines['bottom'].set_linewidth(1.5)
  ax.spines['left'].set_linewidth(1.5)
  ax.spines['right'].set_linewidth(1.5)
  ax.spines['top'].set_linewidth(1.5)

  # plt.plot(x1, y1, 'k--', label="", linewidth=1.0, color='lime')  # 画正态分布曲线
  if lsl !='NA' and usl != 'NA':
      plt.plot([lsl, lsl, ], [0, 1000000, ], 'k--', linewidth=3.0, color='red')  # 画lower limit线，
      plt.plot([usl, usl, ], [0, 1000000, ], 'k--', linewidth=3.0, color='red')  # 画upper limit线，
      # 添加文字
      # plt.text(0.1,20, r'$\mu=100,\ \sigma=15$')
      plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 10, 'color': 'r'})
      plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 10, 'color': 'r'})
      # plt.axis([-0.2, 0.2,0, 100])#x 轴，y 轴

  plt.text(bins_l+range/3, y_ticks*0.78, info, size=10, rotation=0.0, alpha=0.85,fontsize=8,ha="left",
           va="center",bbox=dict(boxstyle="round", ec=('royalblue'),linestyle='-.',lw=1, fc=('white'), ))



  #os.system('mkdir fail')
  if len(column_category_data_list) < 30:    
      plt.legend(loc="upper right",framealpha=1,edgecolor='royalblue',borderaxespad=0.3,fontsize=6)#facecolor ='None',

  # plt.legend(bbox_to_anchor=(1.01, 1), loc=2, borderaxespad=0)

  plt.grid(linestyle=':',c='gray')  # 生成网格
  # path="/Users/rex/PycharmProjects/my/fail/"

  plt.savefig(pic_path,dpi=200)
  plt.draw()

  # plt.show()
  plt.close('all')

  plt.ioff()  



def get_bins(min_num,max_num,lsl,usl,set_bins):
    # print('min_num,max_num,lsl,usl,set_bins=====>',min_num,max_num,lsl,usl,set_bins)

    bins_l = 0
    bins_h = 0
    if lsl == 'NA' or usl == 'NA':
        bins_l,bins_h = min_num,max_num

    else:
        if min_num < lsl and max_num < lsl:
            bins_l = min_num
            bins_h = usl
        elif min_num < lsl and max_num > lsl and max_num <= usl:
            bins_l = min_num
            bins_h = usl
        elif min_num < lsl and max_num > usl:
            bins_l = min_num
            bins_h = max_num
        elif lsl <= min_num and min_num <= usl and max_num <= usl:
            bins_l = lsl
            bins_h = usl
        elif lsl <= min_num and min_num <= usl and max_num > usl:
            bins_l = lsl
            bins_h = max_num
        elif min_num > usl:
            bins_l = lsl
            bins_h = max_num
    range = get_limit_range(bins_l,bins_h)
    range = round((range/set_bins),5)
    bins = np.arange(bins_l, bins_h, range)#必须是单调递增的
    # print('lsl,usl,min_num,max_num,bins_l,bins_h in get_bins=====>',lsl,usl,min_num,max_num,bins_l,bins_h)
    return bins,bins_l,bins_h




def draw_histogram(column_data,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,set_bins,start_time_first,start_time_last):


    bins,bins_l,bins_h = get_bins(min_num,max_num,lsl,usl,set_bins)
    # print(len(column_data),'---->',min(column_data),max(column_data),bins)
    probability_distribution_extend(column_data,bins,0,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,bins_l,bins_h,start_time_first,start_time_last)

    return True





def draw_more_histogram(column_category_data_list,column_data, item_name, lsl, usl, mean, max_num, min_num, stdev, x1, y1,cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last):
    """

    """

    # range = get_limit_range(lsl, usl)
    # range = round((range /set_bins), 5)
    # bins = np.arange(lsl, usl, range)  # 必须是单调递增的
    bins,bins_l,bins_h = get_bins(min_num,max_num,lsl,usl,set_bins)

    # print('9999 column_category_data_list--->',column_category_data_list)
    # print('9999 column_data--->',len(column_data),column_data)
    # print('9999 bins len--->',len(bins))

    # print(len(column_data),'---->',min(column_data),max(column_data),bins)
    probability_distribution_extend_by_color(column_category_data_list,column_data,bins, 0, item_name, lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,bins_l,bins_h,start_time_first,start_time_last)

    return True





def save_image_to_excel(location,pic_path,plot_sheet):
    #save picture to excel
    plot_sheet.insert_image(location,pic_path,{'x_scale': 1.2, 'y_scale': 1.2})
    return True

def list_category(df,color_by_value):
    column_list = df[color_by_value].tolist()
    column_list = column_list[3:]
    category_list = [x for x in set(column_list)]  # index:[column_list.index(x) for x in set(column_list)]
    # UI control = category_list
    return category_list # ALL category




def select_category(category_l):#action

    get_category_list  =  []#UI control
    color_l = ['#0000FF','#FF0000','#008000','#00FFFF','#000000','#8B008B','#B8860B','#FF6347','#A9A9A9','#FFFF00','#A52A2A','#7FFF00','#D2691E','#6495ED','#FF00FF']

    i=0
    for category_value in category_l:
        temp_list=[]
        # c=colors.cnames.values()[i]
        # print('c====>',c)
        temp_list.insert(0,category_value)
        temp_list.insert(1,color_l[i])
        # color_len = len(colors.cnames.values())
        # print('len(colors.cnames.values() ===>',color_len)
        i = i + 1
        if i == 15 :#
            i=0
        get_category_list.append(temp_list)

    # print('get_category_list====>',get_category_list)
    return get_category_list# selected category,including color_set_list



def select_color_by(df):
    color_by_value ='Station ID'#'off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'---From UI
    if color_by_value =='off':
        category_list =None
    else:
        category_list = list_category(df,color_by_value)
    return color_by_value, category_list



def get_project_code(data_df):
    product_l = list(set(data_df['Product'].values.tolist())) #Product first cell

    project_code =''
    n=1
    for i in product_l:
        if n == 1:
            project_code = i
        else:
            project_code = project_code+'/'+i
        n=n+1

    return project_code



def get_build_stage(data_df):
    special_build_name  = data_df['Special Build Name'].values.tolist() #Special Build Name
    # pattern =re.compile(r'.*\-(.+)\-')
    # result=pattern.match(build_stage)
    # build_stage = result.group(1)
    build_stage =''
    build_stage_l = list(set(special_build_name))
    # print('build_stage_l-->',build_stage_l)
    
    temp_l=[]
    for i in build_stage_l:
        pattern = re.compile(r'.*\-(.+)')
        result = pattern.match(i)
        temp_l.append(result.group(1))
    n=1 
    for j in list(set(temp_l)):
        if n==1:
            build_stage = j
        else:
          build_stage = build_stage +'/'+j
        n=n+1
    return build_stage




def get_station_name(data_df):
    # station_name = data_df[1:2].values[0].tolist()[6] #Station ID first cell
    station_name_l = data_df['Station ID'].values.tolist() #Station ID first cell
    station_name_l = list(set(station_name_l))
    station_name =''
    if len(station_name_l) == 0:
        print('data is empty!')
        station_name == 'xxx'
    else:
        for i in station_name_l:
            if i != '':
                pattern =re.compile(r'.*\_([0-9]+)\_(\D+)')
                result=pattern.match(i)
                try:
                    station_name = result.group(2)
                    return station_name
                except Exception as e:
                    print('match station_name error')

    return station_name




def get_project_info(data_df):
    # print('one row data:--->',data_df[1:2].values[0].tolist(),len(data_df[1:2].values[0].tolist()))
    # project_code = data_df[1:2].values[0].tolist()[1] #Product first cell

    project_code = get_project_code(data_df)

    build_stage = get_build_stage(data_df)

    station_name = get_station_name(data_df)

    print('--------------------project_code:'+project_code+'------------------------')
    print('--------------------build_stage:'+build_stage+'-------------------------')
    print('--------------------station_name:'+station_name+'------------------------')

    return project_code,build_stage,station_name





def open_one_item_csv(event,all_csv_path,data_select,remove_fail):

    tmp_lst = []
    with open(all_csv_path, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('fct') != -1:
                # print("FW version---->")
                pass
            elif row[0].lower().find('display name') != -1:
                pass
            elif row[0].lower().find('pdca priority') != -1:
                pass
            elif row[0].lower().find('upper limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('lower limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('measurement unit') != -1:  # "Measurement Unit ----->" in row:
                pass
            elif row[0].lower().find('site') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # print("index---->", tmp_lst[0])
    header_list = tmp_lst[0][:]#
    temp_header_list = tmp_lst[0]


    if header_list[12] == header_list[13]:
        temp_header_list[13] = temp_header_list[13]+'_2'
        temp_df = pd.DataFrame(tmp_lst[1:], columns=temp_header_list)
        correlation_header_df =temp_df[0:2]



    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    header_df =df[0:2]
    # print('header_df before--->', header_df)
    data_df = df[2:]
    correlation_data_df = data_df
    # print('data_df before--->', data_df)
    
    # print('99999888--->',type(correlation_data_df.columns),correlation_data_df.columns[12],correlation_data_df.columns[13])
    if correlation_data_df.columns[12] != correlation_data_df.columns[13]:
        print(' one_item_plot before remove empty--->', len(correlation_data_df.values.tolist()))
        correlation_data_df=correlation_data_df[~correlation_data_df[correlation_data_df.columns[12]].isin([''])]#Remove SN Empty
        correlation_data_df=correlation_data_df[~correlation_data_df[correlation_data_df.columns[13]].isin([''])]#Remove SN Empty
        print(' one_item_plot after remove empty--->', len(correlation_data_df.values.tolist()))


    print('csv data row number before remove SN empty--->', len(correlation_data_df.values.tolist()))
    correlation_data_df=correlation_data_df[~correlation_data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(correlation_data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(correlation_data_df.values.tolist()))
    if remove_fail== 'yes':
        correlation_data_df=correlation_data_df[correlation_data_df['Test Pass/Fail Status'].isin(['PASS'])]
        # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(correlation_data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(correlation_data_df.values.tolist()))
    if data_select == 'first':
        correlation_data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        correlation_data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        correlation_data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(correlation_data_df.values.tolist()))
    correlation_start_time_first,correlation_start_time_last = '',''
    if len(correlation_data_df.values.tolist()) >2:
        correlation_start_time_l  = correlation_data_df['StartTime'].values.tolist() #StartTime
        correlation_start_time_first = min(correlation_start_time_l)
        correlation_start_time_last  = max(correlation_start_time_l)
        print('<correlation first time -- last time>',correlation_start_time_first,correlation_start_time_last)

    if header_list[12] == header_list[13]:
        df_correlation = correlation_header_df.append(correlation_data_df)
    else:
        df_correlation = header_df.append(correlation_data_df)






    print('csv data row number before remove SN empty--->', len(data_df.values.tolist()))
    data_df=data_df[~data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(data_df.values.tolist()))
    if remove_fail== 'yes':
        data_df=data_df[data_df['Test Pass/Fail Status'].isin(['PASS'])]
        # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(data_df.values.tolist()))
    if data_select == 'first':
        data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(data_df.values.tolist()))

    start_time_l  = data_df['StartTime'].values.tolist() #StartTime
    start_time_first = min(start_time_l)
    start_time_last  = max(start_time_l)
    print('<first time -- last time>',start_time_first,start_time_last)

    if header_list[12] == header_list[13]:
        df = correlation_header_df.append(data_df)
    else:
        df = header_df.append(data_df)

    # print("index1---->", header_list)

    # print('df after--->',df_correlation.columns.values.tolist(), df._stat_axis.values.tolist(),df)
    # print('df.values ---->', df.values)#array([[ ]])
    return header_list,df,df_correlation,start_time_first,start_time_last,correlation_start_time_first,correlation_start_time_last








#
#header_list:A list of string types
#df,Dataframe
#

def open_all_csv(event,all_csv_path,data_select,remove_fail):

    tmp_lst = []
    with open(all_csv_path, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('fct') != -1:
                # print("FW version---->")
                pass
            elif row[0].lower().find('display name') != -1:
                pass
            elif row[0].lower().find('pdca priority') != -1:
                pass
            elif row[0].lower().find('upper limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('lower limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('measurement unit') != -1:  # "Measurement Unit ----->" in row:
                pass
            elif row[0].lower().find('site') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # print("index---->", tmp_lst[0])
    header_list = tmp_lst[0]
  



    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    header_df =df[0:2]
    # print('header_df before--->', header_df)
    data_df = df[2:]
    # print('data_df before--->', data_df)
 

    print('csv data row number before remove SN empty--->', len(data_df.values.tolist()))
    data_df=data_df[~data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(data_df.values.tolist()))
    if remove_fail== 'yes':
        data_df=data_df[data_df['Test Pass/Fail Status'].isin(['PASS'])]
		    # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(data_df.values.tolist()))
    if data_select == 'first':
        data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(data_df.values.tolist()))

    if event == 'keynote-report':
        project_code,build_stage,station_name = get_project_info(data_df)
    else:
        project_code,build_stage,station_name = '','',''

    start_time_l  = data_df['StartTime'].values.tolist() #StartTime
    start_time_first = min(start_time_l)
    start_time_last  = max(start_time_l)
    print('<first time -- last time>',start_time_first,start_time_last)

    df = header_df.append(data_df)

 
    # if event != 'one_item_plot':
    #     station_id_l = df['Station ID'].values.tolist()
    #     fixture_channel_id = df['Fixture Channel ID'].values.tolist()
    #     # print('station_id_l:',station_id_l)
    #     # print('fixture_channel_id:',fixture_channel_id)
    #     temp_l = []
    #     for i in range(0,len(station_id_l[2:])):
    #         temp_l.append(station_id_l[i+2]+'_'+fixture_channel_id[i+2])
    #     temp_l.insert(0,'')
    #     temp_l.insert(0,'')
    #     fixture_channel_id = temp_l
    #     # print(fixture_channel_id)
    #     df['Fixture Channel ID'] = pd.DataFrame({'Fixture Channel ID':fixture_channel_id})
    #     # print('fixture channel id:',df['Fixture Channel ID'].values.tolist())

    #     print('*******************************')


    # print('df after--->', df)
    # print('df.values ---->', df.values)#array([[ ]])
    return header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last

def is_number(num):
    pattern = re.compile(r'(.*)\.(.*)\.(.*)')
    if pattern.match(num):
        return False
    return num.replace(".", "").replace('-','').isdigit()



def valid_column(test_item_name,column_list):

    # column_list = df[test_item_name].tolist()
    # print('valid_column item_name-->',test_item_name)
    # print('valid_column--->',len(column_list),column_list)

    # print('column_value_list--->',column_list)
    if test_item_name.lower().find('_cb_count') == -1 and test_item_name.lower().find('fixture channel id') == -1 and test_item_name.lower().find('ss1_num_tri') == -1 and test_item_name.lower().find('fixture vendor_id') == -1:
        pass
    else:
        return 'not_cpk state1'
    if column_list[0] == '0' and column_list[1] == '0' or column_list[0] == '1' and column_list[
        1] == '1' or  column_list[0] == column_list[1] and  column_list[0] !='NA' or len(column_list)< 5:
        # print('====>',column_list[0],column_list[1])
        return "not_cpk state2"
    else:
        # pattern = re.compile(r'^[+-]?[0-9]*\.?[0-9]+$')
        # print('valid_column len----->',test_item_name,len(column_list))
        # print('v====>',column_list[0],column_list[1],column_list,len(column_list))
        j=0
        for i in range(2,len(column_list),1):
            # print(str(i),'valid_column----->',test_item_name,pattern.match(column_list[i]))
            if is_number(column_list[i]):
                j=j+1
                # print(str(j)+',len(set(column_list))-->',j,len(set(column_list)))
                if j>3 and len(set(column_list))>=3:
                    # print('need_cpk')
                    return "need_cpk"
 
        return "not_cpk state3"


def test_value_to_numeric(test_data_list):
    column_list = []
    i = 0
    for x in test_data_list:
        if i ==0 and x == 'NA' or i ==1 and x == 'NA':
            column_list.append(x)
        else:
            try:
              x = float(x)
              column_list.append(x)
            except Exception as e:
                pass
            # print('-------------------- it is not number --------------')
        i = i + 1

    # print('column_list--->',column_list)
    return column_list



def get_target_value(lsl,usl):
    if lsl != 'NA' or usl != 'NA':
        target_value = round(((lsl + usl) / 2.0), 5)
    else:
        target_value = 'Nan'
    return target_value


def get_limit_range(lsl,usl):
    # print('lsl,usl----->', lsl, usl)
    range = 0
    if lsl < 0 and usl <= 0:
        range = abs(lsl) - abs(usl)
    elif lsl < 0 and usl >= 0:
        range = abs(lsl) + usl
    elif lsl >= 0 and (usl > 0):
        range = usl - lsl
    else:
        print('get_limit_range 00000')
    range = round(range, 5)
    # print('range in get_limit_range----->', range)
    return range


def get_ticks(min_num,max_num,lsl,usl):
    #for correlation plot
    # print('min_num,max_num,lsl,usl=====>',min_num,max_num,lsl,usl)

    ticks_l = 0
    ticks_h = 0
    if lsl =='NA' or usl =='NA':
        ticks_l = min_num
        ticks_h = max_num
    else:
        if min_num < lsl and max_num < lsl:
            ticks_l = min_num
            ticks_h = usl
        elif min_num < lsl and max_num > lsl and max_num <= usl:
            ticks_l = min_num
            ticks_h = usl
        elif min_num < lsl and max_num > usl:
            ticks_l = min_num
            ticks_h = max_num
        elif lsl <= min_num and min_num <= usl and max_num <= usl:
            ticks_l = lsl
            ticks_h = usl
        elif lsl <= min_num and min_num <= usl and max_num > usl:
            ticks_l = lsl
            ticks_h = max_num
        elif min_num > usl:
            ticks_l = lsl
            ticks_h = max_num
    # print('ticks_l,ticks_h:',ticks_l,ticks_h)
    range = get_limit_range(ticks_l,ticks_h)
    # print('ticks_range:',range,round((ticks_l-range/5),2),round((ticks_h+range/5),2))

    ticks_l,ticks_h = round((ticks_l-range/5),2),round((ticks_h+range/5),2)
    return ticks_l,ticks_h


def is_empty_list(l):
    temp_l = []
    for value in l:
        if value != '':
            temp_l.append(value)
    n = len(temp_l)
    if n == 0:
        # print('it is empty list!')
        return True
    else:
        return False



'''
color_by:  Special Build Name,Version,Station ID,SerialNumber
select_category : category list
return table_data,table_category_data;table_data for cpk,table_category_data for draw plot
'''

def parse_all_csv(header_list,df,color_by1,selected_category1,event,color_by2,selected_category2):
    

    color_l = ['#0000FF','#FF0000','#008000','#00FFFF','#000000','#8B008B','#B8860B','#FF6347','#A9A9A9','#FFFF00','#A52A2A','#7FFF00','#D2691E','#6495ED','#FF00FF']
    table_data = []#[[]]
    table_category_data = []#[[[]]]
    column_list = []
    n=0
    for item_name in header_list:
        # if event == 'one_item_plot1':
        #     # print('one column--->',type(df.iloc[:, [n]].values),df.iloc[:, [n]].values.tolist())
        #     temp_l=[]
        #     for i in df.iloc[:, [n]].values.tolist():
        #         temp_l.append(i[0])
        #     column_list = temp_l  # 读取指定键值列的所有行
        #     print('column_list--->',column_list)
        #     n=n+1
        # else:
        # print('item_name:',item_name)
        # print('header_list:',header_list)
        # print('df:',df)
        column_list = df[item_name].tolist()
        # print('column_list--->',column_list)

        need_cpk = valid_column(item_name,column_list)
        # print('need_cpk:---->',need_cpk)
        column_num_list = []
        if need_cpk == 'need_cpk':
            column_list = test_value_to_numeric(column_list)
            column_list.insert(0, item_name)  # item name
            usl = column_list[1]
            lsl = column_list[2]
            # 'Off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'/'Product'/'StartTime'/'Special Build Description'
            if color_by1 == 'SerialNumber' or color_by1 == 'Version' or color_by1 == 'Station ID' or color_by1 == 'Special Build Name' or color_by1 == 'Product' or color_by1 == 'StartTime' or color_by1 =='Special Build Description' or color_by1 =='Fixture Channel ID':
                column_temp = []
                first_filter_category_data_len=0
                second_filter_category_data_len=0
                i =0
                for x in selected_category1:
                    if color_by2 == 'SerialNumber' or color_by2 == 'Version' or color_by2 == 'Station ID' or color_by2 == 'Special Build Name' or color_by2 == 'Product' or color_by2 == 'StartTime' or color_by2 =='Special Build Description' or color_by2 =='Fixture Channel ID':
                        
                        for xx in selected_category2:

                            # print('xxxxx:', x[0],x[1],xx[0],xx[1])
                            one_category_list = df.loc[(df[color_by1] == x[0]) & (df[color_by2] == xx[0]), item_name].tolist()  # 
                            
                            if is_empty_list(one_category_list) != True:
                                # print('one category_data with second filter-->', len(one_category_list),one_category_list)
                                second_filter_category_data_len = second_filter_category_data_len +len(one_category_list)
                                one_category_list = test_value_to_numeric(one_category_list)
                                column_temp = column_temp +one_category_list

  
                                # print('usl--->', usl)
                                # print('lsl--->', lsl)
                                category_value = x[0]+'&'+xx[0]
                                # print('category_value--->',category_value)
                                one_category_list.insert(0, category_value)  # insert category
                                # print('i==>',i)
                                if i >14:
                                    i = 0 
                                one_category_list.insert(1, color_l[i])  # insert color
                                one_category_list.insert(2,item_name)  # item_name
                                if lsl != 'NA' or usl != 'NA':
                                    usl = float(usl)
                                    lsl = float(lsl)

                                one_category_list.insert(3, usl)  # usl
                                one_category_list.insert(4, lsl)  # lsl
                                # print('one_category_list-->', one_category_list)
                                column_num_list.append(one_category_list)
                                i=i+1

                    elif color_by2 == 'Off':

                        # print('xxxxx:', x[0],x[1])
                        one_category_list = df.loc[df[color_by1] == x[0], item_name].tolist()  # 
                        if is_empty_list(one_category_list) != True:
                            # print('one category_data with first filter-->', one_category_list)
                            first_filter_category_data_len = first_filter_category_data_len +len(one_category_list)
                            one_category_list = test_value_to_numeric(one_category_list)
                            column_temp = column_temp +one_category_list

                            # usl = column_list[1]
                            # lsl = column_list[2]
                            # print('usl--->', usl)
                            # print('lsl--->', lsl)
                            one_category_list.insert(0, x[0])  # insert category
                            one_category_list.insert(1, x[1])  # insert color
                            one_category_list.insert(2,item_name)  # item_name
                            if lsl != 'NA' or usl != 'NA':
                                usl = float(usl)
                                lsl = float(lsl)

                            one_category_list.insert(3, usl)  # usl
                            one_category_list.insert(4, lsl)  # lsl
                            # print('one_category_list-->', one_category_list)
                            column_num_list.append(one_category_list)
                # print('one category_data total with first filter-->', first_filter_category_data_len)
                # print('one category_data total with second filter-->', second_filter_category_data_len)
                
                column_temp.insert(0,item_name)  # item_name
                if lsl != 'NA' or usl != 'NA':
                    usl = float(usl)
                    lsl = float(lsl)
                column_temp.insert(1, usl)  # usl
                column_temp.insert(2, lsl)  # lsl
                # print('column_temp-->', column_temp)
                table_data.append(column_temp)# one column's all category data []
                # print('column_num_list-->', column_num_list)
                table_category_data.append(column_num_list)# category [[],[],...]]
            elif color_by1 == 'Off':
                table_data.append(column_list)  # item_name,usl,lsl,data
        else:
            pass
    # print('table_data-->',len(table_data[0]))
    return table_data,table_category_data #[[[ ]]]





def calulate_param(header_list,df,csv_path):
    # print('header_list:',type(header_list),header_list,list(df.columns))

    data = [('item_name','BC','P_Val','a_Q','a_irr','3CV'),]
    f = open(csv_path,'w')
    writer = csv.writer(f)

    writer.writerow(('item_name','BC','P_Val','a_Q','a_irr','3CV'))


    column_list = []
    n=0
    for item_name in header_list:
        column_list = df[item_name].tolist()
        # print('column_list--->',column_list)#string list

        # print('item_name:',item_name)
        need_cpk = valid_column(item_name,column_list)
        # print('need_cpk:---->',need_cpk)
        column_num_list = []
        if need_cpk == 'need_cpk':
            column_list = test_value_to_numeric(column_list[2:])
            # print(str(n)+'column_list:',column_list)
            bc,p_val,a_Q,a_irr,three_CV = get_coefficients(column_list)
            value = (item_name,bc,p_val,a_Q,a_irr,three_CV)
            writer.writerow(value)
        n=n+1

    f.close()



@count_time
def draw_overall_cpk(path,excel_file):
    #'/Users/rex/PycharmProjects/my/cpk_report.xlsx'
    all_cpk_plot_path=path
    path=path+'temp/cpk.xlsx'
    df = pd.read_excel(path)  # 读取xlsx中第一个sheet
    # df.loc[index, column_name],选取指定行和列的数据
    # df.loc[0,'name'] # 'Snow'
    # df.loc[0:2, ['name','age']]          #选取第0行到第2行，name列和age列的数据, 注意这里的行选取是包含下标的。
    # df.loc[[2,3],['name','age']]          #选取指定的第2行和第3行，name和age列的数据
    # df.loc[df['gender']=='M','name']      #选取gender列是M，name列的数据
    # df.loc[df['gender']=='M',['name','age']] #选取gender列是M，name和age列的数据
    item_name_list = df.iloc[:, [1]].values  # 读取指定键值列的所有行
    cpk_list = df.iloc[:, [16]].values  # 读取指定键值列的所有行
    # print("item_name_list：\n{0}".format(item_name_list.tolist()[0]))#first item name
    # print("cpk_list：\n{0}".format(cpk_list.tolist()[0]))#first item value

    x_x = []
    y_y = []
    temp_l=[]
    n=0
    for i in cpk_list.tolist():
        if str(i[0]) =='nan':
            temp_l.append(n)
        else:
            cpk_value = round((i[0]), 5)
            y_y.append(cpk_value)
        n = n+1
    n = 0
    for i in item_name_list.tolist():
        if n in temp_l:
            pass
        else:
            x_x.append(str(i[0]))
        n = n+1
    calculated_cpk_num = len(x_x)
    print('calculated cpk item num:',len(x_x),len(y_y))
    x = x_x
    y = y_y
    cpk_max = max(y)
    print('cpk_max:', cpk_max)
    # print('x:',x)
    # print('y:',y)
    plt.ion()  # 开启interactive mode
    fig, axes = plt.subplots(1, 0, figsize=(12.8, 8), facecolor='#ccddef')
    plt.axes([0.035, 0.30, 0.965, 0.60])  # [左, 下, 宽, 高] 规定的矩形区域 （全部是0~1之间的数，表示比例）
    bars = plt.bar(x, y)
    plt.ylim((0, 60))  # 设置y轴scopex

    # -------------------------------------
    # i=1
    # for bar in bars:
    #     # color=random.choice(colors)
    #     color = str(hex(i))
    #     print(color)
    #     color = color[2:]
    #     print(color,6-len(color))
    #     zero_count = 6-len(color)
    #     for j in range(1,zero_count+1,1):
    #         color ="0"+str(color)
    #     i = i+16
    #     color = '#'+color
    #     print('--->',len(color),color)
    # --------------------------------------

    # print('====>',len(colors.cnames.values()))
    # a = range (0,len(colors.cnames.values())+1,1)
    # b = random.sample(a,1)
    # --------------------------------------
    i = 0
    # print('colors counts:', len(colors.cnames))
    for bar in bars:
        # print('i===>',i)
        c = list(colors.cnames.values())[i]
        # print('c====>',c)
        bar.set_facecolor(c)  # '#2ca02c'
        i = i + 1
        if i == len(colors.cnames.values()):
            i = 0

    labels = x
    # -------cpk value---------------
    for a, b in zip(x, y):
        # print('---->', a, b)
        plt.text(a, b, '%.1f' % b, ha='center', va='bottom', fontsize=1)
    # -------------------------------

    l1 = [0.2]
    l2 = list(range(3, 40, 3))
    l3 = list(range(45, 100, 5))
    l4 = list(range(100,150,30))
    y_ticks = l1 + l2 + l3 + l4
    print(y_ticks)
    plt.ylabel('CPK  ', rotation='horizontal', fontsize=8, verticalalignment='center', horizontalalignment='center')
    plt.xlabel('Item Name',rotation='horizontal', fontsize=8, verticalalignment='center', horizontalalignment='center')
    plt.yticks(y_ticks, fontproperties='Times New Roman', size=5)
    plt.xticks(labels, fontproperties='Times New Roman', size=1, rotation=270)
    plt.rcParams['figure.autolayout'] = True  # 解决不能完整显示的问题（比如因为饼图太大，显示窗口太小）

    plt.rcParams['savefig.dpi'] = 1440  # 图片像素
    plt.rcParams['figure.dpi'] = 1440  # 分辨率
    titile_name = str(calculated_cpk_num)
    titile_name = 'CPK of ' + titile_name + ' items'
    plt.title(titile_name, y=1.0, fontsize=12, fontweight='bold', verticalalignment='baseline',horizontalalignment='center', color="royalblue")
    plt.grid(linestyle=':', c='gray', linewidth=0.1, alpha=0.5)  # 生成网格
    # plt.legend([""],loc="upper center")
    plt.draw()
    # plt.subplots_adjust(wspace=0, hspace=50)  # 调整子图间距
    all_cpk_plot_path= all_cpk_plot_path+'whole_cpk.png'
    plt.subplots_adjust(wspace=0, hspace=0)  # 调整子图间距
    plt.margins(0, 0)
    plt.savefig(all_cpk_plot_path, dpi=1440)
    # plt.show()
    plt.close()
    plt.ioff()  # off interactive mode
    print('draw whole CPK plot finished!!')

    return all_cpk_plot_path


def verify_limit(lsl,usl):
      lsl.replace(' ','')
      usl.replace(' ','')
      try:
        lsl = float(eval(lsl))
        usl = float(eval(usl))
      except:
        # print('no valid limit update!')
        return None,None
      if type(lsl) == float and type(usl) == float:
          # print('after verify lsl===>',lsl)
          # print('after verify usl===>',usl)
          return lsl,usl


@count_time
def draw_correlation(xValue,yValue,x_item_name,y_item_name,pic_save_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last):
    pearson,spearman = correlation_coefficient_calc(xValue, yValue, x_item_name, y_item_name)
    plt.ion()  # 开启interactive mode
    # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
    fig, axes = plt.subplots(1, 0, figsize=(8, 6), facecolor='#ccddef')
    plt.axes([0.15, 0.15, 0.75, 0.75])  # [左, 下, 宽, 高] 规定的矩形区域 （全部是0~1之间的数，表示比例）    
    plt.title('Correlation pearson coefficient = ' + str(pearson)+'\n'+str(start_time_first)+' -- '+str(start_time_last),size=14)
    if len(x_item_name) > 55:
        x_item_name = x_item_name[0:55] + '\n' + x_item_name[55:]
    if len(y_item_name) > 55:
        y_item_name = y_item_name[0:55] + '\n' + y_item_name[55:]
    plt.xlabel(x_item_name,size=12)
    plt.ylabel(y_item_name,size=12)

    x_min_num, x_max_num = min(xValue), max(xValue)
    # print('xvalue min,max:', x_min_num, x_max_num)
    x_ticks_l, x_ticks_h = get_ticks(x_min_num, x_max_num, x_lsl, x_usl)
    # print('x_ticks_l x_ticks_h:', x_ticks_l, x_ticks_h)
    plt.xlim(x_ticks_l, x_ticks_h)

    y_min_num, y_max_num = min(yValue), max(yValue)
    # print('yvalue min,max:', y_min_num, y_max_num)
    y_ticks_l, y_ticks_h = get_ticks(y_min_num, y_max_num, y_lsl, y_usl)
    # print('y_ticks_l y_ticks_h:', y_ticks_l, y_ticks_h)
    plt.ylim((y_ticks_l, y_ticks_h))  # 设置y轴scopex
    ax=plt.gca()
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)
    if x_lsl != 'NA' and x_usl != 'NA':
        plt.plot([x_lsl, x_lsl, ], [y_ticks_l, 100000000, ], 'k--', linewidth=3.0, color='red')  # x lower limit线，
        plt.plot([x_usl, x_usl, ], [y_ticks_l, 100000000, ], 'k--', linewidth=3.0, color='red')  # x upper limit线，
    if y_lsl != 'NA' and y_usl != 'NA':
        plt.plot([x_ticks_l, 100000000, ], [y_lsl, y_lsl, ], 'k--', linewidth=3.0, color='red')  # y lower limit线，
        plt.plot([x_ticks_l, 100000000, ], [y_usl, y_usl, ], 'k--', linewidth=3.0, color='red')  # y upper limit线，

    # plt.scatter(x, y, s, c, marker)
    # x: x轴坐标
    # y：y轴坐标
    # s：点的大小/粗细 标量或array_like 默认是 rcParams['lines.markersize'] ** 2
    # c: 点的颜色
    # marker: 标记的样式 默认是 'o'
    # plt.legend()
    plt.rcParams['savefig.dpi'] = 250  # 图片像素
    plt.rcParams['figure.dpi'] = 150  # 分辨率
    plt.scatter(xValue, yValue, s=45,linewidth =1, c="blue", marker='+')
    plt.grid(linestyle=':', c='gray', linewidth=1, alpha=0.6)  # 生成网格
    plt.savefig(pic_save_path, dpi=250)
    plt.draw()
    # plt.show()
    plt.close()
    plt.ioff()


@count_time
def draw_correlation_by_color(xValue,yValue,x_category_value,y_category_value,x_item_name,y_item_name,pic_save_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last):
    pearson, spearman = correlation_coefficient_calc(xValue, yValue, x_item_name, y_item_name)
    plt.ion()  # 开启interactive mode
    # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
    fig, axes = plt.subplots(1, 0, figsize=(8, 6), facecolor='#ccddef')
    plt.axes([0.15, 0.15, 0.75, 0.75])  # [左, 下, 宽, 高] 规定的矩形区域 （全部是0~1之间的数，表示比例）
    plt.title('Correlation pearson coefficient = ' + str(pearson)+'\n'+str(start_time_first)+' -- '+str(start_time_last),size=14)
    if len(x_item_name) > 55:
        x_item_name = x_item_name[0:55] + '\n' + x_item_name[55:]
    if len(y_item_name) > 55:
        y_item_name = y_item_name[0:55] + '\n' + y_item_name[55:]
    plt.xlabel(x_item_name,size=12)
    plt.ylabel(y_item_name,size=12)

    x_min_num, x_max_num = min(xValue), max(xValue)
    # print('xvalue min,max:', x_min_num, x_max_num)
    x_ticks_l, x_ticks_h = get_ticks(x_min_num, x_max_num, x_lsl, x_usl)
    # print('x_ticks_l x_ticks_h:', x_ticks_l, x_ticks_h)
    plt.xlim(x_ticks_l, x_ticks_h)

    y_min_num, y_max_num = min(yValue), max(yValue)
    # print('yvalue min,max:', y_min_num, y_max_num)
    y_ticks_l, y_ticks_h = get_ticks(y_min_num, y_max_num, y_lsl, y_usl)
    # print('y_ticks_l y_ticks_h:', y_ticks_l, y_ticks_h)
    plt.ylim((y_ticks_l, y_ticks_h))  # 设置y轴scopex
    ax=plt.gca()
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)
    if x_lsl != 'NA' and x_usl != 'NA':
        plt.plot([x_lsl, x_lsl, ], [y_ticks_l, 100000000, ], 'k--', linewidth=3.0, color='red')  # x lower limit线，
        plt.plot([x_usl, x_usl, ], [y_ticks_l, 100000000, ], 'k--', linewidth=3.0, color='red')  # x upper limit线，
    if y_lsl != 'NA' and y_usl != 'NA':
        plt.plot([x_ticks_l, 100000000, ], [y_lsl, y_lsl, ], 'k--', linewidth=3.0, color='red')  # y lower limit线，
        plt.plot([x_ticks_l, 100000000, ], [y_usl, y_usl, ], 'k--', linewidth=3.0, color='red')  # y upper limit线，

    # plt.scatter(x, y, s, c, marker)
    # x: x轴坐标
    # y：y轴坐标
    # s：点的大小/粗细 标量或array_like 默认是 rcParams['lines.markersize'] ** 2
    # c: 点的颜色
    # marker: 标记的样式 默认是 'o'
    # plt.legend()
    plt.rcParams['savefig.dpi'] = 250  # 图片像素
    plt.rcParams['figure.dpi'] = 150  # 分辨率
    # print('x_category_value,y_category_value length:---->',len(x_category_value),len(y_category_value),x_category_value[0], y_category_value[0])
    for i in range(0,len(x_category_value),1):
        print(x_category_value[i][1],y_category_value[i][1])
        set_color = x_category_value[i][1]
        plt.scatter(x_category_value[i][5:], y_category_value[i][5:], s=45,linewidth =1, c=set_color, marker='+')
    plt.grid(linestyle=':', c='gray', linewidth=1, alpha=0.6)  # 生成网格
    plt.savefig(pic_save_path, dpi=250)
    plt.draw()
    # plt.show()
    plt.close()

    plt.ioff()


@count_time
def correlation_plot(table_data,table_category_data,pic_path,start_time_first,start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl):


    j=0
    path=pic_path
    info = ''
    # print('show_correlation_plot table_data length:',len(table_data),table_data)
    # print('show_correlation_plot table_category_data length:',len(table_category_data),table_category_data)
    if len(table_data) < 2:
       print('data length short error!')
       return
    else:
        if len(table_data[0])<6 or len(table_data[1])<6:
            print('empty data or too short,pls check!')
            return

    y_item_name = table_data[0][0]
    x_item_name = table_data[1][0]

    # print('y_item_name==>',y_item_name)
    # print('x_item_name==>',x_item_name)

  
    y_usl = table_data[0][1]
    y_lsl = table_data[0][2]

    x_usl = table_data[1][1]
    x_lsl = table_data[1][2]

    yValue = table_data[0][3:]
    xValue = table_data[1][3:]
    # print('show_correlation_plot data lenghth:',len(xValue),len(yValue))
    if len(yValue) != len(xValue):
        print('xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot') 
        return 'xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot'
    new_lsl,new_usl = verify_limit(new_y_lsl,new_y_usl)
    if new_lsl != None and new_lsl != None:
        y_lsl = new_lsl
        y_usl = new_usl

    new_lsl,new_usl = verify_limit(new_x_lsl,new_x_usl)
    if new_lsl != None and new_lsl != None:
        x_lsl = new_lsl
        x_usl = new_usl


    # image_name = item_name.replace('/','_')+".png"
    image_name ='correlation.png'
    pic_path = path +'temp/'
    if not os.path.exists(pic_path):
        os.makedirs(pic_path)
    # os.system('mkdir '+pic_path)
    pic_path = pic_path + image_name
    # print('pic_path--->',pic_path)

    if len(table_category_data) == 0:
        draw_correlation(xValue,yValue,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)
    else:
        #[[],[],[]...],only value category lists
        y_category_value = table_category_data[0]
        x_category_value = table_category_data[1]
        # print('y_category_value,x_category_value:',table_category_data[0],table_category_data[1])
        draw_correlation_by_color(xValue,yValue,x_category_value,y_category_value,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)

    j=j+1
    info = 'correlation plot draw finished!'
    print(info)
    return info



@count_time
def show_correlation_plot(event,header_list,df,color_by1,pic_path,select_category1,start_time_first,start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl,color_by2,select_category2):
    # print('header_list===>',header_list)

    table_data,table_category_data = parse_all_csv(header_list,df,color_by1,select_category1,event,color_by2,select_category2)

    j=0
    path=pic_path
    info = ''
    # print('show_correlation_plot table_data length:',len(table_data),table_data)
    # print('show_correlation_plot table_category_data length:',len(table_category_data),table_category_data)

    if len(table_data) < 2:
       print('data length short error!')
       return
    else:
        if len(table_data[0])<6 or len(table_data[1])<6:
            print('empty data or too short,pls check!')
            return


    y_item_name = table_data[0][0]
    x_item_name = table_data[1][0]

    # print('y_item_name==>',y_item_name)
    # print('x_item_name==>',x_item_name)

  
    y_usl = table_data[0][1]
    y_lsl = table_data[0][2]

    x_usl = table_data[1][1]
    x_lsl = table_data[1][2]

    yValue = table_data[0][3:]
    xValue = table_data[1][3:]
    # print('show_correlation_plot data lenghth:',len(xValue),len(yValue))
    if len(yValue) != len(xValue):
        print('xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot') 
        return 'xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot'
    new_lsl,new_usl = verify_limit(new_y_lsl,new_y_usl)
    if new_lsl != None and new_lsl != None:
        y_lsl = new_lsl
        y_usl = new_usl

    new_lsl,new_usl = verify_limit(new_x_lsl,new_x_usl)
    if new_lsl != None and new_lsl != None:
        x_lsl = new_lsl
        x_usl = new_usl


    # image_name = item_name.replace('/','_')+".png"
    image_name ='correlation.png'
    pic_path = path +'temp/'
    if not os.path.exists(pic_path):
        os.makedirs(pic_path)
    # os.system('mkdir '+pic_path)
    pic_path = pic_path + image_name
    # print('pic_path--->',pic_path)

    if len(table_category_data) == 0:
        draw_correlation(xValue,yValue,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)
    else:
        #[[],[],[]...],only value category lists
        y_category_value = table_category_data[0]
        x_category_value = table_category_data[1]
        # print('y_category_value,x_category_value:',table_category_data[0],table_category_data[1])
        draw_correlation_by_color(xValue,yValue,x_category_value,y_category_value,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)

    j=j+1
    info = 'correlation plot draw finished!'
    print(info)
    return info




@count_time
def cpk_plot(table_data,table_category_data,pic_path,set_bins,select_new_lsl,select_new_usl,start_time_first,start_time_last):


    j=0
    path=pic_path
    info = ''
    # print('table_data length:',len(table_data),table_data)
    # print('table_category_data length:',len(table_category_data),table_category_data)
    if len(table_data) == 0:
        print('data empty,pls check!')
        return
    elif len(table_data) == 1:
        if len(table_data[0]) <6:
            print('empty data or too short,pls check!')
            return
            
    for column_data in table_data:
        # print('column_data length,column_data--->',len(column_data),column_data)
        item_name=column_data[0]
        usl = column_data[1]
        lsl = column_data[2]
        column_data = column_data[3:]
        # print('item_name:',item_name)
        # print('usl:',usl)
        # print('lsl:',lsl)
        # print(str(j)+' column test value:',column_data)
        new_lsl,new_usl = verify_limit(select_new_lsl,select_new_usl)
        if new_lsl != None and new_lsl != None:
            lsl = new_lsl
            usl = new_usl
            # print('new lsl===>',lsl)
            # print('new usl===>',usl)

        mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)
        if stdev == 0:
            info = item_name+'stdv == 0 cannot calculate cpk!'
            print(info)#stdev ==0
            return info
        else:

            # print('cpk:',cpk)
            # image_name = item_name.replace('/','_')+".png"
            image_name ='temp_pic.png'
            pic_path = path +'temp/'
            if not os.path.exists(pic_path):
                os.makedirs(pic_path)
            # os.system('mkdir '+pic_path)
            pic_path = pic_path + image_name

            if len(table_category_data) == 0:
                draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
            else:
                #[[],[],[]...],only value category lists
                draw_more_histogram(table_category_data[j],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)

        j=j+1
        info = item_name+' item cpk and plot draw finished!'
        print(info)
        return info






@count_time
def show_cpk_plot(event,one_item_head_list,df,color_by1,pic_path,select_category_l1,set_bins,select_new_lsl,select_new_usl,start_time_first,start_time_last,color_by2,select_category_l2):

    table_data,table_category_data = parse_all_csv(one_item_head_list,df,color_by1,select_category_l1,event,color_by2,select_category_l2)
    j=0
    path=pic_path
    info = ''
    # print('table_data length:',len(table_data),table_data)
    # print('table_category_data length:',len(table_category_data),table_category_data)
    if len(table_data) == 0:
        print('data empty,pls check!')
        return
    elif len(table_data) == 1:
        if len(table_data[0]) <6:
            print('empty data or too short,pls check!')
            return

    for column_data in table_data:
        # print('column_data length,column_data--->',len(column_data),column_data)
        item_name=column_data[0]
        usl = column_data[1]
        lsl = column_data[2]
        column_data = column_data[3:]
        # print('item_name:',item_name)
        # print('usl:',usl)
        # print('lsl:',lsl)
        # print(str(j)+' column test value:',column_data)
        new_lsl,new_usl = verify_limit(select_new_lsl,select_new_usl)
        if new_lsl != None and new_lsl != None:
            lsl = new_lsl
            usl = new_usl
            # print('new lsl===>',lsl)
            # print('new usl===>',usl)

        mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)
        if stdev == 0:
            info = item_name+'stdv == 0 cannot calculate cpk!'
            print(info)#stdev ==0
            return info
        else:

            # print('cpk:',cpk)
            # image_name = item_name.replace('/','_')+".png"
            image_name ='temp_pic.png'
            pic_path = path +'temp/'
            if not os.path.exists(pic_path):
                os.makedirs(pic_path)
            # os.system('mkdir '+pic_path)
            pic_path = pic_path + image_name

            if len(table_category_data) == 0:
                draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
            else:
                #[[],[],[]...],only value category lists
                draw_more_histogram(table_category_data[j],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)

        j=j+1
        info = item_name+' item cpk and plot draw finished!'
        print(info)
        return info



def creat_excel_report_file(path,file,cpk_lsl,cpk_usl,event,fail_plot_to_excel):
    if event == 'keynote-report':
        excel_file_path = path+'temp/cpk.xlsx'
    else:
        excel_file_path = path + file

    book = xlsxwriter.Workbook(excel_file_path)#'cpk.xlsx'
    report_sheet = book.add_worksheet('report')#'report'
    if event == 'excel-report' and fail_plot_to_excel == 'yes':
        plot_sheet = book.add_worksheet('fail plot')#'plot'
    else:
        plot_sheet = None
    report_sheet.set_column("B:B", 80)  # 设定A列列宽为40
    report_sheet.set_column("M:M", 10)  # 设定A列列宽为10
    report_sheet.set_row(0, 30)    #设置行高度

    cpk_result ='CPK_Result(cpk_lsl:'+str(cpk_lsl)+';cpk_usl:'+str(cpk_usl)+')'
    report_sheet_title = [u'No', u'Item_name',u'BC',u'P_Val',u'a_Q',u'a_irr',u'3CV',u'ORIG LSL',u'Target', u'ORIG USL', u'Min', u'Mean',u'Max', u'Std',u'CPL', u'CPU', u'CPK',cpk_result, u'New LSL',u'New Target',u'New USL',u'New CPL',u'New CPU',u'New CPK','New '+cpk_result]

    format_normal = book.add_format()  # 定义format格式对象
    format_normal.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_normal.set_valign('center') # 定义format_titile对象单元格对齐方式

    format_normal.set_border(1)  # 定义format对象单元格边框加粗的格式
    # format_normal.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    format_normal.set_text_wrap()  # 内容换行



    new_format_pass = book.add_format()  # 定义format格式对象
    new_format_pass.set_align('center')  # 定义format_titile对象单元格对齐方式
    new_format_pass.set_valign('center') # 定义format_titile对象单元格对齐方式
    new_format_pass.set_bg_color('green')  # 定义format_titile对象单元格背景颜色

    new_format_pass.set_border(1)  # 定义format对象单元格边框加粗的格式
    # new_format_pass.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    new_format_pass.set_text_wrap()  # 内容换行


    format_highlight = book.add_format()  # 定义format_title 格式对象
    format_highlight.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    format_highlight.set_bg_color('yellow')  # 定义format_titile对象单元格背景颜色
    format_highlight.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_highlight.set_valign('center') # 定义format_titile对象单元格对齐方式

    # format_highlight.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    format_highlight.set_text_wrap()  # 内容换行



    new_format_fail = book.add_format()  # 定义format_title 格式对象
    new_format_fail.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    new_format_fail.set_bg_color('red')  # 定义format_titile对象单元格背景颜色
    new_format_fail.set_align('center')  # 定义format_titile对象单元格对齐方式
    new_format_fail.set_valign('center') # 定义format_titile对象单元格对齐方式

    # format_fail.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    new_format_fail.set_text_wrap()  # 内容换行


    format_titile = book.add_format()  # 定义format_title 格式对象
    format_titile.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    format_titile.set_bg_color('#cddccc')  # 定义format_titile对象单元格背景颜色
    format_titile.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_titile.set_valign('center')  # 定义format_titile对象单元格对齐方式
    format_titile.set_bold()  # 定义format_titile对象内容加粗
    format_titile.set_text_wrap()  # 内容换行


    format_ave = book.add_format()  # 定义format_ave格式对象
    format_ave.set_border(1)  # 边框加粗的格式
    format_ave.set_num_format('0.00')  # 定义format_ave对象单元格数字类别显示格式
    report_sheet.write_row('A1', report_sheet_title, format_titile)
    print('Excel report head created finished!')
    return book,report_sheet,plot_sheet,format_highlight,format_normal,format_titile,new_format_fail,new_format_pass


def get_one_item_new_limit_from_csv(new_limit_path,item_name):
    new_limit_file = new_limit_path + 'temp/item_limit.csv'
    tmp_lst = []
    # try:
    with open(new_limit_file, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('item') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # except IOError as exc:
    #   print('ERROR:',exc)
    header_list = tmp_lst[0]
    # print('header_list--->',header_list)
    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    # print(df)
    try:
        new_lsl=df.loc[df['item'] == item_name, 'new_lsl'].tolist()[0]
        new_usl=df.loc[df['item'] == item_name, 'new_usl'].tolist()[0]
    except IndexError as e:
        print('item_limit.csv item wrong!')
    return new_lsl,new_usl


def load_excel_table(excel_report_path):
    excel_data = openpyxl.load_workbook(excel_report_path)
    sheetnames = excel_data.get_sheet_names()
    table = excel_data.get_sheet_by_name(sheetnames[0])#[u'report', u'fail plot']
    table = excel_data.active
    print('append one item new limit data to-->',table.title) 
    nrows = table.max_row 
    ncolumns = table.max_column 
    print('load_excel_table nrows--->',nrows)
    print('load_excel_table ncolumns--->',ncolumns)
    return excel_data,table,nrows,ncolumns



def append_new_limit_data_to_excel_report(data,table_name,excel_report_path,row_n,new_lsl,new_usl,new_cpl,new_cpu,new_cpk):
    print('data in append_new_limit_data_to_excel_report--->',data,excel_report_path,row_n,new_lsl,new_usl,new_cpl,new_cpu,new_cpk)
    table_name.cell(row_n+1,14).value = new_lsl
    table_name.cell(row_n+1,15).value = new_usl
    table_name.cell(row_n+1,16).value = new_cpl 
    table_name.cell(row_n+1,17).value = new_cpu
    table_name.cell(row_n+1,18).value = new_cpk
    data.save(excel_report_path)


def save_new_limit_to_report(new_limit_path,excel_report_name):
    new_limit_file = new_limit_path + 'temp/item_limit.csv'
    tmp_lst = []
    # try:
    with open(new_limit_file, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('item') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # except IOError as exc:
    #   print('ERROR:',exc)
    header_list = tmp_lst[0]
    print('header_list--->',header_list)
    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    # new_lsl=df.loc[df['item'] == 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_x', 'new_lsl'].tolist()
    # new_usl=df.loc[df['item'] == 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_x', 'new_usl'].tolist()
    data_df = pd.read_excel(new_limit_path+excel_report_name,names=None)
    # l= list(sheet.values())  
    item_list = data_df.loc[:, ['Item_name']].values.tolist()

    data = openpyxl.load_workbook(new_limit_path+excel_report_name)
    sheetnames = data.get_sheet_names()
    table = data.get_sheet_by_name(sheetnames[0])#[u'report', u'fail plot']
    table = data.active
    print('append new limit to-->',table.title) 
    nrows = table.max_row 
    ncolumns = table.max_column 
    print('nrows--->',nrows)
    print('ncolumns--->',ncolumns)
    i=1
    for x in item_list:
      # print('item_name-->',x[0])
      new_lsl=df.loc[df['item'] == x[0], 'new_lsl'].tolist()[0]
      new_usl=df.loc[df['item'] == x[0], 'new_usl'].tolist()[0]
      table.cell(i+1,14).value = new_lsl
      table.cell(i+1,15).value = new_usl
      i=i+1
    data.save(new_limit_path+excel_report_name)
    print('save new limit to excel report finished!')


def get_file_pic_name(file_dir):
  pic_file_l = []
  for root,dirs,files in os.walk(file_dir):
    # print(root)
    # print('fail picture path:',dirs)
    # print(files)
    for file in files:
      if os.path.splitext(file)[1] == '.png':
          # print('suffix name---->',os.path.splitext(file)[1])
          pic_file_l.insert(0,os.path.splitext(file)[0])
  return pic_file_l

def clear_files(path):
    for root,dirs,files in os.walk(path):
        for file in files:
            os.remove(path+'/'+file)
@count_time
def generate_keynote_report(project_code,station_name,build_stage,dir_path):
    keynote_title_name = project_code+'_'+station_name+'_'+build_stage+'_DATA Review'
    description_info =  'Issue description:\n'
    root_cause_info = 'Root Cause:\n'
    plan_info = 'Next steps:'
    print('dir_path--->',dir_path)
    keynote_save_path = dir_path+'cpk_report_'+datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')+'.key'
    overall_plot_path = dir_path + 'whole_cpk.png'

    doc,keynote = generate_keynote.create_keynote(keynote_title_name)
    print('save title page to keynote report finished!')

    generate_keynote.add_overall_pic(doc,keynote,overall_plot_path)
    print('save whole_cpk plot to keynote report finished!')


    fail_pic_path = dir_path+'fail_plot'
    # print('fail pic path---->',fail_pic_path)
    file_l=[]
    file_l = get_file_pic_name(fail_pic_path)
    # print('file_l --->',file_l)

    for f_name in file_l:
        one_fail_pic_path = fail_pic_path +'/'+f_name+'.png'
        print('fail pic name--->',one_fail_pic_path)
        generate_keynote.add_fail_pic(doc,keynote,one_fail_pic_path,description_info,root_cause_info,plan_info)

    generate_keynote.save_keynote(doc,keynote,keynote_save_path)
    print('save fail_plot to keynote report finished!')


def create_report(event,header_list,df,color_by1,pic_path,select_category1,cpk_lsl,cpk_usl,save_all_cpk_path,set_bins,excel_name,project_code,build_stage,station_name,start_time_first,start_time_last,color_by2,select_category_l2,excel_report_item,fail_plot_to_excel):
    clear_files(pic_path)
    book,report_sheet,plot_sheet,format_highlight,format_normal,format_titile,new_format_fail,new_format_pass = creat_excel_report_file(save_all_cpk_path,excel_name,cpk_lsl,cpk_usl,event,fail_plot_to_excel)
    # excel_report_data,excel_report_table,max_row,max_col = load_excel_table(save_all_cpk_path+excel_name)

    table_data,table_category_data = parse_all_csv(header_list,df,color_by1,select_category1,event,color_by2,select_category_l2)#
    i,j,n,t=0,0,0,0
    path=pic_path
    result='pass'
    # print('table_data length in create_report:',len(table_data),table_data)
    for column_data in table_data:
        # print('column_data length,column_data--->',len(column_data),column_data)
        item_name=column_data[0]
        usl = column_data[1]
        lsl = column_data[2]
        column_data = column_data[3:]
        # print('item_name:',item_name)
        # print('usl:',usl)
        # print('lsl:',lsl)
        # print(str(t)+' column test value:',len(column_data))
        if len(column_data) >4:   
            bc,p_val,a_Q,a_irr,three_CV = get_coefficients(column_data)
            row_data = []
            target_value = 9999999999
            mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)

            result=''

            if stdev == 0:
                n=n+1
                print(item_name+': do not record this item cpk value in excel and draw plot due to stdev == 0!')#stdev ==0
            else:

                location = 'A' + str(j+2)
                # print("location,j--->",location,j)

                # print('cpk:',cpk)
                if cpk == None:
                    result='Nan'
                    target_value = get_target_value(lsl,usl)
                    # print('target_value--->',lsl,usl,target_value)
                    if excel_report_item == 'all' or event == 'keynote-report':#'Fail'
                        row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,lsl,target_value, usl, min_num, mean, max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                        report_sheet.write_row(location,row_data,format_normal)
                    # if not os.path.exists(path):
                    #     os.makedirs(path)
                    # image_name = item_name.replace('/','_')+".png"
                    # pic_path = path + image_name
                    # if len(table_category_data) == 0:
                    #     draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                    # else:

                    #     #[[],[],[]...],only value category lists
                    #     draw_more_histogram(table_category_data[t],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                    # if i==0:
                    #     location='A1'
                    # else:
                    #     location = 'A'+str(29*i)
                    # save_image_to_excel(location,pic_path,plot_sheet)
                    # i = i + 1

                else:
                    if cpk < cpk_lsl or cpk > cpk_usl:
                        # print('----------------------cpk value fail--------------------------')
                        # print('cpk_lsl:',cpk_lsl)
                        # print('cpk_usl:',cpk_usl)
                        result='Fail'
                        target_value = get_target_value(lsl,usl)
                        # print('target_value--->',lsl,usl,target_value)
                        row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,lsl,target_value, usl, min_num, mean, max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                        report_sheet.write_row(location,row_data,format_normal)
                        report_sheet.write(j+1,17,result,new_format_fail)
                        if not os.path.exists(path):
                            os.makedirs(path)
                        image_name = item_name.replace('/','_')+".png"
                        pic_path = path + image_name
                        if len(table_category_data) == 0:
                            draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                        else:

                            #[[],[],[]...],only value category lists
                            draw_more_histogram(table_category_data[t],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                        if i==0:
                            location='A1'
                        else:
                            location = 'A'+str(29*i)
                        if fail_plot_to_excel == 'yes' and event =='excel-report':
                            save_image_to_excel(location,pic_path,plot_sheet)
                            i = i + 1
                    else:
                        if excel_report_item == 'all' or event == 'keynote-report':#'Fail'
                            result='Pass'
                            target_value = get_target_value(lsl,usl)
                            # print('target_value--->',lsl,usl,target_value)

                            row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,lsl,target_value,usl, min_num, mean, max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                            report_sheet.write_row(location,row_data,format_normal)
                            report_sheet.write(j+1,17,result,new_format_pass)
                        else:
                            pass




            #--------------------------------get new limit from csv start------------------------
            new_lsl,new_usl = get_one_item_new_limit_from_csv(save_all_cpk_path,item_name)
            new_lsl,new_usl = verify_limit(new_lsl,new_usl)
            old_mean, old_max_num,old_target_value,old_min_num, old_stdev,old_cpu, old_cpl, old_cpk,old_lsl,old_usl,old_result = mean, max_num,target_value,min_num, stdev, cpu, cpl, cpk,lsl,usl,result
            if new_lsl != None and new_lsl != None:
                lsl = new_lsl
                usl = new_usl
                # print('new lsl===>',lsl)
                # print('new usl===>',usl)
            #--------------------------------get new limit from csv end------------------------
                mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)
                if stdev == 0:
                    n=n+1
                    print(item_name+': do not record this item cpk value in excel and draw plot due to stdev == 0!(new limit)')#stdev ==0
                else:

                    location = 'A' + str(j+2)
                    # print("location,j--->",location,j)

                    # print('cpk:',cpk)
                    if cpk < cpk_lsl or cpk > cpk_usl:
                        print('----------------------cpk value fail--------------------------')
                        # print('cpk_lsl:',cpk_lsl)
                        # print('cpk_usl:',cpk_usl)
                        result='Fail'
                        target_value = get_target_value(lsl,usl)
                        # print('target_value--->',lsl,usl,target_value)
                        # append_new_limit_data_to_excel_report(excel_report_data,excel_report_table,save_all_cpk_path+excel_name,j+2,lsl,usl,cpl,cpu,cpk)
                        row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,old_lsl,old_target_value, old_usl, old_min_num, old_mean, old_max_num, old_stdev, old_cpl, old_cpu, old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                        report_sheet.write_row(location,row_data,format_normal)
                        if old_result == 'Fail':
                            report_sheet.write(j+1,17,old_result,new_format_fail)
                        elif old_result == 'Pass':
                            report_sheet.write(j+1,17,old_result,new_format_pass)
                        report_sheet.write(j+1,18,lsl,format_highlight)
                        report_sheet.write(j+1,19,target_value,format_highlight)
                        report_sheet.write(j+1,20,usl,format_highlight)
                        report_sheet.write(j+1,21,cpl,format_highlight)
                        report_sheet.write(j+1,22,cpu,format_highlight)
                        report_sheet.write(j+1,23,cpk,format_highlight)
                        report_sheet.write(j+1,24,result,new_format_fail)
                        if not os.path.exists(path):
                            os.makedirs(path)
                        image_name = item_name.replace('/','_')+" new_limit.png"
                        pic_path = path + image_name
                        if len(table_category_data) == 0:
                            draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                        else:

                            #[[],[],[]...],only value category lists
                            draw_more_histogram(table_category_data[t],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                        if i==0:
                            location='A1'
                        else:
                            location = 'A'+str(29*i)
                        if fail_plot_to_excel == 'yes' and event =='excel-report':
                            save_image_to_excel(location,pic_path,plot_sheet)
                            i = i + 1
                    else:
                        if excel_report_item == 'all' or event == 'keynote-report':#'Fail'
                            result='Pass'
                            target_value = get_target_value(lsl,usl)
                            # print('target_value--->',lsl,usl,target_value)
                            # append_new_limit_data_to_excel_report(excel_report_data,excel_report_table,save_all_cpk_path+excel_name,j+2,lsl,usl,cpl,cpu,cpk)

                            row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,old_lsl,old_target_value, old_usl, old_min_num, old_mean, old_max_num, old_stdev, old_cpl, old_cpu, old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                            report_sheet.write_row(location,row_data,format_normal)
                            if old_result == 'Fail':
                                report_sheet.write(j+1,17,old_result,new_format_fail)
                            elif old_result == 'Pass':
                                report_sheet.write(j+1,17,old_result,new_format_pass)
                            report_sheet.write(j+1,18,lsl,format_highlight)
                            report_sheet.write(j+1,19,target_value,format_highlight)
                            report_sheet.write(j+1,20,usl,format_highlight)
                            report_sheet.write(j+1,21,cpl,format_highlight)
                            report_sheet.write(j+1,22,cpu,format_highlight)
                            report_sheet.write(j+1,23,cpk,format_highlight)
                            report_sheet.write(j+1,24,result,new_format_pass)


            if stdev != 0 and excel_report_item == 'all' or stdev != 0 and excel_report_item == 'fail' and old_result =='Fail' or stdev != 0 and excel_report_item == 'fail' and result =='Fail' or stdev != 0 and event == 'keynote-report':
                j=j+1
            t = t +1
        else:
            t = t + 1
    print('Total {} items not calulate cpk'.format('%.0f' % n))
    book.close()
    print('All items cpk calulate/draw plots/excel report finished!')
    # save_new_limit_to_report(save_all_cpk_path,excel_name)
    if event == 'keynote-report':
        draw_overall_cpk(save_all_cpk_path,excel_name)
        generate_keynote_report(project_code,station_name,build_stage,save_all_cpk_path)

@count_time
def run(event,one_item_name,all_csv_path,cpk_path,cpk_lsl,cpk_usl,color_by1,selected_category_l1,color_by2,selected_category_l2,data_select,remove_fail,set_bins,new_x_lsl,new_x_usl,new_y_lsl,new_y_usl,excel_report_item,fail_plot_to_excel,excel_report_user,excel_report_stage):
    print('event--->',event,cpk_path,color_by1,selected_category_l1,color_by2,selected_category_l2)


    fail_pic_path =cpk_path+'fail_plot/'
    if event == 'one_item_plot':
        all_csv_path = cpk_path + 'temp/one_item.csv'


    if event != 'one_item_plot':
        start_time1 = datetime.datetime.now()  #
        header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last =  open_all_csv(event,all_csv_path,data_select,remove_fail)
        over_time1 = datetime.datetime.now()   #
        total_time1 = (over_time1-start_time1).total_seconds()
        print('open_all_csv  expend total time is: %s s' % total_time1)
    # print('header_list------>',len(header_list),header_list)
    if event == 'calculate-param':
       csv_path = cpk_path+'temp/calculate_param.csv'
       calulate_param(header_list,df,csv_path)
       print('['+str(datetime.datetime.now())+']','calulate param finished!')
       return
    # color_by_value,category_list = select_color_by(df)
    select_category_l1 =[]
    if color_by1 =='Off':
        select_category_l1 =[]
    else:
        select_category_l1 = select_category(selected_category_l1)
    print('select_category_l1------>',select_category_l1)

    select_category_l2 =[]
    if color_by2 =='Off':
        select_category_l2 =[]
    else:
        select_category_l2 = select_category(selected_category_l2)
    print('select_category_l2------>',select_category_l2)

    if event == 'one_item_plot':
        header_list,df,df_correlation,start_time_first,start_time_last,correlation_start_time_first,correlation_start_time_last =  open_one_item_csv(event,all_csv_path,data_select,remove_fail)

        if header_list[12] == header_list[13]:
            df = df.iloc[:, 0:13]
        one_item_head_list = [one_item_name]
        show_cpk_plot(event,one_item_head_list,df,color_by1,cpk_path,select_category_l1,set_bins,new_y_lsl,new_y_usl,start_time_first,start_time_last,color_by2,select_category_l2)
        if header_list[12] == header_list[13]:
            header_list = [one_item_name,one_item_name]
        else:
            header_list = [header_list[12],header_list[13]]
        if len(df_correlation.values.tolist())>5:
            show_correlation_plot(event,header_list,df_correlation,color_by1,cpk_path,select_category_l1,correlation_start_time_first,correlation_start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl,color_by2,select_category_l2)
            print('['+str(datetime.datetime.now())+']','show plot finished!')

    elif event == 'excel-report' or event == 'keynote-report':
        excel_report_file_name ='Limit_Update_'+str(excel_report_user)+'_'+project_code+'_For_'+excel_report_stage+'_'+datetime.datetime.now().strftime('%Y-%m-%d')+'.xlsx'#%Y-%m-%d_%H-%M-%S
        create_report(event,header_list,df,color_by1,fail_pic_path,select_category_l1,cpk_lsl,cpk_usl,cpk_path,set_bins,excel_report_file_name,project_code,build_stage,station_name,start_time_first,start_time_last,color_by2,select_category_l2,excel_report_item,fail_plot_to_excel)
        print('['+str(datetime.datetime.now())+']','create report finished!')
    else:
        print('param error!')


if __name__ == '__main__':
    '''
    
    
    '''
    #parameter list
    #----------------------------------------------------------------------------------
    # event = 'one_item_plot'#'excel-report'/'one_item_plot'/'keynote-report'
    # one_item_name = 'SystemState Wakeup_Voltage PPVDD_S1_SOC'
    # all_csv_path = '/Users/rex/Desktop/P1_Retest/cpk/222.csv'
    # cpk_path = "/Users/rex/PycharmProjects/my/"
    # cpk_lsl=1.33
    # cpk_usl=9999999999999999
    # color_by='Station ID'#'Off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'/'Product'/'StartTime'/'Special Build Description'
    # selected_category_l = ['CWNJ_C02-2FAP-24_2_FCT', 'CWNJ_C02-2FAP-24_1_FCT', 'CWNJ_C02-2F-REL01_1_FCT', 'CWNJ_C02-2FAP-23_2_FCT'] #[]
    # data_select = 'all'#first/last/no_retest/all --empty sn reminder only for
    # remove_fail = 'yes'
    # set_bins = 250
    # new_y_lsl,new_y_usl,new_x_lsl,new_x_usl = '','','',''
    #----------------------------------------------------------------------------------

    # print('color_by:', color_by)
    # print('cpk_lsl:', cpk_lsl)
    # print('cpk_usl:', cpk_usl)

    # excel_report_file_name ='cpk_report.xlsx'
    # pic_path =cpk_path+'fail/'
    # header_list,df =  open_all_csv(all_csv_path,data_select,remove_fail)
    # color_by_value,category_list = select_color_by(df)
    # if color_by_value =='Off':
    #     select_category_l =[]
    # else:
    #     select_category_l = select_category(category_list)
    # print('select_category_l------>',select_category_l)

    # if event == 'one_item_plot':
    #     header_list = [one_item_name]
    #     show_cpk_plot(event,header_list,df,color_by,cpk_path,select_category_l,set_bins)
    # elif event == 'cpk-report':
    #     create_report(header_list,df,color_by_value,pic_path,select_category_l,cpk_lsl,cpk_usl,cpk_path,set_bins,excel_report_file_name)
    # else:
    #     print('param error!')
    # run(event, one_item_name, all_csv_path, cpk_path, cpk_lsl, cpk_usl, color_by, selected_category_l, data_select,remove_fail,set_bins,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl)
    # draw_overall_cpk('/Users/rex/Desktop/CPK_Log/')
    #------------------------ debug cpk_plot --------------------------------
    # color list: ['#0000FF','#FF0000','#008000','#00FFFF','#9400D3','#8B008B','#B8860B','#FFA500','#A9A9A9','#FFFF00']

    # new_y_lsl = ''
    # new_y_usl = ''
    # new_x_lsl = ''
    # new_x_usl = ''
    # start_time_first = '2020/3/11 14:16'
    # start_time_last = '2020/3/26 9:58'
    # pic_path = '/Users/rex/Desktop/CPK_Log/'
    # correlation_plot(table_data,table_category_data,pic_path,start_time_first,start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl)

    #------------------------ debug correlation_plot end --------------------------------


    #------------------------ debug calulate param --------------------------------
    # header_list = [] 
    # df = #dataframe

    all_csv_path = '/Users/RyanGao/Desktop/cpk/cpk_data_0611/J5xx-FCT.csv'
    data_select = 'all'
    remove_fail = 'yes'
    header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last =  open_all_csv('calculate-param',all_csv_path,data_select,remove_fail)
    csv_path = '/Users/RyanGao/Desktop/CPK_Log/temp/calculate_param.csv'
    calulate_param(header_list,df,csv_path)

    #------------------------ debug calulate param end --------------------------------
