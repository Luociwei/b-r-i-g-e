#! /usr/bin/env python3
# --*-- coding: utf-8 ---*---


import sys,os,time,math,re
import datetime
start_time = datetime.datetime.now()  #
print('-------------- script update on 20200629 ------------')
show_log_flag = True

BASE_DIR=os.path.dirname(os.path.abspath(__file__))
print('BASE_DIR--->',BASE_DIR)
sys.path.insert(0,BASE_DIR+'/site-packages/')

print('python import ----> csv')
try:
    import csv
except Exception as e:
    print('e---->',e)

print('python import ---->matplotlib')
try:
    import matplotlib.pyplot as plt
except Exception as e:
    print('e---->',e)

print('python import ----> matplotlib.colors')
try:
    import matplotlib.colors as colors
except Exception as e:
    print('e---->',e)

print('python import ----> FontProperties')
try:
    from matplotlib.font_manager import FontProperties
except Exception as e:
    print('e---->',e)


print('python import ----> numpy')
try:
    import numpy as np
except Exception as e:
    print('e--->',e)

print('python import ----> pandas')
try:
    import pandas as pd
except Exception as e:
    print('e--->',e)

print('python import ----> openpyxl')
try:
    import openpyxl
except Exception as e:
  print('import openpyxl error:',e)
print('python import ----> xlsxwriter')
try:
    import xlsxwriter
except Exception as e:
  print('import xlsxwriter error:',e)


# print('python import ----> scipy')
# try:
#     from scipy import stats
# except Exception as e:
#     print('import scipy error:',e)


print('python import ----> diptest')
try:
    import diptest
except Exception as e:
    print('import diptest error:',e)

print(sys.getdefaultencoding())



current_dir = os.path.dirname(os.path.realpath(__file__))
keynote_lib_path = current_dir+'/python_keynote'
print('cpk file--->',keynote_lib_path)
# sys.path.append(keynote_lib_path)
# sys.path.insert(0,keynote_lib_path)
cmd = 'cd '+current_dir+'/python_keynote;export PYTHONPATH='+keynote_lib_path
# print('cpk file execute cmd--->',cmd)
print(os.system(cmd))
print('python import ----> python_keynote')
try:
    from python_keynote import generate_keynote
except Exception as e:
  print('e---->',e)



# vals =np.random.normal(5,3,30)#分布的均值,分布的标准差,个数
# print(vals)
over_time = datetime.datetime.now()   #
total_time = (over_time-start_time).total_seconds()
print('import expend total time is: %s s' % total_time)


def count_time(func):
    def int_time(*args, **kwargs):
        start_time = datetime.datetime.now()  #
        func(*args, **kwargs)
        over_time = datetime.datetime.now()   #
        total_time = (over_time-start_time).total_seconds()
        print('Function < '+func.__name__+' > expend total time is: %s s' % total_time)
    return int_time


def list_to_df(column_item_name,column_value_list):
    a = []
    for i in column_value_list:
        b = []
        b.append(i)
        a.append(b)
    #a-->[['a'], ['b'], ['x']]
    df = pd.DataFrame(columns=[column_item_name], data=a)
    print('df--->',df)
    return df
    

def show_log(message):
    if show_log_flag == True:
        print(message)
    return 
    


def cpk(vals,lower,upper=None):
    mean = np.mean(vals)
    std = np.std(vals)
    if upper:
        cpl = (mean -lower)/(3*std)
        cpu = (upper -mean)/(3*std)
        cpk = min(cpl,cpu)
        # print('Mean: {},Standard Deviation:{}'.format(mean,std))
        # print('Cpl:{},Cpu: {}'.format(cpl,cpu))
        # print('Cpk:{}'.format(cpk))
    else:
        cpk = abs(mean-lower)/(3*std)
        # print('Mean:{},Standard Deviation:{}'.format(mean,std))
        # print('Cpk:{}'.format(cpk))
# cpk(vals,3)




def cpk_calc(df_data,lsl,usl):
    """
    :param df_data: list
    :param usl: 数据指标上限
    :param lsl: 数据指标下限
    :return:
    """

    sigma = 3
    # print('limit--->',lsl,usl)
    # 数据平均值
    mean = np.mean(df_data)#
    # print('mean ---->',mean)

    # 数据max值
    max_num = max(df_data)
    # print('max_num ---->',max_num)

    # 数据min值
    min_num = min(df_data)
    # print('min_num ---->',min_num)

    # a = np.array([[1, 2], [3, 4]])
    # print('a--->', type(a))
    # print('gobal std:',np.std(a))#全局标准差
    # print('each line std:',np.std(a, axis=0,ddof=1))
    # print("each row std:",np.std(a, axis=1,ddof=1))

    # 数据标准差
    stdev = np.std(df_data,ddof=1,axis=0)
    # print('stdev ---->',stdev)
    if stdev == 0:#stop count cpk
        return (mean,max_num,min_num,stdev,None,None,None,None,None)
    # 生成横轴数据平均分布
    # x1 = np.linspace(mean - sigma * stdev, mean + sigma * stdev, 1000)
    # print('x1 ---->',x1)
    # 计算正态分布曲线
    # y1 = np.exp(-(x1 - mean) ** 2 / (2 * stdev ** 2)) / (math.sqrt(2 * math.pi) * stdev)
    # print('y1 ---->',y1)
    x1,y1 = None,None
    if lsl == 'NA' or usl == 'NA':
        return (mean,max_num,min_num,stdev,None,None,None,None,None)
    cpu = (usl - mean) / (sigma * stdev)
    cpl = (mean - lsl) / (sigma * stdev)
    # print('cpu ---->',cpu)
    # print('cpl ---->',cpl)

    # 得出cpk
    cpk = min(cpu, cpl)
    return (mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk)

def correlation_coefficient_calc(data_l1,data_l2,item_name1,item_name2):

    '''
    item_name1 : string
    item_name2 : string
    data_l2 :[]
    data_l1 :[]
    '''
    if item_name1 == item_name2:# can not be the same name.
       item_name2 = item_name2+'_2'
    data = pd.DataFrame({item_name1:data_l1,
                       item_name2:data_l2})
    # print('correlation data:',data)
    pearson = data.corr(method='pearson').values[0].tolist()[1]
    spearman = data.corr(method='spearman').values[0].tolist()[1]
    pearson = round(pearson,5)
    spearman = round(spearman,5)
    # print('pearson:',pearson)
    # print('spearman:',spearman)
    return pearson,spearman




def get_coefficients(value_l):
    '''
    param value_l: need float list
    return: bc,p_val,a_Q,a_irr,three_σ_x100_divide_mean
    1σ＝690000／1000000 #fault rate
    2σ＝308000／1000000
    3σ＝66800／1000000
    4σ＝6210／1000000
    5σ＝230／1000000
    6σ＝3.4／1000000 
    7σ＝0／1000000
    '''
    three_σ= 66800/1000000.0
    data = np.array(value_l)
    if len(data) <= 3:
        return '','','','',''
    try:
        dip, p_val = diptest.diptest(data)
    except Exception as e:
        print('calculate dip,p_val error:',e)
        return '','','','',''
    # print('dip,p_val:',dip,p_val)
    item_name ='value1'
    data = pd.DataFrame({item_name:value_l})
    # print('data--->',type(data),data)
    u1 = data[item_name].mean() # 计算均值
    # std1 = data[item_name].std() # 计算标准差
    # t,pval=stats.kstest(data[item_name], 'norm', (u1, std1))
    # print('normality test t,pval:',str(t),str(pval))

    # print('------')
    # 正态性检验 → pvalue >0.05

    n= float(len(value_l))
    #Item (xi-ẍ)^2
    item_l_1 =[]
    item_l_2 = []
    item_l_3 = []
    for i in value_l:
        temp1 = (i-u1)**2
        temp2 = (i-u1)**3
        temp3 = (i-u1)**4
        item_l_1.append(temp1)
        item_l_2.append(temp2)
        item_l_3.append(temp3)
    # print('item_l_1--->',item_l_1)
    # print('item_l_2--->',item_l_2)
    # print('item_l_3--->',item_l_3)
    sum_item_l_1 = sum(item_l_1)
    sum_item_l_2 = sum(item_l_2)
    sum_item_l_3 = sum(item_l_3)
    # print('sum_item_l_1',sum_item_l_1)
    # print('sum_item_l_2',sum_item_l_2)
    # print('sum_item_l_3',sum_item_l_3)
    if n<=3 or sum_item_l_1==0 or sum_item_l_2==0 or sum_item_l_3==0:
        # print('len < 3--->')
        if abs(u1) == 0:
            return 'Nan',round(p_val,6),'Nan','Nan','Nan'
        else:
            three_σ_x100_divide_mean = three_σ*100/abs(u1)
            # print('three_σ_x100_divide_mean:',three_σ_x100_divide_mean)
            return 'Nan',round(p_val,6),'Nan','Nan',round(three_σ_x100_divide_mean,6)
    else:
        try:
            m3 = np.sqrt(n*(n-1))/(n-2)*((1/n*sum_item_l_2)/np.sqrt(1/n*sum_item_l_1)**3)

            # print('m3=',m3)
            # print('d8:',n+1)
            # print('d15:',1/n*sum_item_l_3/(1/n*sum_item_l_1)**2)
            # print('d16:',(n+1)*1/n*sum_item_l_3/(1/n*sum_item_l_1)**2-3*(n-1))
            # print('d6/d7',(n-1)/((n-2)*(n-3)))
            m4 = ((n-1)/((n-2)*(n-3)))*((n+1)*1/n*sum_item_l_3/(1/n*sum_item_l_1)**2-3*(n-1))#(d6/d7)*d16
            # print('m4=',m4)
            #(d14**2+1)/(d17+3*(d10/d7))
            bc =(m3**2+1)/(m4+3*((n-1)**2/((n-2)*(n-3))))
            # print('bc:',bc)
            a_L=0.05
            a_M=0.1
            a_U=0.32
            a_Q = (a_U-a_L)*bc**2+a_L
            # print ('a_Q:',a_Q)
            a_irr = np.sqrt((a_U-a_L)**2*bc)+a_L
            # print('a_irr:',a_irr)
        except Exception as e:
            # print('calculate error',e)
            if abs(u1) == 0:
                return 'Nan',round(p_val,6),'Nan','Nan','Nan'
            else:
                three_σ_x100_divide_mean = three_σ*100/abs(u1)
                # print('three_σ_x100_divide_mean:',three_σ_x100_divide_mean)
                return 'Nan',round(p_val,6),'Nan','Nan',round(three_σ_x100_divide_mean,6)


    if abs(u1) == 0:
        return round(bc,6),round(p_val,6),round(a_Q,6),round(a_irr,6),'Nan'
    else:
        three_σ_x100_divide_mean = three_σ*100/abs(u1)
        # print('three_σ_x100_divide_mean:',three_σ_x100_divide_mean)
        return round(bc,6),round(p_val,6),round(a_Q,6),round(a_irr,6),round(three_σ_x100_divide_mean,6)





def write_row_data_to_excel(location,row_data,format,sheet_name):
    sheet_name.write_row(location, row_data, format)
    # report_sheet.write_row('A1', title, format_titile)

def write_column_data_to_excel(location, column_data, format,sheet_name):
    sheet_name.write_column(location, column_data, format)




# 按照固定区间长度绘制频率分布直方图
# bins_interval 区间的长度
# margin        设定的左边和右边空留的大小

def probability_distribution(data, bins_interval=1, margin=1):
    bins = range(int(min(data)), int(max(data)) + bins_interval - 1, bins_interval)
    print(len(bins))
    for i in range(0, len(bins)):
        print(bins[i])
    plt.xlim(min(data) - margin, max(data) + margin)
    plt.title("Probability-distribution")
    plt.xlabel('Interval')
    plt.ylabel('Probability')
    # 频率分布normed=True，频次分布normed=False
    prob,left,rectangle = plt.hist(x=data, bins=bins, density=False, histtype='bar', color=['r'])
    for x, y in zip(left, prob):
        # 字体上边文字
        # 频率分布数据 normed=True
        # plt.text(x + bins_interval / 2, y + 0.003, '%.2f' % y, ha='center', va='top')
        # 频次分布数据 normed=False
        plt.text(x + bins_interval / 2, y + 0.25, '%.2f' % y, ha='center', va='top')
    plt.show()







def probability_distribution_extend(data,bins,margin,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,bins_l,bins_h,start_time_first,start_time_last):
  bins = sorted(bins)
  length = len(bins)
  intervals = np.zeros(length+1)
  for value in data:
    i = 0
    while i < length and value >= bins[i]:
      i += 1
    intervals[i] += 1
  intervals = intervals / float(len(data))
  plt.ion()  # 开启interactive mode

  plt.figure(1)  # 创建图表1
  plt.xlim(min(bins) - margin, max(bins) + margin)
  bins.insert(0, -999)
  plt.title("probability-distribution")
  plt.bar(bins, intervals,color=['r'], label='')#频率分布
  x_ticks,labels = plt.xticks()
  x_ticks_start=round(x_ticks[0],2)
  x_ticks_end = round(x_ticks[len(x_ticks) - 1],2)
  y_ticks, labels = plt.yticks()
  # print('y_ticks--->',y_ticks,y_ticks[len(y_ticks) - 1],y_ticks[0])
  y_ticks=round(y_ticks[len(y_ticks) - 1],5)
  # print("x_ticks_start,x_ticks_end--->",x_ticks_start,x_ticks_end)
  # print('y_ticks_end--->',y_ticks)
  # plt.show()
  plt.close(1)


  plt.figure(2,dpi=150)  # 创建图表2

  if cpl ==None:
    cpl_value =''
  else:
    cpl_value = str("%.3f" % cpl)

  if cpu ==None:
    cpu_value =''
  else:
    cpu_value = str("%.3f" % cpu)

  if cpk ==None:
    cpk_value =''
  else:
    cpk_value = str("%.3f" % cpk)

  info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
      "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + cpl_value + '\n' + "Cpu: " + cpu_value + '\n' + "Cpk: " + cpk_value
  # info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
  #     "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + str(
  #     "%.3f" % cpl) + '\n' + "Cpu: " + str("%.3f" % cpu) + '\n' + "Cpk: " + str("%.3f" % cpk)
  if len(item_name) > 55:
      item_name = item_name[0:55] + '\n' + item_name[55:]
  # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
  # plt.title(item_name,FontProperties=font)
  plt.title(item_name)
  plt.xlabel(str(start_time_first)+' -- '+str(start_time_last))
  plt.ylabel('Count')
  # plt.title(item_name+"\nCpk={0}".format(str("%.6f" % cpk)))
  # plt.hist(x=data, bins=bins, density=False, histtype='bar', color=['r'])
  bins = [round(x,5) for x in bins]
  bins=sorted(bins)
  # print('----->bins-->',bins)
  plt.hist(data, bins=bins, label=info, histtype='stepfilled',color = 'blue', edgecolor='blue', linewidth=1.5,align='mid',density=False) #time分布
  
  range = get_limit_range(bins_l, bins_h)
  range =round(range/5,5)
  # print('plot bar--->',bins_l,range)
  plt.xlim(bins_l-range, bins_h+range)

  y_ticks = len(data) * y_ticks
  plt.ylim((0, y_ticks))  # 设置y轴scopex
  ax=plt.gca()
  ax.spines['bottom'].set_linewidth(1.5)
  ax.spines['left'].set_linewidth(1.5)
  ax.spines['right'].set_linewidth(1.5)
  ax.spines['top'].set_linewidth(1.5)
  # plt.plot(x1, y1, 'k--', label="", linewidth=1.0, color='lime')  # 画正态分布曲线
  if lsl !='NA' and usl != 'NA':
      plt.plot([lsl, lsl, ], [0, 100000, ], 'k--', linewidth=3.0, color='red')  # 画lower limit线，
      plt.plot([usl, usl, ], [0, 100000, ], 'k--', linewidth=3.0, color='red')  # 画upper limit线，

      # 添加文字
      # plt.text(0.1,20, r'$\mu=100,\ \sigma=15$')
      plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 10, 'color': 'r'})
      plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 10, 'color': 'r'})
      # plt.axis([-0.2, 0.2,0, 100])#x 轴，y 轴

  #os.system('mkdir fail')
  plt.legend(loc="upper right",framealpha=0.85,edgecolor='royalblue',borderaxespad=0.3)
  plt.grid(linestyle=':',c='gray')  # 生成网格
  # path="/Users/rex/PycharmProjects/my/fail/"

  plt.savefig(pic_path,dpi=200)
  plt.draw()

  # plt.show()
  plt.close()

  plt.ioff()  






def probability_distribution_extend_by_color(column_category_data_list,data,bins,margin,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,bins_l,bins_h,start_time_first,start_time_last):
  # print('one column data len:',len(data),data)
  bins = sorted(bins)
  length = len(bins)
  intervals = np.zeros(length+1)
  for value in data:
    i = 0
    while i < length and value >= bins[i]:
      i += 1
    intervals[i] += 1
  intervals = intervals / float(len(data))
  plt.ion()  # 开启interactive mode

  plt.figure(1)  # 创建图表1
  plt.xlim(min(bins) - margin, max(bins) + margin)
  bins.insert(0, -999)
  plt.title("probability-distribution")
  plt.bar(bins, intervals,color=['r'], label='')#频率分布
  x_ticks,labels = plt.xticks()
  x_ticks_start=round(x_ticks[0],2)
  x_ticks_end = round(x_ticks[len(x_ticks) - 1],2)
  y_ticks, labels = plt.yticks()
  # print('y_ticks--->',y_ticks,y_ticks[len(y_ticks) - 1],y_ticks[0])
  y_ticks=round(y_ticks[len(y_ticks) - 1],5)
  # print("x_ticks_start,x_ticks_end--->",x_ticks_start,x_ticks_end)
  # print('y_ticks_end--->',y_ticks)
  # plt.show()
  plt.close(1)


  plt.figure(2,dpi=150)  # 创建图表2,facecolor='blue',edgecolor='black'

  if cpl ==None:
    cpl_value =''
  else:
    cpl_value = str("%.3f" % cpl)

  if cpu ==None:
    cpu_value =''
  else:
    cpu_value = str("%.3f" % cpu)

  if cpk ==None:
    cpk_value =''
  else:
    cpk_value = str("%.3f" % cpk)

  info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
      "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + cpl_value + '\n' + "Cpu: " + cpu_value + '\n' + "Cpk: " + cpk_value
  


  # info = "Sample: " + str("%.f" % len(data)) + '\n' +"Max: " + str("%.3f" % max_num) + '\n' + "Mean: " + str("%.3f" % mean) + '\n' + "Min: " + str(
  #     "%.3f" % min_num) + '\n' + "Std: " + str("%.3f" % stdev) + '\n' + "Cpl: " + str(
  #     "%.3f" % cpl) + '\n' + "Cpu: " + str("%.3f" % cpu) + '\n' + "Cpk: " + str("%.3f" % cpk)
  if len(item_name) > 55:
      item_name = item_name[0:55] + '\n' + item_name[55:]
  # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
  plt.title(item_name)
  plt.xlabel(str(start_time_first)+' -- '+str(start_time_last))
  plt.ylabel('Count')

  # plt.title(item_name+"\nCpk={0}".format(str("%.6f" % cpk)))
  # plt.hist(x=data, bins=bins, density=False, histtype='bar', color=['r'])
  bins = [round(x,5) for x in bins]
  bins=sorted(bins)
  # print('----->bins-->',bins)
  # print(' more draw category data:',len(column_category_data_list),column_category_data_list)
  l=0
  l_len=[]
  for category_data in column_category_data_list:
      # print('category_test_data--->', category_data[5:])
      # print('category_name,category_color--->', category_data[0],str(category_data[1]))
      category_name,category_color = category_data[0],str(category_data[1])
      n=len(category_data[5:])
      l=l+n
      # print('category len:',n)
      l_len.append(n)
      if len(column_category_data_list) == 1:
          plt.hist(category_data[5:], bins=bins, label=category_data[0], color=category_color ,histtype='stepfilled',edgecolor=category_color,linewidth=1.5,align='mid',density=False) #time分布
      else:
          plt.hist(category_data[5:], bins=bins, label=category_data[0], color='white' ,histtype='step',edgecolor=category_color,linewidth=1.5,align='mid',density=False) #time分布
  # print('----->a column len,max in one category:-->',l,max(l_len))

 
  y_ticks = max(l_len) * (y_ticks+0.04)
  range = get_limit_range(bins_l, bins_h)
  range =round(range/5,5)
  # print('plot bar--->',bins_l,range)
  plt.xlim(bins_l-range, bins_h+range)
  plt.ylim((0, y_ticks))  # 设置y轴scopex

  ax=plt.gca()
  ax.spines['bottom'].set_linewidth(1.5)
  ax.spines['left'].set_linewidth(1.5)
  ax.spines['right'].set_linewidth(1.5)
  ax.spines['top'].set_linewidth(1.5)

  # plt.plot(x1, y1, 'k--', label="", linewidth=1.0, color='lime')  # 画正态分布曲线
  if lsl !='NA' and usl != 'NA':
      plt.plot([lsl, lsl, ], [0, 1000000, ], 'k--', linewidth=3.0, color='red')  # 画lower limit线，
      plt.plot([usl, usl, ], [0, 1000000, ], 'k--', linewidth=3.0, color='red')  # 画upper limit线，
      # 添加文字
      # plt.text(0.1,20, r'$\mu=100,\ \sigma=15$')
      plt.text(lsl, y_ticks / 3, ' LSL\n' + ' ' + str(lsl), fontdict={'size': 10, 'color': 'r'})
      plt.text(usl, y_ticks / 2, ' USL\n' + ' ' + str(usl), fontdict={'size': 10, 'color': 'r'})
      # plt.axis([-0.2, 0.2,0, 100])#x 轴，y 轴

  plt.text(bins_l+range/3, y_ticks*0.78, info, size=10, rotation=0.0, alpha=0.85,fontsize=8,ha="left",
           va="center",bbox=dict(boxstyle="round", ec=('royalblue'),linestyle='-.',lw=1, fc=('white'), ))



  #os.system('mkdir fail')
  if len(column_category_data_list) < 24:    
      plt.legend(loc="upper right",framealpha=0.85,edgecolor='royalblue',borderaxespad=0.3,fontsize=8)
  # plt.legend(bbox_to_anchor=(1.01, 1), loc=2, borderaxespad=0)

  plt.grid(linestyle=':',c='gray')  # 生成网格
  # path="/Users/rex/PycharmProjects/my/fail/"

  plt.savefig(pic_path,dpi=200)
  plt.draw()

  # plt.show()
  plt.close()

  plt.ioff()  



def get_bins(min_num,max_num,lsl,usl,set_bins):
    # print('min_num,max_num,lsl,usl,set_bins=====>',min_num,max_num,lsl,usl,set_bins)

    bins_l = 0
    bins_h = 0
    if lsl == 'NA' or usl == 'NA':
        bins_l,bins_h = min_num,max_num

    else:
        if min_num < lsl and max_num < lsl:
            bins_l = min_num
            bins_h = usl
        elif min_num < lsl and max_num > lsl and max_num <= usl:
            bins_l = min_num
            bins_h = usl
        elif min_num < lsl and max_num > usl:
            bins_l = min_num
            bins_h = max_num
        elif lsl <= min_num and min_num <= usl and max_num <= usl:
            bins_l = lsl
            bins_h = usl
        elif lsl <= min_num and min_num <= usl and max_num > usl:
            bins_l = lsl
            bins_h = max_num
        elif min_num > usl:
            bins_l = lsl
            bins_h = max_num
    range = get_limit_range(bins_l,bins_h)
    range = round((range/set_bins),5)
    bins = np.arange(bins_l, bins_h, range)#必须是单调递增的
    # print('lsl,usl,min_num,max_num,bins_l,bins_h in get_bins=====>',lsl,usl,min_num,max_num,bins_l,bins_h)
    return bins,bins_l,bins_h




def draw_histogram(column_data,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,set_bins,start_time_first,start_time_last):


    bins,bins_l,bins_h = get_bins(min_num,max_num,lsl,usl,set_bins)
    # print(len(column_data),'---->',min(column_data),max(column_data),bins)
    probability_distribution_extend(column_data,bins,0,item_name,lsl,usl,mean,max_num,min_num,stdev,x1,y1,cpu,cpl,cpk,pic_path,bins_l,bins_h,start_time_first,start_time_last)

    return True





def draw_more_histogram(column_category_data_list,column_data, item_name, lsl, usl, mean, max_num, min_num, stdev, x1, y1,cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last):
    """

    """

    # range = get_limit_range(lsl, usl)
    # range = round((range /set_bins), 5)
    # bins = np.arange(lsl, usl, range)  # 必须是单调递增的
    bins,bins_l,bins_h = get_bins(min_num,max_num,lsl,usl,set_bins)

    # print('9999 column_category_data_list--->',column_category_data_list)
    # print('9999 column_data--->',len(column_data),column_data)
    # print('9999 bins len--->',len(bins))

    # print(len(column_data),'---->',min(column_data),max(column_data),bins)
    probability_distribution_extend_by_color(column_category_data_list,column_data,bins, 0, item_name, lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,bins_l,bins_h,start_time_first,start_time_last)

    return True





def save_image_to_excel(location,pic_path,plot_sheet):
    #save picture to excel
    plot_sheet.insert_image(location,pic_path,{'x_scale': 1.2, 'y_scale': 1.2})
    return True

def list_category(df,color_by_value):
    column_list = df[color_by_value].tolist()
    column_list = column_list[3:]
    category_list = [x for x in set(column_list)]  # index:[column_list.index(x) for x in set(column_list)]
    # UI control = category_list
    return category_list # ALL category




def select_category(category_l):#action

    get_category_list  =  []#UI control
    color_l = ['#0000FF','#FF0000','#008000','#00FFFF','#9400D3','#8B008B','#B8860B','#FFA500','#A9A9A9','#FFFF00']
    i=0
    for category_value in category_l:
        temp_list=[]
        # c=colors.cnames.values()[i]
        # print('c====>',c)
        temp_list.insert(0,category_value)
        temp_list.insert(1,color_l[i])
        # color_len = len(colors.cnames.values())
        # print('len(colors.cnames.values() ===>',color_len)
        i = i + 1
        if i == 10 :#
            i=0
        get_category_list.append(temp_list)

    # print('get_category_list====>',get_category_list)
    return get_category_list# selected category,including color_set_list



def select_color_by(df):
    color_by_value ='Station ID'#'off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'---From UI
    if color_by_value =='off':
        category_list =None
    else:
        category_list = list_category(df,color_by_value)
    return color_by_value, category_list



def get_project_info(data_df):
    # print('one row data:--->',data_df[1:2].values[0].tolist(),len(data_df[1:2].values[0].tolist()))
    project_code = data_df[1:2].values[0].tolist()[1] #Product first cell
    special_build_name  = data_df['Special Build Name'].values.tolist() #Special Build Name
    station_name = data_df[1:2].values[0].tolist()[6] #Station ID first cell
    print('station_name-->',type(station_name),station_name)
    # pattern =re.compile(r'.*\-(.+)\-')
    # result=pattern.match(build_stage)
    # build_stage = result.group(1)
    build_stage =''
    build_stage_l = list(set(special_build_name))
    print('build_stage_l-->',build_stage_l)
    
    temp_l=[]
    for i in build_stage_l:
        pattern = re.compile(r'.*\-(.+)')
        result = pattern.match(i)
        temp_l.append(result.group(1))
    n=1 
    for j in list(set(temp_l)):
        if n==1:
            build_stage = j
        else:
          build_stage = build_stage +'/'+j
        n=n+1

    if station_name != '':
        pattern =re.compile(r'.*\_([0-9]+)\_(\D+)')
        result=pattern.match(station_name)
        try:
            station_name = result.group(2)
        except Exception as e:
            print('station_name error')
            station_name ='FCT'
    else:
        station_name ='FCT'
    print('--------------------project_code:'+project_code+'------------------------')
    print('--------------------build_stage:'+build_stage+'-------------------------')
    print('--------------------station_name:'+station_name+'------------------------')

    return project_code,build_stage,station_name





def open_one_item_csv(event,all_csv_path,data_select,remove_fail):

    tmp_lst = []
    with open(all_csv_path, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('fct') != -1:
                # print("FW version---->")
                pass
            elif row[0].lower().find('display name') != -1:
                pass
            elif row[0].lower().find('pdca priority') != -1:
                pass
            elif row[0].lower().find('upper limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('lower limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('measurement unit') != -1:  # "Measurement Unit ----->" in row:
                pass
            elif row[0].lower().find('site') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # print("index---->", tmp_lst[0])
    header_list = tmp_lst[0][:]#
    temp_header_list = tmp_lst[0]


    if header_list[12] == header_list[13]:
        temp_header_list[13] = temp_header_list[13]+'_2'
        temp_df = pd.DataFrame(tmp_lst[1:], columns=temp_header_list)
        correlation_header_df =temp_df[0:2]



    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    header_df =df[0:2]
    # print('header_df before--->', header_df)
    data_df = df[2:]
    correlation_data_df = data_df
    # print('data_df before--->', data_df)
    
    # print('99999888--->',type(correlation_data_df.columns),correlation_data_df.columns[12],correlation_data_df.columns[13])
    if correlation_data_df.columns[12] != correlation_data_df.columns[13]:
        print(' one_item_plot before remove empty--->', len(correlation_data_df.values.tolist()))
        correlation_data_df=correlation_data_df[~correlation_data_df[correlation_data_df.columns[12]].isin([''])]#Remove SN Empty
        correlation_data_df=correlation_data_df[~correlation_data_df[correlation_data_df.columns[13]].isin([''])]#Remove SN Empty
        print(' one_item_plot after remove empty--->', len(correlation_data_df.values.tolist()))


    print('csv data row number before remove SN empty--->', len(correlation_data_df.values.tolist()))
    correlation_data_df=correlation_data_df[~correlation_data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(correlation_data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(correlation_data_df.values.tolist()))
    if remove_fail== 'yes':
        correlation_data_df=correlation_data_df[correlation_data_df['Test Pass/Fail Status'].isin(['PASS'])]
        # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(correlation_data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(correlation_data_df.values.tolist()))
    if data_select == 'first':
        correlation_data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        correlation_data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        correlation_data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(correlation_data_df.values.tolist()))
    correlation_start_time_first,correlation_start_time_last = '',''
    if len(correlation_data_df.values.tolist()) >2:
        correlation_start_time_l  = correlation_data_df['StartTime'].values.tolist() #StartTime
        correlation_start_time_first = min(correlation_start_time_l)
        correlation_start_time_last  = max(correlation_start_time_l)
        print('<correlation first time -- last time>',correlation_start_time_first,correlation_start_time_last)

    if header_list[12] == header_list[13]:
        df_correlation = correlation_header_df.append(correlation_data_df)
    else:
        df_correlation = header_df.append(correlation_data_df)






    print('csv data row number before remove SN empty--->', len(data_df.values.tolist()))
    data_df=data_df[~data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(data_df.values.tolist()))
    if remove_fail== 'yes':
        data_df=data_df[data_df['Test Pass/Fail Status'].isin(['PASS'])]
        # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(data_df.values.tolist()))
    if data_select == 'first':
        data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(data_df.values.tolist()))

    start_time_l  = data_df['StartTime'].values.tolist() #StartTime
    start_time_first = min(start_time_l)
    start_time_last  = max(start_time_l)
    print('<first time -- last time>',start_time_first,start_time_last)

    if header_list[12] == header_list[13]:
        df = correlation_header_df.append(data_df)
    else:
        df = header_df.append(data_df)

    # print("index1---->", header_list)

    # print('df after--->',df_correlation.columns.values.tolist(), df._stat_axis.values.tolist(),df)
    # print('df.values ---->', df.values)#array([[ ]])
    return header_list,df,df_correlation,start_time_first,start_time_last,correlation_start_time_first,correlation_start_time_last








#
#header_list:A list of string types
#df,Dataframe
#

def open_all_csv(event,all_csv_path,data_select,remove_fail):

    tmp_lst = []
    with open(all_csv_path, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('fct') != -1:
                # print("FW version---->")
                pass
            elif row[0].lower().find('display name') != -1:
                pass
            elif row[0].lower().find('pdca priority') != -1:
                pass
            elif row[0].lower().find('upper limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('lower limit') != -1:
                tmp_lst.append(row)
            elif row[0].lower().find('measurement unit') != -1:  # "Measurement Unit ----->" in row:
                pass
            elif row[0].lower().find('site') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # print("index---->", tmp_lst[0])
    header_list = tmp_lst[0]
  



    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    header_df =df[0:2]
    # print('header_df before--->', header_df)
    data_df = df[2:]
    # print('data_df before--->', data_df)
 

    print('csv data row number before remove SN empty--->', len(data_df.values.tolist()))
    data_df=data_df[~data_df['SerialNumber'].isin([''])]#Remove SN Empty
    print('csv data row number after remove SN empty--->', len(data_df.values.tolist()))
    print('csv data row number before remove fail--->', len(data_df.values.tolist()))
    if remove_fail== 'yes':
        data_df=data_df[data_df['Test Pass/Fail Status'].isin(['PASS'])]
		    # data_df=data_df[~data_df['Test Pass/Fail Status'].isin(['FAIL'])]
    print('csv data row number after remove fail--->', len(data_df.values.tolist()))
    print('csv data row number before remove retest--->', len(data_df.values.tolist()))
    if data_select == 'first':
        data_df.drop_duplicates(['SerialNumber'],keep='first',inplace=True)
    elif data_select == 'last':
        data_df.drop_duplicates(['SerialNumber'],keep='last',inplace=True)
    elif data_select == 'no_retest':
        data_df.drop_duplicates(['SerialNumber'],keep=False,inplace=True)
    elif data_select == 'all':
        pass
    print('csv data row number after remove retest--->', len(data_df.values.tolist()))

    if event == 'keynote-report':
        project_code,build_stage,station_name = get_project_info(data_df)
    else:
        project_code,build_stage,station_name = '','',''
    start_time_l  = data_df['StartTime'].values.tolist() #StartTime
    start_time_first = min(start_time_l)
    start_time_last  = max(start_time_l)
    print('<first time -- last time>',start_time_first,start_time_last)

    df = header_df.append(data_df)

 
    if event != 'one_item_plot':
        station_id_l = df['Station ID'].values.tolist()
        fixture_channel_id = df['Fixture Channel ID'].values.tolist()
        # print('station_id_l:',station_id_l)
        # print('fixture_channel_id:',fixture_channel_id)
        temp_l = []
        for i in range(0,len(station_id_l[2:])):
            temp_l.append(station_id_l[i+2]+'_'+fixture_channel_id[i+2])
        temp_l.insert(0,'')
        temp_l.insert(0,'')
        fixture_channel_id = temp_l
        # print(fixture_channel_id)
        df['Fixture Channel ID'] = pd.DataFrame({'Fixture Channel ID':fixture_channel_id})
        # print('fixture channel id:',df['Fixture Channel ID'].values.tolist())

        print('*******************************')


    # print('df after--->', df)
    # print('df.values ---->', df.values)#array([[ ]])
    return header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last

def is_number(num):
    pattern = re.compile(r'(.*)\.(.*)\.(.*)')
    if pattern.match(num):
        return False
    return num.replace(".", "").replace('-','').isdigit()



def valid_column(test_item_name,df):

    column_list = df[test_item_name].tolist()
    # print('valid_column item_name-->',test_item_name)
    # print('valid_column--->',len(column_list),column_list)

    # print('column_value_list--->',column_list)
    if test_item_name.lower().find('_cb_count') == -1 and test_item_name.lower().find('fixture channel id') == -1 and test_item_name.lower().find('ss1_num_tri') == -1 and test_item_name.lower().find('fixture vendor_id') == -1:
        pass
    else:
        return 'not_cpk state1'
    if column_list[0] == '0' and column_list[1] == '0' or column_list[0] == '1' and column_list[
        1] == '1' or  column_list[0] == column_list[1] and  column_list[0] !='NA' or len(column_list)< 5:
        # print('====>',column_list[0],column_list[1])
        return "not_cpk state2"
    else:
        # pattern = re.compile(r'^[+-]?[0-9]*\.?[0-9]+$')
        # print('valid_column len----->',test_item_name,len(column_list))
        # print('v====>',column_list[0],column_list[1],column_list,len(column_list))
        j=0
        for i in range(2,len(column_list),1):
            # print(str(i),'valid_column----->',test_item_name,pattern.match(column_list[i]))
            if is_number(column_list[i]):
                j=j+1
                # print(str(j)+',len(set(column_list))-->',j,len(set(column_list)))
                if j>3 and len(set(column_list))>=3:
                    # print('need_cpk')
                    return "need_cpk"
 
        return "not_cpk state3"


def test_value_to_numeric(test_data_list):
    column_list = []
    i = 0
    for x in test_data_list:
        if i ==0 and x == 'NA' or i ==1 and x == 'NA':
            column_list.append(x)
        else:
            try:
              x = float(x)
              column_list.append(x)
            except Exception as e:
                pass
            # print('-------------------- it is not number --------------')
        i = i + 1

    # print('column_list--->',column_list)
    return column_list



def get_target_value(lsl,usl):
    if lsl != 'NA' or usl != 'NA':
        target_value = round(((lsl + usl) / 2.0), 5)
    else:
        target_value = 'Nan'
    return target_value


def get_limit_range(lsl,usl):
    # print('lsl,usl----->', lsl, usl)
    range = 0
    if lsl < 0 and usl <= 0:
        range = abs(lsl) - abs(usl)
    elif lsl < 0 and usl >= 0:
        range = abs(lsl) + usl
    elif lsl >= 0 and (usl > 0):
        range = usl - lsl
    else:
        print('get_limit_range 00000')
    range = round(range, 5)
    # print('range in get_limit_range----->', range)
    return range


def get_ticks(min_num,max_num,lsl,usl):
    #for correlation plot
    # print('min_num,max_num,lsl,usl=====>',min_num,max_num,lsl,usl)

    ticks_l = 0
    ticks_h = 0
    if lsl =='NA' or usl =='NA':
        ticks_l = min_num
        ticks_h = max_num
    else:
        if min_num < lsl and max_num < lsl:
            ticks_l = min_num
            ticks_h = usl
        elif min_num < lsl and max_num > lsl and max_num <= usl:
            ticks_l = min_num
            ticks_h = usl
        elif min_num < lsl and max_num > usl:
            ticks_l = min_num
            ticks_h = max_num
        elif lsl <= min_num and min_num <= usl and max_num <= usl:
            ticks_l = lsl
            ticks_h = usl
        elif lsl <= min_num and min_num <= usl and max_num > usl:
            ticks_l = lsl
            ticks_h = max_num
        elif min_num > usl:
            ticks_l = lsl
            ticks_h = max_num
    # print('ticks_l,ticks_h:',ticks_l,ticks_h)
    range = get_limit_range(ticks_l,ticks_h)
    # print('ticks_range:',range,round((ticks_l-range/5),2),round((ticks_h+range/5),2))

    ticks_l,ticks_h = round((ticks_l-range/5),2),round((ticks_h+range/5),2)
    return ticks_l,ticks_h



'''
color_by:  Special Build Name,Version,Station ID,SerialNumber
select_category : category list
return table_data,table_category_data;table_data for cpk,table_category_data for draw plot
'''

def parse_all_csv(header_list,df,color_by,selected_category,event):
    # print('header_list:',type(header_list),header_list,color_by,selected_category)
    
    table_data = []#[[]]
    table_category_data = []#[[[]]]
    column_list = []
    n=0
    for item_name in header_list:
        # if event == 'one_item_plot1':
        #     # print('one column--->',type(df.iloc[:, [n]].values),df.iloc[:, [n]].values.tolist())
        #     temp_l=[]
        #     for i in df.iloc[:, [n]].values.tolist():
        #         temp_l.append(i[0])
        #     column_list = temp_l  # 读取指定键值列的所有行
        #     print('column_list--->',column_list)
        #     n=n+1
        # else:
        # print('item_name:',item_name)
        # print('header_list:',header_list)
        # print('df:',df)
        column_list = df[item_name].tolist()
        # print('column_list--->',column_list)

        need_cpk = valid_column(item_name,df)
        # print('need_cpk:---->',need_cpk)
        column_num_list = []
        if need_cpk == 'need_cpk':
            column_list = test_value_to_numeric(column_list)
            column_list.insert(0, item_name)  # item name
            # 'Off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'/'Product'/'StartTime'/'Special Build Description'
            if color_by == 'SerialNumber' or color_by == 'Version' or color_by == 'Station ID' or color_by == 'Special Build Name' or color_by == 'Product' or color_by == 'StartTime' or color_by =='Special Build Description' or color_by =='Fixture Channel ID':
                column_temp = []
                i =0
                for x in selected_category:
                    # print('xxxxx:', x[0],x[1])
                    one_category_list = df.loc[df[color_by] == x[0], item_name].tolist()  # 选取gender列是M，name列的数据
                    # print('one category_data-->', one_category_list)
                    one_category_list = test_value_to_numeric(one_category_list)
                    column_temp = column_temp +one_category_list

                    usl = column_list[1]
                    lsl = column_list[2]
                    # print('usl--->', usl)
                    # print('lsl--->', lsl)
                    one_category_list.insert(0, x[0])  # insert category
                    one_category_list.insert(1, x[1])  # insert color
                    one_category_list.insert(2,item_name)  # item_name
                    if lsl != 'NA' or usl != 'NA':
                        usl = float(usl)
                        lsl = float(lsl)

                    one_category_list.insert(3, usl)  # usl
                    one_category_list.insert(4, lsl)  # lsl
                    # print('one_category_list-->', one_category_list)
                    column_num_list.append(one_category_list)
                    i = i+1

                column_temp.insert(0,item_name)  # item_name
                if lsl != 'NA' or usl != 'NA':
                    usl = float(usl)
                    lsl = float(lsl)
                column_temp.insert(1, usl)  # usl
                column_temp.insert(2, lsl)  # lsl
                # print('column_temp-->', column_temp)
                table_data.append(column_temp)# one column's all category data []
                # print('column_num_list-->', column_num_list)
                table_category_data.append(column_num_list)# category [[],[],...]]
            elif color_by == 'Off':
                table_data.append(column_list)  # item_name,usl,lsl,data
        else:
            pass
    # print('table_data-->',table_data,table_category_data)
    return table_data,table_category_data #[[[ ]]]





def calulate_param(header_list,df,csv_path):
    # print('header_list:',type(header_list),header_list)

    data = [('item_name','BC','P_Val','a_Q','a_irr','3CV'),]
    f = open(csv_path,'w')
    writer = csv.writer(f)

    writer.writerow(('item_name','BC','P_Val','a_Q','a_irr','3CV'))


    table_data = []#[[]]
    table_category_data = []#[[[]]]
    column_list = []
    n=0
    for item_name in header_list:
        column_list = df[item_name].tolist()
        # print('column_list--->',column_list)#string list

        # print('item_name:',item_name)
        need_cpk = valid_column(item_name,df)
        # print('need_cpk:---->',need_cpk)
        column_num_list = []
        if need_cpk == 'need_cpk':
            column_list = test_value_to_numeric(column_list[2:])
            # print(str(n)+'column_list:',column_list)
            bc,p_val,a_Q,a_irr,three_σ_x100_divide_mean = get_coefficients(column_list)
            value = (item_name,bc,p_val,a_Q,a_irr,three_σ_x100_divide_mean)
            writer.writerow(value)
        n=n+1

    f.close()



@count_time
def draw_overall_cpk(path,excel_file):
    #'/Users/rex/PycharmProjects/my/cpk_report.xlsx'
    all_cpk_plot_path=path
    path=path+excel_file
    df = pd.read_excel(path)  # 读取xlsx中第一个sheet
    # df.loc[index, column_name],选取指定行和列的数据
    # df.loc[0,'name'] # 'Snow'
    # df.loc[0:2, ['name','age']]          #选取第0行到第2行，name列和age列的数据, 注意这里的行选取是包含下标的。
    # df.loc[[2,3],['name','age']]          #选取指定的第2行和第3行，name和age列的数据
    # df.loc[df['gender']=='M','name']      #选取gender列是M，name列的数据
    # df.loc[df['gender']=='M',['name','age']] #选取gender列是M，name和age列的数据
    item_name_list = df.iloc[:, [1]].values  # 读取指定键值列的所有行
    cpk_list = df.iloc[:, [16]].values  # 读取指定键值列的所有行
    # print("item_name_list：\n{0}".format(item_name_list.tolist()[0]))#first item name
    # print("cpk_list：\n{0}".format(cpk_list.tolist()[0]))#first item value

    x_x = []
    y_y = []
    temp_l=[]
    n=0
    for i in cpk_list.tolist():
        if str(i[0]) =='nan':
            temp_l.append(n)
        else:
            cpk_value = round((i[0]), 5)
            y_y.append(cpk_value)
        n = n+1
    n = 0
    for i in item_name_list.tolist():
        if n in temp_l:
            pass
        else:
            x_x.append(str(i[0]))
        n = n+1
    calculated_cpk_num = len(x_x)
    print('calculated cpk item num:',len(x_x),len(y_y))
    x = x_x
    y = y_y
    cpk_max = max(y)
    print('cpk_max:', cpk_max)
    # print('x:',x)
    # print('y:',y)
    plt.ion()  # 开启interactive mode
    fig, axes = plt.subplots(1, 0, figsize=(12.8, 8), facecolor='#ccddef')
    plt.axes([0.035, 0.30, 0.965, 0.60])  # [左, 下, 宽, 高] 规定的矩形区域 （全部是0~1之间的数，表示比例）
    bars = plt.bar(x, y)
    plt.ylim((0, 60))  # 设置y轴scopex

    # -------------------------------------
    # i=1
    # for bar in bars:
    #     # color=random.choice(colors)
    #     color = str(hex(i))
    #     print(color)
    #     color = color[2:]
    #     print(color,6-len(color))
    #     zero_count = 6-len(color)
    #     for j in range(1,zero_count+1,1):
    #         color ="0"+str(color)
    #     i = i+16
    #     color = '#'+color
    #     print('--->',len(color),color)
    # --------------------------------------

    # print('====>',len(colors.cnames.values()))
    # a = range (0,len(colors.cnames.values())+1,1)
    # b = random.sample(a,1)
    # --------------------------------------
    i = 0
    # print('colors counts:', len(colors.cnames))
    for bar in bars:
        # print('i===>',i)
        c = list(colors.cnames.values())[i]
        # print('c====>',c)
        bar.set_facecolor(c)  # '#2ca02c'
        i = i + 1
        if i == len(colors.cnames.values()):
            i = 0

    labels = x
    # -------cpk value---------------
    for a, b in zip(x, y):
        # print('---->', a, b)
        plt.text(a, b, '%.1f' % b, ha='center', va='bottom', fontsize=1)
    # -------------------------------

    l1 = [0.2]
    l2 = list(range(3, 40, 3))
    l3 = list(range(45, 100, 5))
    l4 = list(range(100,150,30))
    y_ticks = l1 + l2 + l3 + l4
    print(y_ticks)
    plt.ylabel('CPK  ', rotation='horizontal', fontsize=8, verticalalignment='center', horizontalalignment='center')
    plt.xlabel('Item Name',rotation='horizontal', fontsize=8, verticalalignment='center', horizontalalignment='center')
    plt.yticks(y_ticks, fontproperties='Times New Roman', size=5)
    plt.xticks(labels, fontproperties='Times New Roman', size=1, rotation=270)
    plt.rcParams['figure.autolayout'] = True  # 解决不能完整显示的问题（比如因为饼图太大，显示窗口太小）

    plt.rcParams['savefig.dpi'] = 1440  # 图片像素
    plt.rcParams['figure.dpi'] = 1440  # 分辨率
    titile_name = str(calculated_cpk_num)
    titile_name = 'CPK of ' + titile_name + ' items'
    plt.title(titile_name, y=1.0, fontsize=12, fontweight='bold', verticalalignment='baseline',horizontalalignment='center', color="royalblue")
    plt.grid(linestyle=':', c='gray', linewidth=0.1, alpha=0.5)  # 生成网格
    # plt.legend([""],loc="upper center")
    plt.draw()
    # plt.subplots_adjust(wspace=0, hspace=50)  # 调整子图间距
    all_cpk_plot_path= all_cpk_plot_path+'whole_cpk.png'
    plt.subplots_adjust(wspace=0, hspace=0)  # 调整子图间距
    plt.margins(0, 0)
    plt.savefig(all_cpk_plot_path, dpi=1440)
    # plt.show()
    plt.close()
    plt.ioff()  # off interactive mode
    print('draw whole CPK plot finished!!')

    return all_cpk_plot_path


def verify_limit(lsl,usl):
      lsl.replace(' ','')
      usl.replace(' ','')
      try:
        lsl = float(eval(lsl))
        usl = float(eval(usl))
      except:
        # print('no valid limit update!')
        return None,None
      if type(lsl) == float and type(usl) == float:
          print('after verify lsl===>',lsl)
          print('after verify usl===>',usl)
          return lsl,usl


@count_time
def draw_correlation(xValue,yValue,x_item_name,y_item_name,pic_save_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last):
    pearson,spearman = correlation_coefficient_calc(xValue, yValue, x_item_name, y_item_name)
    plt.ion()  # 开启interactive mode
    # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
    fig, axes = plt.subplots(1, 0, figsize=(8, 6), facecolor='#ccddef')
    plt.axes([0.15, 0.15, 0.75, 0.75])  # [左, 下, 宽, 高] 规定的矩形区域 （全部是0~1之间的数，表示比例）    
    plt.title('Correlation pearson coefficient = ' + str(pearson)+'\n'+str(start_time_first)+' -- '+str(start_time_last),size=14)
    if len(x_item_name) > 55:
        x_item_name = x_item_name[0:55] + '\n' + x_item_name[55:]
    if len(y_item_name) > 55:
        y_item_name = y_item_name[0:55] + '\n' + y_item_name[55:]
    plt.xlabel(x_item_name,size=12)
    plt.ylabel(y_item_name,size=12)

    x_min_num, x_max_num = min(xValue), max(xValue)
    print('xvalue min,max:', x_min_num, x_max_num)
    x_ticks_l, x_ticks_h = get_ticks(x_min_num, x_max_num, x_lsl, x_usl)
    print('x_ticks_l x_ticks_h:', x_ticks_l, x_ticks_h)
    plt.xlim(x_ticks_l, x_ticks_h)

    y_min_num, y_max_num = min(yValue), max(yValue)
    print('yvalue min,max:', y_min_num, y_max_num)
    y_ticks_l, y_ticks_h = get_ticks(y_min_num, y_max_num, y_lsl, y_usl)
    print('y_ticks_l y_ticks_h:', y_ticks_l, y_ticks_h)
    plt.ylim((y_ticks_l, y_ticks_h))  # 设置y轴scopex
    ax=plt.gca()
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)
    if x_lsl != 'NA' and x_usl != 'NA':
        plt.plot([x_lsl, x_lsl, ], [y_ticks_l, 1000000, ], 'k--', linewidth=3.0, color='red')  # x lower limit线，
        plt.plot([x_usl, x_usl, ], [y_ticks_l, 1000000, ], 'k--', linewidth=3.0, color='red')  # x upper limit线，
    if y_lsl != 'NA' and y_usl != 'NA':
        plt.plot([x_ticks_l, 1000000, ], [y_lsl, y_lsl, ], 'k--', linewidth=3.0, color='red')  # y lower limit线，
        plt.plot([x_ticks_l, 1000000, ], [y_usl, y_usl, ], 'k--', linewidth=3.0, color='red')  # y upper limit线，

    # plt.scatter(x, y, s, c, marker)
    # x: x轴坐标
    # y：y轴坐标
    # s：点的大小/粗细 标量或array_like 默认是 rcParams['lines.markersize'] ** 2
    # c: 点的颜色
    # marker: 标记的样式 默认是 'o'
    # plt.legend()
    plt.rcParams['savefig.dpi'] = 400  # 图片像素
    plt.rcParams['figure.dpi'] = 150  # 分辨率
    plt.scatter(xValue, yValue, s=40,linewidth =0.6, c="blue", marker='+')
    plt.grid(linestyle=':', c='gray', linewidth=1, alpha=0.9)  # 生成网格
    plt.savefig(pic_save_path, dpi=400)
    plt.draw()
    # plt.show()
    plt.close()
    plt.ioff()


@count_time
def draw_correlation_by_color(xValue,yValue,x_category_value,y_category_value,x_item_name,y_item_name,pic_save_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last):
    pearson, spearman = correlation_coefficient_calc(xValue, yValue, x_item_name, y_item_name)
    plt.ion()  # 开启interactive mode
    # font = FontProperties(fname=r"/Library/Fonts/Songti.ttc", size=12)
    fig, axes = plt.subplots(1, 0, figsize=(8, 6), facecolor='#ccddef')
    plt.axes([0.15, 0.15, 0.75, 0.75])  # [左, 下, 宽, 高] 规定的矩形区域 （全部是0~1之间的数，表示比例）
    plt.title('Correlation pearson coefficient = ' + str(pearson)+'\n'+str(start_time_first)+' -- '+str(start_time_last),size=14)
    if len(x_item_name) > 55:
        x_item_name = x_item_name[0:55] + '\n' + x_item_name[55:]
    if len(y_item_name) > 55:
        y_item_name = y_item_name[0:55] + '\n' + y_item_name[55:]
    plt.xlabel(x_item_name,size=12)
    plt.ylabel(y_item_name,size=12)

    x_min_num, x_max_num = min(xValue), max(xValue)
    print('xvalue min,max:', x_min_num, x_max_num)
    x_ticks_l, x_ticks_h = get_ticks(x_min_num, x_max_num, x_lsl, x_usl)
    print('x_ticks_l x_ticks_h:', x_ticks_l, x_ticks_h)
    plt.xlim(x_ticks_l, x_ticks_h)

    y_min_num, y_max_num = min(yValue), max(yValue)
    print('yvalue min,max:', y_min_num, y_max_num)
    y_ticks_l, y_ticks_h = get_ticks(y_min_num, y_max_num, y_lsl, y_usl)
    print('y_ticks_l y_ticks_h:', y_ticks_l, y_ticks_h)
    plt.ylim((y_ticks_l, y_ticks_h))  # 设置y轴scopex
    ax=plt.gca()
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)
    if x_lsl != 'NA' and x_usl != 'NA':
        plt.plot([x_lsl, x_lsl, ], [y_ticks_l, 1000000, ], 'k--', linewidth=3.0, color='red')  # x lower limit线，
        plt.plot([x_usl, x_usl, ], [y_ticks_l, 1000000, ], 'k--', linewidth=3.0, color='red')  # x upper limit线，
    if y_lsl != 'NA' and y_usl != 'NA':
        plt.plot([x_ticks_l, 1000000, ], [y_lsl, y_lsl, ], 'k--', linewidth=3.0, color='red')  # y lower limit线，
        plt.plot([x_ticks_l, 1000000, ], [y_usl, y_usl, ], 'k--', linewidth=3.0, color='red')  # y upper limit线，

    # plt.scatter(x, y, s, c, marker)
    # x: x轴坐标
    # y：y轴坐标
    # s：点的大小/粗细 标量或array_like 默认是 rcParams['lines.markersize'] ** 2
    # c: 点的颜色
    # marker: 标记的样式 默认是 'o'
    # plt.legend()
    plt.rcParams['savefig.dpi'] = 400  # 图片像素
    plt.rcParams['figure.dpi'] = 150  # 分辨率
    # print('x_category_value,y_category_value length:---->',len(x_category_value),len(y_category_value),x_category_value[0], y_category_value[0])
    for i in range(0,len(x_category_value),1):
        print(x_category_value[i][1],y_category_value[i][1])
        set_color = x_category_value[i][1]
        plt.scatter(x_category_value[i][5:], y_category_value[i][5:], s=40,linewidth =0.6, c=set_color, marker='+')
    plt.grid(linestyle=':', c='gray', linewidth=1, alpha=0.9)  # 生成网格
    plt.savefig(pic_save_path, dpi=400)
    plt.draw()
    # plt.show()
    plt.close()

    plt.ioff()


@count_time
def correlation_plot(table_data,table_category_data,pic_path,start_time_first,start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl):


    j=0
    path=pic_path
    info = ''
    # print('show_correlation_plot table_data length:',len(table_data),table_data)
    # print('show_correlation_plot table_category_data length:',len(table_category_data),table_category_data)


    y_item_name = table_data[0][0]
    x_item_name = table_data[1][0]

    # print('y_item_name==>',y_item_name)
    # print('x_item_name==>',x_item_name)

  
    y_usl = table_data[0][1]
    y_lsl = table_data[0][2]

    x_usl = table_data[1][1]
    x_lsl = table_data[1][2]

    yValue = table_data[0][3:]
    xValue = table_data[1][3:]
    # print('show_correlation_plot data lenghth:',len(xValue),len(yValue))
    if len(yValue) != len(xValue):
        print('xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot') 
        return 'xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot'
    new_lsl,new_usl = verify_limit(new_y_lsl,new_y_usl)
    if new_lsl != None and new_lsl != None:
        y_lsl = new_lsl
        y_usl = new_usl

    new_lsl,new_usl = verify_limit(new_x_lsl,new_x_usl)
    if new_lsl != None and new_lsl != None:
        x_lsl = new_lsl
        x_usl = new_usl


    # image_name = item_name.replace('/','_')+".png"
    image_name ='correlation.png'
    pic_path = path +'temp/'
    if not os.path.exists(pic_path):
        os.makedirs(pic_path)
    # os.system('mkdir '+pic_path)
    pic_path = pic_path + image_name
    print('pic_path--->',pic_path)

    if len(table_category_data) == 0:
        draw_correlation(xValue,yValue,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)
    else:
        #[[],[],[]...],only value category lists
        y_category_value = table_category_data[0]
        x_category_value = table_category_data[1]
        # print('y_category_value,x_category_value:',table_category_data[0],table_category_data[1])
        draw_correlation_by_color(xValue,yValue,x_category_value,y_category_value,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)

    j=j+1
    info = 'correlation plot draw finished!'
    print(info)
    return info



@count_time
def show_correlation_plot(event,header_list,df,color_by,pic_path,select_category,start_time_first,start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl):
    # print('header_list===>',header_list)

    table_data,table_category_data = parse_all_csv(header_list,df,color_by,select_category,event)

    j=0
    path=pic_path
    info = ''
    # print('show_correlation_plot table_data length:',len(table_data),table_data)
    # print('show_correlation_plot table_category_data length:',len(table_category_data),table_category_data)


    y_item_name = table_data[0][0]
    x_item_name = table_data[1][0]

    # print('y_item_name==>',y_item_name)
    # print('x_item_name==>',x_item_name)

  
    y_usl = table_data[0][1]
    y_lsl = table_data[0][2]

    x_usl = table_data[1][1]
    x_lsl = table_data[1][2]

    yValue = table_data[0][3:]
    xValue = table_data[1][3:]
    # print('show_correlation_plot data lenghth:',len(xValue),len(yValue))
    if len(yValue) != len(xValue):
        print('xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot') 
        return 'xValue and yValue are not same length!Can not calculate pearson/generate correlation_plot'
    new_lsl,new_usl = verify_limit(new_y_lsl,new_y_usl)
    if new_lsl != None and new_lsl != None:
        y_lsl = new_lsl
        y_usl = new_usl

    new_lsl,new_usl = verify_limit(new_x_lsl,new_x_usl)
    if new_lsl != None and new_lsl != None:
        x_lsl = new_lsl
        x_usl = new_usl


    # image_name = item_name.replace('/','_')+".png"
    image_name ='correlation.png'
    pic_path = path +'temp/'
    if not os.path.exists(pic_path):
        os.makedirs(pic_path)
    # os.system('mkdir '+pic_path)
    pic_path = pic_path + image_name
    print('pic_path--->',pic_path)

    if len(table_category_data) == 0:
        draw_correlation(xValue,yValue,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)
    else:
        #[[],[],[]...],only value category lists
        y_category_value = table_category_data[0]
        x_category_value = table_category_data[1]
        # print('y_category_value,x_category_value:',table_category_data[0],table_category_data[1])
        draw_correlation_by_color(xValue,yValue,x_category_value,y_category_value,x_item_name,y_item_name,pic_path,x_lsl,x_usl,y_lsl,y_usl,start_time_first,start_time_last)

    j=j+1
    info = 'correlation plot draw finished!'
    print(info)
    return info




@count_time
def cpk_plot(table_data,table_category_data,pic_path,set_bins,select_new_lsl,select_new_usl,start_time_first,start_time_last):


    j=0
    path=pic_path
    info = ''
    # print('table_data length:',len(table_data),table_data)
    # print('table_category_data length:',len(table_category_data),table_category_data)

    for column_data in table_data:
        # print('column_data length,column_data--->',len(column_data),column_data)
        item_name=column_data[0]
        usl = column_data[1]
        lsl = column_data[2]
        column_data = column_data[3:]
        # print('item_name:',item_name)
        # print('usl:',usl)
        # print('lsl:',lsl)
        # print(str(j)+' column test value:',column_data)
        new_lsl,new_usl = verify_limit(select_new_lsl,select_new_usl)
        if new_lsl != None and new_lsl != None:
            lsl = new_lsl
            usl = new_usl
            # print('new lsl===>',lsl)
            # print('new usl===>',usl)

        mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)
        if stdev == 0:
            info = item_name+'stdv == 0 cannot calculate cpk!'
            print(info)#stdev ==0
            return info
        else:

            # print('cpk:',cpk)
            # image_name = item_name.replace('/','_')+".png"
            image_name ='temp_pic.png'
            pic_path = path +'temp/'
            if not os.path.exists(pic_path):
                os.makedirs(pic_path)
            # os.system('mkdir '+pic_path)
            pic_path = pic_path + image_name

            if len(table_category_data) == 0:
                draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
            else:
                #[[],[],[]...],only value category lists
                draw_more_histogram(table_category_data[j],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)

        j=j+1
        info = item_name+' item cpk and plot draw finished!'
        print(info)
        return info






@count_time
def show_cpk_plot(event,header_list,df,color_by,pic_path,select_category,set_bins,select_new_lsl,select_new_usl,start_time_first,start_time_last):
    print(pic_path,type(set_bins),type(select_new_lsl),type(select_new_usl),type(start_time_first),type(start_time_last),set_bins,select_new_lsl,select_new_usl,start_time_first,start_time_last)
    table_data,table_category_data = parse_all_csv(header_list,df,color_by,select_category,event)

    j=0
    path=pic_path
    info = ''
    print('table_data length:',len(table_data),table_data)
    print('table_category_data length:',len(table_category_data),table_category_data)

    for column_data in table_data:
        # print('column_data length,column_data--->',len(column_data),column_data)
        item_name=column_data[0]
        usl = column_data[1]
        lsl = column_data[2]
        column_data = column_data[3:]
        # print('item_name:',item_name)
        # print('usl:',usl)
        # print('lsl:',lsl)
        # print(str(j)+' column test value:',column_data)
        new_lsl,new_usl = verify_limit(select_new_lsl,select_new_usl)
        if new_lsl != None and new_lsl != None:
            lsl = new_lsl
            usl = new_usl
            # print('new lsl===>',lsl)
            # print('new usl===>',usl)

        mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)
        if stdev == 0:
            info = item_name+'stdv == 0 cannot calculate cpk!'
            print(info)#stdev ==0
            return info
        else:

            # print('cpk:',cpk)
            # image_name = item_name.replace('/','_')+".png"
            image_name ='temp_pic.png'
            pic_path = path +'temp/'
            if not os.path.exists(pic_path):
                os.makedirs(pic_path)
            # os.system('mkdir '+pic_path)
            pic_path = pic_path + image_name

            if len(table_category_data) == 0:
                draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
            else:
                #[[],[],[]...],only value category lists
                draw_more_histogram(table_category_data[j],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)

        j=j+1
        info = item_name+' item cpk and plot draw finished!'
        print(info)
        return info



def creat_excel_report_file(path,file,cpk_lsl,cpk_usl):

    excel_file_path = path + file
    book = xlsxwriter.Workbook(excel_file_path)#'cpk.xlsx'
    report_sheet = book.add_worksheet('report')#'report'
    plot_sheet = book.add_worksheet('fail plot')#'plot'
    report_sheet.set_column("B:B", 80)  # 设定A列列宽为40
    report_sheet.set_column("M:M", 10)  # 设定A列列宽为10
    report_sheet.set_row(0, 30)    #设置行高度

    cpk_result ='CPK_Result(cpk_lsl:'+str(cpk_lsl)+';cpk_usl:'+str(cpk_usl)+')'
    report_sheet_title = [u'No', u'Item_name',u'BC',u'P_Val',u'a_Q',u'a_irr',u'3CV',u'LSL',u'Target', u'USL', u'Min', u'Mean',u'Max', u'Std',u'CPL', u'CPU', u'CPK',cpk_result, u'New LSL',u'New Target',u'New USL',u'New CPL',u'New CPU',u'New CPK','New '+cpk_result]

    format_pass = book.add_format()  # 定义format格式对象
    format_pass.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_pass.set_valign('center') # 定义format_titile对象单元格对齐方式

    format_pass.set_border(1)  # 定义format对象单元格边框加粗的格式
    # format_pass.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    format_pass.set_text_wrap()  # 内容换行



    new_format_pass = book.add_format()  # 定义format格式对象
    new_format_pass.set_align('center')  # 定义format_titile对象单元格对齐方式
    new_format_pass.set_valign('center') # 定义format_titile对象单元格对齐方式
    new_format_pass.set_bg_color('cyan')  # 定义format_titile对象单元格背景颜色

    new_format_pass.set_border(1)  # 定义format对象单元格边框加粗的格式
    # new_format_pass.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    new_format_pass.set_text_wrap()  # 内容换行


    format_fail = book.add_format()  # 定义format_title 格式对象
    format_fail.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    format_fail.set_bg_color('yellow')  # 定义format_titile对象单元格背景颜色
    format_fail.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_fail.set_valign('center') # 定义format_titile对象单元格对齐方式

    # format_fail.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    format_fail.set_text_wrap()  # 内容换行



    new_format_fail = book.add_format()  # 定义format_title 格式对象
    new_format_fail.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    new_format_fail.set_bg_color('red')  # 定义format_titile对象单元格背景颜色
    new_format_fail.set_align('center')  # 定义format_titile对象单元格对齐方式
    new_format_fail.set_valign('center') # 定义format_titile对象单元格对齐方式

    # format_fail.set_num_format('0.00')  # 格式化数据格式为小数点后两位
    new_format_fail.set_text_wrap()  # 内容换行


    format_titile = book.add_format()  # 定义format_title 格式对象
    format_titile.set_border(1)  # 定义format_titile 对象单元格边框加粗的格式
    format_titile.set_bg_color('#cddccc')  # 定义format_titile对象单元格背景颜色
    format_titile.set_align('center')  # 定义format_titile对象单元格对齐方式
    format_titile.set_valign('center')  # 定义format_titile对象单元格对齐方式
    format_titile.set_bold()  # 定义format_titile对象内容加粗
    format_titile.set_text_wrap()  # 内容换行


    format_ave = book.add_format()  # 定义format_ave格式对象
    format_ave.set_border(1)  # 边框加粗的格式
    format_ave.set_num_format('0.00')  # 定义format_ave对象单元格数字类别显示格式
    report_sheet.write_row('A1', report_sheet_title, format_titile)
    print('Excel report head created finished!')
    return book,report_sheet,plot_sheet,format_fail,format_pass,format_titile,new_format_fail,new_format_pass


def get_one_item_new_limit_from_csv(new_limit_path,item_name):
    new_limit_file = new_limit_path + 'temp/item_limit.csv'
    tmp_lst = []
    # try:
    with open(new_limit_file, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('item') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # except IOError as exc:
    #   print('ERROR:',exc)
    header_list = tmp_lst[0]
    # print('header_list--->',header_list)
    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    # print(df)
    try:
        new_lsl=df.loc[df['item'] == item_name, 'new_lsl'].tolist()[0]
        new_usl=df.loc[df['item'] == item_name, 'new_usl'].tolist()[0]
    except IndexError as e:
        print('item_limit.csv item wrong!')
    return new_lsl,new_usl


def load_excel_table(excel_report_path):
    excel_data = openpyxl.load_workbook(excel_report_path)
    sheetnames = excel_data.get_sheet_names()
    table = excel_data.get_sheet_by_name(sheetnames[0])#[u'report', u'fail plot']
    table = excel_data.active
    print('append one item new limit data to-->',table.title) 
    nrows = table.max_row 
    ncolumns = table.max_column 
    print('load_excel_table nrows--->',nrows)
    print('load_excel_table ncolumns--->',ncolumns)
    return excel_data,table,nrows,ncolumns



def append_new_limit_data_to_excel_report(data,table_name,excel_report_path,row_n,new_lsl,new_usl,new_cpl,new_cpu,new_cpk):
    print('data in append_new_limit_data_to_excel_report--->',data,excel_report_path,row_n,new_lsl,new_usl,new_cpl,new_cpu,new_cpk)
    table_name.cell(row_n+1,14).value = new_lsl
    table_name.cell(row_n+1,15).value = new_usl
    table_name.cell(row_n+1,16).value = new_cpl 
    table_name.cell(row_n+1,17).value = new_cpu
    table_name.cell(row_n+1,18).value = new_cpk
    data.save(excel_report_path)


def save_new_limit_to_report(new_limit_path,excel_report_name):
    new_limit_file = new_limit_path + 'temp/item_limit.csv'
    tmp_lst = []
    # try:
    with open(new_limit_file, 'r') as f:
        reader = csv.reader(f)
        i = 1
        for row in reader:
            # print(row[0].lower())
            if row[0].lower().find('item') != -1:
                tmp_lst.append(row)
            else:
                tmp_lst.append(row)
            i = i + 1
    # except IOError as exc:
    #   print('ERROR:',exc)
    header_list = tmp_lst[0]
    print('header_list--->',header_list)
    df = pd.DataFrame(tmp_lst[1:], columns=tmp_lst[0])
    # new_lsl=df.loc[df['item'] == 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_x', 'new_lsl'].tolist()
    # new_usl=df.loc[df['item'] == 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_x', 'new_usl'].tolist()
    data_df = pd.read_excel(new_limit_path+excel_report_name,names=None)
    # l= list(sheet.values())  
    item_list = data_df.loc[:, ['Item_name']].values.tolist()

    data = openpyxl.load_workbook(new_limit_path+excel_report_name)
    sheetnames = data.get_sheet_names()
    table = data.get_sheet_by_name(sheetnames[0])#[u'report', u'fail plot']
    table = data.active
    print('append new limit to-->',table.title) 
    nrows = table.max_row 
    ncolumns = table.max_column 
    print('nrows--->',nrows)
    print('ncolumns--->',ncolumns)
    i=1
    for x in item_list:
      # print('item_name-->',x[0])
      new_lsl=df.loc[df['item'] == x[0], 'new_lsl'].tolist()[0]
      new_usl=df.loc[df['item'] == x[0], 'new_usl'].tolist()[0]
      table.cell(i+1,14).value = new_lsl
      table.cell(i+1,15).value = new_usl
      i=i+1
    data.save(new_limit_path+excel_report_name)
    print('save new limit to excel report finished!')


def get_file_pic_name(file_dir):
  pic_file_l = []
  for root,dirs,files in os.walk(file_dir):
    # print(root)
    # print('fail picture path:',dirs)
    # print(files)
    for file in files:
      if os.path.splitext(file)[1] == '.png':
          # print('suffix name---->',os.path.splitext(file)[1])
          pic_file_l.insert(0,os.path.splitext(file)[0])
  return pic_file_l

def clear_files(path):
    for root,dirs,files in os.walk(path):
        for file in files:
            os.remove(path+'/'+file)
@count_time
def generate_keynote_report(project_code,station_name,build_stage,dir_path):
    keynote_title_name = project_code+'_'+station_name+'_'+build_stage+'_DATA Review'
    description_info =  'Issue description:\n'
    root_cause_info = 'Root Cause:\n'
    plan_info = 'Next steps:'
    print('dir_path--->',dir_path)
    keynote_save_path = dir_path+'cpk_report_'+datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')+'.key'
    overall_plot_path = dir_path + 'whole_cpk.png'

    doc,keynote = generate_keynote.create_keynote(keynote_title_name)
    print('save title page to keynote report finished!')

    generate_keynote.add_overall_pic(doc,keynote,overall_plot_path)
    print('save whole_cpk plot to keynote report finished!')


    fail_pic_path = dir_path+'fail_plot'
    print('fail pic path---->',fail_pic_path)
    file_l=[]
    file_l = get_file_pic_name(fail_pic_path)
    # print('file_l --->',file_l)

    for f_name in file_l:
        one_fail_pic_path = fail_pic_path +'/'+f_name+'.png'
        print('fail pic name--->',one_fail_pic_path)
        generate_keynote.add_fail_pic(doc,keynote,one_fail_pic_path,description_info,root_cause_info,plan_info)

    generate_keynote.save_keynote(doc,keynote,keynote_save_path)
    print('save fail_plot to keynote report finished!')


def create_report(event,header_list,df,color_by,pic_path,select_category,cpk_lsl,cpk_usl,save_all_cpk_path,set_bins,excel_name,project_code,build_stage,station_name,start_time_first,start_time_last):
    clear_files(pic_path)
    book,report_sheet,plot_sheet,format_fail,format_pass,format_titile,new_format_fail,new_format_pass = creat_excel_report_file(save_all_cpk_path,excel_name,cpk_lsl,cpk_usl)
    # excel_report_data,excel_report_table,max_row,max_col = load_excel_table(save_all_cpk_path+excel_name)

    table_data,table_category_data = parse_all_csv(header_list,df,color_by,select_category,event)#
    i,j,n,t=0,0,0,0
    path=pic_path
    result='pass'

    # print('table_data length in create_report:',len(table_data),table_data)
    for column_data in table_data:
        # print('column_data length,column_data--->',len(column_data),column_data)
        item_name=column_data[0]
        usl = column_data[1]
        lsl = column_data[2]
        column_data = column_data[3:]
        # print('item_name:',item_name)
        # print('usl:',usl)
        # print('lsl:',lsl)
        # print(str(j)+' column test value:',column_data)
        bc,p_val,a_Q,a_irr,three_CV = get_coefficients(column_data)
        row_data = []
        target_value = 9999999999
        mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)



        result=''

        if stdev == 0:
            n=n+1
            print(item_name+': do not record this item cpk value in excel and draw plot due to stdev == 0!(data limit)')#stdev ==0
        else:

            location = 'A' + str(j+2)
            # print("location,j--->",location,j)

            # print('cpk:',cpk)
            if cpk == None:
                result='Nan'
                target_value = get_target_value(lsl,usl)
                # print('target_value--->',lsl,usl,target_value)
                row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,lsl,target_value, usl, min_num, mean, max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                report_sheet.write_row(location,row_data,format_pass)
                # if not os.path.exists(path):
                #     os.makedirs(path)
                # image_name = item_name.replace('/','_')+".png"
                # pic_path = path + image_name
                # if len(table_category_data) == 0:
                #     draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                # else:

                #     #[[],[],[]...],only value category lists
                #     draw_more_histogram(table_category_data[t],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                # if i==0:
                #     location='A1'
                # else:
                #     location = 'A'+str(29*i)
                # save_image_to_excel(location,pic_path,plot_sheet)
                # i = i + 1

            else:
                if cpk < cpk_lsl or cpk > cpk_usl:
                    # print('----------------------cpk value fail--------------------------')
                    # print('cpk_lsl:',cpk_lsl)
                    # print('cpk_usl:',cpk_usl)
                    result='Fail'
                    target_value = get_target_value(lsl,usl)
                    # print('target_value--->',lsl,usl,target_value)
                    row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,lsl,target_value, usl, min_num, mean, max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                    report_sheet.write_row(location,row_data,format_fail)
                    if not os.path.exists(path):
                        os.makedirs(path)
                    image_name = item_name.replace('/','_')+".png"
                    pic_path = path + image_name
                    if len(table_category_data) == 0:
                        draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                    else:

                        #[[],[],[]...],only value category lists
                        draw_more_histogram(table_category_data[t],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                    if i==0:
                        location='A1'
                    else:
                        location = 'A'+str(29*i)
                    save_image_to_excel(location,pic_path,plot_sheet)
                    i = i + 1
                else:
                    result='Pass'
                    target_value = get_target_value(lsl,usl)
                    # print('target_value--->',lsl,usl,target_value)

                    row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,lsl,target_value,usl, min_num, mean, max_num, stdev, cpl, cpu, cpk, result,'','','','','','','']
                    report_sheet.write_row(location,row_data,format_pass)



        #--------------------------------get new limit from csv start------------------------
        new_lsl,new_usl = get_one_item_new_limit_from_csv(save_all_cpk_path,item_name)
        new_lsl,new_usl = verify_limit(new_lsl,new_usl)
        if new_lsl != None and new_lsl != None:
            old_mean, old_max_num,old_target_value,old_min_num, old_stdev,old_cpu, old_cpl, old_cpk,old_lsl,old_usl,old_result = mean, max_num,target_value,min_num, stdev, cpu, cpl, cpk,lsl,usl,result
            lsl = new_lsl
            usl = new_usl
            # print('new lsl===>',lsl)
            # print('new usl===>',usl)
        #--------------------------------get new limit from csv end------------------------
            mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk = cpk_calc(column_data, lsl, usl)
            if stdev == 0:
                n=n+1
                print(item_name+': do not record this item cpk value in excel and draw plot due to stdev == 0!(new limit)')#stdev ==0
            else:

                location = 'A' + str(j+2)
                # print("location,j--->",location,j)

                # print('cpk:',cpk)
                if cpk < cpk_lsl or cpk > cpk_usl:
                    print('----------------------cpk value fail--------------------------')
                    # print('cpk_lsl:',cpk_lsl)
                    # print('cpk_usl:',cpk_usl)
                    result='Fail'
                    target_value = get_target_value(lsl,usl)
                    # print('target_value--->',lsl,usl,target_value)
                    # append_new_limit_data_to_excel_report(excel_report_data,excel_report_table,save_all_cpk_path+excel_name,j+2,lsl,usl,cpl,cpu,cpk)
                    row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,old_lsl,old_target_value, old_usl, old_min_num, old_mean, old_max_num, old_stdev, old_cpl, old_cpu, old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                    report_sheet.write_row(location,row_data,new_format_fail)
                    if not os.path.exists(path):
                        os.makedirs(path)
                    image_name = item_name.replace('/','_')+" new_limit.png"
                    pic_path = path + image_name
                    if len(table_category_data) == 0:
                        draw_histogram(column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                    else:

                        #[[],[],[]...],only value category lists
                        draw_more_histogram(table_category_data[t],column_data,item_name,lsl, usl, mean, max_num, min_num, stdev, x1, y1, cpu, cpl, cpk, pic_path,set_bins,start_time_first,start_time_last)
                    if i==0:
                        location='A1'
                    else:
                        location = 'A'+str(29*i)
                    save_image_to_excel(location,pic_path,plot_sheet)
                    i = i + 1
                else:
                    result='Pass'
                    target_value = get_target_value(lsl,usl)
                    # print('target_value--->',lsl,usl,target_value)
                    # append_new_limit_data_to_excel_report(excel_report_data,excel_report_table,save_all_cpk_path+excel_name,j+2,lsl,usl,cpl,cpu,cpk)

                    row_data = [j + 1, item_name,bc,p_val,a_Q,a_irr,three_CV,old_lsl,old_target_value, old_usl, old_min_num, old_mean, old_max_num, old_stdev, old_cpl, old_cpu, old_cpk, old_result,lsl,target_value,usl,cpl,cpu,cpk,result]
                    report_sheet.write_row(location,row_data,new_format_pass)


        if stdev != 0:
            j=j+1
        t = t +1
    print('Total {} items not calulate cpk'.format('%.0f' % n))
    book.close()
    print('All items cpk calulate/draw plots/excel report finished!')
    # save_new_limit_to_report(save_all_cpk_path,excel_name)
    if event == 'keynote-report':
        draw_overall_cpk(save_all_cpk_path,excel_name)
        generate_keynote_report(project_code,station_name,build_stage,save_all_cpk_path)

@count_time
def run(event,one_item_name,all_csv_path,cpk_path,cpk_lsl,cpk_usl,color_by,selected_category_l,data_select,remove_fail,set_bins,new_x_lsl,new_x_usl,new_y_lsl,new_y_usl):
    print('event--->',event,cpk_path)
    


    fail_pic_path =cpk_path+'fail_plot/'
    if event == 'one_item_plot':
        all_csv_path = cpk_path + 'temp/one_item.csv'


    if event != 'one_item_plot':
        start_time1 = datetime.datetime.now()  #
        header_list,df,project_code,build_stage,station_name,start_time_first,start_time_last =  open_all_csv(event,all_csv_path,data_select,remove_fail)
        over_time1 = datetime.datetime.now()   #
        total_time1 = (over_time1-start_time1).total_seconds()
        print('open_all_csv  expend total time is: %s s' % total_time1)
    # print('header_list------>',len(header_list),header_list)
    if event == 'calculate-param':
       csv_path = cpk_path+'temp/calculate_param.csv'
       calulate_param(header_list,df,csv_path)
       print('['+str(datetime.datetime.now())+']','calulate param finished!')
       return
    # color_by_value,category_list = select_color_by(df)
    select_category_l =[]
    if color_by =='Off':
        select_category_l =[]
    else:
        select_category_l = select_category(selected_category_l)

    print('select_category_l------>',select_category_l)

    if event == 'one_item_plot':
        header_list,df,df_correlation,start_time_first,start_time_last,correlation_start_time_first,correlation_start_time_last =  open_one_item_csv(event,all_csv_path,data_select,remove_fail)

        if header_list[12] == header_list[13]:
            df = df.iloc[:, 0:13]
        one_item_head_list = [one_item_name]
        show_cpk_plot(event,one_item_head_list,df,color_by,cpk_path,select_category_l,set_bins,new_y_lsl,new_y_usl,start_time_first,start_time_last)
        if header_list[12] == header_list[13]:
            header_list = [one_item_name,one_item_name]
        else:
            header_list = [header_list[12],header_list[13]]
        if len(df_correlation.values.tolist())>5:
            show_correlation_plot(event,header_list,df_correlation,color_by,cpk_path,select_category_l,correlation_start_time_first,correlation_start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl)
            print('['+str(datetime.datetime.now())+']','show plot finished!')

    elif event == 'excel-report' or event == 'keynote-report':
        excel_report_file_name ='cpk_report_'+datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')+'.xlsx'
        create_report(event,header_list,df,color_by,fail_pic_path,select_category_l,cpk_lsl,cpk_usl,cpk_path,set_bins,excel_report_file_name,project_code,build_stage,station_name,start_time_first,start_time_last)
        print('['+str(datetime.datetime.now())+']','create report finished!')
    else:
        print('param error!')


if __name__ == '__main__':
    '''
    
    
    '''
    #parameter list
    #----------------------------------------------------------------------------------
    # event = 'one_item_plot'#'excel-report'/'one_item_plot'/'keynote-report'
    # one_item_name = 'SystemState Wakeup_Voltage PPVDD_S1_SOC'
    # all_csv_path = '/Users/rex/Desktop/P1_Retest/cpk/222.csv'
    # cpk_path = "/Users/rex/PycharmProjects/my/"
    # cpk_lsl=1.33
    # cpk_usl=9999999999999999
    # color_by='Station ID'#'Off'/'SerialNumber'/'Version'/'Station ID'/'Special Build Name'/'Product'/'StartTime'/'Special Build Description'
    # selected_category_l = ['CWNJ_C02-2FAP-24_2_FCT', 'CWNJ_C02-2FAP-24_1_FCT', 'CWNJ_C02-2F-REL01_1_FCT', 'CWNJ_C02-2FAP-23_2_FCT'] #[]
    # data_select = 'all'#first/last/no_retest/all --empty sn reminder only for
    # remove_fail = 'yes'
    # set_bins = 250
    # new_y_lsl,new_y_usl,new_x_lsl,new_x_usl = '','','',''
    #----------------------------------------------------------------------------------

    # print('color_by:', color_by)
    # print('cpk_lsl:', cpk_lsl)
    # print('cpk_usl:', cpk_usl)

    # excel_report_file_name ='cpk_report.xlsx'
    # pic_path =cpk_path+'fail/'
    # header_list,df =  open_all_csv(all_csv_path,data_select,remove_fail)
    # color_by_value,category_list = select_color_by(df)
    # if color_by_value =='Off':
    #     select_category_l =[]
    # else:
    #     select_category_l = select_category(category_list)
    # print('select_category_l------>',select_category_l)

    # if event == 'one_item_plot':
    #     header_list = [one_item_name]
    #     show_cpk_plot(event,header_list,df,color_by,cpk_path,select_category_l,set_bins)
    # elif event == 'cpk-report':
    #     create_report(header_list,df,color_by_value,pic_path,select_category_l,cpk_lsl,cpk_usl,cpk_path,set_bins,excel_report_file_name)
    # else:
    #     print('param error!')
    # run(event, one_item_name, all_csv_path, cpk_path, cpk_lsl, cpk_usl, color_by, selected_category_l, data_select,remove_fail,set_bins,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl)
    # draw_overall_cpk('/Users/rex/Desktop/CPK_Log/')
    #------------------------ debug cpk_plot --------------------------------
    # color list: ['#0000FF','#FF0000','#008000','#00FFFF','#9400D3','#8B008B','#B8860B','#FFA500','#A9A9A9','#FFFF00']

    table_data = [['Accelerometer AVG-FS8g_ODR200HZ_Zup accel_normal_average_x', 0.118, -0.118, 0.00875, 0.00072, 0.00518, 0.00557, 0.01365, 0.01179, -0.00121, -0.00049, 0.00526, 0.00965, -0.00106, 0.00051, 0.00852, 0.00428, 0.00519, 0.00496, 0.005, 0.00248, 0.00018, 0.00764, 0.01148, 0.00595, -0.00527, 0.00503, 0.01028, 0.00534, -0.00191, -0.00129, 0.00439, 0.00432, -0.00367, 0.00559, -0.00143, -0.00065, -0.00367, -0.00043, 0.00315, 0.00388, -0.00243, -0.00054, 0.00497, 0.00243, 0.00117, 0.01055, 0.00327, 0.00151, 0.00636, 0.00099, 0.00127, 0.00827, 0.00287, -0.00182, 0.00511, 0.0093, 0.00242, 0.00183, 0.00194, -0.00507, -0.00036, -0.00039, 0.00368, 0.00366, -0.0031, 0.00762, 0.00058, 0.00178, -0.00225, 0.00183, 0.00567, 0.00214, 0.0012, -0.00293, -0.0005, 0.00338, 0.00502, 0.00443, 0.0033, 0.00361, -0.00357, 0.00893, 0.00096, 0.00365, -0.00223, 0.0, -0.00328, -0.00542, -0.00548, -0.00518, 0.00198, 0.0015, -0.00436, 0.00074, 0.00138, 0.00552, -0.00177, 0.00191, 0.00676, 0.0099, 0.00314, -0.00677, 0.00819, 0.00945, 0.00082, 0.00697, 0.0122, 0.00035, 0.00974, 0.00239, 0.0014, 0.0019, 0.00311, 0.00384, 0.00639, 0.00892, 0.004, -0.00119, 0.00178, -0.00302, -0.00977, 0.0021, 0.00203, 0.00083, -0.00107, 0.00389, -0.00053, -0.00197, -0.00033, -0.00049, 0.00606, 0.00895, 0.00455, 0.00439, 0.00142, 0.00067, 0.00887, 0.0102, 0.00646, 0.00691, 0.00553, 0.00781, -0.00123, 0.00055, 0.00351, 0.00455, 0.00847, 0.00327, 0.00153, 0.0035, 0.0026, 0.00848, 0.00787, 0.00123, 0.00322, -0.00556, 0.00712, -0.00121, 0.00447, -0.00488, 0.00298, 0.00297, -0.00032, 0.00304, 0.00035, 0.00993, -0.00212, 0.00247, 0.00455, 0.0041, 0.00084, 0.01233, 0.00204, 0.01306, 0.00516, -0.00209, 0.00211, 0.00103, 0.00104, 0.00355, 0.00283, 0.00667, 0.00876, -0.00113, 0.00068, 0.00419, -0.00218, -0.002, 0.01113, 0.00554, 0.00565, -0.00136, 0.00227, 0.00212, 0.0047, -0.00134, 0.00019, 0.00649, 0.00318, 0.00215, 0.01619, 0.0, 0.00771, 0.00591, 0.00334, 0.00762, 0.0065, -0.00175, 0.01151, -0.00248, 0.0, -0.00556, -0.0032, 0.01536, 0.00111, 0.0, -0.00366, 0.00797, 0.00591, 0.00599, 0.00426, 0.00271, 0.00653, -0.00788, 0.00677, 0.00132, 0.00886, 0.00357, 0.00498, 0.00723, 0.00269, -0.00103, 0.00027, 0.00012, -0.00061, 0.00578, 0.00399, -0.00108, 0.01251, -0.00014, 0.0, 0.00984, -0.00232, 0.00253, 0.00401, -0.00397, 0.00075, 0.00189, 0.00784, -0.00113, -0.0015, 0.00247, 0.0024, -0.00329, -0.00318, 0.0055, 0.01047, 0.00128, -0.00191, 0.00118, 0.00627, 0.00707, -0.00645, -0.00263, 0.00165, -0.00501, 0.00981, -0.00138, 0.00897, 0.01069, -0.00348, -0.00012, 0.00391, 0.0032, 0.00252, 0.00384, 0.00428, -0.00075, -0.0006, 0.00819, -0.00143, 0.00364, 0.00462, 0.00568, 0.00716, -0.00244, -0.00197, 0.00307, 0.00022, 0.00054, 0.00928, 0.0088, 0.00781, 0.0016, 0.00226, 0.01081, -0.00241, -0.00135, 0.00544, 0.00406, 0.00222, 0.00878, 0.00536, 0.00996, 0.00435, -0.00011, 0.0022, 0.01029, 0.00737, 0.00277, 0.00526, 0.00963, 0.00767, 0.00582, 0.00771, -0.00221, 0.00054, -0.00336, 0.00551, 0.00238, 0.00134, -0.00063, 0.00241, 0.00757, 0.00249, -0.00662, 0.00273, 0.00451, -0.00323, 0.00172, 0.00843, 0.00793, 0.00357, 0.00308, 0.00367, 0.00166, 0.00431, -0.00311, 0.00185, 0.00043, 0.00367, -0.00021, 0.0025, 0.01089, 0.00215, 0.00421, 0.00332, 0.00423, 0.0098, 0.00606, 0.00078, 0.00267, 0.00264, -0.00228, 0.00018, 0.00327, -0.00085, -0.00036, 0.00522, -0.00739, 0.00371, 0.00092, 0.00419, 0.00376, -0.00135, 0.00128, 0.00467, 0.00258, -0.00135, 0.0, 0.00597, 0.00585, -0.00105, 0.00124, 0.00069, -0.00287, 0.00414, 0.00399, 0.00416, 0.00983, -0.00291, -0.00259, 0.0072, 0.00357, 0.00435, -0.0005, -0.00043, 0.00458, 0.00247, 0.00096, 0.00196, 0.00291, -0.00186, 0.0, 0.00079, 0.00339, 0.00185, -0.00176, -0.0031, -0.00521, 0.00719, 0.00399, -0.00114, 0.00971, 0.00365, 0.00145, 0.00768, 0.00078, -0.00888, 0.0044, -0.00357, -0.00056, 0.00196, -0.0045, -0.00073, 0.00105, 0.00579, 0.00432, 0.00327, 0.00585, 0.00034, -0.00057, 0.01771, -0.00921, 0.00812, 0.00149, 0.00284, 0.0078, 0.00894, 0.00582, -0.00047, 0.00835, 0.00485, 0.00192, -0.00314, -0.00488, -0.00053, -0.00093, 0.00124, 0.00261, 0.01084, 0.00348, -0.00029, 0.00116, -0.00098, 0.00347, 0.0032, 0.00548, 0.0042, 0.00402, 0.0023, -0.00346, 0.00643, 0.00291, 0.00351, 0.00144, 0.00065, -0.00122, -0.00333, 0.00151, -0.00192, 0.00755, -0.00083, 0.00541, -0.00179, 0.00787, 0.00095, 0.00946, 0.00824, -0.00154, 0.00437, 0.0037, -0.00067, 0.00789, 0.00601, -0.00475, -0.00322, -0.00275, -0.00265, 0.00067, 0.0028, 0.00603, -0.00111, -0.00265, -0.0031, 0.00501, 0.00158, 0.00271, 0.001, 0.00032, 0.0082, -0.0002, 0.00744, 0.00376, 0.00141, 0.00213, 0.00447, -0.00552, 0.00268, 0.00911, 0.00821, -0.00092, 0.00199, 0.00172, 0.00402, 0.00758, -0.00021, 0.00107, 0.00216, -0.00197, 0.00248, -0.00195, 0.0081, -0.00049, 0.00185, 0.00434, 0.0087, 0.00259, -0.00014, 0.00112, 0.00741, -0.00685, -0.0015, -0.00587, 0.00307, 0.00212, -0.00749, 0.0062, 0.00332, 0.00272, -0.00107, 0.00336, 0.01341, 0.00788, 0.00632, -0.00883, -0.00226, 0.0, 0.00094, -0.00049, 0.00164, 0.00258, 0.00584, 0.00456, 0.00034, -0.00264, -0.00067, 0.0087, 0.00306, 0.00547, 0.00212, 0.00761, 0.00625, 0.01027, 0.0029, 0.00311, 0.00086, 0.00828, 0.0, 0.00522, 0.00261, 0.00692, -0.00212, 0.00019, 0.00081, -0.00266, -0.00439, 0.00694, 0.00332, 0.00461, 0.00299, 0.0029, -0.00195, 0.00139, 0.0013, 0.00649, 0.00808, 0.00329, -0.00024, 0.00603, -0.00043, -0.00962, 0.00142, 0.00735, 0.00189, 0.00156, 0.0051, 0.00626, -0.0032, 0.00445, -0.00117, 0.00025, 0.00363, 0.00632, 0.0061, -0.00481, 0.00425, 0.00205, -0.00291, 0.00648, 0.00266, -0.00453, -0.00364, 0.00078, 0.00574, -0.00375, -0.00065, 0.00798, 0.00473, -0.00243, 0.00481, 0.00208, 0.00314, 0.00211, 0.00063, 0.01095, -0.0014, 0.00167, 0.00309, -0.00172, -0.00492, 0.01533, 0.00777, 0.00297, 0.00174, 0.0026, 0.00509, -0.00263, 0.00081, 0.00064, 0.00056, 0.01097, 0.00678, 0.00319, 0.00486, -0.00047, 0.0069, -0.00358, 0.00849, -0.00189, 0.00635, 0.00152, 0.00155, -0.00072, 0.00825, 0.00077, 0.00217, 0.0015, 0.00743, 0.00372, 0.00938, -0.00219, 0.00026, -0.00275, 0.01011, 0.00066, 0.00178, 0.0, 0.0043, 0.00092, -0.00581, 0.00058, 0.00172, 0.00545, -0.00048, 0.00544, -0.00105, 0.00159, 0.00653, 0.00148, 0.00034, 0.00087, 0.00137, 0.00282, -0.00142, 0.00033, -0.00736, -0.00123, 0.0081, 0.00205, 0.00116, -0.00519, 0.00673, -0.00191, 0.00569, -0.00581, -0.00523, 0.00611, 0.00362, 0.00188, 0.0042, 0.00493, 0.0057, -0.00566, 0.00557, 0.00383, 0.00339, -0.00343, 0.00245, -0.00094, 0.00056, 0.00454, 0.01121, 0.00459, -0.00768, -0.00978, 0.01015, 0.00968, -0.0013, -0.01728, 0.01083, 0.00996, -0.01279, -0.01211, 0.01378, 0.00953, 0.00228, 0.00254, 0.01177, -0.00642, -0.0092, 0.01164, 0.01548, 0.00204, 0.00114, 0.00586, 0.00322, -0.00396, 0.00066, 0.0152, 0.00822, -0.002, -0.00976, 0.00776, 0.01003, -0.00367, -0.0072, 0.00086, 0.00443, -0.00067, -0.00029, 0.00087, 0.01087, -0.00288, 0.0, -0.0027, -0.00787, 0.00457, -0.00311, -0.00062, 0.00588, 0.00623, -0.00081, -0.00165, 0.01097, 0.01587, 0.0, 0.00111, 0.00177, 0.00069, 0.00351, 0.00254, 0.00743, 0.00208, -0.005, -0.00457, 0.00093, 0.0053, 0.00345, 0.00339, -0.00152, -0.00101, -0.00065, 0.00048, 0.01086, 0.006, 0.00118, 0.00162, 0.00717, 0.00371, 0.00887, -0.00597, -0.00253, 0.00893, -0.001, -0.00502, 0.00125, 0.00271, 0.00121, -0.00334, 0.00235, -0.00103, 0.00645, -0.00121, 0.00625, 0.0, 0.00772, -0.00362, 0.00366, 0.01088, 0.00238, -0.0015, 0.00934, -0.00462, -0.00424, 0.00335, -0.00081, 0.00158, -0.00317, -0.00115, 0.00588, 0.00806, 0.00079, 0.00526, -0.00065, 0.00318, 0.00383, 0.00771, -0.00015, -0.00245, 0.0075, 0.00563, -0.00223, 0.00368, -0.00295, -0.00199, 0.00276, 0.00489, -0.00113, 0.0, -0.002, -0.00065, 0.00238, -0.0019, 0.00514, -0.0025, 0.00133, 0.0111, 0.00416, 0.00022, -0.00205, -0.00531, 0.00639, -0.00343, -0.00206, -0.00274, 0.00279, 0.00173, -0.00114, 0.00677, 0.00895, 0.017404, 0.014196, -0.012434, -0.007955, 0.021675, 0.016083, -0.004455, -0.003674, 0.010693, 0.020249, -0.013457, -0.004792, 0.014762, 0.011479, -0.015235, -0.013161, 0.021875, 0.017191, -0.012668, -0.009655, 0.014624, 0.015017, -0.00997, -0.007622, 0.015532, 0.015681, -0.009978, -0.007865, 0.019603, 0.010732, -0.010212, -0.0135, 0.016777, 0.013657, -0.014978, -0.008793, 0.015982, 0.013434, -0.008279, -0.010875, 0.014077, 0.015478, -0.014007, -0.007138, 0.004797, 0.011938, -0.015699, -0.01629, 0.01604, 0.011166, -0.015168, -0.000265, 0.019898, 0.021764, -0.008146, -0.00903, 0.01403, 0.012649, -0.016166, 0.000295, 0.016997, 0.010756, -0.009021, -0.016898, 0.020278, 0.018457, -0.006079, -0.020667, 0.007023, 0.009995, -0.012875, -0.014259, 0.016171, 0.010556, -0.008689, -0.007465, 0.019792, 0.013226, -0.01304, -0.004704, 0.015627, 0.018039, -0.007436, -0.007188, 0.008269, 0.01666, -0.00479, -0.015327, 0.017256, 0.010406, -0.018305, -0.009765, 0.01971, 0.009813, -0.007063, -0.018571, 0.01331, 0.011267, -0.006494, -0.004936, 0.005678, 0.01708, 0.012821, 0.019666, -0.009574, -0.005175, -0.00153, 0.008066, 0.016601, 0.015412, 0.014075, -0.020463, -0.003049, 0.000584, -0.012575, 0.013302, 0.011694, -0.000616, -0.013676, 0.018919, 0.009672, -0.005222, -0.009169, 0.013405, 0.016721, -0.014268, -0.016521, 0.0078, 0.019311, 0.01302, 0.013647, -0.006352, -0.01144, 0.012963, 0.015951, -0.012729, -0.009787, 0.008951, 0.014204, 0.014488, 0.015052, -0.009985, -0.015122, -0.01276, -0.000652, 0.016413, 0.012819, -0.017922, 0.025687, -0.008644, -0.014013, 0.012122, 0.012221, 0.007084, -0.017099, -0.01473, 0.017407, 0.013321, -0.01029, -0.009717, 0.013975, 0.019156, -0.006644, -0.015818, 0.020065, 0.000983, -0.005487, 0.004946, 0.013232, -0.009826, -0.014203, 0.013308, 0.00936, -0.010505, -0.012281, 0.022151, 0.016451, -0.011882, -0.014841, 0.014814, 0.013488, 0.015504, 0.013662, -0.01152, -0.009514, 0.011774, 0.009316, 0.010922, -0.009506, -0.008972, -0.009671, 0.016089, 0.009199, -0.013384, -0.014392, 0.008068, 0.010853, -0.017433, -0.004365, -0.007553, 0.013398, 0.010908, -0.009884, -0.001649, -0.011921, 0.003544, 0.016423, -0.013159, -0.019088, 0.011322, 0.018793, -0.018654, 0.001358, 0.005798, 0.010099, 0.009448, 0.016237, -0.012746, -0.011605, -0.012845, -0.011443, 0.019311, 0.008419, -0.013151, 0.011319, 0.009714, 0.018898, 0.014658, 0.009429, -0.017882, -0.009682, -0.008455, -0.00627, 0.015141, 0.01356, 0.01799, 0.014892, -0.023739, -0.019709, -0.009079, -0.005859, 0.01504, 0.013439, -0.005076, -0.012888, 0.013833, 0.021414, 0.016767, -0.008977, -0.005835, 0.011184, 0.011467, -0.016347, -0.008913, 0.013016, 0.014582, 0.005197, 0.01626, -0.006674, -0.016643, -0.010268, -0.015226, 0.018815, 0.016538, -0.005496, -0.010821, 0.021801, 0.012361, -0.018638, -0.009224, 0.0091, 0.017824, -0.008444, -0.0079, 0.015085, 0.022338, 0.014109, 0.013418, -0.011328, -0.009606, -0.010026, -0.012702, 0.010906, 0.016271, -0.007043, -0.011943, 0.014273, 0.013792, -0.009978, -0.00541, 0.011517, 0.013871, -0.014162, -0.01135, 0.009841, 0.011348, 0.015298, 0.014283, -0.012895, -0.016879, -0.018977, -0.013698, 0.014278, 0.017141, -0.005903, -0.009323, 0.010201, 0.011017, 0.017148, 0.016069, -0.017962, -0.0116, -0.008851, -0.020297, 0.02036, 0.021338, -0.007109, -0.008963, 0.015869, 0.019204, -0.015683, -0.005508, 0.017911, 0.022761, -0.012897, -0.004865, 0.019637, 0.000981, -0.003359, -0.010809, 0.019248, 0.012052, -0.017785, -0.013383, 0.010867, 0.018989, -0.011699, -0.015185, 0.010751, 0.010922, -0.009472, -0.008005, 0.013425, 0.016135, -0.007531, -0.011175, 0.01431, 0.015078, -0.013314, -0.018636, 0.021225, 0.012631, -0.016181, -0.005383, 0.019151, 0.013063, -0.017009, -0.001878, 0.011674, 0.016596, -0.011049, -0.015275, 0.013413, 0.014413, -0.003504, -0.008839, 0.015021, 0.013623, -0.01319, -0.010771, 0.017324, 0.01708, -0.010084, -0.013879, 0.015605, 0.020139, 0.013688, 0.018161, -0.009152, -7e-05, -0.013015, -0.012744, 0.009079, 0.013814, -0.012939, -0.007712, 0.008774, 0.01687, -0.00862, -0.017491, 0.013229, -0.007701, -0.015576, 0.008447, 0.008625, -0.01352, -0.016887, 0.016541, 0.008571, -0.009172, -0.005388, 0.021455, 0.014362, -0.013007, -0.010394, 0.015207, 0.008032, -0.006139, -0.008479, 0.01568, 0.01622, -0.019857, -0.009089, 0.008291, 0.01321, -0.011315, -0.011052, 0.009992, 0.017169, 0.003452, 0.009137, -0.015139, -0.004009, -0.01421, -0.0162, 0.01434, 0.013302, -0.01176, -0.007204, 0.024537, 0.013526, 0.021396, 0.01094, -0.008247, -0.014914, -0.007155, -0.00968, 0.010674, 0.020761, -0.015249, -0.011201, 0.017619, 0.013657, 0.007113, 0.014772, -0.013454, -0.009335, -0.011443, -0.007451, 0.014658, 0.00961, -0.006491, -0.017819, 0.022414, 0.014914, -0.013454, -0.021103, 0.014539, 0.015424, 0.020031, -0.016945, -0.007233, 0.019654, 0.008776, -0.012226, -0.014172, 0.009676, 0.012934, -0.011152, -0.012083, 0.014409, 0.010019, 0.009504, 0.015351, -0.005617, -0.013079, -0.015625, 0.017431, 0.015668, -0.01208, -0.009274, 0.013984, 0.014143, -0.014997, -0.00823, 0.02049, 0.015979, -0.005458, -0.015476, 0.016337, 0.013773, -0.013309, -0.007783, 0.015078, 0.015463, -0.004062, -0.000352, 0.022444, 0.012983, -0.007341, -0.014851, 0.020106, 0.022714, -0.013667, -0.004533, 0.009614, 0.005935, -0.004612, -0.012656, 0.018032, 0.013287, -0.016967, -0.014018, 0.01263, 0.01145, -0.014691, -0.009584, 0.01173, 0.012285, -0.011657, -0.018963, 0.0227, 0.02078, -0.011786, -0.007478, 0.008413, 0.013366, -0.007202, -0.014248, 0.008694, 0.015194, -0.011701, -0.014245, 0.018979, -0.009533, -0.012788, 0.016584, 0.010268, -0.00935, -0.013945, 0.012121, 0.016217, -0.01102, -0.013838, 0.01684, 0.008598, -0.013657, -0.011266, 0.014526, 0.01467, -0.01726, -0.007937, 0.017222, 0.007382, -0.015161, -0.015114, 0.013278, 0.017006, -0.006681, -0.015229, 0.00971, 0.01444, -0.012702, -0.011858, 0.019864, 0.007723, -0.007626, -0.011165, 0.013452, 0.008481, -0.005228, -0.010882, 0.007231, 0.018217, -0.010766, -0.016394, 0.016203, 0.016638, -0.012156, -0.011166, 0.01258, 0.010326, -0.011694, -0.008395, 0.008346, 0.009333, -0.004641, -0.02234, 0.010166, 0.016654, -0.008631, -0.015625, 0.012521, 0.012707, -0.015992, 0.023522, 0.010642, -0.014064, -0.004486, 0.020866, 0.017437, -0.015941, -0.009978, 0.020217, 0.015085, -0.010478, -0.016151, 0.01218, 0.02072, -0.017928, -0.002727, 0.019215, 0.00373, 0.018122, 0.015987, -0.01567, -0.017198, -0.01215, -0.013337, 0.014679, 0.021782, -0.009011, -0.006761, 0.015318, 0.011429, 0.016838, 0.013747, -0.011083, -0.01135, -0.00941, -0.003193, 0.016084, 0.01916, -0.024938, -0.011623, 0.006318, 0.012097, -0.002636, -0.005912, 0.01692, 0.014957, -0.017239, -0.007536, 0.018438, 0.01809, -0.011728, -0.011093, 0.010688, 0.014743, -0.014041, -0.023282, 0.016728, 0.012648, 0.012453, 0.010577, -0.021187, -0.012364, -0.014401, -0.013889, 0.015438, 0.01173, -2.6e-05, -0.010872, 0.007167, 0.010578, -0.010163, -0.014691, 0.00955, 0.02206, -0.0151, -0.009576, 0.004211, 0.014974, -0.007312, -0.010935, 0.011945, 0.013972, -0.005595, -0.01634, 0.011923, 0.016057, 0.01342, 0.010461, -0.009763, -0.014339, -0.010366, -0.01904, 0.013751, 0.01918, -0.01405, -0.011667, 0.020102, 0.028811, -0.015146, -0.008537, 0.015334, 0.018378, -0.00613, -0.016461, 0.012504, 0.015173, -0.012765, -0.010471, 0.0189, 0.005475, 0.006025, 0.019633, -0.013845, -0.012331, -0.00917, -0.014321, 0.014521, 0.016711, -0.006458, -0.012131, 0.018912, 0.010836, -0.0129, -0.011146, 0.017002, -0.002293, -0.008117, -0.011555, 0.020125, 0.009572, -0.015441, -0.008946, 0.012161, 0.013369, -0.011015, -0.010227, 0.024008, 0.0129, 0.011039, -0.009995, -0.007536, 0.011602, 0.013072, -0.01321, -0.008854, 0.015476, 0.009995, -0.005017, -0.013567, 0.012182, 0.015471, -0.007893, -0.010384, 0.01538, 0.017663, -0.008498, -0.006901, 0.015479, 0.011996, -0.005756, -0.004274, 0.016608, -0.01414, -0.009316, 0.00772, 0.011999, -0.015136, -0.010031, 0.018936, 0.019519, 0.015211, 0.013725, -0.011929, -0.0126, -0.012985, -0.019511, 0.008068, 0.013877, 0.009743, 0.010869, -0.014399, -0.007907, -0.013471, -0.017246, 0.013674, 0.009692, 0.01561, 0.009074, 0.009871, 0.007824, -0.017016, -0.015878, -0.010249, -0.008022, -0.003408, -0.006772, 0.017675, 0.01072, 0.010686, -0.011594, -0.007077, -0.013171, 0.020005, 0.011191, -0.014892, -0.008486, 0.014341, 0.015522, 0.012356, 0.015078, -0.014546, -0.014128, -0.00999, -0.009838, 0.011397, 0.010398, -0.013021, -0.00188, 0.015229, 0.006291, -0.009294, -0.006665, 0.009619, 0.012762, -0.010307, -0.000939, 0.012446, -0.013048, -0.004717, 0.012284, 0.016257, -0.016921, -0.010683, 0.013657, 0.01326, -0.014334, -0.008852, 0.005698, 0.019047, 0.011119, 0.017243, -0.017225, -0.011653, -0.010604, -0.018416, 0.014135, 0.01216, 0.000652, -0.008581, 0.013263, -0.008025, -0.006533, 0.01489, 0.020683, -0.010336, -0.017967, 0.014791, 0.017304, -0.015535, -0.00773, 0.01205, 0.013945, -0.013553, -0.008425, 0.021269, 0.021108, -0.007897, -0.006273, 0.007258, 0.019306, -0.009436, -0.005595, 0.01752, 0.015146, -0.011694, -0.016185, 0.00979, 0.01144, -0.001305, -0.00742, 0.01142, 0.00795, -0.00734, -0.00135, 0.00936, 0.01138, -0.00434, 0.00143, 0.00945, 0.0143, -0.00782, -0.01126, 0.00906, 0.01103, -0.01359, -0.00649, 0.0106, 0.01149, -0.00448, -0.00643, 0.0061, 0.02318, -0.00332, -0.00185, 0.01097, 0.01734, -0.00988, -0.01043, 0.01431, 0.0119, -0.01029, -0.00709, 0.00751, 0.00407, 0.01442, -0.00635, -0.0082, 0.01799, 0.0128, -0.01826, -0.00595, 0.0098, 0.01624, -0.00806, -0.00536, 0.00961, -0.0034, -0.00539, 0.01388, 0.01297, -0.0091, -0.00389, 0.00804, 0.00824, 0.01673, 0.00895, 0.0, -0.00639, -0.01321, -0.00733, 0.01227, -0.0094, -0.00601, 0.02058, 0.01407, -0.00142, -0.0089, 0.00636, -0.00792, -0.00777, 0.01061, 0.00541, 0.0231, 0.01452, 0.00176, -0.00605, -0.00304, -0.00744, 0.01015, 0.01389, -0.00721, -0.00708, 0.01222, 0.01147, -0.0175, -0.00289, 0.01586, 0.00932, 0.0087, -0.0116, -0.00989, 0.00875, 0.01195, -0.00548, -0.01423, 0.01238, 0.00985, -0.00478, -0.01207, 0.01404, -0.01004, -0.00699, 0.00677, 0.02095, -0.00857, -0.00946, 0.00673, 0.00996, 0.00884, 0.00884, -0.00398, -0.0004, -0.01556, -0.00415, 0.01861, 0.0051, -0.00539, -0.01194, 0.0147, 0.01291, 0.01272, 0.01651, -0.00674, -0.00945, -0.00667, -0.01304, 0.00997, 0.01358, 0.01151, 0.00987, -0.01265, -0.00924, -0.00792, -0.00531, 0.00709, 0.01422, -0.00541, -0.00761, 0.01196, 0.01786, 0.01321, -0.00342, -0.00318, -0.0036, -0.01221, 0.01051, 0.01449, -0.00277, 0.00702, 0.01262, 0.01994, -0.00107, -0.00507, 0.01204, 0.01984, -0.01438, -0.00494, 0.01307, 0.01251, -0.01252, -0.01152, 0.00545, 0.00928, 0.01118, 0.00647, -0.00744, -0.00342, -0.00272, -0.00749, 0.0, 0.0015, 0.01722, 0.01038, -0.01478, -0.00871, -0.00692, -0.00859, 0.01829, 0.01261, -0.00932, 0.00896, 0.01366, -0.00199, -0.0019, 0.00581, 0.01433, -0.01397, -0.00671, 0.01056, 0.01255, -0.00905, -0.01491, 0.01586, 0.01214, -0.00854, -0.00058, 0.01375, 0.0067, -0.01011, -0.00771, 0.00903, 0.01565, -0.0021, -0.00911, 0.01381, 0.0234, -0.00462, -0.001, 0.00663, 0.0157, -0.00511, -0.00753, 0.0116, 0.01214, -0.00295, -0.00425, 0.01053, 0.01516, -0.0152, -0.00957, 0.01268, 0.01122, -0.01779, -0.01105, 0.01546, 0.00469, -0.00918, -0.00731, 0.01057, 0.00196, -0.00665, -0.00963, 0.00974, 0.00871, -0.00924, -0.00854, 0.00976, 0.01322, -0.0145, -0.00952, 0.00659, 0.00786, -0.00442, 0.00547, 0.00594, 0.00023, -0.00093, -0.00056, 0.00669, -0.00594, 0.00941, 0.00562, -0.00156, 0.01068, 0.00355, -0.00198, 0.0056, 0.00201, 0.01173, 0.00024, 0.0016, 0.00303, 0.00309, 0.00096, -0.00149, 0.01289, -0.00719, 0.00504, 0.00369, 0.00645, -0.00356, 0.0, -0.00223, 0.01216, 0.00311, 0.00157, -0.00472, -0.00102, 0.00669, 0.00036, 0.00754, 0.00016, -0.00287, 0.00124, 0.00203, -0.00039, -0.0131, 0.00632, 0.00655, 0.00533, 0.00146, -0.00101, 0.00531, 0.00296, -0.00456, 0.00243, 0.00285, 0.0, 0.00445, 0.00287, -0.00134, -0.0021, 0.00071, 0.00637, 0.00722, -0.00021, 0.00098, -0.0033, 0.00515, -0.00061, -0.00168, 0.004, 0.00389, 0.00427, -0.00076, 0.01102, -0.00415, -0.00145, 0.0079, -0.00057, 0.0003, -0.00217, -0.00105, -0.00049, 0.01096, 0.00487, 0.00128, 0.00029, 0.00085, 0.00119, 0.00121, 0.00556, 0.00513, -0.00163, -0.00122, -0.00639, 0.00326, 0.00515, 0.00044, 0.01472, 0.00217, -0.00571, -0.00347, 0.01083, -0.00042, 0.00117, -0.00141, 0.00458, 0.00447, 0.0, 0.00602, 0.01734, 0.00559, 0.00014, 0.00407, 0.00401, -0.00221, 0.00572, 0.00041, 0.01441, -0.0023, 0.00165, 0.00488, -0.00538, -0.00131, -0.0013, -0.00232, 0.00046, 0.0, 0.00618, 0.00521, 0.00229, -0.0004, -0.0009, 0.00227, 0.00093, 0.00263, -0.00485, -0.00346, 0.00657, 0.00105, -0.00566, 0.00452, -0.0005, -0.00106, 0.00373, 0.00389, -0.00093, 0.00383, -0.00136, -0.00289, 0.00139, 0.00039, 0.00231, -0.00024, -0.00402, 0.00114, -0.00633, -0.01035, -0.00131, -0.0063, 0.00115, 0.00796, 0.00328, 0.00061, 0.00022, 0.00372, 0.00227, 0.00415, 0.00162, 0.00012, -0.00288, 0.00066, -0.00517, -0.00446, -0.00099, 0.00711, -0.00031, -0.0008, 0.00972, 0.00069, -0.00096, 0.00039, -0.00584, -0.00089, -0.00208, 0.00212, 0.00054, -0.00029, -0.00344, 0.00254, 0.00116, 0.00143, -0.00296, -0.00368, 0.00026, -0.0003, -0.0031, -0.00516, 0.006, 0.0, -0.00305, -0.00191, -0.0021, 0.0, 0.00229, 0.0031, 0.00343, 0.00681, -0.00258, -0.00263, 0.00415, 0.00822, 0.00416, 0.00419, -0.00068, -0.00621, -0.00625, 0.0027, 0.00283, -0.00811, 0.00055, 0.00469, 0.00614, 0.00559, 0.00676, -0.00042, -0.00522, -0.00039, 0.005, 0.01043, -0.00686, 0.00501, 0.00433, 0.00272, -0.00639, 0.0, 0.00142, 0.00652, -0.00137, -0.00302, 0.00611, 0.00088, -0.00196, -0.00564, 0.00812, -0.00154, 0.00051, -0.00445, 0.01202, -0.00014, -0.00174, 0.00692, 0.00163, 0.00728, -0.00161, -0.00353, 0.00486, -0.00264, -0.00444, -0.0022, 0.00309, 0.00089, 0.00103, -0.0043, 0.00611, -0.00646, -0.00319, 0.00354, 0.00654, 0.00126, 0.00175, 0.00194, 0.00708, -0.00751, 0.00241, -0.0028, 0.00403, 0.0049, -0.00104, 0.00279, 0.0025, 0.00435, -0.00327, -0.00083, 0.00126, 0.00396, 0.00328, 0.00659, 0.00821, 0.00137, 0.0045, 0.00321, 0.01291, 0.00036, -0.0018, 0.0108, 0.00038, 0.00827, 0.00556, 0.00886, 0.0, 0.00118, 0.00345, 0.00341, 0.00073, 0.00193, 0.00352, 0.00205, 0.00122, -0.00243, -0.00075, 0.00723, -0.00182, -0.00055, 0.00511, 0.00219, 0.00916, 0.0024, 0.00586, 0.00129, -0.00376, 0.00177, 0.00345, 0.00557, -0.00059, -0.00614, 0.01533, 0.0072, -0.00379, -0.00351, 0.00733, 0.00337, 0.00345, -0.00296, 0.00474, 0.01143, 0.00184, -0.00183, 0.00326, 0.00403, -0.00295, -0.00378, -0.00209, -0.00098, 0.00679, 0.00554, 0.01901, -0.00372, -0.00014, -0.00341, -0.00111, 0.01099, 0.00855, 0.00177, 0.0, -0.00052, 0.00416, 0.01309, -0.0006, -0.00422, 0.01259, -0.00124, 0.00738, -0.00339, 0.00731, 0.00836, -0.01253, -0.006, 0.00505, 0.01272, 0.0073, -0.00264, -0.00319, 0.00577, -0.00594, -0.00095, 0.00826, 0.00467, 0.00153, -0.00573, -0.00474, 0.00523, -0.006, -0.00335, 0.00314, 0.00585, 0.0096, -0.00178, -0.00258, -0.0016, 0.00685, -0.00339, 0.00102, 0.00422, 0.00528, -0.00171, -0.00465, 0.00852, -0.00138, -0.00528, 0.00178, 0.00938, 0.00655, 0.00198, -0.006, 0.01373, 0.00242, -0.00279, -0.00744, 0.00597, 0.00813, -0.00037, 0.0, 0.00784, 0.00769, -0.0035, 0.00262, -0.00292, -0.002, -0.00769, 0.00816, 0.00511, -0.00605, 0.00575, 0.00626, -0.00811, -0.00163, 0.00808, 0.00389, -0.00062, 0.00198, 0.01455, -0.00205, -0.00063, 0.01408, 0.00233, -0.00209, -0.00804, 0.0117, 0.00164, -0.00801, -0.00347, 0.01128, 0.01589, -0.00276, -0.00079, 0.00548, 0.00886, -0.00605, -0.00117, -0.00032, 0.00165, -0.00229, 0.00808, 0.00205, -0.00276, -0.00754, 0.00337, 0.00691, -0.00303, 0.00315, 0.00892, 0.00961, 0.00172, -0.00169, -0.00194, 0.00827, 0.00158, 0.01111, 0.00835, 0.00732, -0.00541, 0.00263, 0.01327, 0.00563, 0.0, 0.00172, 0.00554, 0.00345, 0.00041, 0.01466, 0.00693, 0.00319, -0.00033, 0.00156, 0.00904, 0.01463, 0.00311, 0.00317, 0.00414, 0.0, 0.00319, 0.00854, 0.00803, 0.00118, -0.00639, 0.0084, 0.00614, 0.00494, 0.00845, 0.00343, 0.0134, 0.00359, 0.0, -0.001, 0.00589, 0.00368, -0.00252, 0.00203, 0.0, 0.00698, 0.00733, 0.01077, 0.00558, 0.01181, 0.0012, -0.00871, 0.01012, 0.00319, -0.00118, -0.00656, -0.00977, -0.00169, 0.0071, 0.00766, 0.00279, 0.0018, 0.00016, -0.01383, 0.00805, 0.00621, 0.00147, 0.00422, -0.00229, 0.01081, 0.00793, -0.00894, 0.00779, 0.01012, 0.01486, -0.00242, 0.00176, 0.0074, 0.0034, 0.00101, -0.0003, 0.01712, 0.00349, -0.00491, 0.00034, -0.00111, 0.00462, -0.002, 0.00696, 0.00179, 0.00898, -0.00163, 0.01015, 0.00746, 0.01057, 0.015025, 0.015447, -0.016106, -0.025225, 0.026282, 0.020133, -0.021225, -0.023918, 0.014392, 0.028044, -0.025925, -0.016173, 0.024125, 0.018965, -0.013222, -0.006369, 0.029306, 0.021462, -0.018178, -0.018433, 0.027439, 0.025036, -0.009913, -0.016166, 0.030425, 0.01933, -0.023227, -0.014423, 0.015931, 0.014113, -0.021182, -0.017395, 0.019111, 0.016135, -0.004167, -0.025869, 0.021066, 0.013959, -0.019687, -0.012765, 0.025561, 0.026064, -0.025811, -0.021694, 0.030285, -0.013229, -0.023425, 0.026882, 0.024605, -0.017768, -0.017851, 0.024551, 0.015946, -0.016894, -0.01774, 0.022514, 0.019697, -0.010645, -0.020522, 0.022071, 0.016552, -0.020749, -0.021909, 0.018806, 0.0235, -0.018229, -0.014546, 0.022944, 0.015796, -0.017287, -0.018252, 0.021199, 0.027759, -0.020882, -0.017189, 0.024929, 0.020074, -0.02161, -0.022748, 0.018576, 0.026618, -0.01518, -0.017343, 0.015886, 0.01869, -0.0151, -0.021831, 0.017017, 0.016483, 0.019992, 0.020415, -0.011095, -0.02443, -0.026594, -0.016574, 0.020321, 0.012941, -0.013376, -0.017155, 0.017549, 0.020756, 0.01766, 0.015131, -0.026362, -0.010957, -0.013899, -0.016208, 0.024191, 0.030875, -0.009122, -0.018005, 0.020212, -0.020307, -0.017535, 0.020859, 0.018315, -0.011757, -0.011479, 0.019672, -0.023562, -0.016564, 0.030086, 0.02436, -0.018679, -0.019801, 0.015037, 0.018122, -0.021343, -0.020072, -0.010101, 0.024238, 0.016794, 0.024399, -0.021457, -0.017707, 0.015988, 0.020925, -0.013971, -0.015683, 0.024571, 0.020012, -0.01547, -0.018532, 0.03049, 0.021508, -0.01918, -0.020947, 0.022961, 0.016367, -0.014885, -0.012114, 0.028942, 0.013908, -0.013188, -0.019685, 0.026768, 0.017065, -0.011047, -0.006843, 0.01999, 0.020727, -0.014438, 0.017253, 0.021622, 0.018666, -0.018578, -0.01538, 0.024707, 0.017812, -0.017288, -0.017231, 0.027399, 0.022108, -0.013706, -0.019299, 0.02132, 0.025097, -0.017241, -0.024194, 0.0229, 0.023959, -0.022555, -0.019787, 0.018994, 0.025129, -0.007822, -0.005902, 0.028815, 0.019418, 0.024421, -0.014599, -0.013862, 0.028117, 0.024208, -0.014904, -0.017989, 0.020469, 0.021532, -0.020437, -0.015052, 0.019441, 0.01426, -0.016277, -0.024056, 0.021594, 0.027009, -0.014258, -0.016642, 0.022414, 0.024453, 0.016932, 0.02239, -0.018501, -0.019299, -0.0178, -0.021074, 0.012509, 0.018327, 0.012766, 0.018063, -0.018351, -0.024355, 0.02488, 0.021357, -0.017456, -0.021022, 0.016483, 0.018415, -0.018967, -0.019147, 0.033156, 0.021764, -0.012973, -0.017097, 0.018726, 0.01897, -0.016152, -0.02039, 0.018412, 0.016891, -0.014282, -0.01672, 0.025539, 0.015686, -0.015059, -0.021165, 0.019652, 0.020597, -0.018776, -0.015434, 0.018896, 0.02445, -0.019398, -0.021859, 0.019045, 0.021782, -0.012248, -0.01582, -0.021537, 0.019284, 0.015375, 0.018487, 0.014583, -0.019062, -0.016005, -0.020878, -0.029555, 0.022103, 0.021399, -0.012501, 0.025441, 0.017536, -0.016018, -0.025622, 0.014832, 0.019995, -0.022004, -0.019241, 0.020546, 0.029693, 0.02508, 0.018951, -0.022248, -0.018445, -0.016928, -0.021469, 0.018991, 0.018647, -0.021341, -0.016939, 0.019234, 0.016038, -0.015794, -0.014756, 0.020179, 0.024592, -0.01363, -0.021437, 0.022257, 0.021367, 0.022767, 0.025441, -0.021137, -0.021303, -0.02056, -0.017019, 0.019543, 0.014926, -0.013244, -0.01457, 0.018229, 0.019526, -0.021107, -0.019855, 0.01118, 0.01571, -0.019868, -0.019008, 0.021728, 0.023154, 0.027029, 0.021025, -0.013803, -0.023418, -0.02838, -0.024759, 0.026814, 0.029512, -0.017996, -0.0132, 0.027693, 0.017167, -0.029775, -0.024558, 0.013947, 0.025726, 0.01758, 0.023766, -0.020273, -0.025158, -0.018361, -0.012958, 0.01457, 0.027892, -0.02284, -0.024818, 0.018008, 0.01822, -0.017643, -0.017255, 0.0257, 0.023591, -0.020148, -0.022537, 0.025324, 0.020971, -0.026115, -0.016296, 0.013242, 0.020937, -0.010524, -0.014069, -0.013947, 0.013559, 0.015883, -0.027299, -0.024411, 0.024074, 0.01589, -0.023796, -0.024855, 0.027292, 0.017701, -0.022502, -0.01481, 0.01875, 0.018408, -0.014521, -0.019233, 0.011413, 0.011403, -0.013676, -0.020984, 0.017203, 0.021958, -0.018784, -0.016666, 0.018115, 0.016009, -0.024097, -0.016236, 0.018815, 0.017798, -0.012726, -0.020771, 0.018484, 0.024049, -0.02815, -0.014851, 0.02123, 0.018929, -0.016927, -0.025698, 0.020187, 0.015673, -0.024058, -0.00963, 0.019965, 0.020327, -0.024588, -0.019642, 0.015727, 0.027846, -0.020405, -0.013558, 0.016931, 0.022702, -0.013435, -0.019756, 0.02757, 0.012898, -0.017136, -0.018042, 0.021897, 0.018896, -0.020095, -0.015913, 0.014689, 0.011459, -0.023949, -0.018714, 0.021506, 0.018632, -0.025764, -0.018149, 0.015992, 0.016417, -0.017531, -0.023853, 0.025275, 0.027643, -0.028738, -0.021032, 0.022175, 0.020786, 0.025861, -0.019682, -0.016879, 0.022801, 0.023118, -0.010077, -0.021533, 0.022168, 0.023806, -0.023876, -0.012661, 0.024273, 0.019267, -0.015866, -0.012092, 0.020101, 0.024841, -0.011119, -0.024044, -0.019958, 0.017993, 0.015987, -0.0154, -0.013744, 0.021289, 0.017045, -0.015528, -0.024706, 0.025562, 0.021851, -0.01602, -0.016152, 0.023945, 0.024609, -0.01197, -0.025393, 0.019069, 0.019765, -0.03066, -0.013911, 0.022867, 0.028455, -0.021347, 0.015593, 0.01798, 0.015114, -0.01343, -0.013496, 0.021058, 0.020622, -0.011263, 0.020299, 0.020043, -0.017773, -0.01156, 0.025269, 0.026955, -0.025375, -0.022231, 0.022289, 0.017875, 0.022226, 0.021772, -0.020493, -0.017123, -0.015668, 0.017669, 0.020701, -0.012789, -0.015424, 0.011118, 0.018926, -0.029719, -0.011933, 0.019426, 0.022883, -0.020682, -0.018184, 0.016843, 0.025537, -0.013002, -0.020615, 0.027131, 0.01925, -0.015206, -0.018619, 0.018054, 0.015837, -0.018579, -0.01808, 0.026906, 0.027031, -0.020683, -0.017817, 0.018728, 0.030551, -0.015412, -0.022562, 0.01932, 0.012792, -0.011675, -0.021228, 0.021281, -0.020185, -0.022364, 0.023754, 0.040502, -0.012041, -0.025573, 0.02138, 0.019304, -0.011489, -0.017928, 0.025888, 0.01519, -0.022326, -0.023107, 0.017258, 0.023198, -0.015971, -0.019287, 0.022609, 0.029824, -0.020152, -0.02153, 0.023217, 0.014945, -0.020894, -0.021323, 0.023132, 0.013667, 0.025673, -0.014247, -0.026031, 0.019853, 0.018666, 0.015129, 0.019792, 0.021213, 0.01815, -0.018461, -0.021322, -0.016475, -0.015139, -0.020043, -0.013432, 0.02385, 0.017029, -0.018443, -0.018027, -0.012246, 0.025049, 0.018105, 0.019748, 0.028271, -0.017701, -0.019553, -0.016584, -0.020991, 0.021585, 0.022998, -0.012243, -0.006586, 0.015864, 0.014248, 0.030117, 0.024406, -0.016074, -0.020676, -0.017401, -0.017746, 0.01967, 0.022204, 0.019407, -0.01737, -0.019399, 0.016724, 0.020031, -0.021591, -0.018738, 0.01984, 0.02283, -0.017656, -0.020435, 0.018465, 0.019429, -0.022789, -0.019023, 0.015603, 0.027667, -0.015063, -0.013645, 0.019082, 0.021972, -0.016882, -0.021764, 0.025823, 0.011044, -0.020073, -0.020887, 0.023618, 0.015698, -0.013745, -0.019579, 0.00842, 0.00548, 0.00585, 0.00818, 0.0097, 0.0121, 0.00242, 0.00308, 0.01044, 0.00755, 0.00959, -0.01006, 0.01121, -0.00498, 0.01415, -0.0077, 0.01118, 0.01036, 0.01076, -0.00466, 0.01155, 0.00925, -0.00468, 0.01308, 0.00875, 0.00924, -0.00801, -0.0058, 0.00693, -0.01083, 0.00933, 0.01103, 0.00642, -0.00971, -0.01527, -0.01309, -0.0086, 0.00506, 0.02145, 0.01839, 0.01358, -0.00442, 0.01769, -0.00072, -0.00437, -0.00289, -0.00608, 0.01492, 0.00877, -0.01034, 0.01192, -0.00302, -0.00387, 0.01044, -0.01045, -0.01073, 0.00927, -0.0061, 0.00593, -0.01001, -0.00724, 0.01197, 0.01468, -0.00365, -0.00501, -0.00202, 0.00941, -0.00601, -0.00748, 0.01012, -0.0059, 0.011, 0.01582, -0.00197, -0.01057, -0.00754, -0.00924, 0.00829, -0.00411, -0.00642, 0.00431, 0.00396, -0.01067, -0.00871, -0.00551, -0.00391, 0.0157, -0.0024, -0.0147, -0.01276, -0.00033, 0.00101, 0.00651, 0.00586, -0.0085, 0.01039, 0.0174, -0.00137, 0.00861, 0.01175, 0.015, 0.01511, 0.01041, 0.01381, 0.0222, 0.0223, 0.00603, -0.01459, 0.00731, -0.00748, 0.01356, 0.01745, 0.00529, -0.01145, 0.00836, -0.01117, -0.00878, -0.00754, 0.00522, 0.0255, 0.01629, 0.01636, 0.01267, -0.00167, 0.01348, 0.01467, 0.0169, 0.01404, 0.00822, 0.00719, 0.00038, 0.00159, 0.00694, 0.00956, 0.00845, -0.00986, -0.00085, -0.0006, 0.00691, -0.01071, 0.00553, -0.0111, 0.01131, 0.01002, 0.01314, 0.01547, -0.00551, 0.013, -0.01085, -0.00843, 0.0109, 0.0091, -0.00665, -0.0034, 0.01371, 0.01129, -0.00764, -0.00912, 0.00069, 0.01363, 0.01186, 0.00941, 0.00959, 0.01278, 0.00618, 0.00635, 0.0139, -0.00298, 0.00947, -0.00743, -0.01002, -0.00828, -0.01238, 0.00398, 0.00967, -0.00829, -0.00251, 0.01505, 0.01162, -0.00343, 0.01704, 0.01957, -0.0095, -0.00747, -0.00789, 0.01098, -0.0078, 0.00797, -0.00765, 0.01203, 0.01239, -0.00696, -0.00408, -0.01033, 0.00789, 0.01533, 0.01409, 0.01217, -0.0039, -0.00694, 0.00982, -0.01157, -0.00894, -0.00634, 0.0107, 0.00843, -0.00918, 0.01022, 0.01113, 0.01365, 0.01634, -0.01446, -0.01073, -0.00364, -0.00403, 0.00549, -0.00811, -0.01134, -0.00801, 0.01653, -0.00157, -0.00679, 0.0108, -0.00352, 0.01316, 0.01247, -0.00677, 0.00781, -0.00727, -0.00189, 0.01588, 0.00996, 0.00958, 0.0104, 0.01134, 0.01132, -0.00368, 0.0113, 0.01117, 0.01181, -0.00542, 0.00835, -0.00738, 0.00107, 0.00483, 0.00242, -0.00062, 0.00498, 0.00586, -0.00226, -0.00222, -0.00799, -0.00691, 0.0133, -0.00155, -0.00048, -0.0146, 0.0057, 0.00823, -0.00885, -0.00676, 0.00449, 0.0244, 0.00783, -0.00873, 0.00499, 0.00541, 0.0093, -0.007, 0.01165, -0.00718, 0.00449, -0.01565, 0.01782, -0.00148, 0.01208, 0.01194, -0.0119, -0.01153, -0.00952, -0.00984, -0.00793, 0.01043, 0.01357, 0.01024, 0.00528, -0.01039, -0.01073, 0.0071, -0.00939, 0.005, 0.01296, 0.01249, 0.01869, 0.02078, 0.00715, -0.012, -0.01247, -0.01188, 0.01283, -0.00335, 0.01293, -0.00423, -0.00963, -0.01005, 0.0076, -0.01045, 0.00975, 0.00765, 0.00507, -0.00756, -0.00928, -0.00977, 0.01206, -0.00413, 0.01142, 0.01268, 0.00909, -0.01087, 0.00602, 0.00798, -0.00815, -0.00792, 0.00956, 0.00445, -0.01342, 0.01065, -0.00777, 0.00357, 0.00607, 0.00342, 0.00104, -0.00752, -0.01048, 0.01134, -0.00533, 0.0097, -0.00876, 0.01335, 0.01444, 0.00988, -0.00733, 0.00898, 0.00814, 0.0106, 0.0122, -0.01196, -0.01221, 0.01076, -0.007, 0.01519, 0.01566, 0.01039, 0.01196, -0.01362, -0.00471, 0.01145, 0.0119, -0.00096, 0.00785, -0.01586, 0.01052, 0.00941, 0.00114, -0.01453, 0.00634, 0.00905, 0.0076, 0.00896, 0.00821, -0.00622, -0.00727, 0.00735, 0.00378, 0.00988, 0.00502, 0.00788, -0.01107, -0.00288, 0.00945, 0.00685, 0.00905, 0.01608, 0.00964, 0.01737, 0.01063, -0.0068, 0.00557, 0.0096, -0.0153, 0.02054, 0.00944, -0.01226, 0.01355, 0.00331, 0.00527, -0.00895, -0.01289, -0.00861, 0.01089, -0.00757, 0.0082, 0.01178, -0.01569, 0.01792, 0.01013, -0.00934, 0.01362, 0.00915, -0.00268, 0.00269, 0.00331, 0.01112, 0.01319, 0.0118, -0.00525, 0.00806, 0.00983, 0.00147, -0.00935, 0.0094, 0.00925, -0.00596, -0.0136, -0.00636, 0.00822, 0.01102, 0.00535, -0.01007, -0.00235, -0.00382, 0.00755, 0.00748, 0.00645, 0.00577, 0.00141, 0.01054, -0.00492, 0.00734, 0.00776, 0.01, 0.0076, 0.01041, 0.00376, 0.01049, -0.00897, 0.0166, 0.00893, 0.01048, 0.01082, -0.01051, -0.00963, -0.00499, 0.00967, 0.00697, -0.00647, 0.00901, 0.00885, -0.01084, 0.00633, 0.01372, 0.00863, 0.00865, 0.00691, -0.00884, 0.01155, -0.0086, -0.00641, -0.00386, -0.00731, -0.01379, 0.00979, -0.01373, 0.00766, 0.01155, 0.01643, 0.00071, 0.01127, 0.00598, 0.01141, -0.00529, -0.0095, -0.00957, 0.00779, 0.00212, 0.0073, 0.00693, -0.00634, 0.00384, 0.01015, 0.0024, 0.01024, 0.01021, 0.00975, -0.00945, 0.00465, -0.00913, 0.01077, -0.01076, -0.00699, -0.01393, -0.01214, -0.00901, 0.00293, 0.00075, 0.00504, -0.01088, -0.00814, -0.00398, -0.00859, -0.01142, 0.01089, 0.01228, -0.0132, -0.01441, -0.01774, -0.0039, 0.01151, 0.00978, 0.00857, 0.00319, 0.00119, -0.00961, 0.00623, -0.02012, 0.01009, 0.0068, -0.01114, 0.01187, -0.01103, -0.01016, 0.00484, 0.00755, -0.00578, -0.00624, -0.01314, -0.00929, -0.01918, 0.01302, 0.00771, -0.01755, -0.00599, 0.01414, 0.00518, -0.00706, 0.01622, -0.00624, 0.00755, -0.00989, 0.00379, 0.00576, 0.01419, -0.01273, -0.01128, -0.00684, 0.00678, 0.0063, -0.01053, -0.00252, 0.01227, 0.00344, 0.00462, -0.00874, 0.00542, 0.00946, 0.00916, -0.00863, -0.00979, -0.0064, 0.00753, 0.00742, 0.01124, 0.00621, 0.00428, 0.01055, 0.0, 0.01206, -0.01485, -0.00395, -0.01145, -0.01271, -0.01275, 0.01277, -0.00761, 0.00128, -0.00988, -0.00758, -0.01175, -0.01091, -0.00438, -0.0075, 0.01417, -0.00972, -0.0067, -0.00705, -0.00661, 0.00081, 0.0105, -0.00874, 0.00572, 0.00244, -0.00315, 0.0043, -0.00824, 0.00604, 0.00549, -0.01385, -0.00359, -0.01208, -0.01466, 0.00266, 0.00783, 0.01007, 0.00991, 0.00435, -0.0069, 0.00795, 0.00529, 0.00879, -0.01021, -0.00957, -0.00726, -0.01041, -0.00884, 0.00245, 0.00711, -0.00653, 0.01373, 0.00522, -0.00826, 0.00364, 0.00571, -0.00755, -0.00829, -0.00644, 0.00636, -0.00639, 0.00664, 0.00843, 0.00542, 0.01236, -0.00896, -0.00733]]
    table_category_data = [[['CWNJ_C02-2FAP-23_2_FCT', '#0000FF', 'Accelerometer AVG-FS8g_ODR200HZ_Zup accel_normal_average_x', 0.118, -0.118, 0.00875, 0.00072, 0.00518, 0.00557, 0.01365, 0.01179, -0.00121, -0.00049, 0.00526, 0.00965, -0.00106, 0.00051, 0.00852, 0.00428, 0.00519, 0.00496, 0.005, 0.00248, 0.00018, 0.00764, 0.01148, 0.00595, -0.00527, 0.00503, 0.01028, 0.00534, -0.00191, -0.00129, 0.00439, 0.00432, -0.00367, 0.00559, -0.00143, -0.00065, -0.00367, -0.00043, 0.00315, 0.00388, -0.00243, -0.00054, 0.00497, 0.00243, 0.00117, 0.01055, 0.00327, 0.00151, 0.00636, 0.00099, 0.00127, 0.00827, 0.00287, -0.00182, 0.00511, 0.0093, 0.00242, 0.00183, 0.00194, -0.00507, -0.00036, -0.00039, 0.00368, 0.00366, -0.0031, 0.00762, 0.00058, 0.00178, -0.00225, 0.00183, 0.00567, 0.00214, 0.0012, -0.00293, -0.0005, 0.00338, 0.00502, 0.00443, 0.0033, 0.00361, -0.00357, 0.00893, 0.00096, 0.00365, -0.00223, 0.0, -0.00328, -0.00542, -0.00548, -0.00518, 0.00198, 0.0015, -0.00436, 0.00074, 0.00138, 0.00552, -0.00177, 0.00191, 0.00676, 0.0099, 0.00314, -0.00677, 0.00819, 0.00945, 0.00082, 0.00697, 0.0122, 0.00035, 0.00974, 0.00239, 0.0014, 0.0019, 0.00311, 0.00384, 0.00639, 0.00892, 0.004, -0.00119, 0.00178, -0.00302, -0.00977, 0.0021, 0.00203, 0.00083, -0.00107, 0.00389, -0.00053, -0.00197, -0.00033, -0.00049, 0.00606, 0.00895, 0.00455, 0.00439, 0.00142, 0.00067, 0.00887, 0.0102, 0.00646, 0.00691, 0.00553, 0.00781, -0.00123, 0.00055, 0.00351, 0.00455, 0.00847, 0.00327, 0.00153, 0.0035, 0.0026, 0.00848, 0.00787, 0.00123, 0.00322, -0.00556, 0.00712, -0.00121, 0.00447, -0.00488, 0.00298, 0.00297, -0.00032, 0.00304, 0.00035, 0.00993, -0.00212, 0.00247, 0.00455, 0.0041, 0.00084, 0.01233, 0.00204, 0.01306, 0.00516, -0.00209, 0.00211, 0.00103, 0.00104, 0.00355, 0.00283, 0.00667, 0.00876, -0.00113, 0.00068, 0.00419, -0.00218, -0.002, 0.01113, 0.00554, 0.00565, -0.00136, 0.00227, 0.00212, 0.0047, -0.00134, 0.00019, 0.00649, 0.00318, 0.00215, 0.01619, 0.0, 0.00771, 0.00591, 0.00334, 0.00762, 0.0065, -0.00175, 0.01151, -0.00248, 0.0, -0.00556, -0.0032, 0.01536, 0.00111, 0.0, -0.00366, 0.00797, 0.00591, 0.00599, 0.00426, 0.00271, 0.00653, -0.00788, 0.00677, 0.00132, 0.00886, 0.00357, 0.00498, 0.00723, 0.00269, -0.00103, 0.00027, 0.00012, -0.00061, 0.00578, 0.00399, -0.00108, 0.01251, -0.00014, 0.0, 0.00984, -0.00232, 0.00253, 0.00401, -0.00397, 0.00075, 0.00189, 0.00784, -0.00113, -0.0015, 0.00247, 0.0024, -0.00329, -0.00318, 0.0055, 0.01047, 0.00128, -0.00191, 0.00118, 0.00627, 0.00707, -0.00645, -0.00263, 0.00165, -0.00501, 0.00981, -0.00138, 0.00897, 0.01069, -0.00348, -0.00012, 0.00391, 0.0032, 0.00252, 0.00384, 0.00428, -0.00075, -0.0006, 0.00819, -0.00143, 0.00364, 0.00462, 0.00568, 0.00716, -0.00244, -0.00197, 0.00307, 0.00022, 0.00054, 0.00928, 0.0088, 0.00781, 0.0016, 0.00226, 0.01081, -0.00241, -0.00135, 0.00544, 0.00406, 0.00222, 0.00878, 0.00536, 0.00996, 0.00435, -0.00011, 0.0022, 0.01029, 0.00737, 0.00277, 0.00526, 0.00963, 0.00767, 0.00582, 0.00771, -0.00221, 0.00054, -0.00336, 0.00551, 0.00238, 0.00134, -0.00063, 0.00241, 0.00757, 0.00249, -0.00662, 0.00273, 0.00451, -0.00323, 0.00172, 0.00843, 0.00793, 0.00357, 0.00308, 0.00367, 0.00166, 0.00431, -0.00311, 0.00185, 0.00043, 0.00367, -0.00021, 0.0025, 0.01089, 0.00215, 0.00421, 0.00332, 0.00423, 0.0098, 0.00606, 0.00078, 0.00267, 0.00264, -0.00228, 0.00018, 0.00327, -0.00085, -0.00036, 0.00522, -0.00739, 0.00371, 0.00092, 0.00419, 0.00376, -0.00135, 0.00128, 0.00467, 0.00258, -0.00135, 0.0, 0.00597, 0.00585, -0.00105, 0.00124, 0.00069, -0.00287, 0.00414, 0.00399, 0.00416, 0.00983, -0.00291, -0.00259, 0.0072, 0.00357, 0.00435, -0.0005, -0.00043, 0.00458, 0.00247, 0.00096, 0.00196, 0.00291, -0.00186, 0.0, 0.00079, 0.00339, 0.00185, -0.00176, -0.0031, -0.00521, 0.00719, 0.00399, -0.00114, 0.00971, 0.00365, 0.00145, 0.00768, 0.00078, -0.00888, 0.0044, -0.00357, -0.00056, 0.00196, -0.0045, -0.00073, 0.00105, 0.00579, 0.00432, 0.00327, 0.00585, 0.00034, -0.00057, 0.01771, -0.00921, 0.00812, 0.00149, 0.00284, 0.0078, 0.00894, 0.00582, -0.00047, 0.00835, 0.00485, 0.00192, -0.00314, -0.00488, -0.00053, -0.00093, 0.00124, 0.00261, 0.01084, 0.00348, -0.00029, 0.00116, -0.00098, 0.00347, 0.0032, 0.00548, 0.0042, 0.00402, 0.0023, -0.00346, 0.00643, 0.00291, 0.00351, 0.00144, 0.00065, -0.00122, -0.00333, 0.00151, -0.00192, 0.00755, -0.00083, 0.00541, -0.00179, 0.00787, 0.00095, 0.00946, 0.00824, -0.00154, 0.00437, 0.0037, -0.00067, 0.00789, 0.00601, -0.00475, -0.00322, -0.00275, -0.00265, 0.00067, 0.0028, 0.00603, -0.00111, -0.00265, -0.0031, 0.00501, 0.00158, 0.00271, 0.001, 0.00032, 0.0082, -0.0002, 0.00744, 0.00376, 0.00141, 0.00213, 0.00447, -0.00552, 0.00268, 0.00911, 0.00821, -0.00092, 0.00199, 0.00172, 0.00402, 0.00758, -0.00021, 0.00107, 0.00216, -0.00197, 0.00248, -0.00195, 0.0081, -0.00049, 0.00185, 0.00434, 0.0087, 0.00259, -0.00014, 0.00112, 0.00741, -0.00685, -0.0015, -0.00587, 0.00307, 0.00212, -0.00749, 0.0062, 0.00332, 0.00272, -0.00107, 0.00336, 0.01341, 0.00788, 0.00632, -0.00883, -0.00226, 0.0, 0.00094, -0.00049, 0.00164, 0.00258, 0.00584, 0.00456, 0.00034, -0.00264, -0.00067, 0.0087, 0.00306, 0.00547, 0.00212, 0.00761, 0.00625, 0.01027, 0.0029, 0.00311, 0.00086, 0.00828, 0.0, 0.00522, 0.00261, 0.00692, -0.00212, 0.00019, 0.00081, -0.00266, -0.00439, 0.00694, 0.00332, 0.00461, 0.00299, 0.0029, -0.00195, 0.00139, 0.0013, 0.00649, 0.00808, 0.00329, -0.00024, 0.00603, -0.00043, -0.00962, 0.00142, 0.00735, 0.00189, 0.00156, 0.0051, 0.00626, -0.0032, 0.00445, -0.00117, 0.00025, 0.00363, 0.00632, 0.0061, -0.00481, 0.00425, 0.00205, -0.00291, 0.00648, 0.00266, -0.00453, -0.00364, 0.00078, 0.00574, -0.00375, -0.00065, 0.00798, 0.00473, -0.00243, 0.00481, 0.00208, 0.00314, 0.00211, 0.00063, 0.01095, -0.0014, 0.00167, 0.00309, -0.00172, -0.00492, 0.01533, 0.00777, 0.00297, 0.00174, 0.0026, 0.00509, -0.00263, 0.00081, 0.00064, 0.00056, 0.01097, 0.00678, 0.00319, 0.00486, -0.00047, 0.0069, -0.00358, 0.00849, -0.00189, 0.00635, 0.00152, 0.00155, -0.00072, 0.00825, 0.00077, 0.00217, 0.0015, 0.00743, 0.00372, 0.00938, -0.00219, 0.00026, -0.00275, 0.01011, 0.00066, 0.00178, 0.0, 0.0043, 0.00092, -0.00581, 0.00058, 0.00172, 0.00545, -0.00048, 0.00544, -0.00105, 0.00159, 0.00653, 0.00148, 0.00034, 0.00087, 0.00137, 0.00282, -0.00142, 0.00033, -0.00736, -0.00123, 0.0081, 0.00205, 0.00116, -0.00519, 0.00673, -0.00191, 0.00569, -0.00581, -0.00523, 0.00611, 0.00362, 0.00188, 0.0042, 0.00493, 0.0057, -0.00566, 0.00557, 0.00383, 0.00339, -0.00343, 0.00245, -0.00094, 0.00056, 0.00454, 0.01121, 0.00459, -0.00768, -0.00978, 0.01015, 0.00968, -0.0013, -0.01728, 0.01083, 0.00996, -0.01279, -0.01211, 0.01378, 0.00953, 0.00228, 0.00254, 0.01177, -0.00642, -0.0092, 0.01164, 0.01548, 0.00204, 0.00114, 0.00586, 0.00322, -0.00396, 0.00066, 0.0152, 0.00822, -0.002, -0.00976, 0.00776, 0.01003, -0.00367, -0.0072, 0.00086, 0.00443, -0.00067, -0.00029, 0.00087, 0.01087, -0.00288, 0.0, -0.0027, -0.00787, 0.00457, -0.00311, -0.00062, 0.00588, 0.00623, -0.00081, -0.00165, 0.01097, 0.01587, 0.0, 0.00111, 0.00177, 0.00069, 0.00351, 0.00254, 0.00743, 0.00208, -0.005, -0.00457, 0.00093, 0.0053, 0.00345, 0.00339, -0.00152, -0.00101, -0.00065, 0.00048, 0.01086, 0.006, 0.00118, 0.00162, 0.00717, 0.00371, 0.00887, -0.00597, -0.00253, 0.00893, -0.001, -0.00502, 0.00125, 0.00271, 0.00121, -0.00334, 0.00235, -0.00103, 0.00645, -0.00121, 0.00625, 0.0, 0.00772, -0.00362, 0.00366, 0.01088, 0.00238, -0.0015, 0.00934, -0.00462, -0.00424, 0.00335, -0.00081, 0.00158, -0.00317, -0.00115, 0.00588, 0.00806, 0.00079, 0.00526, -0.00065, 0.00318, 0.00383, 0.00771, -0.00015, -0.00245, 0.0075, 0.00563, -0.00223, 0.00368, -0.00295, -0.00199, 0.00276, 0.00489, -0.00113, 0.0, -0.002, -0.00065, 0.00238, -0.0019, 0.00514, -0.0025, 0.00133, 0.0111, 0.00416, 0.00022, -0.00205, -0.00531, 0.00639, -0.00343, -0.00206, -0.00274, 0.00279, 0.00173, -0.00114, 0.00677, 0.00895], ['CWNJ_C02-2FAP-24_1_FCT', '#FF0000', 'Accelerometer AVG-FS8g_ODR200HZ_Zup accel_normal_average_x', 0.118, -0.118, 0.017404, 0.014196, -0.012434, -0.007955, 0.021675, 0.016083, -0.004455, -0.003674, 0.010693, 0.020249, -0.013457, -0.004792, 0.014762, 0.011479, -0.015235, -0.013161, 0.021875, 0.017191, -0.012668, -0.009655, 0.014624, 0.015017, -0.00997, -0.007622, 0.015532, 0.015681, -0.009978, -0.007865, 0.019603, 0.010732, -0.010212, -0.0135, 0.016777, 0.013657, -0.014978, -0.008793, 0.015982, 0.013434, -0.008279, -0.010875, 0.014077, 0.015478, -0.014007, -0.007138, 0.004797, 0.011938, -0.015699, -0.01629, 0.01604, 0.011166, -0.015168, -0.000265, 0.019898, 0.021764, -0.008146, -0.00903, 0.01403, 0.012649, -0.016166, 0.000295, 0.016997, 0.010756, -0.009021, -0.016898, 0.020278, 0.018457, -0.006079, -0.020667, 0.007023, 0.009995, -0.012875, -0.014259, 0.016171, 0.010556, -0.008689, -0.007465, 0.019792, 0.013226, -0.01304, -0.004704, 0.015627, 0.018039, -0.007436, -0.007188, 0.008269, 0.01666, -0.00479, -0.015327, 0.017256, 0.010406, -0.018305, -0.009765, 0.01971, 0.009813, -0.007063, -0.018571, 0.01331, 0.011267, -0.006494, -0.004936, 0.005678, 0.01708, 0.012821, 0.019666, -0.009574, -0.005175, -0.00153, 0.008066, 0.016601, 0.015412, 0.014075, -0.020463, -0.003049, 0.000584, -0.012575, 0.013302, 0.011694, -0.000616, -0.013676, 0.018919, 0.009672, -0.005222, -0.009169, 0.013405, 0.016721, -0.014268, -0.016521, 0.0078, 0.019311, 0.01302, 0.013647, -0.006352, -0.01144, 0.012963, 0.015951, -0.012729, -0.009787, 0.008951, 0.014204, 0.014488, 0.015052, -0.009985, -0.015122, -0.01276, -0.000652, 0.016413, 0.012819, -0.017922, 0.025687, -0.008644, -0.014013, 0.012122, 0.012221, 0.007084, -0.017099, -0.01473, 0.017407, 0.013321, -0.01029, -0.009717, 0.013975, 0.019156, -0.006644, -0.015818, 0.020065, 0.000983, -0.005487, 0.004946, 0.013232, -0.009826, -0.014203, 0.013308, 0.00936, -0.010505, -0.012281, 0.022151, 0.016451, -0.011882, -0.014841, 0.014814, 0.013488, 0.015504, 0.013662, -0.01152, -0.009514, 0.011774, 0.009316, 0.010922, -0.009506, -0.008972, -0.009671, 0.016089, 0.009199, -0.013384, -0.014392, 0.008068, 0.010853, -0.017433, -0.004365, -0.007553, 0.013398, 0.010908, -0.009884, -0.001649, -0.011921, 0.003544, 0.016423, -0.013159, -0.019088, 0.011322, 0.018793, -0.018654, 0.001358, 0.005798, 0.010099, 0.009448, 0.016237, -0.012746, -0.011605, -0.012845, -0.011443, 0.019311, 0.008419, -0.013151, 0.011319, 0.009714, 0.018898, 0.014658, 0.009429, -0.017882, -0.009682, -0.008455, -0.00627, 0.015141, 0.01356, 0.01799, 0.014892, -0.023739, -0.019709, -0.009079, -0.005859, 0.01504, 0.013439, -0.005076, -0.012888, 0.013833, 0.021414, 0.016767, -0.008977, -0.005835, 0.011184, 0.011467, -0.016347, -0.008913, 0.013016, 0.014582, 0.005197, 0.01626, -0.006674, -0.016643, -0.010268, -0.015226, 0.018815, 0.016538, -0.005496, -0.010821, 0.021801, 0.012361, -0.018638, -0.009224, 0.0091, 0.017824, -0.008444, -0.0079, 0.015085, 0.022338, 0.014109, 0.013418, -0.011328, -0.009606, -0.010026, -0.012702, 0.010906, 0.016271, -0.007043, -0.011943, 0.014273, 0.013792, -0.009978, -0.00541, 0.011517, 0.013871, -0.014162, -0.01135, 0.009841, 0.011348, 0.015298, 0.014283, -0.012895, -0.016879, -0.018977, -0.013698, 0.014278, 0.017141, -0.005903, -0.009323, 0.010201, 0.011017, 0.017148, 0.016069, -0.017962, -0.0116, -0.008851, -0.020297, 0.02036, 0.021338, -0.007109, -0.008963, 0.015869, 0.019204, -0.015683, -0.005508, 0.017911, 0.022761, -0.012897, -0.004865, 0.019637, 0.000981, -0.003359, -0.010809, 0.019248, 0.012052, -0.017785, -0.013383, 0.010867, 0.018989, -0.011699, -0.015185, 0.010751, 0.010922, -0.009472, -0.008005, 0.013425, 0.016135, -0.007531, -0.011175, 0.01431, 0.015078, -0.013314, -0.018636, 0.021225, 0.012631, -0.016181, -0.005383, 0.019151, 0.013063, -0.017009, -0.001878, 0.011674, 0.016596, -0.011049, -0.015275, 0.013413, 0.014413, -0.003504, -0.008839, 0.015021, 0.013623, -0.01319, -0.010771, 0.017324, 0.01708, -0.010084, -0.013879, 0.015605, 0.020139, 0.013688, 0.018161, -0.009152, -7e-05, -0.013015, -0.012744, 0.009079, 0.013814, -0.012939, -0.007712, 0.008774, 0.01687, -0.00862, -0.017491, 0.013229, -0.007701, -0.015576, 0.008447, 0.008625, -0.01352, -0.016887, 0.016541, 0.008571, -0.009172, -0.005388, 0.021455, 0.014362, -0.013007, -0.010394, 0.015207, 0.008032, -0.006139, -0.008479, 0.01568, 0.01622, -0.019857, -0.009089, 0.008291, 0.01321, -0.011315, -0.011052, 0.009992, 0.017169, 0.003452, 0.009137, -0.015139, -0.004009, -0.01421, -0.0162, 0.01434, 0.013302, -0.01176, -0.007204, 0.024537, 0.013526, 0.021396, 0.01094, -0.008247, -0.014914, -0.007155, -0.00968, 0.010674, 0.020761, -0.015249, -0.011201, 0.017619, 0.013657, 0.007113, 0.014772, -0.013454, -0.009335, -0.011443, -0.007451, 0.014658, 0.00961, -0.006491, -0.017819, 0.022414, 0.014914, -0.013454, -0.021103, 0.014539, 0.015424, 0.020031, -0.016945, -0.007233, 0.019654, 0.008776, -0.012226, -0.014172, 0.009676, 0.012934, -0.011152, -0.012083, 0.014409, 0.010019, 0.009504, 0.015351, -0.005617, -0.013079, -0.015625, 0.017431, 0.015668, -0.01208, -0.009274, 0.013984, 0.014143, -0.014997, -0.00823, 0.02049, 0.015979, -0.005458, -0.015476, 0.016337, 0.013773, -0.013309, -0.007783, 0.015078, 0.015463, -0.004062, -0.000352, 0.022444, 0.012983, -0.007341, -0.014851, 0.020106, 0.022714, -0.013667, -0.004533, 0.009614, 0.005935, -0.004612, -0.012656, 0.018032, 0.013287, -0.016967, -0.014018, 0.01263, 0.01145, -0.014691, -0.009584, 0.01173, 0.012285, -0.011657, -0.018963, 0.0227, 0.02078, -0.011786, -0.007478, 0.008413, 0.013366, -0.007202, -0.014248, 0.008694, 0.015194, -0.011701, -0.014245, 0.018979, -0.009533, -0.012788, 0.016584, 0.010268, -0.00935, -0.013945, 0.012121, 0.016217, -0.01102, -0.013838, 0.01684, 0.008598, -0.013657, -0.011266, 0.014526, 0.01467, -0.01726, -0.007937, 0.017222, 0.007382, -0.015161, -0.015114, 0.013278, 0.017006, -0.006681, -0.015229, 0.00971, 0.01444, -0.012702, -0.011858, 0.019864, 0.007723, -0.007626, -0.011165, 0.013452, 0.008481, -0.005228, -0.010882, 0.007231, 0.018217, -0.010766, -0.016394, 0.016203, 0.016638, -0.012156, -0.011166, 0.01258, 0.010326, -0.011694, -0.008395, 0.008346, 0.009333, -0.004641, -0.02234, 0.010166, 0.016654, -0.008631, -0.015625, 0.012521, 0.012707, -0.015992, 0.023522, 0.010642, -0.014064, -0.004486, 0.020866, 0.017437, -0.015941, -0.009978, 0.020217, 0.015085, -0.010478, -0.016151, 0.01218, 0.02072, -0.017928, -0.002727, 0.019215, 0.00373, 0.018122, 0.015987, -0.01567, -0.017198, -0.01215, -0.013337, 0.014679, 0.021782, -0.009011, -0.006761, 0.015318, 0.011429, 0.016838, 0.013747, -0.011083, -0.01135, -0.00941, -0.003193, 0.016084, 0.01916, -0.024938, -0.011623, 0.006318, 0.012097, -0.002636, -0.005912, 0.01692, 0.014957, -0.017239, -0.007536, 0.018438, 0.01809, -0.011728, -0.011093, 0.010688, 0.014743, -0.014041, -0.023282, 0.016728, 0.012648, 0.012453, 0.010577, -0.021187, -0.012364, -0.014401, -0.013889, 0.015438, 0.01173, -2.6e-05, -0.010872, 0.007167, 0.010578, -0.010163, -0.014691, 0.00955, 0.02206, -0.0151, -0.009576, 0.004211, 0.014974, -0.007312, -0.010935, 0.011945, 0.013972, -0.005595, -0.01634, 0.011923, 0.016057, 0.01342, 0.010461, -0.009763, -0.014339, -0.010366, -0.01904, 0.013751, 0.01918, -0.01405, -0.011667, 0.020102, 0.028811, -0.015146, -0.008537, 0.015334, 0.018378, -0.00613, -0.016461, 0.012504, 0.015173, -0.012765, -0.010471, 0.0189, 0.005475, 0.006025, 0.019633, -0.013845, -0.012331, -0.00917, -0.014321, 0.014521, 0.016711, -0.006458, -0.012131, 0.018912, 0.010836, -0.0129, -0.011146, 0.017002, -0.002293, -0.008117, -0.011555, 0.020125, 0.009572, -0.015441, -0.008946, 0.012161, 0.013369, -0.011015, -0.010227, 0.024008, 0.0129, 0.011039, -0.009995, -0.007536, 0.011602, 0.013072, -0.01321, -0.008854, 0.015476, 0.009995, -0.005017, -0.013567, 0.012182, 0.015471, -0.007893, -0.010384, 0.01538, 0.017663, -0.008498, -0.006901, 0.015479, 0.011996, -0.005756, -0.004274, 0.016608, -0.01414, -0.009316, 0.00772, 0.011999, -0.015136, -0.010031, 0.018936, 0.019519, 0.015211, 0.013725, -0.011929, -0.0126, -0.012985, -0.019511, 0.008068, 0.013877, 0.009743, 0.010869, -0.014399, -0.007907, -0.013471, -0.017246, 0.013674, 0.009692, 0.01561, 0.009074, 0.009871, 0.007824, -0.017016, -0.015878, -0.010249, -0.008022, -0.003408, -0.006772, 0.017675, 0.01072, 0.010686, -0.011594, -0.007077, -0.013171, 0.020005, 0.011191, -0.014892, -0.008486, 0.014341, 0.015522, 0.012356, 0.015078, -0.014546, -0.014128, -0.00999, -0.009838, 0.011397, 0.010398, -0.013021, -0.00188, 0.015229, 0.006291, -0.009294, -0.006665, 0.009619, 0.012762, -0.010307, -0.000939, 0.012446, -0.013048, -0.004717, 0.012284, 0.016257, -0.016921, -0.010683, 0.013657, 0.01326, -0.014334, -0.008852, 0.005698, 0.019047, 0.011119, 0.017243, -0.017225, -0.011653, -0.010604, -0.018416, 0.014135, 0.01216, 0.000652, -0.008581, 0.013263, -0.008025, -0.006533, 0.01489, 0.020683, -0.010336, -0.017967, 0.014791, 0.017304, -0.015535, -0.00773, 0.01205, 0.013945, -0.013553, -0.008425, 0.021269, 0.021108, -0.007897, -0.006273, 0.007258, 0.019306, -0.009436, -0.005595, 0.01752, 0.015146, -0.011694, -0.016185, 0.00979, 0.01144, -0.001305, -0.00742], ['CWNJ_C02-2FAP-23_1_FCT', '#008000', 'Accelerometer AVG-FS8g_ODR200HZ_Zup accel_normal_average_x', 0.118, -0.118, 0.01142, 0.00795, -0.00734, -0.00135, 0.00936, 0.01138, -0.00434, 0.00143, 0.00945, 0.0143, -0.00782, -0.01126, 0.00906, 0.01103, -0.01359, -0.00649, 0.0106, 0.01149, -0.00448, -0.00643, 0.0061, 0.02318, -0.00332, -0.00185, 0.01097, 0.01734, -0.00988, -0.01043, 0.01431, 0.0119, -0.01029, -0.00709, 0.00751, 0.00407, 0.01442, -0.00635, -0.0082, 0.01799, 0.0128, -0.01826, -0.00595, 0.0098, 0.01624, -0.00806, -0.00536, 0.00961, -0.0034, -0.00539, 0.01388, 0.01297, -0.0091, -0.00389, 0.00804, 0.00824, 0.01673, 0.00895, 0.0, -0.00639, -0.01321, -0.00733, 0.01227, -0.0094, -0.00601, 0.02058, 0.01407, -0.00142, -0.0089, 0.00636, -0.00792, -0.00777, 0.01061, 0.00541, 0.0231, 0.01452, 0.00176, -0.00605, -0.00304, -0.00744, 0.01015, 0.01389, -0.00721, -0.00708, 0.01222, 0.01147, -0.0175, -0.00289, 0.01586, 0.00932, 0.0087, -0.0116, -0.00989, 0.00875, 0.01195, -0.00548, -0.01423, 0.01238, 0.00985, -0.00478, -0.01207, 0.01404, -0.01004, -0.00699, 0.00677, 0.02095, -0.00857, -0.00946, 0.00673, 0.00996, 0.00884, 0.00884, -0.00398, -0.0004, -0.01556, -0.00415, 0.01861, 0.0051, -0.00539, -0.01194, 0.0147, 0.01291, 0.01272, 0.01651, -0.00674, -0.00945, -0.00667, -0.01304, 0.00997, 0.01358, 0.01151, 0.00987, -0.01265, -0.00924, -0.00792, -0.00531, 0.00709, 0.01422, -0.00541, -0.00761, 0.01196, 0.01786, 0.01321, -0.00342, -0.00318, -0.0036, -0.01221, 0.01051, 0.01449, -0.00277, 0.00702, 0.01262, 0.01994, -0.00107, -0.00507, 0.01204, 0.01984, -0.01438, -0.00494, 0.01307, 0.01251, -0.01252, -0.01152, 0.00545, 0.00928, 0.01118, 0.00647, -0.00744, -0.00342, -0.00272, -0.00749, 0.0, 0.0015, 0.01722, 0.01038, -0.01478, -0.00871, -0.00692, -0.00859, 0.01829, 0.01261, -0.00932, 0.00896, 0.01366, -0.00199, -0.0019, 0.00581, 0.01433, -0.01397, -0.00671, 0.01056, 0.01255, -0.00905, -0.01491, 0.01586, 0.01214, -0.00854, -0.00058, 0.01375, 0.0067, -0.01011, -0.00771, 0.00903, 0.01565, -0.0021, -0.00911, 0.01381, 0.0234, -0.00462, -0.001, 0.00663, 0.0157, -0.00511, -0.00753, 0.0116, 0.01214, -0.00295, -0.00425, 0.01053, 0.01516, -0.0152, -0.00957, 0.01268, 0.01122, -0.01779, -0.01105, 0.01546, 0.00469, -0.00918, -0.00731, 0.01057, 0.00196, -0.00665, -0.00963, 0.00974, 0.00871, -0.00924, -0.00854, 0.00976, 0.01322, -0.0145, -0.00952, 0.00659, 0.00786, -0.00442, 0.00547, 0.00594, 0.00023, -0.00093, -0.00056, 0.00669, -0.00594, 0.00941, 0.00562, -0.00156, 0.01068, 0.00355, -0.00198, 0.0056, 0.00201, 0.01173, 0.00024, 0.0016, 0.00303, 0.00309, 0.00096, -0.00149, 0.01289, -0.00719, 0.00504, 0.00369, 0.00645, -0.00356, 0.0, -0.00223, 0.01216, 0.00311, 0.00157, -0.00472, -0.00102, 0.00669, 0.00036, 0.00754, 0.00016, -0.00287, 0.00124, 0.00203, -0.00039, -0.0131, 0.00632, 0.00655, 0.00533, 0.00146, -0.00101, 0.00531, 0.00296, -0.00456, 0.00243, 0.00285, 0.0, 0.00445, 0.00287, -0.00134, -0.0021, 0.00071, 0.00637, 0.00722, -0.00021, 0.00098, -0.0033, 0.00515, -0.00061, -0.00168, 0.004, 0.00389, 0.00427, -0.00076, 0.01102, -0.00415, -0.00145, 0.0079, -0.00057, 0.0003, -0.00217, -0.00105, -0.00049, 0.01096, 0.00487, 0.00128, 0.00029, 0.00085, 0.00119, 0.00121, 0.00556, 0.00513, -0.00163, -0.00122, -0.00639, 0.00326, 0.00515, 0.00044, 0.01472, 0.00217, -0.00571, -0.00347, 0.01083, -0.00042, 0.00117, -0.00141, 0.00458, 0.00447, 0.0, 0.00602, 0.01734, 0.00559, 0.00014, 0.00407, 0.00401, -0.00221, 0.00572, 0.00041, 0.01441, -0.0023, 0.00165, 0.00488, -0.00538, -0.00131, -0.0013, -0.00232, 0.00046, 0.0, 0.00618, 0.00521, 0.00229, -0.0004, -0.0009, 0.00227, 0.00093, 0.00263, -0.00485, -0.00346, 0.00657, 0.00105, -0.00566, 0.00452, -0.0005, -0.00106, 0.00373, 0.00389, -0.00093, 0.00383, -0.00136, -0.00289, 0.00139, 0.00039, 0.00231, -0.00024, -0.00402, 0.00114, -0.00633, -0.01035, -0.00131, -0.0063, 0.00115, 0.00796, 0.00328, 0.00061, 0.00022, 0.00372, 0.00227, 0.00415, 0.00162, 0.00012, -0.00288, 0.00066, -0.00517, -0.00446, -0.00099, 0.00711, -0.00031, -0.0008, 0.00972, 0.00069, -0.00096, 0.00039, -0.00584, -0.00089, -0.00208, 0.00212, 0.00054, -0.00029, -0.00344, 0.00254, 0.00116, 0.00143, -0.00296, -0.00368, 0.00026, -0.0003, -0.0031, -0.00516, 0.006, 0.0, -0.00305, -0.00191, -0.0021, 0.0, 0.00229, 0.0031, 0.00343, 0.00681, -0.00258, -0.00263, 0.00415, 0.00822, 0.00416, 0.00419, -0.00068, -0.00621, -0.00625, 0.0027, 0.00283, -0.00811, 0.00055, 0.00469, 0.00614, 0.00559, 0.00676, -0.00042, -0.00522, -0.00039, 0.005, 0.01043, -0.00686, 0.00501, 0.00433, 0.00272, -0.00639, 0.0, 0.00142, 0.00652, -0.00137, -0.00302, 0.00611, 0.00088, -0.00196, -0.00564, 0.00812, -0.00154, 0.00051, -0.00445, 0.01202, -0.00014, -0.00174, 0.00692, 0.00163, 0.00728, -0.00161, -0.00353, 0.00486, -0.00264, -0.00444, -0.0022, 0.00309, 0.00089, 0.00103, -0.0043, 0.00611, -0.00646, -0.00319, 0.00354, 0.00654, 0.00126, 0.00175, 0.00194, 0.00708, -0.00751, 0.00241, -0.0028, 0.00403, 0.0049, -0.00104, 0.00279, 0.0025, 0.00435, -0.00327, -0.00083, 0.00126, 0.00396, 0.00328, 0.00659, 0.00821, 0.00137, 0.0045, 0.00321, 0.01291, 0.00036, -0.0018, 0.0108, 0.00038, 0.00827, 0.00556, 0.00886, 0.0, 0.00118, 0.00345, 0.00341, 0.00073, 0.00193, 0.00352, 0.00205, 0.00122, -0.00243, -0.00075, 0.00723, -0.00182, -0.00055, 0.00511, 0.00219, 0.00916, 0.0024, 0.00586, 0.00129, -0.00376, 0.00177, 0.00345, 0.00557, -0.00059, -0.00614, 0.01533, 0.0072, -0.00379, -0.00351, 0.00733, 0.00337, 0.00345, -0.00296, 0.00474, 0.01143, 0.00184, -0.00183, 0.00326, 0.00403, -0.00295, -0.00378, -0.00209, -0.00098, 0.00679, 0.00554, 0.01901, -0.00372, -0.00014, -0.00341, -0.00111, 0.01099, 0.00855, 0.00177, 0.0, -0.00052, 0.00416, 0.01309, -0.0006, -0.00422, 0.01259, -0.00124, 0.00738, -0.00339, 0.00731, 0.00836, -0.01253, -0.006, 0.00505, 0.01272, 0.0073, -0.00264, -0.00319, 0.00577, -0.00594, -0.00095, 0.00826, 0.00467, 0.00153, -0.00573, -0.00474, 0.00523, -0.006, -0.00335, 0.00314, 0.00585, 0.0096, -0.00178, -0.00258, -0.0016, 0.00685, -0.00339, 0.00102, 0.00422, 0.00528, -0.00171, -0.00465, 0.00852, -0.00138, -0.00528, 0.00178, 0.00938, 0.00655, 0.00198, -0.006, 0.01373, 0.00242, -0.00279, -0.00744, 0.00597, 0.00813, -0.00037, 0.0, 0.00784, 0.00769, -0.0035, 0.00262, -0.00292, -0.002, -0.00769, 0.00816, 0.00511, -0.00605, 0.00575, 0.00626, -0.00811, -0.00163, 0.00808, 0.00389, -0.00062, 0.00198, 0.01455, -0.00205, -0.00063, 0.01408, 0.00233, -0.00209, -0.00804, 0.0117, 0.00164, -0.00801, -0.00347, 0.01128, 0.01589, -0.00276, -0.00079, 0.00548, 0.00886, -0.00605, -0.00117, -0.00032, 0.00165, -0.00229, 0.00808, 0.00205, -0.00276, -0.00754, 0.00337, 0.00691, -0.00303, 0.00315, 0.00892, 0.00961, 0.00172, -0.00169, -0.00194, 0.00827, 0.00158, 0.01111, 0.00835, 0.00732, -0.00541, 0.00263, 0.01327, 0.00563, 0.0, 0.00172, 0.00554, 0.00345, 0.00041, 0.01466, 0.00693, 0.00319, -0.00033, 0.00156, 0.00904, 0.01463, 0.00311, 0.00317, 0.00414, 0.0, 0.00319, 0.00854, 0.00803, 0.00118, -0.00639, 0.0084, 0.00614, 0.00494, 0.00845, 0.00343, 0.0134, 0.00359, 0.0, -0.001, 0.00589, 0.00368, -0.00252, 0.00203, 0.0, 0.00698, 0.00733, 0.01077, 0.00558, 0.01181, 0.0012, -0.00871, 0.01012, 0.00319, -0.00118, -0.00656, -0.00977, -0.00169, 0.0071, 0.00766, 0.00279, 0.0018, 0.00016, -0.01383, 0.00805, 0.00621, 0.00147, 0.00422, -0.00229, 0.01081, 0.00793, -0.00894, 0.00779, 0.01012, 0.01486, -0.00242, 0.00176, 0.0074, 0.0034, 0.00101, -0.0003, 0.01712, 0.00349, -0.00491, 0.00034, -0.00111, 0.00462, -0.002, 0.00696, 0.00179, 0.00898, -0.00163, 0.01015, 0.00746, 0.01057], ['CWNJ_C02-2FAP-24_2_FCT', '#00FFFF', 'Accelerometer AVG-FS8g_ODR200HZ_Zup accel_normal_average_x', 0.118, -0.118, 0.015025, 0.015447, -0.016106, -0.025225, 0.026282, 0.020133, -0.021225, -0.023918, 0.014392, 0.028044, -0.025925, -0.016173, 0.024125, 0.018965, -0.013222, -0.006369, 0.029306, 0.021462, -0.018178, -0.018433, 0.027439, 0.025036, -0.009913, -0.016166, 0.030425, 0.01933, -0.023227, -0.014423, 0.015931, 0.014113, -0.021182, -0.017395, 0.019111, 0.016135, -0.004167, -0.025869, 0.021066, 0.013959, -0.019687, -0.012765, 0.025561, 0.026064, -0.025811, -0.021694, 0.030285, -0.013229, -0.023425, 0.026882, 0.024605, -0.017768, -0.017851, 0.024551, 0.015946, -0.016894, -0.01774, 0.022514, 0.019697, -0.010645, -0.020522, 0.022071, 0.016552, -0.020749, -0.021909, 0.018806, 0.0235, -0.018229, -0.014546, 0.022944, 0.015796, -0.017287, -0.018252, 0.021199, 0.027759, -0.020882, -0.017189, 0.024929, 0.020074, -0.02161, -0.022748, 0.018576, 0.026618, -0.01518, -0.017343, 0.015886, 0.01869, -0.0151, -0.021831, 0.017017, 0.016483, 0.019992, 0.020415, -0.011095, -0.02443, -0.026594, -0.016574, 0.020321, 0.012941, -0.013376, -0.017155, 0.017549, 0.020756, 0.01766, 0.015131, -0.026362, -0.010957, -0.013899, -0.016208, 0.024191, 0.030875, -0.009122, -0.018005, 0.020212, -0.020307, -0.017535, 0.020859, 0.018315, -0.011757, -0.011479, 0.019672, -0.023562, -0.016564, 0.030086, 0.02436, -0.018679, -0.019801, 0.015037, 0.018122, -0.021343, -0.020072, -0.010101, 0.024238, 0.016794, 0.024399, -0.021457, -0.017707, 0.015988, 0.020925, -0.013971, -0.015683, 0.024571, 0.020012, -0.01547, -0.018532, 0.03049, 0.021508, -0.01918, -0.020947, 0.022961, 0.016367, -0.014885, -0.012114, 0.028942, 0.013908, -0.013188, -0.019685, 0.026768, 0.017065, -0.011047, -0.006843, 0.01999, 0.020727, -0.014438, 0.017253, 0.021622, 0.018666, -0.018578, -0.01538, 0.024707, 0.017812, -0.017288, -0.017231, 0.027399, 0.022108, -0.013706, -0.019299, 0.02132, 0.025097, -0.017241, -0.024194, 0.0229, 0.023959, -0.022555, -0.019787, 0.018994, 0.025129, -0.007822, -0.005902, 0.028815, 0.019418, 0.024421, -0.014599, -0.013862, 0.028117, 0.024208, -0.014904, -0.017989, 0.020469, 0.021532, -0.020437, -0.015052, 0.019441, 0.01426, -0.016277, -0.024056, 0.021594, 0.027009, -0.014258, -0.016642, 0.022414, 0.024453, 0.016932, 0.02239, -0.018501, -0.019299, -0.0178, -0.021074, 0.012509, 0.018327, 0.012766, 0.018063, -0.018351, -0.024355, 0.02488, 0.021357, -0.017456, -0.021022, 0.016483, 0.018415, -0.018967, -0.019147, 0.033156, 0.021764, -0.012973, -0.017097, 0.018726, 0.01897, -0.016152, -0.02039, 0.018412, 0.016891, -0.014282, -0.01672, 0.025539, 0.015686, -0.015059, -0.021165, 0.019652, 0.020597, -0.018776, -0.015434, 0.018896, 0.02445, -0.019398, -0.021859, 0.019045, 0.021782, -0.012248, -0.01582, -0.021537, 0.019284, 0.015375, 0.018487, 0.014583, -0.019062, -0.016005, -0.020878, -0.029555, 0.022103, 0.021399, -0.012501, 0.025441, 0.017536, -0.016018, -0.025622, 0.014832, 0.019995, -0.022004, -0.019241, 0.020546, 0.029693, 0.02508, 0.018951, -0.022248, -0.018445, -0.016928, -0.021469, 0.018991, 0.018647, -0.021341, -0.016939, 0.019234, 0.016038, -0.015794, -0.014756, 0.020179, 0.024592, -0.01363, -0.021437, 0.022257, 0.021367, 0.022767, 0.025441, -0.021137, -0.021303, -0.02056, -0.017019, 0.019543, 0.014926, -0.013244, -0.01457, 0.018229, 0.019526, -0.021107, -0.019855, 0.01118, 0.01571, -0.019868, -0.019008, 0.021728, 0.023154, 0.027029, 0.021025, -0.013803, -0.023418, -0.02838, -0.024759, 0.026814, 0.029512, -0.017996, -0.0132, 0.027693, 0.017167, -0.029775, -0.024558, 0.013947, 0.025726, 0.01758, 0.023766, -0.020273, -0.025158, -0.018361, -0.012958, 0.01457, 0.027892, -0.02284, -0.024818, 0.018008, 0.01822, -0.017643, -0.017255, 0.0257, 0.023591, -0.020148, -0.022537, 0.025324, 0.020971, -0.026115, -0.016296, 0.013242, 0.020937, -0.010524, -0.014069, -0.013947, 0.013559, 0.015883, -0.027299, -0.024411, 0.024074, 0.01589, -0.023796, -0.024855, 0.027292, 0.017701, -0.022502, -0.01481, 0.01875, 0.018408, -0.014521, -0.019233, 0.011413, 0.011403, -0.013676, -0.020984, 0.017203, 0.021958, -0.018784, -0.016666, 0.018115, 0.016009, -0.024097, -0.016236, 0.018815, 0.017798, -0.012726, -0.020771, 0.018484, 0.024049, -0.02815, -0.014851, 0.02123, 0.018929, -0.016927, -0.025698, 0.020187, 0.015673, -0.024058, -0.00963, 0.019965, 0.020327, -0.024588, -0.019642, 0.015727, 0.027846, -0.020405, -0.013558, 0.016931, 0.022702, -0.013435, -0.019756, 0.02757, 0.012898, -0.017136, -0.018042, 0.021897, 0.018896, -0.020095, -0.015913, 0.014689, 0.011459, -0.023949, -0.018714, 0.021506, 0.018632, -0.025764, -0.018149, 0.015992, 0.016417, -0.017531, -0.023853, 0.025275, 0.027643, -0.028738, -0.021032, 0.022175, 0.020786, 0.025861, -0.019682, -0.016879, 0.022801, 0.023118, -0.010077, -0.021533, 0.022168, 0.023806, -0.023876, -0.012661, 0.024273, 0.019267, -0.015866, -0.012092, 0.020101, 0.024841, -0.011119, -0.024044, -0.019958, 0.017993, 0.015987, -0.0154, -0.013744, 0.021289, 0.017045, -0.015528, -0.024706, 0.025562, 0.021851, -0.01602, -0.016152, 0.023945, 0.024609, -0.01197, -0.025393, 0.019069, 0.019765, -0.03066, -0.013911, 0.022867, 0.028455, -0.021347, 0.015593, 0.01798, 0.015114, -0.01343, -0.013496, 0.021058, 0.020622, -0.011263, 0.020299, 0.020043, -0.017773, -0.01156, 0.025269, 0.026955, -0.025375, -0.022231, 0.022289, 0.017875, 0.022226, 0.021772, -0.020493, -0.017123, -0.015668, 0.017669, 0.020701, -0.012789, -0.015424, 0.011118, 0.018926, -0.029719, -0.011933, 0.019426, 0.022883, -0.020682, -0.018184, 0.016843, 0.025537, -0.013002, -0.020615, 0.027131, 0.01925, -0.015206, -0.018619, 0.018054, 0.015837, -0.018579, -0.01808, 0.026906, 0.027031, -0.020683, -0.017817, 0.018728, 0.030551, -0.015412, -0.022562, 0.01932, 0.012792, -0.011675, -0.021228, 0.021281, -0.020185, -0.022364, 0.023754, 0.040502, -0.012041, -0.025573, 0.02138, 0.019304, -0.011489, -0.017928, 0.025888, 0.01519, -0.022326, -0.023107, 0.017258, 0.023198, -0.015971, -0.019287, 0.022609, 0.029824, -0.020152, -0.02153, 0.023217, 0.014945, -0.020894, -0.021323, 0.023132, 0.013667, 0.025673, -0.014247, -0.026031, 0.019853, 0.018666, 0.015129, 0.019792, 0.021213, 0.01815, -0.018461, -0.021322, -0.016475, -0.015139, -0.020043, -0.013432, 0.02385, 0.017029, -0.018443, -0.018027, -0.012246, 0.025049, 0.018105, 0.019748, 0.028271, -0.017701, -0.019553, -0.016584, -0.020991, 0.021585, 0.022998, -0.012243, -0.006586, 0.015864, 0.014248, 0.030117, 0.024406, -0.016074, -0.020676, -0.017401, -0.017746, 0.01967, 0.022204, 0.019407, -0.01737, -0.019399, 0.016724, 0.020031, -0.021591, -0.018738, 0.01984, 0.02283, -0.017656, -0.020435, 0.018465, 0.019429, -0.022789, -0.019023, 0.015603, 0.027667, -0.015063, -0.013645, 0.019082, 0.021972, -0.016882, -0.021764, 0.025823, 0.011044, -0.020073, -0.020887, 0.023618, 0.015698, -0.013745, -0.019579], ['CWNJ_C02-2F-REL01_1_FCT', '#9400D3', 'Accelerometer AVG-FS8g_ODR200HZ_Zup accel_normal_average_x', 0.118, -0.118, 0.00842, 0.00548, 0.00585, 0.00818, 0.0097, 0.0121, 0.00242, 0.00308, 0.01044, 0.00755, 0.00959, -0.01006, 0.01121, -0.00498, 0.01415, -0.0077, 0.01118, 0.01036, 0.01076, -0.00466, 0.01155, 0.00925, -0.00468, 0.01308, 0.00875, 0.00924, -0.00801, -0.0058, 0.00693, -0.01083, 0.00933, 0.01103, 0.00642, -0.00971, -0.01527, -0.01309, -0.0086, 0.00506, 0.02145, 0.01839, 0.01358, -0.00442, 0.01769, -0.00072, -0.00437, -0.00289, -0.00608, 0.01492, 0.00877, -0.01034, 0.01192, -0.00302, -0.00387, 0.01044, -0.01045, -0.01073, 0.00927, -0.0061, 0.00593, -0.01001, -0.00724, 0.01197, 0.01468, -0.00365, -0.00501, -0.00202, 0.00941, -0.00601, -0.00748, 0.01012, -0.0059, 0.011, 0.01582, -0.00197, -0.01057, -0.00754, -0.00924, 0.00829, -0.00411, -0.00642, 0.00431, 0.00396, -0.01067, -0.00871, -0.00551, -0.00391, 0.0157, -0.0024, -0.0147, -0.01276, -0.00033, 0.00101, 0.00651, 0.00586, -0.0085, 0.01039, 0.0174, -0.00137, 0.00861, 0.01175, 0.015, 0.01511, 0.01041, 0.01381, 0.0222, 0.0223, 0.00603, -0.01459, 0.00731, -0.00748, 0.01356, 0.01745, 0.00529, -0.01145, 0.00836, -0.01117, -0.00878, -0.00754, 0.00522, 0.0255, 0.01629, 0.01636, 0.01267, -0.00167, 0.01348, 0.01467, 0.0169, 0.01404, 0.00822, 0.00719, 0.00038, 0.00159, 0.00694, 0.00956, 0.00845, -0.00986, -0.00085, -0.0006, 0.00691, -0.01071, 0.00553, -0.0111, 0.01131, 0.01002, 0.01314, 0.01547, -0.00551, 0.013, -0.01085, -0.00843, 0.0109, 0.0091, -0.00665, -0.0034, 0.01371, 0.01129, -0.00764, -0.00912, 0.00069, 0.01363, 0.01186, 0.00941, 0.00959, 0.01278, 0.00618, 0.00635, 0.0139, -0.00298, 0.00947, -0.00743, -0.01002, -0.00828, -0.01238, 0.00398, 0.00967, -0.00829, -0.00251, 0.01505, 0.01162, -0.00343, 0.01704, 0.01957, -0.0095, -0.00747, -0.00789, 0.01098, -0.0078, 0.00797, -0.00765, 0.01203, 0.01239, -0.00696, -0.00408, -0.01033, 0.00789, 0.01533, 0.01409, 0.01217, -0.0039, -0.00694, 0.00982, -0.01157, -0.00894, -0.00634, 0.0107, 0.00843, -0.00918, 0.01022, 0.01113, 0.01365, 0.01634, -0.01446, -0.01073, -0.00364, -0.00403, 0.00549, -0.00811, -0.01134, -0.00801, 0.01653, -0.00157, -0.00679, 0.0108, -0.00352, 0.01316, 0.01247, -0.00677, 0.00781, -0.00727, -0.00189, 0.01588, 0.00996, 0.00958, 0.0104, 0.01134, 0.01132, -0.00368, 0.0113, 0.01117, 0.01181, -0.00542, 0.00835, -0.00738, 0.00107, 0.00483, 0.00242, -0.00062, 0.00498, 0.00586, -0.00226, -0.00222, -0.00799, -0.00691, 0.0133, -0.00155, -0.00048, -0.0146, 0.0057, 0.00823, -0.00885, -0.00676, 0.00449, 0.0244, 0.00783, -0.00873, 0.00499, 0.00541, 0.0093, -0.007, 0.01165, -0.00718, 0.00449, -0.01565, 0.01782, -0.00148, 0.01208, 0.01194, -0.0119, -0.01153, -0.00952, -0.00984, -0.00793, 0.01043, 0.01357, 0.01024, 0.00528, -0.01039, -0.01073, 0.0071, -0.00939, 0.005, 0.01296, 0.01249, 0.01869, 0.02078, 0.00715, -0.012, -0.01247, -0.01188, 0.01283, -0.00335, 0.01293, -0.00423, -0.00963, -0.01005, 0.0076, -0.01045, 0.00975, 0.00765, 0.00507, -0.00756, -0.00928, -0.00977, 0.01206, -0.00413, 0.01142, 0.01268, 0.00909, -0.01087, 0.00602, 0.00798, -0.00815, -0.00792, 0.00956, 0.00445, -0.01342, 0.01065, -0.00777, 0.00357, 0.00607, 0.00342, 0.00104, -0.00752, -0.01048, 0.01134, -0.00533, 0.0097, -0.00876, 0.01335, 0.01444, 0.00988, -0.00733, 0.00898, 0.00814, 0.0106, 0.0122, -0.01196, -0.01221, 0.01076, -0.007, 0.01519, 0.01566, 0.01039, 0.01196, -0.01362, -0.00471, 0.01145, 0.0119, -0.00096, 0.00785, -0.01586, 0.01052, 0.00941, 0.00114, -0.01453, 0.00634, 0.00905, 0.0076, 0.00896, 0.00821, -0.00622, -0.00727, 0.00735, 0.00378, 0.00988, 0.00502, 0.00788, -0.01107, -0.00288, 0.00945, 0.00685, 0.00905, 0.01608, 0.00964, 0.01737, 0.01063, -0.0068, 0.00557, 0.0096, -0.0153, 0.02054, 0.00944, -0.01226, 0.01355, 0.00331, 0.00527, -0.00895, -0.01289, -0.00861, 0.01089, -0.00757, 0.0082, 0.01178, -0.01569, 0.01792, 0.01013, -0.00934, 0.01362, 0.00915, -0.00268, 0.00269, 0.00331, 0.01112, 0.01319, 0.0118, -0.00525, 0.00806, 0.00983, 0.00147, -0.00935, 0.0094, 0.00925, -0.00596, -0.0136, -0.00636, 0.00822, 0.01102, 0.00535, -0.01007, -0.00235, -0.00382, 0.00755, 0.00748, 0.00645, 0.00577, 0.00141, 0.01054, -0.00492, 0.00734, 0.00776, 0.01, 0.0076, 0.01041, 0.00376, 0.01049, -0.00897, 0.0166, 0.00893, 0.01048, 0.01082, -0.01051, -0.00963, -0.00499, 0.00967, 0.00697, -0.00647, 0.00901, 0.00885, -0.01084, 0.00633, 0.01372, 0.00863, 0.00865, 0.00691, -0.00884, 0.01155, -0.0086, -0.00641, -0.00386, -0.00731, -0.01379, 0.00979, -0.01373, 0.00766, 0.01155, 0.01643, 0.00071, 0.01127, 0.00598, 0.01141, -0.00529, -0.0095, -0.00957, 0.00779, 0.00212, 0.0073, 0.00693, -0.00634, 0.00384, 0.01015, 0.0024, 0.01024, 0.01021, 0.00975, -0.00945, 0.00465, -0.00913, 0.01077, -0.01076, -0.00699, -0.01393, -0.01214, -0.00901, 0.00293, 0.00075, 0.00504, -0.01088, -0.00814, -0.00398, -0.00859, -0.01142, 0.01089, 0.01228, -0.0132, -0.01441, -0.01774, -0.0039, 0.01151, 0.00978, 0.00857, 0.00319, 0.00119, -0.00961, 0.00623, -0.02012, 0.01009, 0.0068, -0.01114, 0.01187, -0.01103, -0.01016, 0.00484, 0.00755, -0.00578, -0.00624, -0.01314, -0.00929, -0.01918, 0.01302, 0.00771, -0.01755, -0.00599, 0.01414, 0.00518, -0.00706, 0.01622, -0.00624, 0.00755, -0.00989, 0.00379, 0.00576, 0.01419, -0.01273, -0.01128, -0.00684, 0.00678, 0.0063, -0.01053, -0.00252, 0.01227, 0.00344, 0.00462, -0.00874, 0.00542, 0.00946, 0.00916, -0.00863, -0.00979, -0.0064, 0.00753, 0.00742, 0.01124, 0.00621, 0.00428, 0.01055, 0.0, 0.01206, -0.01485, -0.00395, -0.01145, -0.01271, -0.01275, 0.01277, -0.00761, 0.00128, -0.00988, -0.00758, -0.01175, -0.01091, -0.00438, -0.0075, 0.01417, -0.00972, -0.0067, -0.00705, -0.00661, 0.00081, 0.0105, -0.00874, 0.00572, 0.00244, -0.00315, 0.0043, -0.00824, 0.00604, 0.00549, -0.01385, -0.00359, -0.01208, -0.01466, 0.00266, 0.00783, 0.01007, 0.00991, 0.00435, -0.0069, 0.00795, 0.00529, 0.00879, -0.01021, -0.00957, -0.00726, -0.01041, -0.00884, 0.00245, 0.00711, -0.00653, 0.01373, 0.00522, -0.00826, 0.00364, 0.00571, -0.00755, -0.00829, -0.00644, 0.00636, -0.00639, 0.00664, 0.00843, 0.00542, 0.01236, -0.00896, -0.00733]]]
    # table_category_data = []#it's empty when color_by == 'Off'
    set_bins = 250 # 100/50
    select_new_lsl = ''
    select_new_usl = ''
    start_time_first = '2020/3/11 14:16'
    start_time_last = '2020/3/26 9:58'
    pic_path = '/Users/rex/Desktop/CPK_Log/'
    cpk_plot(table_data,table_category_data,pic_path,set_bins,select_new_lsl,select_new_usl,start_time_first,start_time_last)

    #------------------------ debug cpk_plot end --------------------------------



    #------------------------ debug correlation_plot --------------------------------
    table_data = [['Accelerometer AVG-FS8g_ODR800HZ_Zup accel_nmDC_average_y', 0.118, -0.118, -0.00105, -0.02004, 0.00256, -0.00017, -0.01117, -0.0076, 0.01392, 0.00581, -0.01061, 0.00613, 0.0049, 0.01169, -0.01533, -0.00912, 0.01449, 0.00366, -0.01347, -0.01472, 0.01017, 0.00695, -0.0176, -0.0094, 0.007, 0.00895, -0.01107, -0.00646, 0.0066, 0.00174, -0.00493, -0.00753, 0.01071, 0.00707, -0.01432, -0.00978, 0.01122, -0.00098, -0.02688, -0.00502, 0.01419, 0.00945, -0.01053, -0.01183, 0.0041, 0.00546, -0.01279, -0.00168, 0.01797, 0.00999, -0.0194, -0.0171, 0.00237, -0.00164, -0.0121, -0.01011, 0.01913, 0.00969, -0.00876, -0.00871, 0.00863, 0.00138, -0.00538, -0.0123, 0.01535, 0.00489, -0.0155, -0.00583, 0.00739, 0.01304, -0.01108, -0.01778, 0.01384, -0.00069, -0.00792, -0.00925, 0.00607, 0.01332, -0.01421, -0.00233, 0.01275, 0.00194, -0.01237, -0.00907, 0.00231, 0.01254, -0.00714, -0.00452, 0.00485, -0.00174, -0.01322, -0.00696, 0.01683, 0.01282, -0.01217, -0.01399, 0.02481, -0.01912, 0.02141, 0.01038, 0.02488, -0.01697, 0.01464, 0.01004, -0.02587, -0.0098, 0.01263, 0.01503, -0.02905, -0.01406, 0.01295, 0.01591, -0.01722, -0.01471, 0.01461, 0.01183, -0.01065, -0.01545, 0.01076, 0.00592, -0.01647, -0.01043, 0.00906, 0.00764, -0.02134, -0.0062, 0.01184, 0.01396, -0.00851, -0.01067, 0.01694, 0.00862, -0.01117, -0.0209, 0.02176, 0.01039, -0.01613, -0.0073, 0.00572, 0.00179, -0.01294, -0.00434, 0.01434, 0.01292, -0.01166, -0.00651, 0.02725, 0.01804, -0.01558, 0.01852, 0.02277, -0.01227, -0.01207, 0.01672, 0.00742, -0.00866, -0.00866, 0.03073, 0.00881, -0.0124, -0.00044, 0.01364, 0.02148, -0.01234, -0.00813, 0.0207, 0.004, -0.00535, -0.01961, 0.00626, 0.01013, -0.01034, -0.0116, 0.01233, 0.0137, -0.01865, -0.00635, 0.01144, 0.00732, -0.01074, -0.00616, 0.0171, 0.01488, -0.00979, -0.02108, 0.02023, -0.0014, -0.02369, -0.00552, 0.0203, 0.0176, -0.0164, -0.02013, 0.0074, 0.00413, -0.0162, -0.00944, 0.01826, 0.00754, -0.01613, -0.01638, 0.01281, 0.00613, -0.01202, -0.01155, -0.01635, -0.0121, 0.02292, 0.01943, 0.0111, 0.02202, -0.02173, -0.00902, 0.01094, 0.00799, -0.01721, -0.01456, 0.01876, 0.00446, -0.01882, -0.00801, 0.02149, 0.01383, -0.01387, -0.01488, -0.00812, -0.01436, 0.02472, 0.00996, 0.02179, 0.00825, -0.00495, -0.01665, 0.01334, 0.01172, -0.01148, -0.01546, 0.02173, 0.02255, -0.01849, -0.01487, 0.02418, 0.01251, -0.0187, -0.01557, -0.01751, -0.01324, 0.02085, 0.01702, 0.00812, -0.02393, -0.01034, 0.03133, 0.01674, -0.01226, -0.01344, -0.01815, -0.01019, 0.00905, 0.01061, 0.02095, 0.01245, -0.01218, -0.02069, 0.00679, 0.00879, -0.01887, -0.02076, 0.02419, 0.01828, -0.01577, 0.02042, 0.01989, -0.01952, -0.02086, 0.01043, 0.00877, -0.01376, -0.01314, 0.01564, 0.00891, -0.00675, -0.01104, 0.01707, 0.01746, -0.00474, -0.01264, 0.01968, 0.01605, -0.00691, -0.02159, 0.00152, 0.01552, -0.02149, -0.00843, 0.00893, 0.00483, -0.01832, -0.0112, 0.01559, 0.00147, -0.0162, -0.02137, 0.01908, 0.00923, -0.0148, -0.00969, 0.02171, 0.02151, -0.00913, -0.01047, 0.01014, 0.01495, -0.0207, -0.00613, 0.02006, 0.00536, -0.01313, -0.02111, 0.01105, 0.01482, -0.02017, -0.00919, 0.01281, 0.00787, -0.01399, -0.00941, 0.01523, 0.01349, -0.01659, -0.01698, 0.01345, -0.01431, -0.00617, 0.0113, 0.02372, -0.01436, -0.00938, 0.00925, 0.00132, -0.01418, -0.01402, -0.0037, -0.00524, 0.00815, 0.01063, 0.01425, 0.01503, -0.00812, -0.00973, -0.01497, -0.01497, 0.0214, 0.00639, 0.01462, 0.01846, -0.01666, -0.0082, -0.0154, -0.00815, 0.01167, 0.01791, 0.018, 0.00632, -0.0077, -0.01209, 0.00522, 0.0069, -0.01625, -0.00705, -0.01938, -0.0156, 0.01139, 0.02218, 0.01004, 0.01089, -0.01721, -0.01281, -0.01259, -0.01542, -0.01556, -0.01582, -0.01134, 0.00586, 0.00482, 0.01088, 0.01086, -0.01067, -0.00836, 0.01376, 0.01108, -0.01954, -0.01059, 0.00951, 0.013, -0.01187, -0.01031, -0.01498, -0.01247, 0.0091, 0.01376, 0.01505, 0.01136, -0.00606, -0.01092, 0.01149, 0.01102, 0.01441, -0.01633, -0.01254, -0.00624, 0.00843, 0.02254, 0.00928, 0.01065, -0.01085, -0.00948, 0.0321, 0.01443, -0.01436, 0.00967, -0.01059, 0.01706, 0.01495, -0.01278, -0.00879, -0.01371, -0.00875, 0.01277, 0.01377, 0.02353, 0.0167, -0.00764, -0.01373, 0.02202, 0.00761, -0.01646, -0.02106, 0.00936, 0.01128, -0.00887, -0.00636, -0.01485, 0.01737, -0.01382, 0.00355, -0.01284, 0.01699, -0.01654, -0.00788, 0.00343, 0.01655, -0.01262, -0.01525, -0.01799, 0.01152, 0.01284, -0.01102, -0.01788, 0.01399, 0.00656, -0.00553, -0.01033, 0.01518, 0.01103, -0.00805, -0.01886, 0.0204, 0.00563, -0.01421, -0.00839, 0.01713, 0.01344, -0.01159, -0.01042, 0.00867, 0.01016, -0.01053, -0.0134, 0.01598, 0.01039, -0.00848, -0.01961, 0.01044, 0.00755, -0.0162, -0.00638, 0.01786, 0.00777, -0.0186, -0.01167, -0.0073, 0.01056, 0.00623, -0.01714, -0.01258, -0.01256, 0.01568, 0.01582, -0.01571, -0.00989, -0.00756, 0.01903, 0.01267, 0.0011, -0.01714, -0.01178, -0.00674, 0.01847, 0.01406, 0.01359, 0.02021, -0.00753, -0.01432, -0.01532, -0.00885, -0.00612, -0.01723, 0.01148, 0.01514, 0.00489, 0.01014, 0.02088, 0.00251, -0.01315, -0.00504, 0.01673, 0.01213, -0.01965, 0.00422, 0.0076, -0.02049, -0.00462, 0.02657, 0.01227, -0.01782, -0.01425, 0.01746, -0.00418, -0.02096, -0.02243, -0.01485, -0.01296, 0.01138, -0.00176, 0.00972, 0.01168, -0.01761, -0.00938, 0.01759, 0.00669, -0.01894, -0.0168, 0.02418, 0.01446, -0.01161, -0.02294, 0.02191, 0.00788, -0.01147, -0.0049, 0.01596, 0.0144, -0.01215, -0.01257, 0.02651, 0.00877, -0.00878, -0.00965, -0.01713, -0.01221, 0.01419, 0.0134, 0.0176, 0.01424, -0.0136, -0.02171, -0.01624, -0.0074, 0.01619, 0.01591, 0.01263, 0.01381, -0.01321, -0.00731, 0.0139, 0.01641, 0.00218, -0.01333, 0.01353, 0.00721, -0.00228, -0.00323, 0.01531, 0.00325, -0.00327, -0.00315, -0.00216, -0.00026, -0.01519, -0.01152, 0.00512, 0.00786, -0.00537, -0.00701, 0.00631, 0.00263, -0.0066, -0.00231, -0.00094, -0.00434, 0.0, -0.00297, 0.00897, -0.01735, -0.00473, -0.01763, -0.01778, 0.01757, 0.00712, 0.01449, 0.01083, -0.00407, -0.01785, -0.01098, -0.01893, 0.00958, 0.007, 0.00271, 0.0118, -0.01292, -0.00664, -0.01324, -0.00738, 0.01378, 0.01642, -0.0007, 0.01583, -0.00633, -0.01471, 0.01061, 0.01297, -0.01588, -0.0135, -0.01058, -0.02435, 0.01825, 0.0156, -0.01534, 0.01398, -0.01716, -0.00963, 0.00965, 0.01264, -0.01459, -0.01398, 0.01583, 0.01683, -0.01755, -0.00903, -0.01077, -0.01565, 0.01642, 0.00875, 0.01698, 0.01391, -0.01246, -0.01304, -0.01187, -0.02077, 0.02154, 0.0161, 0.01789, 0.01736, -0.01478, -0.02276, 0.00926, 0.00489, -0.00323, -0.01781, 0.01518, 0.00403, -0.01454, -0.01242, 0.00695, 0.0099, -0.01803, -0.01187, 0.01347, 0.01274, -0.00808, -0.00853, 0.01256, 0.0142, -0.01446, -0.00729, 0.00099, 0.0104, -0.0184, -0.01331, 0.01543, 0.01497, -0.00521, -0.00776, 0.00881, 0.00855, -0.01555, -0.0155, 0.00761, 0.00767, -0.01167, -0.00917, 0.00278, 0.00783, -0.01997, -0.00492, 0.00115, -0.00802, -0.00894, 0.00703, 0.00344, -0.01152, -0.00024, 0.00208, 0.01736, -0.00279, -0.01458, 0.00212, 0.00331, -0.01687, -0.00542, 0.00074, 0.01083, -0.01457, 0.00048, 0.00299, 0.00468, 0.0126, -0.01211, -0.00848, 0.01041, 0.00535, -0.00113, 0.00333, 0.01551, 0.01152, -0.00134, -0.01316, 0.00252, 0.00702, -0.00988, -0.00723, 0.01078, 0.00628, -0.01408, -0.00535, 0.00632, 0.00977, -0.00283, -0.00163, 0.00414, 0.00957, -0.00408, -0.01258, 0.00143, 0.01101, -0.00729, -0.00412, 0.00797, -0.00786, 0.00022, -0.0046, 0.00986, 0.00387, -0.01238, -0.0172, 0.01202, 0.00579, -0.00443, -0.00614, 0.00812, 0.00707, -0.01272, -0.00727, -0.0002, 0.00516, -0.00066, 0.00078, 0.01554, 0.01021, -0.01847, -0.0018, 0.01142, -0.00443, -0.00389, -0.02235, 0.01418, 0.00961, -0.01435, -0.01414, 0.0, -0.001, -0.00901, -0.00313, -0.00475, -0.00648, 0.00597, 0.0042, -0.01672, -0.01399, 0.00869, 0.00539, -0.00412, -0.00614, 0.00536, -0.00552, -0.01126, -0.00959, 0.0019, -0.00982, -0.01677, -0.00107, -0.00733, -0.00291, -0.00087, -0.01563, -0.00785, -0.00434, -0.00298, -0.00627, -0.00919, 0.00025, -0.00714, -0.00719, -0.01034, -0.0008, -0.01342, -0.0127, -0.01105, -0.018, -0.00389, 0.00131, -0.01922, -0.01225, -0.00199, -0.00394, -0.02298, -0.01126, -0.00837, 0.00623, -0.01149, -0.01002, -0.004784, -0.004183, 0.005602, -0.009599, -0.010599, 0.006937, -0.003276, -0.010383, -0.011383, 0.002014, -0.008719, -0.023569, -0.00729, 0.003051, -0.000618, -0.009226, -0.014455, -0.010385, -0.006469, -0.012288, -0.010561, 0.000408, -0.003281, -0.013213, -0.015061, -0.01145, -0.008162, -0.003817, -0.011477, -0.002931, -0.002641, -0.00962, -0.008303, -0.005853, -0.001762, -0.015197, -0.017743, -0.008202, -0.013215, -0.022393, -0.019039, 0.003245, 0.002261, -0.02026, -0.015423, 0.003422, -0.016795, -0.026817, -0.013022, -0.002421, -0.006563, -0.024116, -0.017668, -0.00449, 0.000461, -0.013312, -0.009719, 0.005094, -0.000287, -0.019095, -0.02376, 0.006142, 0.002628, -0.01565, -0.010705, -0.003142, 0.000484, -0.016752, -0.012101, 0.004668, -0.004429, -0.010781, -0.021791, -0.001748, -0.005075, -0.016394, -0.008919, 0.006871, -0.0106, -0.01313, -0.004398, -0.00523, -0.002703, -0.014615, -0.007359, -0.009628, -0.004858, -0.011959, -0.017628, 0.003078, 0.010342, -0.010801, -0.004591, 0.00032, -0.007471, -0.011368, -0.01167, -0.009179, -0.00353, -0.028292, -0.024105, -0.012227, -0.016483, -0.005227, -0.002409, -0.00557, -0.014273, -0.005641, -0.030162, -0.017904, 0.000226, 0.000859, -4.6e-05, -0.004343, -0.013792, -0.007122, 0.001618, -0.001019, -0.01806, -0.014621, -0.002348, 0.006127, -0.021649, -0.020339, 0.002575, -0.006118, -0.006449, -0.0154, -0.020551, -0.013468, -0.007644, -0.006377, -0.0194, -0.012347, 0.0028, -0.007288, -0.022295, -0.014395, -0.010895, -0.00853, -0.012239, -0.00335, -0.008344, -0.00412, -0.006059, -0.007919, -0.005348, -0.01564, 0.002845, -0.003628, -0.016215, -0.022032, -0.016136, -0.007489, 0.004027, -0.021179, -0.014721, 0.005314, -0.001821, -0.016084, -0.008021, 0.007244, -5.4e-05, -0.00647, 0.002061, -0.003268, -0.02526, -0.016022, -0.007247, -0.002363, -0.023768, -0.02247, -0.009482, 6.3e-05, -0.024167, -0.021387, -0.000739, -0.008883, -0.014065, -0.017096, -0.018336, -0.012384, -0.002617, 0.008962, -0.02348, -0.011488, -0.018308, -0.001655, 0.003334, -0.005262, -0.019078, -0.020252, 0.001825, -0.006194, -0.02208, -0.016715, -0.001092, -0.000855, -0.006779, -0.019829, -0.014152, -0.014125, -0.003427, -0.011534, -0.015656, -0.016625, -0.004681, -0.010007, -0.018977, -0.017392, -0.001987, 0.003123, -0.008877, -0.019216, -0.016266, -0.019749, -0.003643, -0.00468, -0.000562, 0.00038, -0.014478, -0.027559, 0.004325, 0.00228, -0.023392, -0.022846, -0.014756, -0.024076, 0.002437, 0.00446, -0.000793, 0.003132, -0.025726, -0.014923, -0.023985, -0.018391, -0.010153, -0.008497, -0.0027, 0.002362, -0.020962, -0.014085, -0.007489, -0.009991, -0.016373, -0.011452, -0.009226, 0.001469, 0.001287, -0.037242, -0.022186, 0.004152, 0.006335, -0.01873, -0.022383, -0.015842, -0.014863, -0.011765, -0.006921, -0.00588, -0.012788, -0.017841, -0.014788, 0.001264, -0.002788, -0.013978, -0.022869, -0.001634, 0.001787, -0.020125, -0.012178, -0.002966, 0.00207, -0.01101, -0.013503, -0.019331, -0.011314, -0.003044, 0.010666, -0.000702, -0.001546, -0.022786, -0.014715, 0.001134, 0.009413, -0.023911, -0.011462, -0.00326, 0.001864, -0.022526, -0.024709, -0.004014, -0.008703, -0.006202, -0.018831, -0.022917, -0.025552, 0.003422, -0.003844, -0.005336, -0.003351, -0.013176, -0.01431, -0.003151, -0.00187, -0.017543, -0.014523, -0.018582, -0.006378, -0.010014, -0.004245, 0.002287, -0.003445, -0.030831, -0.01267, 0.002786, 0.003481, -0.009417, -0.022259, 0.001241, 0.003976, -0.019917, -0.000661, -0.005348, -0.001373, -0.010885, -0.008288, 0.003928, 0.00076, -0.018842, -0.007341, -0.003084, -0.008867, -0.023929, -0.013614, 0.005338, -0.003117, -0.011466, -0.010271, -0.001251, 0.001522, -0.005551, -0.004674, 0.001537, 0.001537, -0.009002, -0.006766, -0.007737, -0.007265, -0.010757, -0.006131, -0.005506, 0.002664, -0.014571, -0.014946, 0.000476, 0.001383, -0.018831, -0.017189, -0.002911, -0.003265, -0.011749, -0.017577, -0.001352, 0.003732, -0.015282, -0.015751, -0.003225, -0.006276, -0.008054, -0.010017, 0.001179, -0.003953, -0.014942, -0.015998, -0.017423, -0.006262, 0.006304, 0.002675, -0.004908, -0.006501, -0.017383, -0.012896, 0.006376, -0.006742, -0.024391, -0.011626, 0.002644, -0.003916, -0.014527, -0.00356, -0.009639, -0.02465, -0.006309, 0.001052, -0.00313, -0.018865, -0.011402, 0.004129, 0.003814, -0.017598, -0.022779, 0.001376, 0.004796, -0.021439, -0.00832, -5e-05, -0.001214, -0.012897, -0.015729, -0.002587, -0.000368, -0.022866, -0.013553, -0.00245, -0.008319, -0.017808, -0.01733, -0.009346, -0.008819, -0.001909, 0.002375, -0.008694, 0.007126, -0.005214, -0.006602, 0.002932, -0.005352, -0.018302, -0.0127, -0.014825, -0.012654, -0.003696, -0.002984, 0.000191, 0.003746, -0.010908, -0.007225, -0.011439, -0.002519, -0.010082, -0.011524, -0.009021, -0.011392, -0.005632, -0.004465, -0.00142, 0.004129, -0.011361, -0.009652, 0.003733, 0.000981, -0.020263, -0.013646, 0.002234, 0.003614, -0.013683, -0.021306, -0.003726, 0.001506, 0.00627, -0.021354, -0.010762, -0.005524, -0.002502, -0.023239, -0.01172, 0.002656, -0.000313, -0.013707, -0.004241, -0.013911, -0.007788, 0.003329, 0.000169, -0.002957, -0.016353, -0.013106, -0.003283, -0.010263, -0.016093, -0.007155, -0.000985, -0.008603, -0.0141, -0.011815, -0.00557, -0.000119, -0.011885, -0.01534, -0.001366, -0.001002, -0.012303, -0.008099, -0.002652, 0.001241, -0.01914, -0.011098, -0.000678, -0.009808, -0.017761, -0.012397, -0.005164, -0.001978, -0.008339, -0.011475, -0.004496, -0.004401, -0.009993, -0.002129, 0.004545, -0.000966, -0.008742, -0.007239, -0.002886, 0.003488, -0.010936, -0.008853, 0.000423, -0.00403, -0.024184, 0.001819, -0.003204, 0.003113, -0.020829, -0.011226, -0.003953, -0.000878, -0.016515, -0.011338, 0.009098, -0.000973, -0.0214, 0.001079, -0.004502, -0.004743, -0.008382, 0.001509, -0.004741, -0.016766, -0.008501, 0.007374, 0.002875, -0.015518, -0.015091, 0.002376, -0.001871, -0.006028, -0.017441, 0.006766, -0.001274, -0.017266, -0.004746, 0.015177, 0.003478, -0.011636, -0.008906, -0.000293, -1.7e-05, -0.021636, -0.02519, -0.002155, -0.000436, -0.015833, -0.012087, -0.001185, -0.009472, -0.014721, -0.0273, -0.007356, -0.013969, -0.023236, -0.009442, 0.000382, 0.004503, -0.017312, -0.021732, -0.000133, -0.001478, -0.008219, -0.010795, 0.001234, -0.00648, -0.003895, -0.006921, -0.006035, -0.019624, -0.018953, -0.018062, -0.013421, -0.003181, -0.01085, -0.012904, -0.000368, -0.010297, -0.012591, -0.013615, 0.00905, -0.020045, -0.017424, 0.003984, -0.008908, -0.012216, -0.012426, 0.003081, 0.00479, -0.010764, -0.005197, -0.005753, -0.009413, -0.016108, -0.017849, -0.013168, -0.004491, -0.006863, -0.001186, 0.006827, -0.00091, -0.012031, -0.012248, 0.001713, -0.008153, -0.013552, -0.011683, -0.012606, -0.007911, 0.004398, -0.008895, 0.000536, -0.00354, -0.014719, -0.007817, 0.008242, 0.001703, -0.023719, -0.013731, 0.00039, -0.000321, -0.016942, -0.008706, 0.004762, 0.003257, -0.018786, -0.010692, -0.00276, -0.000633, -0.016848, -0.007244, 0.002339, -0.009825, -0.013902, -0.015826, -0.013781, -0.012154, -0.005815, -0.004351, -0.002861, 0.005668, -0.017971, -0.019288, 0.004453, -0.0034, -0.024205, -0.009884, 0.002069, -0.006735, -0.018094, -0.016388, 0.002498, -0.002114, -0.02101, -0.012054, 0.00324, -0.007366, -0.005639, -0.011376, 0.000704, 0.000488, -0.008591, -0.009447, -0.004431, -0.014143, -0.000984, -0.010717, -0.000624, -0.003584, -0.012511, -0.015902, -0.004614, 0.001231, -0.008702, -0.013876, -0.000505, -0.000155, -0.003931, -0.006362, -0.008527, -0.007029, -0.01459, -0.015439, 0.000591, 0.002062, -0.012744, -0.009075, -0.009279, -0.005888, -0.00726, 0.002675, -0.006365, -0.004441, -0.015098, -0.01068, -0.008795, 0.001441, -0.018174, -0.011044, -0.004083, -0.000696, -0.008701, -0.014982, -0.005898, 0.005803, -0.00925, -0.01153, -0.002838, -0.003172, -0.017763, -0.005842, -0.006168, -0.000653, -0.017203, -0.006252, -0.007104, 0.002148, -0.000221, -0.013435, -0.016491, 0.002826, -0.00927, -0.012502, -0.009817, -0.000284, -0.002192, -0.011563, -0.006484, -0.005575, -0.003886, -0.017702, -0.009486, 0.00519, 0.00135, -0.009652, -0.014097, -0.003335, -0.00249, -0.014882, -0.004813, 0.004932, -0.017763, -0.009524, 0.001709, -0.001748, -0.021127, -0.012279, -0.00602, -0.001888, -0.001097, -0.002936, 0.003371, -0.0101, -0.019975, -0.013625, -0.015517, -0.009977, 0.00463, -0.010936, -0.001167, -0.005819, -0.005968, -0.004698, -0.00221, 0.000768, -0.020031, -0.010239, -0.001822, 0.00394, -0.003045, -0.005622, 0.002435, -0.003208, -0.007468, -0.020215, -0.009363, -0.006586, -0.000635, 0.001531, -0.019743, -0.005663, -0.006706, 0.0015, -0.012802, -0.018765, -0.010231, -0.010786, 0.000839, 0.002368, -0.00176, -0.000191, -0.011411, -0.023509, 0.000483, -0.006626, -0.016469, -0.013214, 0.003059, -0.000207, -0.014115, -0.006741, 0.006077, 0.007291, -0.008461, 0.003038, -0.003502, -0.016063, -0.011124, -0.004156, -0.002878, -0.006145, -0.009769, -0.000968, 0.009137, -0.011103, -0.00544, -0.01518, -0.001723, -0.00331, -0.001428, 0.001251, -0.001796, -0.011388, -0.00378, 0.003916, 0.000339, -0.00844, -0.001162, -0.002144, -0.01637, -0.012658, -0.003905, -0.003794, -0.015528, -0.018593, -0.006225, -0.004998, -0.016376, -0.009038, 0.005964, 0.00528, -0.013404, -0.024403, -0.002238, -0.009583, -0.013741, -0.02219, -0.00118, -0.004543, -0.023298, -0.012204, 0.008461, -0.000682, -0.011688, -0.00671, -0.002107, -0.001865, -0.01209, 0.01632, 0.0046, -0.00552, -0.00378, -0.00872, -0.00278, -0.00052, -0.01427, -0.00595, -0.00145, -0.00434, -0.00604, -0.00162, -0.0006, -0.00301, -0.00621, 0.00843, 0.00392, 0.00436, -0.00257, -0.00031, 0.00894, 0.00461, 0.00257, 0.00237, 0.00545, -0.01154, 0.00151, 0.00458, -0.00793, -0.00029, -0.00227, -0.00348, -0.00153, -0.00116, -0.00244, -0.01268, 0.0074, -0.00937, 0.00134, -0.00039, -0.00362, -0.00506, -0.00271, -0.00698, -0.002, -0.00215, 0.00083, -0.00159, 0.00434, -0.00237, -0.0026, -0.0142, 0.00421, -0.00345, -0.00863, -0.00491, -0.00453, -0.00586, -0.00426, 0.00356, 0.00165, -0.00194, -0.01124, -0.00728, 0.00663, 0.00052, 0.00821, -0.00882, -0.00429, -0.01112, 0.00285, -0.00061, 0.01125, -0.00398, -0.00971, -0.00262, -0.00228, 0.00424, -0.00045, 0.00483, -0.00478, -0.00291, 0.00247, -0.00328, 0.00324, -0.00366, 0.00777, -0.00932, -0.00682, -0.01156, 0.00666, 0.0025, -0.00211, -0.00199, -0.00421, -0.00383, -0.00045, 0.00242, -0.00706, -0.00301, -0.01015, 0.00543, 0.00589, -0.00366, 0.002, -0.00127, -0.00773, 0.00425, -0.00081, -0.00464, 0.00183, -0.0068, -0.00065, 0.00186, 0.00121, -0.0084, 0.00346, -0.0057, -0.00562, -0.00215, 0.00734, 0.00562, 0.006, 0.00283, -0.00694, 0.00749, 0.00768, 0.01192, -0.00484, 0.00329, 0.00194, -0.01286, 0.00061, 0.0, 0.0043, 0.00886, 0.0004, 0.00506, -0.00544, 0.00887, 0.00254, -0.00114, -0.00458, -0.00689, -0.00486, -0.00486, -0.00477, -0.00158, 0.00634, -0.00321, -0.00096, -0.00124, 0.01123, -0.00247, -0.00722, -0.01089, 0.00361, -0.01878, -0.00487, -0.00356, -0.00192, 0.00572, 0.00461, -0.01161, -0.00825, 0.00083, -0.00046, -0.00957, -0.00639, -0.00436, 0.00559, -0.00657, 0.00728, -0.01119, -0.0035, -0.01572, -0.00052, -0.00199, 0.00106, 0.00214, 0.0, -0.00725, -0.0007, 0.00039, 0.0088, -0.00165, 0.00445, -0.0079, 0.00189, -0.01237, -0.00907, 0.00644, -0.00113, -0.01236, 0.00434, 0.0006, -0.00641, -0.00554, -0.00343, 0.00352, 0.00262, 0.00105, 0.00853, 0.00582, -0.0051, -0.00547, -0.00428, 0.00611, -0.00128, 0.00491, 0.00991, -0.00537, 0.00221, -0.00501, -0.00785, -0.00016, -0.00174, -0.00604, -0.00279, -0.00054, -0.00551, 0.00127, -0.00021, -0.01437, -0.00632, 0.00786, -0.004, 0.00923, 0.00431, 0.00117, 0.0013, 0.00209, 0.00751, -0.00695, 0.00218, 0.00016, 0.00516, 0.00243, 0.00113, -0.01135, -0.01012, 0.00051, 0.00075, -0.00849, 0.005, -0.00506, 0.0, -0.00487, -0.00193, 0.00871, 0.00734, 0.00622, 0.00169, -0.01011, -0.00137, 0.00794, 0.00612, -0.00336, 0.00512, -0.00096, -0.0026, -0.00034, 0.00742, -0.01261, -0.00262, -0.00531, 0.00446, 0.00487, -0.00611, -0.00411, 0.00329, -0.00411, 0.00194, 0.00394, -0.01217, -0.00264, -0.01695, -0.0045, 0.00311, -0.00282, 0.00473, 0.0012, -0.00737, -0.00167, -0.00357, 0.00014, -0.00041, -0.00041, 0.00213, -0.00606, 0.00537, -0.00433, -0.01031, -0.00405, -0.00163, 0.00666, -0.0008, -0.00103, 0.00553, 0.00126, -0.00659, 0.00362, -0.00366, 0.00283, 0.00334, 0.00377, -0.00792, -0.00515, 0.00142, -0.00511, -0.00334, -0.00763, 0.00162, -0.00447, -0.00285, -0.00162, -0.00165, 0.00793, -0.00036, -0.00528, 0.00167, 0.00433, 0.00455, 0.00911, 0.00166, 0.00198, 0.00177, -0.00516, -0.00528, -0.00918, -0.00901, 0.00073, 0.00098, -0.00956, -0.00988, 0.00674, -0.00724, -0.01124, 0.0018, -0.00548, 0.01679, 0.00069, 0.00229, -0.00154, 0.00232, 0.00706, -0.00371, -0.00686, -0.00163, -0.00225, -0.00333, 0.0, 0.00747, 0.00024, 0.00071, -0.01039, -0.00066, 0.00649, -0.00088, -0.00685, -0.00812, -0.00249, 0.0, -0.00867, -0.0028, 0.00731, -0.00414, -0.00336, -0.00459, 0.00206, -0.01238, -0.00391, -0.00629, -0.00733, -0.01016, -0.00696, 0.00112, 0.00476, -0.01405, 0.00233, 0.00164, -0.00306, 0.00379, -0.00277, 0.0028, 0.00369, 0.00849, 0.00506, 0.00116, -0.02965, -0.00247, 0.0031, 0.01262, -0.00929, 0.00191, -0.00467, -0.00258, 0.00748, -0.02233, -0.00326, -0.00095, -0.00844, -0.00806, -0.00294, -0.0116, 0.00138, 0.00063, -0.0033, -0.00041, 0.00448, -0.00261, -0.00167, -0.00659, -0.00193, 0.00441, -0.006, -0.0159, 0.00017, 0.00178, 0.01157, -0.00826, -0.00511, -0.00219, 0.00244, -0.00543, -0.01021, -0.00703, -0.00414, -0.00168, 0.0061, -0.00585, -0.00413, 0.00208, -0.00028, -0.00836, -0.00145, -0.00402, -0.01088, -0.00123, -0.00113, 0.00822, 0.00824, 0.00256, 0.00154, 0.00275, 0.00312, -0.00684, 0.00514, 0.00545, 0.00728, -0.0029, -0.00259, -0.00086, -0.00338, -0.00028, -0.00305, -0.00474, 0.00423, -0.00185, 0.00026, 0.00062, 0.0052, -0.00048, -0.00258, 0.00025, -0.00265, 0.01128, 0.00307, 0.00554, 0.00434, 0.00165, -0.00028, -0.00254, -0.00334, -0.00044, -0.0064, -0.00286, -0.00315, -0.00214, -0.0041, -0.00916, -0.00717, 0.00018, 0.00589, -0.00576, 0.00489, -0.00379, -0.00335, -0.00193, -0.00493, 0.00446, 0.00192, 0.0005, -0.00207, -0.00145, -0.00447, -0.01432, 0.00091, 0.00436, -0.01487, -0.00179, 0.00742, -0.00294, 0.00717, -0.00592, -0.00107, -0.00797, -0.00532, 0.00217, -0.00105, -0.00819, -0.00175, -0.00014, -0.00283, -0.00763, -0.00813, -0.00399, 0.0033, 0.0074, -0.00492, 0.0, 0.00515, -0.00164, -0.00684, 0.00277, 0.00369, -0.00379, -0.00348, 0.0077, 0.0133, -0.00177, -0.00505, 0.00357, -0.00528, -0.00749, 0.00438, -0.0019, -0.00515, -0.00168, 0.00074, -0.00771, -0.00283, -0.00182, -0.00716, -0.00337, -0.0033, 0.00088, 0.0071, 0.002, -0.00014, -0.00318, -0.00138, -0.0033, -0.00289, 0.00614, 0.00709, 0.0024, -0.00529, -0.01176, -0.0078, 0.00573, -0.00928, -0.0102, -0.00895, 0.00755, 0.00639, 0.00155, -0.00594, 0.00557, -0.0062, -0.01457, -0.00019, 0.01235, -0.00506, -0.00614, 0.00557, -0.00626, -0.0029, -0.0067, -0.00348, 0.00493, 0.00032, 0.00095, -0.01219, -0.00876, -0.01093, -0.0106, -0.00464, -0.00342, 0.00146, 0.0, -0.00768, -0.00752, -0.00376, -0.00528, -0.01817, -0.01056, 0.0048, -0.00105, -0.01418, -0.0066, -0.00881, -0.0047, -0.01204, -0.0111, -0.01032, -0.00631, -0.01313, -0.01183, -0.0109, -0.00932, -0.00743, -0.00964, -0.00253, -0.00343, -0.01172, -0.00092, -0.01219, -0.00179, 0.0043, -0.00362, -0.01571, -0.00542, 0.0032, -0.00306, 0.00083, -0.00219, -0.00306, -0.0059, -0.00793, 0.00359, 0.00491, -0.01205, -0.0207, -0.00426, -0.00094, -0.01896, -0.01376, -0.00727, 0.00214, -0.01614, -0.01899, -0.00297, -0.00494, -0.00668, -0.02279, -0.00664, -0.0031, -0.01247, -0.0108, -0.00438, 0.00799, -0.01674, -0.01299, 0.00087, -0.0009, -0.01379, -0.01112, -0.0094, -0.01286, -0.00359, 0.00048, -0.01856, -0.02039, -0.01204, -0.01398, -0.01035, -0.00777, -0.007, -0.00695, -0.00928, -0.01247, -0.00351, -0.01011, -0.01537, -0.00506, 0.00734, -0.00664, -0.00123, 0.00756, -0.00104, -0.01032, -0.00537, 0.00348, -0.00793, -0.00444, -0.01705, -0.0066, -0.00741, -0.01487, -0.00552, -0.00356, -0.00527, -0.01445, -0.00638, -0.0134, -0.01092, -0.0103, -0.01564, -0.00694, -0.01901, -0.00434, -0.01227, -0.01201, -0.0054, -0.01939, -0.01352, -0.00302, -0.00929, -0.0183, 0.0, -0.00236, -0.00643, -0.01197, -0.01405, 0.00413, -0.01283, -0.0119, -0.00675, -0.00276, -0.01432, -0.01139, -0.00317, 0.00048, -0.01579, -0.00384, -0.00688, 0.00166, -0.01389, -0.01239, -0.00919, -0.01141, -0.00237, -0.01511, -0.01783, -0.00843, -0.01397, -0.00758, -0.00591, -0.00984, -0.01282, -0.0025, -0.00255, -0.00385, -0.00728, -0.00745, -0.00576, -0.00951, 0.00249, -0.00382, -0.00869, -0.0137, 0.00636, -0.00132, -0.00782, -0.02429, -0.0163, -0.00474, -0.00979, -0.00569, -0.00318, -0.00755, -0.0136, -0.01564, -0.00208, -0.01688, -0.00796, -0.00621, -0.00432, -0.00947, -0.00733, 0.00239, -0.00748, -0.01468, -0.00744, 0.00158, -0.00048, -0.00951, -0.00553, -0.01062, -0.01249, -0.01843, -0.00085, -0.01053, -0.00355, -0.01121, -0.01128, -0.01255, -0.01006, -0.00388, -0.00489, -0.00984, -0.00057, -0.01817, -0.01465, -0.00412, -0.0143, -0.0138, -0.00218, -0.013964, -0.000321, -0.004692, -0.0025, 0.007696, 0.001936, -0.001229, -0.007474, -0.004071, -0.004585, -0.004381, -0.014232, -0.001326, -0.01104, 0.003955, -0.007101, -0.011034, -0.00415, -0.01197, -0.01336, -0.010088, -0.013282, -0.00555, -0.01534, 0.001752, -0.006785, 0.0035, -0.005784, -0.007929, -0.00757, -0.006511, 0.001193, -0.008909, -0.005919, 0.009912, -1.9e-05, -0.007143, -0.01648, -0.007352, -0.012278, -0.00523, 0.000941, 0.002715, -0.002446, -0.004557, -0.001727, -0.006573, -0.010529, 0.001064, -0.007575, -0.005593, -0.001148, -0.009806, 0.004737, 0.002417, -0.007084, -0.001154, 0.003137, -0.01172, -0.012237, 0.001608, -0.006458, -0.010103, -0.006977, -0.004052, -0.012586, -0.014656, -0.009572, -0.002845, -0.010888, -0.00405, -0.013561, -0.010006, -0.005212, -0.01222, -0.004777, -0.004542, 0.002614, -0.005777, -0.004966, -0.001148, -0.007078, -0.017365, 0.001323, -0.008876, 0.00153, -0.014523, -0.001731, -0.005198, -0.010162, -0.015433, -0.008891, -0.005351, -0.009541, 0.005014, -0.001202, -0.014548, -0.002528, 0.000588, 0.002603, -0.002825, -0.011836, -0.004227, -0.00604, 0.000296, -0.014162, -0.014428, -0.005901, -0.01412, -0.017753, -0.004605, -0.01233, -0.00165, -0.009253, -0.004949, -0.003171, 0.003803, -0.005824, 0.000887, 0.000306, -0.004179, -0.002379, -0.008476, -0.012575, -0.013251, -0.010312, -0.006218, -0.000567, -0.007819, -0.017152, -0.005879, -0.006253, -0.008308, -0.00171, -0.007385, -0.00801, -0.007299, -0.0042, -0.018528, -0.005322, -0.007545, -0.009477, -0.007745, -0.009069, -0.00848, 0.002887, -0.008652, -0.012732, -0.006868, -7.3e-05, 0.00207, -0.000289, -0.00139, -0.006466, -0.013998, -0.004984, -0.009423, -0.018283, -0.008807, -0.003751, -0.012931, -0.002775, -0.012117, -0.011167, -0.017764, -0.00124, -0.01087, -0.006561, -0.000416, 0.007074, -0.001657, -0.003877, -0.001692, -0.006494, 0.001315, -0.000666, -0.012383, -0.000407, -0.003373, 4.6e-05, -0.010767, -0.009957, -0.00603, -0.006695, 0.008388, -0.00062, -0.004976, -0.007232, -0.015771, -0.017215, 0.004363, 0.000663, -0.007419, -0.014939, -0.002221, -0.008964, -0.001842, 0.002399, -0.010365, -0.00861, -0.003903, -0.005317, 0.000517, -0.006399, -0.00925, -0.011284, -0.008134, -0.007317, -0.016594, -0.016161, -0.004234, -0.005883, -0.011042, -0.006601, -0.001237, -0.016281, -0.014527, -0.013618, 0.000723, -0.012025, -0.003795, -0.000812, -0.001898, -0.003397, -0.006526, -0.005319, -0.000471, -0.002018, 0.002567, -0.003879, -0.009655, 0.001571, 0.002663, -0.013268, -0.000476, -0.00506, -0.003416, -0.003797, -0.006404, -0.01918, -0.005008, -0.006306, -0.004393, -0.010251, 0.001548, 0.0081, -0.004036, -0.017332, -0.008007, -0.004384, -0.008378, 0.003308, 0.000699, 0.004781, -0.011023, -0.018384, -0.004753, -0.005761, -0.006867, -0.006102, -0.007374, -0.005273, 0.004407, 0.003671, 0.000436, -0.001928, -0.011078, -0.011645, -0.00072, 0.000567, -0.005141, -0.003668, -0.001453, -0.009159, -0.012249, -0.000381, -0.007354, -0.002816, -0.005402, 0.00271, -0.008473, -0.013626, -0.009368, -0.005964, -0.012201, 0.0054, -0.000457, -0.010647, -0.002675, -0.003751, -0.003965, -0.002168, 0.000404, -0.009696, -0.007217, -0.009371, -0.001047, -0.013418, -0.003382, 0.001837, -0.004665, -0.013212, -0.009747, -0.013779, 0.004446, -0.001136, 0.001135, -0.015508, -0.006395, -0.001898, -0.01235, -0.004545, -0.000671, -0.010387, -0.001441, -0.007653, -0.000158, -0.000673, -0.004628, -0.008367, -0.011071, -0.011865, 0.000943, -0.001609, -0.003883, -0.007708, 0.002445, -0.002627, -0.016417, 0.00246, -0.003865, -0.008151, -0.003717, -0.011208, -0.021246, 0.001793, -0.00144, 0.001157, -0.003413, 0.00412, -0.006281, 0.003048, -0.013156, -0.012205, -0.008813, 0.001851, 0.003303, -0.007285, -0.012283, -0.002458, -0.001335, -0.019185, -0.00574, -0.005005, -0.002063, -0.012185, 0.000894, -0.001597, -0.003145, -0.007871, -0.001745, -0.009808, -0.009453, 1e-06, -0.010537, -0.00651, -0.002409, 0.000234, -0.011944, -0.003228, -0.001461, -0.000605, -0.002449, -0.001102, -0.004429, 0.009575, -0.010405, -0.002502, -0.001872, -0.003623, -0.008857, 0.002597, -0.001822, -0.005769, -0.00829, -0.001395, -0.005834, -0.005965, -0.002305, -0.001479, -0.006387, -0.010963, -0.001006, 0.003068, 0.000588, -0.000574, 0.002538, -0.002215, -0.005189, -0.009723, -0.01391, 0.004314, -0.012319, -0.017737, -0.006871, -0.005466, -0.008339, -0.006705, -0.003977, -0.005899, -0.0078, -0.010176, -0.00865, -0.002835, -0.00537, 0.003122, -0.002824, -0.001614, -0.01352, -0.001342, 0.001238, -0.010113, 0.006063, -4.6e-05, -0.003457, 0.004366, -0.005628, -0.009351, -0.003863, 0.001972, -0.007676, -0.00986, -0.007052, -0.003161, 0.004623, -0.00235, -0.011452, 9.4e-05, -0.010663, -0.000386, 0.000972, -0.005464, 0.003899, -0.002766, -0.005131, 0.00092, 0.003428, -0.006334, -0.003338, -0.003802, -0.012208, -0.009357, -0.005827, -0.006067, -0.00299, -0.001781, -0.008454, -0.000754, 0.002866, -0.007146, 0.000665, -0.003526, 0.001019, -0.00988, -0.006289, -0.007721, -0.010317, -0.009647, -0.004955, -0.00783, -0.00063, 0.007876, 0.001857, -0.009998, -0.008619, -0.007988, -0.009349, -0.004505, -0.013599, 0.00721, -0.006831, -0.015792, -0.008854, -0.003397, 0.001202, -0.000793, -0.006395, 0.011051, -0.009147, -0.006865, 0.00011, -0.002317, -0.005088, -0.009004, -0.007318, -0.003621, -0.007628, -0.003565, -0.00345, 0.001071, -0.010404, -0.002593, -0.005152, -0.007037, -0.001421, 0.000144, -0.016907, -0.00257, -0.00782, 0.007906, -0.012253, -0.005063, -9.1e-05, -0.003802, -0.002763, -0.013917, -0.003387, -0.010492, -0.010215, -0.005586, -0.002326, -0.005819, 0.000175, -0.002324, -0.011132, -0.001896, -0.006423, -0.010126, -0.012311, 0.001111, -0.004598, -0.001586, -0.009051, -0.003557, -0.005143, -0.007975, -0.003528, -0.004498, -0.001913, -0.002131, -0.008947, -0.009592, -0.018292, -0.008484, -0.001147, -0.005855, -0.00108, -0.001167, -0.007697, -0.009375, 2e-05, -0.015293, -0.000962, -0.011906, -0.007761, -0.009102, -0.007881, -0.009613, -0.004747, -0.006706, -0.00823, -0.002419, -0.006116, -0.002431, -0.0097, -0.009481, -0.001273, -0.005699, -0.002942, -0.001955, -0.012377, 0.002117, -0.00359, -9.7e-05, -0.004179, -0.009861, -0.02148, -0.002187, -0.004824, -0.008395, -0.008017, -0.005364, -0.006339, 0.000652, -0.005698, -0.001312, 0.008458, -0.015583, -0.00385, -0.004444, -0.003281, -0.006638, -0.005562, -0.009787, -0.002753, -0.011835, 0.003215, -0.002192, -0.00272, -0.004975, -0.001701, 0.009147, -0.015003, -0.006995, -0.008278, -0.003232, -0.000762, -0.005539, -0.012603, -0.003327, -0.006514, 0.000139, 0.001126, 0.001883, -0.009425, -0.002391, -0.003555, -0.006622, -0.013033, -0.010824, -0.003866, -0.007913, -0.004618, -0.002437, -0.003394, -0.006364, -0.003462, -0.000178, -0.006276, -0.00468, -0.003358, -0.016316, -0.001473, -0.005788, -0.003478, 0.002821, -0.004787, 0.00137, -0.000463, 0.009948, 0.001019, -0.006959, -0.004642, -0.001381, 0.003631, -0.004681, 0.010685, -0.01263, -0.00469, -0.01299, -0.00621, -0.01716, -0.01351, -0.01346, -0.00879, -0.00754, -0.00406, -0.01265, -0.0053, -0.01696, 0.00529, -0.00291, 0.02042, -0.01007, 0.01035, -0.00473, -0.00047, -0.01416, 0.00443, -0.01074, -0.00363, 0.00944, -0.00294, -0.01431, -0.01052, 0.01089, 0.01193, -0.01823, 0.00519, -0.00629, -0.0054, -0.01318, 0.00803, 0.00402, 0.01274, 0.00479, -0.00701, -0.0111, -0.00499, -0.00696, 0.01198, -0.00667, 0.01672, 0.00762, 0.01019, 0.00704, -0.00716, -0.01843, 0.00597, -0.00845, 0.00756, 0.00124, -0.01003, 0.00559, 0.00905, -0.01254, 0.00444, -0.01567, -0.00108, 0.00461, -0.00778, -0.01551, 0.00459, 0.00166, 0.00251, -0.01014, 0.00818, 0.00468, -0.00925, 0.01663, 0.00075, -0.01296, 0.0074, 0.01, 0.01343, 0.01196, -0.00133, 0.00934, 0.01364, -0.01073, -0.00908, 0.0108, 0.01293, -0.00151, 0.00116, -0.01224, 0.00701, 0.00167, 0.00412, -0.00077, 0.00134, -0.01423, -0.01212, 0.0048, -0.00915, -0.01051, 0.00863, -0.00582, -0.00012, -0.01431, -0.00474, -0.00154, 0.00299, -0.02589, -0.01926, -0.0133, 0.00695, -0.01206, 0.00422, -0.00675, -0.00042, -0.00017, 0.01363, -0.01388, 0.00538, -0.00318, -0.00094, -0.00057, -0.00954, -0.01007, -0.00404, -0.00783, 0.0081, -0.01236, -0.01094, -0.00779, 0.00287, -0.00361, 0.00727, 0.00969, 0.01014, -0.00783, -0.0038, -0.00488, 0.01665, -0.01496, -0.01259, -0.00689, 0.01766, -0.00946, 0.00501, -0.01066, -0.00732, -0.01233, -0.01144, 0.01166, -0.00155, -0.00095, 0.00247, -0.01513, -0.0117, 0.00566, 0.00755, -0.00598, 0.00036, 0.0, 0.00492, 0.00743, -0.00436, -0.01112, -0.01117, -0.02, -0.01568, -0.009, -0.00392, -0.01428, 0.00275, -0.01447, 0.0061, 0.00231, 0.00672, 0.00083, -0.0114, -0.00327, 0.01898, 0.00045, -0.01035, -0.01238, 0.00792, -0.01772, -0.01443, 0.00574, 0.00867, -0.00292, -0.01631, 0.01601, 0.00311, 0.01818, -0.00987, -0.00936, -0.00397, -0.0039, 0.00684, -0.00465, -0.01093, -0.00931, -0.00822, 0.00857, 0.00628, -0.00188, 0.00694, 0.01035, 0.00124, -0.01051, -0.01589, 0.00535, -0.01023, -0.00787, -0.00765, -0.01008, 0.00077, 0.00609, 0.00892, 0.01304, -0.0119, 0.00537, 0.00643, 0.00811, -0.01368, 0.00561, 0.00574, -0.00571, 0.00381, -0.00694, -0.0067, 0.01196, -0.01274, 0.00627, 0.00862, -0.00271, -0.01387, -0.00953, -0.00258, 0.00274, -0.01312, 0.00281, -0.01535, -0.00248, -0.01274, 0.01097, -0.01076, 0.00477, -0.01551, -0.01043, -0.00952, -0.00417, -0.00937, -0.009, 0.00714, 0.01364, 0.00866, 0.01344, -0.01307, 0.00343, -0.0192, -0.0054, -0.0065, -0.00458, 0.00121, 0.00441, 0.01065, -0.00331, -0.00356, 0.01435, -0.01179, -0.01244, -0.01434, 0.0073, -0.0157, 0.0047, -0.00194, 0.01669, -0.00314, 0.018, -0.01228, -0.00436, 0.00406, 0.00683, 0.00669, 0.01089, -0.00164, -0.01088, -0.01257, -0.00638, -0.01108, 0.01232, 0.00358, -0.00975, 0.00433, -0.0036, -0.01252, -0.01041, -0.01537, -0.01444, -0.01487, 0.00579, -0.00189, 0.00245, -0.01094, 0.01143, -0.00815, 0.01229, 0.00538, 0.00986, -0.00288, 0.01634, -0.00856, -0.00226, -0.00599, 0.00989, -0.00031, 0.00493, -0.00942, 0.00283, -0.00618, -0.00683, 0.00087, -0.0039, -0.01856, -0.01129, 0.0099, 0.00676, -0.00469, -0.01332, 0.00822, -0.01423, 0.00624, -0.00904, -0.00619, -0.00974, -0.0067, 0.00623, 0.00938, -0.01645, 0.00285, -0.0107, 0.00572, -0.01823, -0.01129, -0.01725, 0.00131, -0.00735, -0.0023, -0.00487, -0.00461, 0.00665, 0.01341, -0.00828, 0.00721, -0.00166, -0.00115, -0.0119, -0.00442, 0.00154, 0.00483, -0.01164, -0.01293, 0.00648, -0.01128, -0.00111, -0.00499, -0.00772, -0.00949, 0.00546, -0.01254, -0.01038, -0.01032, -0.01125, -0.01208, 0.00805, -0.00143, -0.00629, -0.01139, -0.00723, -0.01485, -0.01584, 0.00281, 0.00215, -0.00656, -0.00705, -0.00757, -0.00019, -0.00428, -0.00644, -0.00475, 0.00686, -0.01239, -0.01158, 0.00423, -0.00723, -0.00905, -0.00053, -0.01457, -0.01441, -0.01238, 0.01101, 0.00392, 0.00383, -0.00184, 0.01509, -0.00803, -0.01135, 0.00142, -0.01128, -0.01009, 0.00383, -0.01215, -0.01324, 0.00654, -0.01288, -0.00832, -0.0117, -0.01033, -0.01448, 0.00583, -0.01498, -0.0053, -0.00861, 0.00231, -0.0135, -0.01769, -0.00094, -0.00074, 0.00582, -0.0199, -0.01385, -0.01945, -0.00216, 0.01314, 0.01599, -0.01437, -0.00762, -0.01731, -0.00557, 0.0033, -0.00031, 0.00077, -0.01191, -0.01381, -0.01463, -0.01645, -0.02053, -0.01234, -0.00935, 0.00531, -0.00742, -0.00483, -0.00657, -0.01238, 0.00927, 0.00602, 0.01237, -0.0142, -0.01016, -0.00155, -0.00909, -0.01311, 0.00652, -0.00622, -0.00889, -0.0046, -0.01979, -0.00708, 0.00775, -0.01087, 0.00333, -0.00296, 0.00191, 0.00092, 0.00831, -0.02261, 0.00117, -0.01279, -0.02103, -0.01291, -0.01488, -0.00914, -0.01417, -0.0096, 0.00883, 0.009, -0.00149, -0.00565, -0.00949, -0.00801, -0.01051, 0.00401, -0.00564, -0.01268, -0.01149, -0.00978, -0.00987, -0.01099, 0.00231, -0.00416, 0.00441, -0.00755, 0.00844, 0.0009, 0.00815, 0.01255, -0.00044, -0.01181, -0.01328, -0.01369, 0.00084, 0.01021, 0.01062, 0.0008, 0.00128, -0.01253, -0.01389, 0.00368, 0.00555, -0.00029, 0.00971, -0.00106, -0.01273, -0.00794, -0.01418, -0.00889, -0.00138, -0.03984, 0.00267, -0.01093, -0.01591, 0.00113, -0.0086, 0.01066, 0.00759, -0.01848, -0.00629, 0.0043, 0.00696, 0.00188, 0.0114, 0.0013, -0.01347, -0.00846, 0.0098, 0.01054, -0.01128, -0.01122, 0.00441, -0.00468, -0.00359, -0.01263, 0.00081, -0.01166, -0.00939, -0.0103, 0.00146, 0.00463, 0.00386, -0.01111, -0.0092, 0.00495, 0.00585, -0.00524, -0.01006, -0.00733, 0.00096, -0.01156, -0.0153, -0.00901, 0.00529, 0.013, -0.00018, -0.01308, -0.01011, -0.01123, -0.00727, -0.00908, -0.00975, -0.00023, -0.01613, -0.0097, 0.00854, 0.0039, 0.00277, 0.00961, -0.01063, 0.00427, -0.01543, -0.00054, -0.00569, 0.00186, 0.00946, 0.00656, -0.00137, -0.00895, 0.00222, 0.0053, 0.00676, 0.00235, -0.00851, -0.01202, 0.00617, -0.00709, -0.00289, 0.0025, -0.00579, 0.00101, -0.0158, -0.00776, 0.00382, 0.00527, 0.00952, 0.00135, -0.01321, -0.00768, -0.01436, -0.01791, -0.00711, 0.00557, -0.00612, -0.01543, -0.00859, 0.00564, 0.00558, -0.0013, 0.00224, 0.01061, 0.0, -0.01388, 0.01211, -0.01135, -0.01037, 0.00756, -0.0118, -0.0094, 0.01114, -0.00038, 0.00804, -0.01618, 0.00481, -0.01521, -0.01603, -0.01014, -0.00851, 0.00704, 0.00369], ['Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_y', 0.118, -0.118, -0.0017, -0.01955, 0.00291, 0.00056, -0.01089, -0.00734, 0.01384, 0.00581, -0.01075, 0.00595, 0.00462, 0.01176, -0.01536, -0.00899, 0.01482, 0.00359, -0.01349, -0.01528, 0.01051, 0.00691, -0.01731, -0.00986, 0.00678, 0.00883, -0.01111, -0.00655, 0.00665, 0.00161, -0.0049, -0.00767, 0.0111, 0.00711, -0.0142, -0.00988, 0.01133, -0.00086, -0.02724, -0.00546, 0.01423, 0.00928, -0.01065, -0.01225, 0.00406, 0.00577, -0.01347, -0.00246, 0.01839, 0.01007, -0.01958, -0.01702, 0.00237, -0.00177, -0.0124, -0.01008, 0.01904, 0.00968, -0.00923, -0.00885, 0.00883, 0.00164, -0.00561, -0.01244, 0.0154, 0.00504, -0.01553, -0.00581, 0.00758, 0.0127, -0.01141, -0.01827, 0.01402, -0.00049, -0.00807, -0.00938, 0.00613, 0.01329, -0.01427, -0.00215, 0.01276, 0.00196, -0.01205, -0.00917, 0.00223, 0.01271, -0.00704, -0.0046, 0.00475, -0.00169, -0.01333, -0.00703, 0.0166, 0.01277, -0.01283, -0.01438, 0.02533, -0.01885, 0.02124, 0.01019, 0.02475, -0.01665, 0.01447, 0.00981, -0.0257, -0.00958, 0.01236, 0.01473, -0.02896, -0.01409, 0.01292, 0.01597, -0.01708, -0.01453, 0.01469, 0.01152, -0.01084, -0.01531, 0.01071, 0.00583, -0.01654, -0.01048, 0.00916, 0.00758, -0.02137, -0.00623, 0.01179, 0.01395, -0.00825, -0.01049, 0.01685, 0.00841, -0.01099, -0.02101, 0.02169, 0.01042, -0.01602, -0.00752, 0.00565, 0.00192, -0.01292, -0.00414, 0.01431, 0.01291, -0.0118, -0.0066, 0.02706, 0.01806, -0.01572, 0.01867, 0.02262, -0.01244, -0.01204, 0.01686, 0.00749, -0.00892, -0.00869, 0.03079, 0.00864, -0.01264, -0.00038, 0.01368, 0.02137, -0.01225, -0.00778, 0.02023, 0.00398, -0.00541, -0.01972, 0.00628, 0.01035, -0.01049, -0.01166, 0.0123, 0.01367, -0.0186, -0.00641, 0.01154, 0.00724, -0.01071, -0.00621, 0.01716, 0.01454, -0.00966, -0.02083, 0.02018, -0.00178, -0.02366, -0.00569, 0.0202, 0.01764, -0.01631, -0.0202, 0.00716, 0.00414, -0.01616, -0.00934, 0.01788, 0.00725, -0.01637, -0.01653, 0.01263, 0.00621, -0.0118, -0.01157, -0.01631, -0.01225, 0.02295, 0.01984, 0.01105, 0.02229, -0.02154, -0.00911, 0.01089, 0.00795, -0.01713, -0.01449, 0.01885, 0.00453, -0.01898, -0.00788, 0.02139, 0.01386, -0.01389, -0.01491, -0.00829, -0.01464, 0.02481, 0.00981, 0.02177, 0.00825, -0.0049, -0.017, 0.01354, 0.01159, -0.01158, -0.0155, 0.02176, 0.02241, -0.01879, -0.01494, 0.0243, 0.01254, -0.01899, -0.01555, -0.01775, -0.01303, 0.02091, 0.01687, 0.00794, -0.02397, -0.01042, 0.03153, 0.01647, -0.01229, -0.01344, -0.01819, -0.0101, 0.00905, 0.01062, 0.02095, 0.01256, -0.01232, -0.02072, 0.00686, 0.0087, -0.01891, -0.02079, 0.02433, 0.0181, -0.01573, 0.02057, 0.01986, -0.01963, -0.02079, 0.01061, 0.00862, -0.01371, -0.01292, 0.01557, 0.00886, -0.00666, -0.01103, 0.01697, 0.01729, -0.00468, -0.0127, 0.01961, 0.01604, -0.00663, -0.02168, 0.00161, 0.01556, -0.02154, -0.00865, 0.00891, 0.00506, -0.01837, -0.01141, 0.01571, 0.00126, -0.01616, -0.0213, 0.01905, 0.00912, -0.01465, -0.0098, 0.02156, 0.02164, -0.0091, -0.01052, 0.01008, 0.01472, -0.02066, -0.00596, 0.02007, 0.00526, -0.01326, -0.02099, 0.01118, 0.01481, -0.02003, -0.00936, 0.0129, 0.00797, -0.01404, -0.00945, 0.01524, 0.0137, -0.01678, -0.01682, 0.01342, -0.01428, -0.00629, 0.01125, 0.02373, -0.01438, -0.00937, 0.00927, 0.00126, -0.01402, -0.01379, -0.00379, -0.00546, 0.00829, 0.01066, 0.01419, 0.01481, -0.00812, -0.00978, -0.01523, -0.01511, 0.02133, 0.00632, 0.01457, 0.01804, -0.01679, -0.00812, -0.01538, -0.00814, 0.01166, 0.01792, 0.01802, 0.00621, -0.00769, -0.01208, 0.0052, 0.00672, -0.01624, -0.00687, -0.01971, -0.01558, 0.01143, 0.02223, 0.01019, 0.01077, -0.01739, -0.01262, -0.01263, -0.0151, -0.01559, -0.01579, -0.01139, 0.00581, 0.00471, 0.01088, 0.01093, -0.01049, -0.0085, 0.01352, 0.01102, -0.01954, -0.01076, 0.00943, 0.01293, -0.01172, -0.01046, -0.01512, -0.01254, 0.00896, 0.01372, 0.01493, 0.01147, -0.00643, -0.01101, 0.01131, 0.01087, 0.01428, -0.01646, -0.01258, -0.00636, 0.0085, 0.0228, 0.00932, 0.01075, -0.01113, -0.00969, 0.03221, 0.01446, -0.01439, 0.00944, -0.01037, 0.01696, 0.01473, -0.01271, -0.00869, -0.01372, -0.00877, 0.01282, 0.01378, 0.02352, 0.01663, -0.00765, -0.01364, 0.02203, 0.00777, -0.01661, -0.02119, 0.0091, 0.01132, -0.00848, -0.00602, -0.01484, 0.01746, -0.01387, 0.00349, -0.01298, 0.01707, -0.0165, -0.00801, 0.00337, 0.01654, -0.01243, -0.01493, -0.01804, 0.01157, 0.01286, -0.01118, -0.01805, 0.01391, 0.00644, -0.00555, -0.01024, 0.01521, 0.01092, -0.0081, -0.01883, 0.0202, 0.0054, -0.01431, -0.00854, 0.01719, 0.01367, -0.01153, -0.01027, 0.0088, 0.01021, -0.01043, -0.01344, 0.01613, 0.01043, -0.00839, -0.01989, 0.01046, 0.00771, -0.01648, -0.00625, 0.01782, 0.00767, -0.01851, -0.01186, -0.00739, 0.01082, 0.00633, -0.0171, -0.01252, -0.01257, 0.01544, 0.01583, -0.0159, -0.00989, -0.00724, 0.01899, 0.0128, 0.00124, -0.01709, -0.01184, -0.00694, 0.01851, 0.01428, 0.01367, 0.02027, -0.00749, -0.01415, -0.01528, -0.00884, -0.00608, -0.01714, 0.01133, 0.01534, 0.00488, 0.01018, 0.0209, 0.00231, -0.01323, -0.00513, 0.01684, 0.01209, -0.01962, 0.004, 0.00761, -0.02052, -0.00443, 0.02668, 0.01229, -0.01802, -0.01414, 0.0178, -0.00408, -0.02103, -0.02224, -0.01498, -0.01316, 0.01137, -0.00177, 0.00976, 0.01165, -0.01756, -0.00937, 0.01745, 0.00683, -0.01877, -0.0168, 0.02414, 0.0145, -0.01149, -0.02286, 0.02191, 0.00819, -0.01159, -0.00467, 0.01577, 0.01453, -0.01228, -0.01247, 0.02694, 0.00859, -0.00875, -0.00981, -0.01723, -0.01223, 0.01433, 0.01338, 0.01764, 0.01447, -0.01381, -0.02168, -0.01634, -0.00751, 0.01627, 0.0157, 0.01255, 0.0139, -0.01318, -0.00737, 0.01378, 0.01608, 0.00208, -0.01329, 0.01347, 0.00737, -0.00239, -0.00322, 0.01515, 0.00324, -0.00332, -0.00305, -0.00225, -0.00033, -0.01488, -0.01107, 0.00519, 0.0075, -0.00568, -0.00687, 0.00609, 0.00275, -0.00668, -0.0022, -0.00101, -0.00443, 0.0, -0.00286, 0.0088, -0.01732, -0.00455, -0.01737, -0.01773, 0.0177, 0.00722, 0.01448, 0.01066, -0.00415, -0.01805, -0.01107, -0.01893, 0.00952, 0.00693, 0.00279, 0.01168, -0.01277, -0.00665, -0.01333, -0.0074, 0.01362, 0.01631, -0.0009, 0.0157, -0.00641, -0.01472, 0.0106, 0.01295, -0.01481, -0.01326, -0.00973, -0.02479, 0.01759, 0.01562, -0.01495, 0.01412, -0.01714, -0.00957, 0.00956, 0.01273, -0.01454, -0.01396, 0.01581, 0.01662, -0.01757, -0.00915, -0.01071, -0.01552, 0.01652, 0.00875, 0.01669, 0.01392, -0.01245, -0.0131, -0.01196, -0.02116, 0.02162, 0.0165, 0.01785, 0.01763, -0.01474, -0.02265, 0.00908, 0.00469, -0.00323, -0.01791, 0.01529, 0.00392, -0.01466, -0.01257, 0.00715, 0.01009, -0.01823, -0.01169, 0.01338, 0.0126, -0.00813, -0.00855, 0.01256, 0.01409, -0.01453, -0.00744, 0.00108, 0.01047, -0.01857, -0.01342, 0.01535, 0.01513, -0.00526, -0.00771, 0.00891, 0.00869, -0.01583, -0.01557, 0.00759, 0.00753, -0.01177, -0.00908, 0.00265, 0.00767, -0.02006, -0.00496, 0.00124, -0.00804, -0.00897, 0.00714, 0.00346, -0.01156, -0.00031, 0.00194, 0.01735, -0.00267, -0.01469, 0.00203, 0.00316, -0.01688, -0.00559, 0.00069, 0.01059, -0.0147, 0.00055, 0.00308, 0.00467, 0.01267, -0.01108, -0.00751, 0.01047, 0.00482, -0.00107, 0.00327, 0.01544, 0.01175, -0.00136, -0.01328, 0.00248, 0.0071, -0.00978, -0.00717, 0.01103, 0.00626, -0.01413, -0.00523, 0.00619, 0.01007, -0.00321, -0.00154, 0.004, 0.0096, -0.0042, -0.01268, 0.00157, 0.01092, -0.00735, -0.00415, 0.00778, -0.00786, 0.00034, -0.00457, 0.00999, 0.00384, -0.01242, -0.01719, 0.01188, 0.00578, -0.00456, -0.006, 0.00803, 0.00687, -0.01262, -0.00718, -0.00027, 0.0052, -0.00068, 0.00061, 0.01561, 0.01021, -0.01847, -0.00178, 0.01113, -0.00451, -0.00407, -0.02228, 0.01422, 0.00944, -0.01435, -0.01446, -0.00027, -0.00108, -0.00917, -0.0032, -0.00457, -0.00644, 0.00594, 0.00404, -0.01683, -0.01398, 0.00858, 0.00541, -0.00408, -0.00613, 0.00556, -0.00545, -0.01103, -0.00969, 0.00189, -0.00981, -0.01676, -0.00097, -0.00752, -0.00313, -0.00084, -0.01554, -0.00769, -0.00451, -0.00286, -0.00631, -0.00903, 0.00041, -0.00715, -0.00707, -0.01049, -0.00104, -0.01338, -0.01292, -0.01095, -0.01779, -0.00416, 0.00133, -0.01941, -0.01212, -0.0018, -0.00403, -0.02301, -0.01115, -0.0083, 0.00633, -0.0114, -0.010068, -0.004843, -0.004042, 0.005639, -0.009609, -0.010717, 0.007167, -0.003115, -0.010679, -0.011284, 0.001923, -0.008916, -0.023735, -0.007338, 0.003154, -0.000537, -0.008813, -0.014443, -0.010224, -0.006352, -0.012431, -0.010327, 0.000214, -0.003437, -0.013305, -0.015332, -0.01124, -0.008061, -0.00372, -0.011484, -0.002802, -0.002456, -0.009067, -0.007993, -0.006528, -0.002167, -0.015184, -0.017685, -0.008271, -0.013331, -0.022377, -0.019129, 0.003236, 0.002016, -0.020507, -0.015239, 0.003298, -0.016499, -0.02686, -0.012612, -0.00249, -0.006625, -0.024198, -0.017856, -0.004741, 0.000375, -0.013125, -0.009633, 0.004931, -0.000576, -0.019589, -0.024052, 0.005959, 0.002807, -0.01581, -0.010888, -0.003417, 0.000693, -0.016875, -0.01206, 0.004702, -0.004951, -0.01083, -0.021879, -0.001761, -0.005058, -0.016435, -0.009125, 0.007233, -0.010693, -0.012661, -0.004326, -0.005208, -0.002729, -0.014726, -0.007421, -0.00977, -0.004741, -0.012149, -0.017592, 0.003081, 0.010185, -0.010766, -0.004873, 0.000277, -0.007377, -0.011264, -0.011772, -0.009462, -0.003701, -0.028408, -0.024243, -0.01236, -0.016797, -0.005337, -0.002575, -0.005834, -0.014389, -0.005532, -0.030092, -0.018631, 4.8e-05, 0.001181, -9e-05, -0.004471, -0.013974, -0.007207, 0.001708, -0.001005, -0.018032, -0.014653, -0.002265, 0.006025, -0.021801, -0.020039, 0.002288, -0.006025, -0.006557, -0.015371, -0.020781, -0.013312, -0.007885, -0.006416, -0.01957, -0.012534, 0.003095, -0.007182, -0.022397, -0.014438, -0.010878, -0.008616, -0.012402, -0.003378, -0.008334, -0.004101, -0.006084, -0.007978, -0.005385, -0.015815, 0.00268, -0.003642, -0.016103, -0.02202, -0.015917, -0.007255, 0.00381, -0.021401, -0.015014, 0.005097, -0.001528, -0.016175, -0.007929, 0.006989, -0.000191, -0.006743, 0.001845, -0.003422, -0.025307, -0.016074, -0.007187, -0.002573, -0.023896, -0.02229, -0.009628, 0.000366, -0.024384, -0.021333, -0.000914, -0.008886, -0.014409, -0.01718, -0.018383, -0.012333, -0.002924, 0.008842, -0.023403, -0.011646, -0.018544, -0.001567, 0.003159, -0.005275, -0.019179, -0.020148, 0.001737, -0.006069, -0.022241, -0.016665, -0.001172, -0.000854, -0.008266, -0.019946, -0.014165, -0.014404, -0.003547, -0.011782, -0.015556, -0.016655, -0.004624, -0.009648, -0.019042, -0.017617, -0.001953, 0.003054, -0.008955, -0.019291, -0.016533, -0.019892, -0.003692, -0.004734, -0.000765, 0.000214, -0.01478, -0.027519, 0.004272, 0.002021, -0.023417, -0.022705, -0.01471, -0.024078, 0.002158, 0.004538, -0.000502, 0.002924, -0.025947, -0.01478, -0.023994, -0.018305, -0.010097, -0.008627, -0.002592, 0.002397, -0.020639, -0.014083, -0.007329, -0.010341, -0.016464, -0.011606, -0.009204, 0.001416, 0.001704, -0.037387, -0.021943, 0.004217, 0.00618, -0.018908, -0.022304, -0.015937, -0.014575, -0.011835, -0.007036, -0.005698, -0.012819, -0.017597, -0.014848, 0.001582, -0.002875, -0.014365, -0.022714, -0.001464, 0.001494, -0.020373, -0.01227, -0.002778, 0.002309, -0.01123, -0.013437, -0.019311, -0.011268, -0.002988, 0.010927, -0.000866, -0.001541, -0.022685, -0.014868, 0.001335, 0.009516, -0.023872, -0.011464, -0.002958, 0.002036, -0.022299, -0.025166, -0.003999, -0.009116, -0.006201, -0.018793, -0.023017, -0.025404, 0.003657, -0.004165, -0.005347, -0.003695, -0.013247, -0.01413, -0.003159, -0.00185, -0.017511, -0.014575, -0.01853, -0.006654, -0.010141, -0.004212, 0.00226, -0.003015, -0.030507, -0.01288, 0.002675, 0.003245, -0.009009, -0.022285, 0.001596, 0.004047, -0.019837, -0.000805, -0.005409, -0.001201, -0.010895, -0.007924, 0.004054, 0.000871, -0.018999, -0.007324, -0.003058, -0.008808, -0.023892, -0.013457, 0.005213, -0.00333, -0.011723, -0.010387, -0.001455, 0.001694, -0.0058, -0.004571, 0.001517, 0.001488, -0.008678, -0.007084, -0.008003, -0.007391, -0.010654, -0.006225, -0.005561, 0.002949, -0.014485, -0.015092, 0.000468, 0.000971, -0.01892, -0.017573, -0.002851, -0.003437, -0.011852, -0.017697, -0.001536, 0.003417, -0.015098, -0.015683, -0.003325, -0.006586, -0.007993, -0.010029, 0.001132, -0.003911, -0.014711, -0.015964, -0.017903, -0.006284, 0.006328, 0.002618, -0.004833, -0.00666, -0.017458, -0.012758, 0.006669, -0.00708, -0.024466, -0.011611, 0.002416, -0.00415, -0.014433, -0.003476, -0.009612, -0.02479, -0.006274, 0.001113, -0.003256, -0.018857, -0.011064, 0.004033, 0.003471, -0.017611, -0.022749, 0.001402, 0.004584, -0.021494, -0.008398, -0.000148, -0.001489, -0.012781, -0.015756, -0.002766, -0.000488, -0.022817, -0.013745, -0.002374, -0.008339, -0.0181, -0.017204, -0.009306, -0.008966, -0.001929, 0.002416, -0.008664, 0.007592, -0.0056, -0.006773, 0.002993, -0.005586, -0.018291, -0.012802, -0.015053, -0.012768, -0.00351, -0.002967, 0.000122, 0.00394, -0.010804, -0.007451, -0.01144, -0.00236, -0.010287, -0.01145, -0.009243, -0.01122, -0.005678, -0.004453, -0.001596, 0.003793, -0.011513, -0.00977, 0.003666, 0.000991, -0.019946, -0.013564, 0.002597, 0.00311, -0.013442, -0.021655, -0.003725, 0.001528, 0.005854, -0.021474, -0.010938, -0.005419, -0.002565, -0.02351, -0.011948, 0.002573, -0.000341, -0.01332, -0.004131, -0.013959, -0.008139, 0.00331, -0.000205, -0.002867, -0.016245, -0.012817, -0.003369, -0.010587, -0.016274, -0.00726, -0.000932, -0.008872, -0.014238, -0.01206, -0.005776, -0.000302, -0.01184, -0.015543, -0.001752, -0.001215, -0.012119, -0.008315, -0.002573, 0.001098, -0.018857, -0.011386, -0.000655, -0.009851, -0.017978, -0.01269, -0.005478, -0.001982, -0.008374, -0.011596, -0.00482, -0.004562, -0.010052, -0.00223, 0.004692, -0.000722, -0.008984, -0.007353, -0.002822, 0.003823, -0.010971, -0.00871, 0.000219, -0.004033, -0.024443, 0.001796, -0.003312, 0.00303, -0.020795, -0.011137, -0.004379, -0.001035, -0.016692, -0.011303, 0.009199, -0.001079, -0.021342, 0.000698, -0.004711, -0.004892, -0.008349, 0.001313, -0.004538, -0.016697, -0.008295, 0.007451, 0.002939, -0.015327, -0.015263, 0.002309, -0.0019, -0.005839, -0.017207, 0.006552, -0.001254, -0.017276, -0.005029, 0.015058, 0.003574, -0.01184, -0.009355, -0.00011, 0.000254, -0.021621, -0.025078, -0.001987, -0.00059, -0.01596, -0.012426, -0.001416, -0.009311, -0.014926, -0.027416, -0.007495, -0.014174, -0.023159, -0.009443, 0.000424, 0.00456, -0.017539, -0.021542, -0.000312, -0.001665, -0.008238, -0.010861, 0.001279, -0.006572, -0.003745, -0.006762, -0.006396, -0.019837, -0.018837, -0.017956, -0.013365, -0.002954, -0.010839, -0.012729, -0.000439, -0.01034, -0.012368, -0.01371, 0.009057, -0.020129, -0.017241, 0.004204, -0.009106, -0.012422, -0.012221, 0.002792, 0.004447, -0.010579, -0.005664, -0.006285, -0.009492, -0.016161, -0.018188, -0.01337, -0.004765, -0.006616, -0.00133, 0.006718, -0.001025, -0.011728, -0.012216, 0.001801, -0.007797, -0.013696, -0.011742, -0.012377, -0.007978, 0.004433, -0.009169, 0.000312, -0.003686, -0.014848, -0.007875, 0.008378, 0.001616, -0.023686, -0.013413, 0.000385, -0.00061, -0.017089, -0.0085, 0.00477, 0.003145, -0.019125, -0.010659, -0.003295, -0.000888, -0.016972, -0.007353, 0.002456, -0.01, -0.014125, -0.015797, -0.013979, -0.012077, -0.005883, -0.004492, -0.002919, 0.005771, -0.018061, -0.019469, 0.004528, -0.003374, -0.024555, -0.010136, 0.002241, -0.006606, -0.018217, -0.016422, 0.002565, -0.002068, -0.021001, -0.012133, 0.003369, -0.007499, -0.005708, -0.01124, 0.000532, 0.000986, -0.008491, -0.009404, -0.004165, -0.014145, -0.001069, -0.010791, -0.000981, -0.0037, -0.012812, -0.01581, -0.004853, 0.001123, -0.008227, -0.014062, -0.000312, -0.000136, -0.004116, -0.006655, -0.008611, -0.007124, -0.014467, -0.015302, 0.000847, 0.002294, -0.012846, -0.009228, -0.009365, -0.005717, -0.007426, 0.002622, -0.006669, -0.004552, -0.015352, -0.010922, -0.008876, 0.001742, -0.018256, -0.011103, -0.004291, -0.000689, -0.008511, -0.015043, -0.005551, 0.006194, -0.009257, -0.011381, -0.002977, -0.003063, -0.017989, -0.006186, -0.006635, -0.000426, -0.017167, -0.006414, -0.007429, 0.00226, 1.4e-05, -0.013442, -0.016596, 0.002924, -0.009194, -0.01269, -0.009564, -0.000273, -0.002363, -0.011772, -0.006503, -0.00561, -0.004025, -0.017983, -0.009648, 0.005063, 0.001464, -0.009847, -0.01404, -0.003598, -0.002776, -0.015017, -0.005043, 0.005063, -0.017683, -0.00977, 0.001372, -0.002111, -0.021376, -0.012329, -0.006005, -0.001586, -0.001079, -0.00288, 0.003283, -0.010234, -0.01976, -0.013535, -0.015429, -0.010195, 0.004755, -0.011035, -0.001132, -0.006079, -0.005991, -0.004687, -0.002197, 0.000415, -0.020019, -0.010273, -0.002094, 0.003833, -0.003037, -0.005537, 0.002446, -0.003515, -0.007739, -0.020143, -0.00943, -0.006914, -0.000512, 0.001541, -0.019956, -0.005739, -0.006538, 0.001513, -0.012812, -0.018842, -0.010071, -0.010766, 0.000622, 0.00228, -0.001669, -0.000224, -0.011577, -0.023657, 0.000395, -0.006298, -0.016323, -0.013178, 0.002939, -0.000288, -0.014289, -0.006902, 0.005782, 0.007143, -0.008388, 0.002829, -0.003632, -0.015644, -0.011191, -0.004414, -0.003027, -0.006093, -0.009765, -0.001148, 0.009326, -0.011083, -0.005429, -0.01517, -0.00168, -0.003359, -0.001425, 0.001367, -0.002101, -0.011635, -0.003896, 0.003949, 0.000302, -0.00852, -0.001186, -0.002075, -0.016289, -0.012753, -0.003437, -0.003599, -0.015688, -0.018393, -0.006304, -0.00541, -0.016684, -0.008984, 0.006275, 0.005209, -0.013466, -0.024687, -0.002099, -0.009848, -0.013837, -0.022341, -0.001391, -0.004936, -0.022866, -0.012421, 0.008466, -0.000732, -0.011826, -0.006621, -0.001871, -0.001518, -0.01208, 0.01629, 0.00439, -0.00563, -0.00374, -0.00868, -0.00249, -0.00036, -0.01437, -0.00595, -0.00152, -0.00438, -0.0061, -0.00155, -0.00046, -0.0032, -0.00619, 0.00828, 0.00406, 0.00445, -0.00242, -0.00057, 0.00911, 0.00484, 0.0027, 0.00227, 0.00525, -0.01162, 0.0016, 0.00445, -0.00802, -0.0003, -0.00238, -0.00421, -0.00148, -0.00118, -0.00233, -0.01256, 0.00761, -0.00925, 0.00124, -0.0005, -0.00354, -0.00506, -0.00277, -0.00698, -0.00202, -0.00216, 0.0008, -0.00145, 0.00418, -0.00217, -0.00261, -0.01398, 0.00403, -0.00326, -0.0086, -0.00488, -0.00439, -0.00569, -0.00435, 0.00355, 0.0016, -0.0019, -0.01115, -0.00725, 0.00658, 0.00019, 0.00837, -0.009, -0.00431, -0.01104, 0.00252, -0.00045, 0.01141, -0.00404, -0.01001, -0.00249, -0.00256, 0.00445, -0.0005, 0.00467, -0.00498, -0.00296, 0.00241, -0.00338, 0.00327, -0.00375, 0.00767, -0.0092, -0.00665, -0.01129, 0.00666, 0.00271, -0.00211, -0.00201, -0.00416, -0.0037, -0.00033, 0.0025, -0.00691, -0.00343, -0.01024, 0.00528, 0.00584, -0.0034, 0.00198, -0.00114, -0.00769, 0.00436, -0.00082, -0.00459, 0.0019, -0.00674, -0.00073, 0.00198, 0.00113, -0.00857, 0.00348, -0.00567, -0.0057, -0.00216, 0.00738, 0.00548, 0.00574, 0.00275, -0.0068, 0.00764, 0.00745, 0.01188, -0.00498, 0.00324, 0.0019, -0.01285, 0.00054, -0.00018, 0.00439, 0.00899, 0.00026, 0.00508, -0.00544, 0.00895, 0.00256, -0.00137, -0.00452, -0.00667, -0.00497, -0.0048, -0.00474, -0.00146, 0.00642, -0.00311, -0.00083, -0.00122, 0.01127, -0.00249, -0.00713, -0.01089, 0.00358, -0.01884, -0.00503, -0.00348, -0.00185, 0.00554, 0.00464, -0.0115, -0.00808, 0.00082, -0.00056, -0.00957, -0.00656, -0.00449, 0.00563, -0.0067, 0.00743, -0.01106, -0.00354, -0.01586, -0.00038, -0.00199, 0.00084, 0.00194, 0.00033, -0.00694, -0.00057, 0.00024, 0.00876, -0.00127, 0.00477, -0.00748, 0.00138, -0.01269, -0.00904, 0.00631, -0.00123, -0.01192, 0.00465, 0.00052, -0.00661, -0.00556, -0.00354, 0.0035, 0.00253, 0.00086, 0.00858, 0.00574, -0.00482, -0.00548, -0.00406, 0.00606, -0.00118, 0.00492, 0.00975, -0.00562, 0.0022, -0.00498, -0.00774, -0.00027, -0.00178, -0.00598, -0.00281, -0.00067, -0.00539, 0.00128, 0.0, -0.01414, -0.00643, 0.00778, -0.00408, 0.00922, 0.00419, 0.00121, 0.00141, 0.00221, 0.00749, -0.00691, 0.00205, 0.00015, 0.00522, 0.00241, 0.00106, -0.01122, -0.01, 0.00042, 0.00074, -0.00863, 0.00489, -0.00523, 0.0, -0.00514, -0.00202, 0.0087, 0.00755, 0.00623, 0.00181, -0.00987, -0.0013, 0.00783, 0.0061, -0.00341, 0.0051, -0.00096, -0.00235, -0.00034, 0.00733, -0.01231, -0.00252, -0.00503, 0.00441, 0.00476, -0.00652, -0.00355, 0.00357, -0.0042, 0.00237, 0.00381, -0.01215, -0.00281, -0.01696, -0.00456, 0.00306, -0.00297, 0.00462, 0.00133, -0.00733, -0.00154, -0.00354, -0.00018, -0.00084, 0.0, 0.00195, -0.00608, 0.00538, -0.00438, -0.01031, -0.00408, -0.0017, 0.00656, -0.00086, -0.00096, 0.0054, 0.00117, -0.00645, 0.00345, -0.00372, 0.00317, 0.00337, 0.00366, -0.00827, -0.0052, 0.00142, -0.00512, -0.00344, -0.00771, 0.00166, -0.00467, -0.00296, -0.00175, -0.00181, 0.00787, -0.00041, -0.00524, 0.00171, 0.00415, 0.00442, 0.0091, 0.00169, 0.00191, 0.00164, -0.00515, -0.00532, -0.0091, -0.009, 0.00064, 0.00095, -0.00968, -0.00989, 0.00681, -0.00717, -0.01132, 0.0017, -0.00549, 0.01664, 0.00072, 0.0023, -0.00154, 0.00233, 0.00708, -0.00399, -0.00688, -0.00166, -0.00228, -0.00364, -0.00018, 0.00736, 0.00051, 0.00054, -0.01047, -0.00072, 0.00662, -0.00117, -0.00655, -0.00811, -0.00242, -0.00011, -0.00873, -0.00275, 0.00731, -0.00434, -0.00321, -0.0043, 0.00209, -0.01229, -0.00402, -0.00642, -0.00732, -0.01025, -0.00698, 0.001, 0.00473, -0.01367, 0.00272, 0.0013, -0.00344, 0.00368, -0.00279, 0.00277, 0.00368, 0.00839, 0.00503, 0.0011, -0.02952, -0.00227, 0.00298, 0.01251, -0.00931, 0.00173, -0.00471, -0.00222, 0.0073, -0.02232, -0.00338, -0.00082, -0.00846, -0.00797, -0.00283, -0.01163, 0.00124, 0.00081, -0.00348, -0.00034, 0.00433, -0.00255, -0.00177, -0.00658, -0.0021, 0.0044, -0.00614, -0.01655, 0.00028, 0.00175, 0.0116, -0.00843, -0.00519, -0.00198, 0.00236, -0.00547, -0.01008, -0.00717, -0.00412, -0.00161, 0.00615, -0.00569, -0.00419, 0.00203, -0.0003, -0.00841, -0.00138, -0.00411, -0.01103, -0.00131, -0.00106, 0.00828, 0.00847, 0.00261, 0.00162, 0.0026, 0.00303, -0.00662, 0.00505, 0.00536, 0.00743, -0.00288, -0.00264, -0.00109, -0.00333, -0.00041, -0.00316, -0.00462, 0.0041, -0.00194, 0.00041, 0.00056, 0.00508, -0.00048, -0.00266, 0.00022, -0.00284, 0.01149, 0.0032, 0.00562, 0.00442, 0.00153, -0.00018, -0.00247, -0.00339, -0.00036, -0.00624, -0.00296, -0.00334, -0.00209, -0.00399, -0.00916, -0.00709, 0.00012, 0.00596, -0.00597, 0.00505, -0.00375, -0.00355, -0.00193, -0.00505, 0.00433, 0.00189, 0.00025, -0.0019, -0.00129, -0.00462, -0.01412, 0.00079, 0.00439, -0.01506, -0.00196, 0.00758, -0.00287, 0.00713, -0.00592, -0.00106, -0.00806, -0.00529, 0.00229, -0.00107, -0.00825, -0.00164, 0.0, -0.003, -0.00772, -0.00809, -0.00405, 0.00342, 0.00733, -0.00497, 0.0, 0.00504, -0.00147, -0.00683, 0.00275, 0.00376, -0.00366, -0.00357, 0.00781, 0.01318, -0.00198, -0.00513, 0.00348, -0.00543, -0.00776, 0.00422, -0.00184, -0.00522, -0.00159, 0.00069, -0.00782, -0.00285, -0.00167, -0.00729, -0.00354, -0.00323, 0.00087, 0.00703, 0.00217, -0.00023, -0.00324, -0.0014, -0.00334, -0.00283, 0.00597, 0.00688, 0.00247, -0.00524, -0.01168, -0.00797, 0.00582, -0.00916, -0.0102, -0.00901, 0.00737, 0.00628, 0.00177, -0.006, 0.00549, -0.00614, -0.01483, -0.00034, 0.01229, -0.0051, -0.00613, 0.00544, -0.00618, -0.00297, -0.00677, -0.00362, 0.00505, 0.0, 0.00127, -0.01243, -0.00886, -0.01075, -0.01057, -0.00466, -0.00348, 0.00145, 0.00017, -0.00766, -0.00747, -0.00374, -0.00515, -0.01826, -0.01083, 0.00487, -0.00121, -0.01431, -0.00646, -0.00879, -0.00482, -0.01233, -0.01101, -0.01042, -0.00631, -0.01294, -0.0118, -0.01105, -0.00944, -0.00749, -0.00969, -0.00294, -0.00376, -0.01165, -0.0005, -0.01192, -0.00175, 0.00432, -0.00347, -0.01584, -0.00557, 0.00304, -0.00307, 0.00063, -0.00229, -0.00291, -0.006, -0.00782, 0.00365, 0.0049, -0.01185, -0.02077, -0.00415, -0.00097, -0.019, -0.01401, -0.00716, 0.00234, -0.01634, -0.01902, -0.00298, -0.00501, -0.0065, -0.02269, -0.00674, -0.00289, -0.01247, -0.01099, -0.00439, 0.0084, -0.01665, -0.01295, 0.0007, -0.00109, -0.01376, -0.01102, -0.00937, -0.01284, -0.00348, 0.00048, -0.01841, -0.02052, -0.01211, -0.01396, -0.0105, -0.00779, -0.007, -0.00686, -0.00914, -0.01175, -0.00305, -0.01004, -0.01483, -0.00501, 0.00753, -0.00645, -0.00149, 0.00746, -0.00122, -0.01047, -0.00546, 0.00353, -0.00797, -0.00433, -0.01687, -0.00664, -0.00736, -0.01477, -0.00545, -0.00351, -0.0052, -0.0146, -0.00654, -0.01326, -0.01086, -0.01035, -0.01572, -0.00703, -0.019, -0.00438, -0.01235, -0.01217, -0.00529, -0.01931, -0.01358, -0.00311, -0.00954, -0.0182, 0.0, -0.00241, -0.0066, -0.01176, -0.0141, 0.00429, -0.01271, -0.01196, -0.00686, -0.0029, -0.01433, -0.01169, -0.00337, 0.0004, -0.01598, -0.00391, -0.00696, 0.00163, -0.01377, -0.01252, -0.00917, -0.01135, -0.00224, -0.01518, -0.01783, -0.00863, -0.01388, -0.00758, -0.00608, -0.00988, -0.01289, -0.00249, -0.00261, -0.00395, -0.00738, -0.0077, -0.00575, -0.00944, 0.00245, -0.00371, -0.00855, -0.01358, 0.00636, -0.00126, -0.008, -0.02402, -0.01654, -0.0049, -0.00984, -0.00561, -0.0032, -0.00769, -0.01344, -0.01572, -0.00202, -0.01666, -0.00812, -0.00627, -0.00432, -0.00947, -0.0073, 0.00212, -0.00766, -0.0147, -0.00743, 0.00158, -0.00055, -0.00963, -0.00568, -0.01055, -0.01272, -0.01839, -0.00076, -0.01046, -0.00372, -0.01129, -0.01126, -0.01254, -0.01, -0.00392, -0.00487, -0.0098, -0.00057, -0.01797, -0.01439, -0.00423, -0.01425, -0.01374, -0.00223, -0.013793, -0.000385, -0.004653, -0.002612, 0.0079, 0.00224, -0.001425, -0.007597, -0.004185, -0.004736, -0.004379, -0.014423, -0.001523, -0.010976, 0.004218, -0.00684, -0.011086, -0.004126, -0.011733, -0.013398, -0.01019, -0.013351, -0.005512, -0.015385, 0.002364, -0.006826, 0.003305, -0.006438, -0.008233, -0.007666, -0.006491, 0.001181, -0.00894, -0.005878, 0.010068, -3.9e-05, -0.007113, -0.016314, -0.007495, -0.012255, -0.004863, 0.001025, 0.002929, -0.002578, -0.004726, -0.001733, -0.006773, -0.010546, 0.000976, -0.00746, -0.005541, -0.001206, -0.009918, 0.004848, 0.002333, -0.007011, -0.001181, 0.003476, -0.011982, -0.012163, 0.00144, -0.006907, -0.010063, -0.006796, -0.004169, -0.01269, -0.014763, -0.009794, -0.003095, -0.010737, -0.003916, -0.013815, -0.009789, -0.005126, -0.012011, -0.00436, -0.004715, 0.002637, -0.005815, -0.004922, -0.001196, -0.007171, -0.017491, 0.001284, -0.009018, 0.001732, -0.014394, -0.002089, -0.005126, -0.0102, -0.015434, -0.009114, -0.00528, -0.010078, 0.005122, -0.001181, -0.014658, -0.002766, 0.000961, 0.002537, -0.003002, -0.01178, -0.00458, -0.005854, 0.000366, -0.01392, -0.014248, -0.006074, -0.013881, -0.017963, -0.004824, -0.012373, -0.001567, -0.009334, -0.005073, -0.003427, 0.003632, -0.005471, 0.001142, 0.000419, -0.004477, -0.002626, -0.008857, -0.012671, -0.013481, -0.01038, -0.005864, -0.000406, -0.008085, -0.017314, -0.005786, -0.005996, -0.008579, -0.001743, -0.007211, -0.008066, -0.00725, -0.00434, -0.018608, -0.005428, -0.0077, -0.009497, -0.007651, -0.008994, -0.008632, 0.002895, -0.008168, -0.012432, -0.006743, -0.000607, 0.001958, -0.000161, -0.001206, -0.006557, -0.01435, -0.00516, -0.00944, -0.018779, -0.008918, -0.002583, -0.011347, -0.003876, -0.011694, -0.011311, -0.0181, -0.00123, -0.010917, -0.006562, -0.000263, 0.006941, -0.001533, -0.00373, -0.001546, -0.006689, 0.001292, -0.000888, -0.012382, -0.000283, -0.0033, 0.000141, -0.010751, -0.010141, -0.006098, -0.006649, 0.008593, -0.00059, -0.004951, -0.007209, -0.015585, -0.017113, 0.004257, 0.000646, -0.007397, -0.015219, -0.002312, -0.008984, -0.002109, 0.00206, -0.010263, -0.008579, -0.004054, -0.005045, 0.000373, -0.00664, -0.009531, -0.011196, -0.008195, -0.007094, -0.016547, -0.015979, -0.003964, -0.006079, -0.010976, -0.006401, -0.001362, -0.016367, -0.014571, -0.013623, 0.0006, -0.012192, -0.003853, -0.000859, -0.001743, -0.003652, -0.006218, -0.00517, -0.00019, -0.002338, 0.002919, -0.003657, -0.009609, 0.001679, 0.002646, -0.013158, -0.000114, -0.005107, -0.003232, -0.003798, -0.00665, -0.019449, -0.005126, -0.006515, -0.004628, -0.010038, 0.001665, 0.008212, -0.004155, -0.017099, -0.007939, -0.004413, -0.008198, 0.003379, 0.001049, 0.004571, -0.011225, -0.018203, -0.005034, -0.00581, -0.007285, -0.006098, -0.007519, -0.005371, 0.0046, 0.003476, 0.000378, -0.001855, -0.01119, -0.012016, -0.000689, 0.000421, -0.004973, -0.003486, -0.001391, -0.008789, -0.012368, -0.000351, -0.007519, -0.003017, -0.005639, 0.002805, -0.008779, -0.013735, -0.009531, -0.006036, -0.012041, 0.005351, -0.000156, -0.010581, -0.002807, -0.003738, -0.00373, -0.002264, 0.000339, -0.009614, -0.007443, -0.009243, -0.001053, -0.013359, -0.003354, 0.001943, -0.004672, -0.013303, -0.009638, -0.013833, 0.004316, -0.001088, 0.00104, -0.015566, -0.006269, -0.001835, -0.012568, -0.004506, -0.000903, -0.01059, -0.00166, -0.007861, 0.000185, -0.000639, -0.004858, -0.008505, -0.010948, -0.011843, 0.000922, -0.001608, -0.003608, -0.007949, 0.002168, -0.002723, -0.016256, 0.002543, -0.003973, -0.008544, -0.003857, -0.011274, -0.021186, 0.001823, -0.001274, 0.001318, -0.00352, 0.00414, -0.006235, 0.003237, -0.013279, -0.012329, -0.008798, 0.001875, 0.003432, -0.00726, -0.012321, -0.002573, -0.001321, -0.019287, -0.005712, -0.004839, -0.002109, -0.012377, 0.001323, -0.00144, -0.003251, -0.008027, -0.00206, -0.009727, -0.00957, -0.000167, -0.010229, -0.006679, -0.002592, 0.000166, -0.012006, -0.003046, -0.001328, -0.000693, -0.002273, -0.000986, -0.004487, 0.009689, -0.010444, -0.002392, -0.001748, -0.00371, -0.008906, 0.002548, -0.001756, -0.005893, -0.008549, -0.001622, -0.005864, -0.005932, -0.002519, -0.001498, -0.00651, -0.010957, -0.001431, 0.002875, 0.000634, -0.000439, 0.002814, -0.002397, -0.005234, -0.009784, -0.013984, 0.004136, -0.012617, -0.017695, -0.006977, -0.005849, -0.008593, -0.006928, -0.004052, -0.005727, -0.008114, -0.010014, -0.008691, -0.002905, -0.005538, 0.002944, -0.003011, -0.001826, -0.013633, -0.00133, 0.001035, -0.010004, 0.006215, 0.00019, -0.003608, 0.004101, -0.00561, -0.009389, -0.004165, 0.001821, -0.007617, -0.009926, -0.007241, -0.003198, 0.004555, -0.002377, -0.011263, 0.000317, -0.010844, -0.000483, 0.001396, -0.005677, 0.004016, -0.002699, -0.005165, 0.000927, 0.00359, -0.006767, -0.003496, -0.003652, -0.01225, -0.009477, -0.0058, -0.006206, -0.002905, -0.002065, -0.008382, -0.000791, 0.003046, -0.007446, 0.000639, -0.003714, 0.001105, -0.009885, -0.006015, -0.007558, -0.010311, -0.009497, -0.004863, -0.008022, -0.000634, 0.007769, 0.001806, -0.010043, -0.009077, -0.008516, -0.008969, -0.004586, -0.013515, 0.007405, -0.007075, -0.01623, -0.008803, -0.003344, 0.001225, -0.000938, -0.006472, 0.011059, -0.009223, -0.006845, -1.4e-05, -0.002314, -0.005371, -0.009057, -0.007358, -0.003891, -0.007458, -0.003403, -0.003494, 0.000825, -0.010395, -0.002556, -0.004995, -0.00707, -0.00124, 5.3e-05, -0.016831, -0.003022, -0.007869, 0.008047, -0.012148, -0.005232, 0.000181, -0.003803, -0.002829, -0.01392, -0.003432, -0.010747, -0.010282, -0.005874, -0.001948, -0.005747, 0.000488, -0.002431, -0.011308, -0.002101, -0.00642, -0.009824, -0.012216, 0.001054, -0.004609, -0.001665, -0.009219, -0.003754, -0.005423, -0.008149, -0.003881, -0.004458, -0.002089, -0.002221, -0.008564, -0.009746, -0.018238, -0.008707, -0.001264, -0.005712, -0.001313, -0.001157, -0.007993, -0.009375, -0.000175, -0.015395, -0.00124, -0.01168, -0.007597, -0.009218, -0.007646, -0.009487, -0.004741, -0.006855, -0.008671, -0.002524, -0.006347, -0.002578, -0.009946, -0.009584, -0.001293, -0.005751, -0.003236, -0.001929, -0.012578, 0.002412, -0.003632, 5.3e-05, -0.004301, -0.009951, -0.021474, -0.002294, -0.004946, -0.008359, -0.0082, -0.005356, -0.006474, 0.000654, -0.005649, -0.001372, 0.008291, -0.015473, -0.004013, -0.004609, -0.003261, -0.006494, -0.005653, -0.010311, -0.002827, -0.012041, 0.003317, -0.002421, -0.002958, -0.005004, -0.001918, 0.008925, -0.014921, -0.007281, -0.00812, -0.003266, -0.000629, -0.00561, -0.012543, -0.003305, -0.006611, 0.000253, 0.001328, 0.001811, -0.009665, -0.00225, -0.003535, -0.00686, -0.013037, -0.010737, -0.003535, -0.007929, -0.004814, -0.002656, -0.003471, -0.006425, -0.003515, -0.00017, -0.006235, -0.004547, -0.003291, -0.016161, -0.001464, -0.006049, -0.003422, 0.00289, -0.004799, 0.00143, -0.000214, 0.009975, 0.000766, -0.006904, -0.00456, -0.001289, 0.003857, -0.004902, 0.010336, -0.012709, -0.004833, -0.0128, -0.00623, -0.01704, -0.01349, -0.01247, -0.00883, -0.00771, -0.00404, -0.01261, -0.00529, -0.01706, 0.00519, -0.00298, 0.02047, -0.01036, 0.01062, -0.00476, -0.00049, -0.0142, 0.00431, -0.01226, -0.00427, 0.00933, -0.00331, -0.0145, -0.00933, 0.01094, 0.01213, -0.01819, 0.00512, -0.00716, -0.00588, -0.01296, 0.00798, 0.00368, 0.01265, 0.00513, -0.00724, -0.01116, -0.00504, -0.00695, 0.01174, -0.00573, 0.01672, 0.00754, 0.00948, 0.00725, -0.0072, -0.01835, 0.00584, -0.00811, 0.00772, 0.00133, -0.01023, 0.00561, 0.00893, -0.01231, 0.00444, -0.01554, -0.00111, 0.00478, -0.00783, -0.01557, 0.00502, 0.00161, 0.00255, -0.01042, 0.00812, 0.00552, -0.00934, 0.0165, 0.00071, -0.01304, 0.00768, 0.00991, 0.01338, 0.0117, -0.0014, 0.01098, 0.01338, -0.01056, -0.00927, 0.01071, 0.01315, -0.0015, 0.00108, -0.01229, 0.00695, 0.00155, 0.00396, -0.00082, 0.00117, -0.015, -0.01232, 0.00396, -0.00912, -0.01107, 0.00876, -0.0057, 0.0, -0.01414, -0.00461, -0.00164, 0.00302, -0.02617, -0.01916, -0.01395, 0.0072, -0.01186, 0.00432, -0.00687, -0.00046, -0.00021, 0.01353, -0.01379, 0.00571, -0.00318, -0.00079, -0.00057, -0.00952, -0.01015, -0.00376, -0.00767, 0.00812, -0.01252, -0.01104, -0.00774, 0.00293, -0.00362, 0.00722, 0.00975, 0.01004, -0.00792, -0.00401, -0.00455, 0.01658, -0.01508, -0.01256, -0.00733, 0.01782, -0.0095, 0.00489, -0.01049, -0.00714, -0.01217, -0.01153, 0.01158, -0.00144, -0.0009, 0.0026, -0.01518, -0.01193, 0.00575, 0.00808, -0.00606, 0.0, -0.00042, 0.00476, 0.00726, -0.00438, -0.01125, -0.01127, -0.02003, -0.01584, -0.00914, -0.00377, -0.0142, 0.00266, -0.0144, 0.00602, 0.00204, 0.0068, 0.00103, -0.01132, -0.00438, 0.01863, 0.00033, -0.01041, -0.01244, 0.00781, -0.01785, -0.01449, 0.00563, 0.00838, -0.00349, -0.01639, 0.01594, 0.00305, 0.01875, -0.0097, -0.00941, -0.00385, -0.00401, 0.00678, -0.00492, -0.01185, -0.0088, -0.00818, 0.00856, 0.00662, -0.00181, 0.0068, 0.0113, 0.00115, -0.01072, -0.01587, 0.00551, -0.01029, -0.00784, -0.00781, -0.01017, 0.0006, 0.00604, 0.00884, 0.01268, -0.01208, 0.00527, 0.00588, 0.00804, -0.01359, 0.00562, 0.00589, -0.00586, 0.00382, -0.0069, -0.00698, 0.01182, -0.01279, 0.00595, 0.0086, -0.0027, -0.01367, -0.00947, -0.00263, 0.00249, -0.01309, 0.00277, -0.01518, -0.00251, -0.01234, 0.01074, -0.01082, 0.00488, -0.01546, -0.01041, -0.00975, -0.00444, -0.00933, -0.00942, 0.00716, 0.01354, 0.00854, 0.01346, -0.01293, 0.00355, -0.01931, -0.00557, -0.00648, -0.00457, 0.00097, 0.00456, 0.01033, -0.0033, -0.00355, 0.01416, -0.01178, -0.01249, -0.01433, 0.00736, -0.01563, 0.00472, -0.00189, 0.01659, -0.00308, 0.01798, -0.0123, -0.00457, 0.00401, 0.00688, 0.00697, 0.01055, -0.00192, -0.01079, -0.01273, -0.00621, -0.01114, 0.01227, 0.00349, -0.00955, 0.0042, -0.00313, -0.01222, -0.01082, -0.01539, -0.01456, -0.01518, 0.00604, -0.00169, 0.00228, -0.01093, 0.01115, -0.00835, 0.01268, 0.00557, 0.00982, -0.00279, 0.01643, -0.00817, -0.00229, -0.00606, 0.00982, -0.00139, 0.00472, -0.00966, 0.00315, -0.00635, -0.00677, 0.00113, -0.00388, -0.01861, -0.01122, 0.00983, 0.00692, -0.00489, -0.0133, 0.00817, -0.01433, 0.00618, -0.00902, -0.00621, -0.00961, -0.00677, 0.00615, 0.00922, -0.01663, 0.00297, -0.01081, 0.00569, -0.0184, -0.01136, -0.0175, 0.0013, -0.00743, -0.00225, -0.00476, -0.00487, 0.00659, 0.01373, -0.00826, 0.00774, -0.0018, -0.00099, -0.01198, -0.0046, 0.00135, 0.00476, -0.01167, -0.01305, 0.00643, -0.01133, -0.00133, -0.00504, -0.00788, -0.00927, 0.0054, -0.01251, -0.01032, -0.01023, -0.01076, -0.0121, 0.00892, -0.00162, -0.00636, -0.01148, -0.00744, -0.01505, -0.01577, 0.0027, 0.00216, -0.00661, -0.00705, -0.00738, -0.0006, -0.00403, -0.00648, -0.00503, 0.00701, -0.01266, -0.01172, 0.00442, -0.00723, -0.00914, -0.00072, -0.01448, -0.01414, -0.01238, 0.01096, 0.00349, 0.00377, -0.00193, 0.01497, -0.00782, -0.01155, 0.00181, -0.01159, -0.01017, 0.00387, -0.01218, -0.01328, 0.0064, -0.01331, -0.00834, -0.01159, -0.01053, -0.01453, 0.00593, -0.01542, -0.0057, -0.0086, 0.00233, -0.01363, -0.01779, -0.00094, -0.00075, 0.00578, -0.01994, -0.01385, -0.01964, -0.00207, 0.01314, 0.01568, -0.01423, -0.00746, -0.01732, -0.00559, 0.00359, -0.0001, 0.00052, -0.01208, -0.01365, -0.01464, -0.01655, -0.02025, -0.01238, -0.00948, 0.00521, -0.00757, -0.0047, -0.00652, -0.01231, 0.00925, 0.0062, 0.01217, -0.01429, -0.01042, -0.00145, -0.00879, -0.01361, 0.00651, -0.00613, -0.00892, -0.00451, -0.01997, -0.00742, 0.00773, -0.01103, 0.00328, -0.00295, 0.00211, 0.00084, 0.00822, -0.02271, 0.00168, -0.01284, -0.02105, -0.0127, -0.01466, -0.00913, -0.01466, -0.00968, 0.00866, 0.00894, -0.00106, -0.00589, -0.00978, -0.0082, -0.01072, 0.00416, -0.00564, -0.01294, -0.01143, -0.00992, -0.00991, -0.01152, 0.00233, -0.00428, 0.00434, -0.00756, 0.00846, 0.00071, 0.00831, 0.0121, -0.00039, -0.01177, -0.01331, -0.01348, 0.00077, 0.01022, 0.01055, 0.00053, 0.00127, -0.01282, -0.01394, 0.00352, 0.00562, -0.0003, 0.00972, -0.00115, -0.01284, -0.00805, -0.01432, -0.00932, -0.00153, -0.03988, 0.00274, -0.01153, -0.016, 0.00113, -0.00851, 0.01053, 0.00764, -0.01857, -0.00617, 0.0044, 0.00664, 0.00205, 0.01155, 0.00136, -0.01342, -0.00851, 0.00981, 0.01046, -0.01151, -0.0111, 0.0045, -0.00481, -0.00366, -0.0129, 0.00075, -0.0114, -0.00944, -0.0103, 0.00135, 0.00451, 0.00403, -0.01112, -0.00925, 0.00475, 0.00582, -0.00562, -0.0102, -0.00742, 0.0009, -0.01134, -0.01551, -0.0089, 0.00545, 0.01304, -0.00014, -0.013, -0.01018, -0.01126, -0.00707, -0.00915, -0.00948, -0.00023, -0.01616, -0.00985, 0.00852, 0.00367, 0.0028, 0.00933, -0.0108, 0.00436, -0.01551, -0.00096, -0.00583, 0.00188, 0.0096, 0.0069, -0.00168, -0.00899, 0.00235, 0.0053, 0.00688, 0.00264, -0.00853, -0.01204, 0.00616, -0.00717, -0.00297, 0.00231, -0.00592, 0.00092, -0.01574, -0.00786, 0.00368, 0.00449, 0.00936, 0.00133, -0.01327, -0.00776, -0.01451, -0.01802, -0.00735, 0.00544, -0.00624, -0.01543, -0.00873, 0.00549, 0.00541, -0.00109, 0.00219, 0.01081, 0.00082, -0.01423, 0.01225, -0.01077, -0.01086, 0.00734, -0.01147, -0.00931, 0.01072, -0.00044, 0.00813, -0.01622, 0.00481, -0.01533, -0.01575, -0.00994, -0.00854, 0.00713, 0.00367]]
    table_category_data = [[['CWNJ_C02-2FAP-23_2_FCT', '#0000FF', 'Accelerometer AVG-FS8g_ODR800HZ_Zup accel_nmDC_average_y', 0.118, -0.118, -0.00105, -0.02004, 0.00256, -0.00017, -0.01117, -0.0076, 0.01392, 0.00581, -0.01061, 0.00613, 0.0049, 0.01169, -0.01533, -0.00912, 0.01449, 0.00366, -0.01347, -0.01472, 0.01017, 0.00695, -0.0176, -0.0094, 0.007, 0.00895, -0.01107, -0.00646, 0.0066, 0.00174, -0.00493, -0.00753, 0.01071, 0.00707, -0.01432, -0.00978, 0.01122, -0.00098, -0.02688, -0.00502, 0.01419, 0.00945, -0.01053, -0.01183, 0.0041, 0.00546, -0.01279, -0.00168, 0.01797, 0.00999, -0.0194, -0.0171, 0.00237, -0.00164, -0.0121, -0.01011, 0.01913, 0.00969, -0.00876, -0.00871, 0.00863, 0.00138, -0.00538, -0.0123, 0.01535, 0.00489, -0.0155, -0.00583, 0.00739, 0.01304, -0.01108, -0.01778, 0.01384, -0.00069, -0.00792, -0.00925, 0.00607, 0.01332, -0.01421, -0.00233, 0.01275, 0.00194, -0.01237, -0.00907, 0.00231, 0.01254, -0.00714, -0.00452, 0.00485, -0.00174, -0.01322, -0.00696, 0.01683, 0.01282, -0.01217, -0.01399, 0.02481, -0.01912, 0.02141, 0.01038, 0.02488, -0.01697, 0.01464, 0.01004, -0.02587, -0.0098, 0.01263, 0.01503, -0.02905, -0.01406, 0.01295, 0.01591, -0.01722, -0.01471, 0.01461, 0.01183, -0.01065, -0.01545, 0.01076, 0.00592, -0.01647, -0.01043, 0.00906, 0.00764, -0.02134, -0.0062, 0.01184, 0.01396, -0.00851, -0.01067, 0.01694, 0.00862, -0.01117, -0.0209, 0.02176, 0.01039, -0.01613, -0.0073, 0.00572, 0.00179, -0.01294, -0.00434, 0.01434, 0.01292, -0.01166, -0.00651, 0.02725, 0.01804, -0.01558, 0.01852, 0.02277, -0.01227, -0.01207, 0.01672, 0.00742, -0.00866, -0.00866, 0.03073, 0.00881, -0.0124, -0.00044, 0.01364, 0.02148, -0.01234, -0.00813, 0.0207, 0.004, -0.00535, -0.01961, 0.00626, 0.01013, -0.01034, -0.0116, 0.01233, 0.0137, -0.01865, -0.00635, 0.01144, 0.00732, -0.01074, -0.00616, 0.0171, 0.01488, -0.00979, -0.02108, 0.02023, -0.0014, -0.02369, -0.00552, 0.0203, 0.0176, -0.0164, -0.02013, 0.0074, 0.00413, -0.0162, -0.00944, 0.01826, 0.00754, -0.01613, -0.01638, 0.01281, 0.00613, -0.01202, -0.01155, -0.01635, -0.0121, 0.02292, 0.01943, 0.0111, 0.02202, -0.02173, -0.00902, 0.01094, 0.00799, -0.01721, -0.01456, 0.01876, 0.00446, -0.01882, -0.00801, 0.02149, 0.01383, -0.01387, -0.01488, -0.00812, -0.01436, 0.02472, 0.00996, 0.02179, 0.00825, -0.00495, -0.01665, 0.01334, 0.01172, -0.01148, -0.01546, 0.02173, 0.02255, -0.01849, -0.01487, 0.02418, 0.01251, -0.0187, -0.01557, -0.01751, -0.01324, 0.02085, 0.01702, 0.00812, -0.02393, -0.01034, 0.03133, 0.01674, -0.01226, -0.01344, -0.01815, -0.01019, 0.00905, 0.01061, 0.02095, 0.01245, -0.01218, -0.02069, 0.00679, 0.00879, -0.01887, -0.02076, 0.02419, 0.01828, -0.01577, 0.02042, 0.01989, -0.01952, -0.02086, 0.01043, 0.00877, -0.01376, -0.01314, 0.01564, 0.00891, -0.00675, -0.01104, 0.01707, 0.01746, -0.00474, -0.01264, 0.01968, 0.01605, -0.00691, -0.02159, 0.00152, 0.01552, -0.02149, -0.00843, 0.00893, 0.00483, -0.01832, -0.0112, 0.01559, 0.00147, -0.0162, -0.02137, 0.01908, 0.00923, -0.0148, -0.00969, 0.02171, 0.02151, -0.00913, -0.01047, 0.01014, 0.01495, -0.0207, -0.00613, 0.02006, 0.00536, -0.01313, -0.02111, 0.01105, 0.01482, -0.02017, -0.00919, 0.01281, 0.00787, -0.01399, -0.00941, 0.01523, 0.01349, -0.01659, -0.01698, 0.01345, -0.01431, -0.00617, 0.0113, 0.02372, -0.01436, -0.00938, 0.00925, 0.00132, -0.01418, -0.01402, -0.0037, -0.00524, 0.00815, 0.01063, 0.01425, 0.01503, -0.00812, -0.00973, -0.01497, -0.01497, 0.0214, 0.00639, 0.01462, 0.01846, -0.01666, -0.0082, -0.0154, -0.00815, 0.01167, 0.01791, 0.018, 0.00632, -0.0077, -0.01209, 0.00522, 0.0069, -0.01625, -0.00705, -0.01938, -0.0156, 0.01139, 0.02218, 0.01004, 0.01089, -0.01721, -0.01281, -0.01259, -0.01542, -0.01556, -0.01582, -0.01134, 0.00586, 0.00482, 0.01088, 0.01086, -0.01067, -0.00836, 0.01376, 0.01108, -0.01954, -0.01059, 0.00951, 0.013, -0.01187, -0.01031, -0.01498, -0.01247, 0.0091, 0.01376, 0.01505, 0.01136, -0.00606, -0.01092, 0.01149, 0.01102, 0.01441, -0.01633, -0.01254, -0.00624, 0.00843, 0.02254, 0.00928, 0.01065, -0.01085, -0.00948, 0.0321, 0.01443, -0.01436, 0.00967, -0.01059, 0.01706, 0.01495, -0.01278, -0.00879, -0.01371, -0.00875, 0.01277, 0.01377, 0.02353, 0.0167, -0.00764, -0.01373, 0.02202, 0.00761, -0.01646, -0.02106, 0.00936, 0.01128, -0.00887, -0.00636, -0.01485, 0.01737, -0.01382, 0.00355, -0.01284, 0.01699, -0.01654, -0.00788, 0.00343, 0.01655, -0.01262, -0.01525, -0.01799, 0.01152, 0.01284, -0.01102, -0.01788, 0.01399, 0.00656, -0.00553, -0.01033, 0.01518, 0.01103, -0.00805, -0.01886, 0.0204, 0.00563, -0.01421, -0.00839, 0.01713, 0.01344, -0.01159, -0.01042, 0.00867, 0.01016, -0.01053, -0.0134, 0.01598, 0.01039, -0.00848, -0.01961, 0.01044, 0.00755, -0.0162, -0.00638, 0.01786, 0.00777, -0.0186, -0.01167, -0.0073, 0.01056, 0.00623, -0.01714, -0.01258, -0.01256, 0.01568, 0.01582, -0.01571, -0.00989, -0.00756, 0.01903, 0.01267, 0.0011, -0.01714, -0.01178, -0.00674, 0.01847, 0.01406, 0.01359, 0.02021, -0.00753, -0.01432, -0.01532, -0.00885, -0.00612, -0.01723, 0.01148, 0.01514, 0.00489, 0.01014, 0.02088, 0.00251, -0.01315, -0.00504, 0.01673, 0.01213, -0.01965, 0.00422, 0.0076, -0.02049, -0.00462, 0.02657, 0.01227, -0.01782, -0.01425, 0.01746, -0.00418, -0.02096, -0.02243, -0.01485, -0.01296, 0.01138, -0.00176, 0.00972, 0.01168, -0.01761, -0.00938, 0.01759, 0.00669, -0.01894, -0.0168, 0.02418, 0.01446, -0.01161, -0.02294, 0.02191, 0.00788, -0.01147, -0.0049, 0.01596, 0.0144, -0.01215, -0.01257, 0.02651, 0.00877, -0.00878, -0.00965, -0.01713, -0.01221, 0.01419, 0.0134, 0.0176, 0.01424, -0.0136, -0.02171, -0.01624, -0.0074, 0.01619, 0.01591, 0.01263, 0.01381, -0.01321, -0.00731, 0.0139, 0.01641, 0.00218, -0.01333, 0.01353, 0.00721, -0.00228, -0.00323, 0.01531, 0.00325, -0.00327, -0.00315, -0.00216, -0.00026, -0.01519, -0.01152, 0.00512, 0.00786, -0.00537, -0.00701, 0.00631, 0.00263, -0.0066, -0.00231, -0.00094, -0.00434, 0.0, -0.00297, 0.00897, -0.01735, -0.00473, -0.01763, -0.01778, 0.01757, 0.00712, 0.01449, 0.01083, -0.00407, -0.01785, -0.01098, -0.01893, 0.00958, 0.007, 0.00271, 0.0118, -0.01292, -0.00664, -0.01324, -0.00738, 0.01378, 0.01642, -0.0007, 0.01583, -0.00633, -0.01471, 0.01061, 0.01297, -0.01588, -0.0135, -0.01058, -0.02435, 0.01825, 0.0156, -0.01534, 0.01398, -0.01716, -0.00963, 0.00965, 0.01264, -0.01459, -0.01398, 0.01583, 0.01683, -0.01755, -0.00903, -0.01077, -0.01565, 0.01642, 0.00875, 0.01698, 0.01391, -0.01246, -0.01304, -0.01187, -0.02077, 0.02154, 0.0161, 0.01789, 0.01736, -0.01478, -0.02276, 0.00926, 0.00489, -0.00323, -0.01781, 0.01518, 0.00403, -0.01454, -0.01242, 0.00695, 0.0099, -0.01803, -0.01187, 0.01347, 0.01274, -0.00808, -0.00853, 0.01256, 0.0142, -0.01446, -0.00729, 0.00099, 0.0104, -0.0184, -0.01331, 0.01543, 0.01497, -0.00521, -0.00776, 0.00881, 0.00855, -0.01555, -0.0155, 0.00761, 0.00767, -0.01167, -0.00917, 0.00278, 0.00783, -0.01997, -0.00492, 0.00115, -0.00802, -0.00894, 0.00703, 0.00344, -0.01152, -0.00024, 0.00208, 0.01736, -0.00279, -0.01458, 0.00212, 0.00331, -0.01687, -0.00542, 0.00074, 0.01083, -0.01457, 0.00048, 0.00299, 0.00468, 0.0126, -0.01211, -0.00848, 0.01041, 0.00535, -0.00113, 0.00333, 0.01551, 0.01152, -0.00134, -0.01316, 0.00252, 0.00702, -0.00988, -0.00723, 0.01078, 0.00628, -0.01408, -0.00535, 0.00632, 0.00977, -0.00283, -0.00163, 0.00414, 0.00957, -0.00408, -0.01258, 0.00143, 0.01101, -0.00729, -0.00412, 0.00797, -0.00786, 0.00022, -0.0046, 0.00986, 0.00387, -0.01238, -0.0172, 0.01202, 0.00579, -0.00443, -0.00614, 0.00812, 0.00707, -0.01272, -0.00727, -0.0002, 0.00516, -0.00066, 0.00078, 0.01554, 0.01021, -0.01847, -0.0018, 0.01142, -0.00443, -0.00389, -0.02235, 0.01418, 0.00961, -0.01435, -0.01414, 0.0, -0.001, -0.00901, -0.00313, -0.00475, -0.00648, 0.00597, 0.0042, -0.01672, -0.01399, 0.00869, 0.00539, -0.00412, -0.00614, 0.00536, -0.00552, -0.01126, -0.00959, 0.0019, -0.00982, -0.01677, -0.00107, -0.00733, -0.00291, -0.00087, -0.01563, -0.00785, -0.00434, -0.00298, -0.00627, -0.00919, 0.00025, -0.00714, -0.00719, -0.01034, -0.0008, -0.01342, -0.0127, -0.01105, -0.018, -0.00389, 0.00131, -0.01922, -0.01225, -0.00199, -0.00394, -0.02298, -0.01126, -0.00837, 0.00623, -0.01149], ['CWNJ_C02-2FAP-24_1_FCT', '#FF0000', 'Accelerometer AVG-FS8g_ODR800HZ_Zup accel_nmDC_average_y', 0.118, -0.118, -0.01002, -0.004784, -0.004183, 0.005602, -0.009599, -0.010599, 0.006937, -0.003276, -0.010383, -0.011383, 0.002014, -0.008719, -0.023569, -0.00729, 0.003051, -0.000618, -0.009226, -0.014455, -0.010385, -0.006469, -0.012288, -0.010561, 0.000408, -0.003281, -0.013213, -0.015061, -0.01145, -0.008162, -0.003817, -0.011477, -0.002931, -0.002641, -0.00962, -0.008303, -0.005853, -0.001762, -0.015197, -0.017743, -0.008202, -0.013215, -0.022393, -0.019039, 0.003245, 0.002261, -0.02026, -0.015423, 0.003422, -0.016795, -0.026817, -0.013022, -0.002421, -0.006563, -0.024116, -0.017668, -0.00449, 0.000461, -0.013312, -0.009719, 0.005094, -0.000287, -0.019095, -0.02376, 0.006142, 0.002628, -0.01565, -0.010705, -0.003142, 0.000484, -0.016752, -0.012101, 0.004668, -0.004429, -0.010781, -0.021791, -0.001748, -0.005075, -0.016394, -0.008919, 0.006871, -0.0106, -0.01313, -0.004398, -0.00523, -0.002703, -0.014615, -0.007359, -0.009628, -0.004858, -0.011959, -0.017628, 0.003078, 0.010342, -0.010801, -0.004591, 0.00032, -0.007471, -0.011368, -0.01167, -0.009179, -0.00353, -0.028292, -0.024105, -0.012227, -0.016483, -0.005227, -0.002409, -0.00557, -0.014273, -0.005641, -0.030162, -0.017904, 0.000226, 0.000859, -4.6e-05, -0.004343, -0.013792, -0.007122, 0.001618, -0.001019, -0.01806, -0.014621, -0.002348, 0.006127, -0.021649, -0.020339, 0.002575, -0.006118, -0.006449, -0.0154, -0.020551, -0.013468, -0.007644, -0.006377, -0.0194, -0.012347, 0.0028, -0.007288, -0.022295, -0.014395, -0.010895, -0.00853, -0.012239, -0.00335, -0.008344, -0.00412, -0.006059, -0.007919, -0.005348, -0.01564, 0.002845, -0.003628, -0.016215, -0.022032, -0.016136, -0.007489, 0.004027, -0.021179, -0.014721, 0.005314, -0.001821, -0.016084, -0.008021, 0.007244, -5.4e-05, -0.00647, 0.002061, -0.003268, -0.02526, -0.016022, -0.007247, -0.002363, -0.023768, -0.02247, -0.009482, 6.3e-05, -0.024167, -0.021387, -0.000739, -0.008883, -0.014065, -0.017096, -0.018336, -0.012384, -0.002617, 0.008962, -0.02348, -0.011488, -0.018308, -0.001655, 0.003334, -0.005262, -0.019078, -0.020252, 0.001825, -0.006194, -0.02208, -0.016715, -0.001092, -0.000855, -0.006779, -0.019829, -0.014152, -0.014125, -0.003427, -0.011534, -0.015656, -0.016625, -0.004681, -0.010007, -0.018977, -0.017392, -0.001987, 0.003123, -0.008877, -0.019216, -0.016266, -0.019749, -0.003643, -0.00468, -0.000562, 0.00038, -0.014478, -0.027559, 0.004325, 0.00228, -0.023392, -0.022846, -0.014756, -0.024076, 0.002437, 0.00446, -0.000793, 0.003132, -0.025726, -0.014923, -0.023985, -0.018391, -0.010153, -0.008497, -0.0027, 0.002362, -0.020962, -0.014085, -0.007489, -0.009991, -0.016373, -0.011452, -0.009226, 0.001469, 0.001287, -0.037242, -0.022186, 0.004152, 0.006335, -0.01873, -0.022383, -0.015842, -0.014863, -0.011765, -0.006921, -0.00588, -0.012788, -0.017841, -0.014788, 0.001264, -0.002788, -0.013978, -0.022869, -0.001634, 0.001787, -0.020125, -0.012178, -0.002966, 0.00207, -0.01101, -0.013503, -0.019331, -0.011314, -0.003044, 0.010666, -0.000702, -0.001546, -0.022786, -0.014715, 0.001134, 0.009413, -0.023911, -0.011462, -0.00326, 0.001864, -0.022526, -0.024709, -0.004014, -0.008703, -0.006202, -0.018831, -0.022917, -0.025552, 0.003422, -0.003844, -0.005336, -0.003351, -0.013176, -0.01431, -0.003151, -0.00187, -0.017543, -0.014523, -0.018582, -0.006378, -0.010014, -0.004245, 0.002287, -0.003445, -0.030831, -0.01267, 0.002786, 0.003481, -0.009417, -0.022259, 0.001241, 0.003976, -0.019917, -0.000661, -0.005348, -0.001373, -0.010885, -0.008288, 0.003928, 0.00076, -0.018842, -0.007341, -0.003084, -0.008867, -0.023929, -0.013614, 0.005338, -0.003117, -0.011466, -0.010271, -0.001251, 0.001522, -0.005551, -0.004674, 0.001537, 0.001537, -0.009002, -0.006766, -0.007737, -0.007265, -0.010757, -0.006131, -0.005506, 0.002664, -0.014571, -0.014946, 0.000476, 0.001383, -0.018831, -0.017189, -0.002911, -0.003265, -0.011749, -0.017577, -0.001352, 0.003732, -0.015282, -0.015751, -0.003225, -0.006276, -0.008054, -0.010017, 0.001179, -0.003953, -0.014942, -0.015998, -0.017423, -0.006262, 0.006304, 0.002675, -0.004908, -0.006501, -0.017383, -0.012896, 0.006376, -0.006742, -0.024391, -0.011626, 0.002644, -0.003916, -0.014527, -0.00356, -0.009639, -0.02465, -0.006309, 0.001052, -0.00313, -0.018865, -0.011402, 0.004129, 0.003814, -0.017598, -0.022779, 0.001376, 0.004796, -0.021439, -0.00832, -5e-05, -0.001214, -0.012897, -0.015729, -0.002587, -0.000368, -0.022866, -0.013553, -0.00245, -0.008319, -0.017808, -0.01733, -0.009346, -0.008819, -0.001909, 0.002375, -0.008694, 0.007126, -0.005214, -0.006602, 0.002932, -0.005352, -0.018302, -0.0127, -0.014825, -0.012654, -0.003696, -0.002984, 0.000191, 0.003746, -0.010908, -0.007225, -0.011439, -0.002519, -0.010082, -0.011524, -0.009021, -0.011392, -0.005632, -0.004465, -0.00142, 0.004129, -0.011361, -0.009652, 0.003733, 0.000981, -0.020263, -0.013646, 0.002234, 0.003614, -0.013683, -0.021306, -0.003726, 0.001506, 0.00627, -0.021354, -0.010762, -0.005524, -0.002502, -0.023239, -0.01172, 0.002656, -0.000313, -0.013707, -0.004241, -0.013911, -0.007788, 0.003329, 0.000169, -0.002957, -0.016353, -0.013106, -0.003283, -0.010263, -0.016093, -0.007155, -0.000985, -0.008603, -0.0141, -0.011815, -0.00557, -0.000119, -0.011885, -0.01534, -0.001366, -0.001002, -0.012303, -0.008099, -0.002652, 0.001241, -0.01914, -0.011098, -0.000678, -0.009808, -0.017761, -0.012397, -0.005164, -0.001978, -0.008339, -0.011475, -0.004496, -0.004401, -0.009993, -0.002129, 0.004545, -0.000966, -0.008742, -0.007239, -0.002886, 0.003488, -0.010936, -0.008853, 0.000423, -0.00403, -0.024184, 0.001819, -0.003204, 0.003113, -0.020829, -0.011226, -0.003953, -0.000878, -0.016515, -0.011338, 0.009098, -0.000973, -0.0214, 0.001079, -0.004502, -0.004743, -0.008382, 0.001509, -0.004741, -0.016766, -0.008501, 0.007374, 0.002875, -0.015518, -0.015091, 0.002376, -0.001871, -0.006028, -0.017441, 0.006766, -0.001274, -0.017266, -0.004746, 0.015177, 0.003478, -0.011636, -0.008906, -0.000293, -1.7e-05, -0.021636, -0.02519, -0.002155, -0.000436, -0.015833, -0.012087, -0.001185, -0.009472, -0.014721, -0.0273, -0.007356, -0.013969, -0.023236, -0.009442, 0.000382, 0.004503, -0.017312, -0.021732, -0.000133, -0.001478, -0.008219, -0.010795, 0.001234, -0.00648, -0.003895, -0.006921, -0.006035, -0.019624, -0.018953, -0.018062, -0.013421, -0.003181, -0.01085, -0.012904, -0.000368, -0.010297, -0.012591, -0.013615, 0.00905, -0.020045, -0.017424, 0.003984, -0.008908, -0.012216, -0.012426, 0.003081, 0.00479, -0.010764, -0.005197, -0.005753, -0.009413, -0.016108, -0.017849, -0.013168, -0.004491, -0.006863, -0.001186, 0.006827, -0.00091, -0.012031, -0.012248, 0.001713, -0.008153, -0.013552, -0.011683, -0.012606, -0.007911, 0.004398, -0.008895, 0.000536, -0.00354, -0.014719, -0.007817, 0.008242, 0.001703, -0.023719, -0.013731, 0.00039, -0.000321, -0.016942, -0.008706, 0.004762, 0.003257, -0.018786, -0.010692, -0.00276, -0.000633, -0.016848, -0.007244, 0.002339, -0.009825, -0.013902, -0.015826, -0.013781, -0.012154, -0.005815, -0.004351, -0.002861, 0.005668, -0.017971, -0.019288, 0.004453, -0.0034, -0.024205, -0.009884, 0.002069, -0.006735, -0.018094, -0.016388, 0.002498, -0.002114, -0.02101, -0.012054, 0.00324, -0.007366, -0.005639, -0.011376, 0.000704, 0.000488, -0.008591, -0.009447, -0.004431, -0.014143, -0.000984, -0.010717, -0.000624, -0.003584, -0.012511, -0.015902, -0.004614, 0.001231, -0.008702, -0.013876, -0.000505, -0.000155, -0.003931, -0.006362, -0.008527, -0.007029, -0.01459, -0.015439, 0.000591, 0.002062, -0.012744, -0.009075, -0.009279, -0.005888, -0.00726, 0.002675, -0.006365, -0.004441, -0.015098, -0.01068, -0.008795, 0.001441, -0.018174, -0.011044, -0.004083, -0.000696, -0.008701, -0.014982, -0.005898, 0.005803, -0.00925, -0.01153, -0.002838, -0.003172, -0.017763, -0.005842, -0.006168, -0.000653, -0.017203, -0.006252, -0.007104, 0.002148, -0.000221, -0.013435, -0.016491, 0.002826, -0.00927, -0.012502, -0.009817, -0.000284, -0.002192, -0.011563, -0.006484, -0.005575, -0.003886, -0.017702, -0.009486, 0.00519, 0.00135, -0.009652, -0.014097, -0.003335, -0.00249, -0.014882, -0.004813, 0.004932, -0.017763, -0.009524, 0.001709, -0.001748, -0.021127, -0.012279, -0.00602, -0.001888, -0.001097, -0.002936, 0.003371, -0.0101, -0.019975, -0.013625, -0.015517, -0.009977, 0.00463, -0.010936, -0.001167, -0.005819, -0.005968, -0.004698, -0.00221, 0.000768, -0.020031, -0.010239, -0.001822, 0.00394, -0.003045, -0.005622, 0.002435, -0.003208, -0.007468, -0.020215, -0.009363, -0.006586, -0.000635, 0.001531, -0.019743, -0.005663, -0.006706, 0.0015, -0.012802, -0.018765, -0.010231, -0.010786, 0.000839, 0.002368, -0.00176, -0.000191, -0.011411, -0.023509, 0.000483, -0.006626, -0.016469, -0.013214, 0.003059, -0.000207, -0.014115, -0.006741, 0.006077, 0.007291, -0.008461, 0.003038, -0.003502, -0.016063, -0.011124, -0.004156, -0.002878, -0.006145, -0.009769, -0.000968, 0.009137, -0.011103, -0.00544, -0.01518, -0.001723, -0.00331, -0.001428, 0.001251, -0.001796, -0.011388, -0.00378, 0.003916, 0.000339, -0.00844, -0.001162, -0.002144, -0.01637, -0.012658, -0.003905, -0.003794, -0.015528, -0.018593, -0.006225, -0.004998, -0.016376, -0.009038, 0.005964, 0.00528, -0.013404, -0.024403, -0.002238, -0.009583, -0.013741, -0.02219, -0.00118, -0.004543, -0.023298, -0.012204, 0.008461, -0.000682, -0.011688, -0.00671, -0.002107, -0.001865], ['CWNJ_C02-2FAP-23_1_FCT', '#008000', 'Accelerometer AVG-FS8g_ODR800HZ_Zup accel_nmDC_average_y', 0.118, -0.118, -0.01209, 0.01632, 0.0046, -0.00552, -0.00378, -0.00872, -0.00278, -0.00052, -0.01427, -0.00595, -0.00145, -0.00434, -0.00604, -0.00162, -0.0006, -0.00301, -0.00621, 0.00843, 0.00392, 0.00436, -0.00257, -0.00031, 0.00894, 0.00461, 0.00257, 0.00237, 0.00545, -0.01154, 0.00151, 0.00458, -0.00793, -0.00029, -0.00227, -0.00348, -0.00153, -0.00116, -0.00244, -0.01268, 0.0074, -0.00937, 0.00134, -0.00039, -0.00362, -0.00506, -0.00271, -0.00698, -0.002, -0.00215, 0.00083, -0.00159, 0.00434, -0.00237, -0.0026, -0.0142, 0.00421, -0.00345, -0.00863, -0.00491, -0.00453, -0.00586, -0.00426, 0.00356, 0.00165, -0.00194, -0.01124, -0.00728, 0.00663, 0.00052, 0.00821, -0.00882, -0.00429, -0.01112, 0.00285, -0.00061, 0.01125, -0.00398, -0.00971, -0.00262, -0.00228, 0.00424, -0.00045, 0.00483, -0.00478, -0.00291, 0.00247, -0.00328, 0.00324, -0.00366, 0.00777, -0.00932, -0.00682, -0.01156, 0.00666, 0.0025, -0.00211, -0.00199, -0.00421, -0.00383, -0.00045, 0.00242, -0.00706, -0.00301, -0.01015, 0.00543, 0.00589, -0.00366, 0.002, -0.00127, -0.00773, 0.00425, -0.00081, -0.00464, 0.00183, -0.0068, -0.00065, 0.00186, 0.00121, -0.0084, 0.00346, -0.0057, -0.00562, -0.00215, 0.00734, 0.00562, 0.006, 0.00283, -0.00694, 0.00749, 0.00768, 0.01192, -0.00484, 0.00329, 0.00194, -0.01286, 0.00061, 0.0, 0.0043, 0.00886, 0.0004, 0.00506, -0.00544, 0.00887, 0.00254, -0.00114, -0.00458, -0.00689, -0.00486, -0.00486, -0.00477, -0.00158, 0.00634, -0.00321, -0.00096, -0.00124, 0.01123, -0.00247, -0.00722, -0.01089, 0.00361, -0.01878, -0.00487, -0.00356, -0.00192, 0.00572, 0.00461, -0.01161, -0.00825, 0.00083, -0.00046, -0.00957, -0.00639, -0.00436, 0.00559, -0.00657, 0.00728, -0.01119, -0.0035, -0.01572, -0.00052, -0.00199, 0.00106, 0.00214, 0.0, -0.00725, -0.0007, 0.00039, 0.0088, -0.00165, 0.00445, -0.0079, 0.00189, -0.01237, -0.00907, 0.00644, -0.00113, -0.01236, 0.00434, 0.0006, -0.00641, -0.00554, -0.00343, 0.00352, 0.00262, 0.00105, 0.00853, 0.00582, -0.0051, -0.00547, -0.00428, 0.00611, -0.00128, 0.00491, 0.00991, -0.00537, 0.00221, -0.00501, -0.00785, -0.00016, -0.00174, -0.00604, -0.00279, -0.00054, -0.00551, 0.00127, -0.00021, -0.01437, -0.00632, 0.00786, -0.004, 0.00923, 0.00431, 0.00117, 0.0013, 0.00209, 0.00751, -0.00695, 0.00218, 0.00016, 0.00516, 0.00243, 0.00113, -0.01135, -0.01012, 0.00051, 0.00075, -0.00849, 0.005, -0.00506, 0.0, -0.00487, -0.00193, 0.00871, 0.00734, 0.00622, 0.00169, -0.01011, -0.00137, 0.00794, 0.00612, -0.00336, 0.00512, -0.00096, -0.0026, -0.00034, 0.00742, -0.01261, -0.00262, -0.00531, 0.00446, 0.00487, -0.00611, -0.00411, 0.00329, -0.00411, 0.00194, 0.00394, -0.01217, -0.00264, -0.01695, -0.0045, 0.00311, -0.00282, 0.00473, 0.0012, -0.00737, -0.00167, -0.00357, 0.00014, -0.00041, -0.00041, 0.00213, -0.00606, 0.00537, -0.00433, -0.01031, -0.00405, -0.00163, 0.00666, -0.0008, -0.00103, 0.00553, 0.00126, -0.00659, 0.00362, -0.00366, 0.00283, 0.00334, 0.00377, -0.00792, -0.00515, 0.00142, -0.00511, -0.00334, -0.00763, 0.00162, -0.00447, -0.00285, -0.00162, -0.00165, 0.00793, -0.00036, -0.00528, 0.00167, 0.00433, 0.00455, 0.00911, 0.00166, 0.00198, 0.00177, -0.00516, -0.00528, -0.00918, -0.00901, 0.00073, 0.00098, -0.00956, -0.00988, 0.00674, -0.00724, -0.01124, 0.0018, -0.00548, 0.01679, 0.00069, 0.00229, -0.00154, 0.00232, 0.00706, -0.00371, -0.00686, -0.00163, -0.00225, -0.00333, 0.0, 0.00747, 0.00024, 0.00071, -0.01039, -0.00066, 0.00649, -0.00088, -0.00685, -0.00812, -0.00249, 0.0, -0.00867, -0.0028, 0.00731, -0.00414, -0.00336, -0.00459, 0.00206, -0.01238, -0.00391, -0.00629, -0.00733, -0.01016, -0.00696, 0.00112, 0.00476, -0.01405, 0.00233, 0.00164, -0.00306, 0.00379, -0.00277, 0.0028, 0.00369, 0.00849, 0.00506, 0.00116, -0.02965, -0.00247, 0.0031, 0.01262, -0.00929, 0.00191, -0.00467, -0.00258, 0.00748, -0.02233, -0.00326, -0.00095, -0.00844, -0.00806, -0.00294, -0.0116, 0.00138, 0.00063, -0.0033, -0.00041, 0.00448, -0.00261, -0.00167, -0.00659, -0.00193, 0.00441, -0.006, -0.0159, 0.00017, 0.00178, 0.01157, -0.00826, -0.00511, -0.00219, 0.00244, -0.00543, -0.01021, -0.00703, -0.00414, -0.00168, 0.0061, -0.00585, -0.00413, 0.00208, -0.00028, -0.00836, -0.00145, -0.00402, -0.01088, -0.00123, -0.00113, 0.00822, 0.00824, 0.00256, 0.00154, 0.00275, 0.00312, -0.00684, 0.00514, 0.00545, 0.00728, -0.0029, -0.00259, -0.00086, -0.00338, -0.00028, -0.00305, -0.00474, 0.00423, -0.00185, 0.00026, 0.00062, 0.0052, -0.00048, -0.00258, 0.00025, -0.00265, 0.01128, 0.00307, 0.00554, 0.00434, 0.00165, -0.00028, -0.00254, -0.00334, -0.00044, -0.0064, -0.00286, -0.00315, -0.00214, -0.0041, -0.00916, -0.00717, 0.00018, 0.00589, -0.00576, 0.00489, -0.00379, -0.00335, -0.00193, -0.00493, 0.00446, 0.00192, 0.0005, -0.00207, -0.00145, -0.00447, -0.01432, 0.00091, 0.00436, -0.01487, -0.00179, 0.00742, -0.00294, 0.00717, -0.00592, -0.00107, -0.00797, -0.00532, 0.00217, -0.00105, -0.00819, -0.00175, -0.00014, -0.00283, -0.00763, -0.00813, -0.00399, 0.0033, 0.0074, -0.00492, 0.0, 0.00515, -0.00164, -0.00684, 0.00277, 0.00369, -0.00379, -0.00348, 0.0077, 0.0133, -0.00177, -0.00505, 0.00357, -0.00528, -0.00749, 0.00438, -0.0019, -0.00515, -0.00168, 0.00074, -0.00771, -0.00283, -0.00182, -0.00716, -0.00337, -0.0033, 0.00088, 0.0071, 0.002, -0.00014, -0.00318, -0.00138, -0.0033, -0.00289, 0.00614, 0.00709, 0.0024, -0.00529, -0.01176, -0.0078, 0.00573, -0.00928, -0.0102, -0.00895, 0.00755, 0.00639, 0.00155, -0.00594, 0.00557, -0.0062, -0.01457, -0.00019, 0.01235, -0.00506, -0.00614, 0.00557, -0.00626, -0.0029, -0.0067, -0.00348, 0.00493, 0.00032, 0.00095, -0.01219, -0.00876, -0.01093, -0.0106, -0.00464, -0.00342, 0.00146, 0.0, -0.00768, -0.00752, -0.00376, -0.00528, -0.01817, -0.01056, 0.0048, -0.00105, -0.01418, -0.0066, -0.00881, -0.0047, -0.01204, -0.0111, -0.01032, -0.00631, -0.01313, -0.01183, -0.0109, -0.00932, -0.00743, -0.00964, -0.00253, -0.00343, -0.01172, -0.00092, -0.01219, -0.00179, 0.0043, -0.00362, -0.01571, -0.00542, 0.0032, -0.00306, 0.00083, -0.00219, -0.00306, -0.0059, -0.00793, 0.00359, 0.00491, -0.01205, -0.0207, -0.00426, -0.00094, -0.01896, -0.01376, -0.00727, 0.00214, -0.01614, -0.01899, -0.00297, -0.00494, -0.00668, -0.02279, -0.00664, -0.0031, -0.01247, -0.0108, -0.00438, 0.00799, -0.01674, -0.01299, 0.00087, -0.0009, -0.01379, -0.01112, -0.0094, -0.01286, -0.00359, 0.00048, -0.01856, -0.02039, -0.01204, -0.01398, -0.01035, -0.00777, -0.007, -0.00695, -0.00928, -0.01247, -0.00351, -0.01011, -0.01537, -0.00506, 0.00734, -0.00664, -0.00123, 0.00756, -0.00104, -0.01032, -0.00537, 0.00348, -0.00793, -0.00444, -0.01705, -0.0066, -0.00741, -0.01487, -0.00552, -0.00356, -0.00527, -0.01445, -0.00638, -0.0134, -0.01092, -0.0103, -0.01564, -0.00694, -0.01901, -0.00434, -0.01227, -0.01201, -0.0054, -0.01939, -0.01352, -0.00302, -0.00929, -0.0183, 0.0, -0.00236, -0.00643, -0.01197, -0.01405, 0.00413, -0.01283, -0.0119, -0.00675, -0.00276, -0.01432, -0.01139, -0.00317, 0.00048, -0.01579, -0.00384, -0.00688, 0.00166, -0.01389, -0.01239, -0.00919, -0.01141, -0.00237, -0.01511, -0.01783, -0.00843, -0.01397, -0.00758, -0.00591, -0.00984, -0.01282, -0.0025, -0.00255, -0.00385, -0.00728, -0.00745, -0.00576, -0.00951, 0.00249, -0.00382, -0.00869, -0.0137, 0.00636, -0.00132, -0.00782, -0.02429, -0.0163, -0.00474, -0.00979, -0.00569, -0.00318, -0.00755, -0.0136, -0.01564, -0.00208, -0.01688, -0.00796, -0.00621, -0.00432, -0.00947, -0.00733, 0.00239, -0.00748, -0.01468, -0.00744, 0.00158, -0.00048, -0.00951, -0.00553, -0.01062, -0.01249, -0.01843, -0.00085, -0.01053, -0.00355, -0.01121, -0.01128, -0.01255, -0.01006, -0.00388, -0.00489, -0.00984, -0.00057, -0.01817, -0.01465, -0.00412, -0.0143, -0.0138, -0.00218], ['CWNJ_C02-2FAP-24_2_FCT', '#00FFFF', 'Accelerometer AVG-FS8g_ODR800HZ_Zup accel_nmDC_average_y', 0.118, -0.118, -0.013964, -0.000321, -0.004692, -0.0025, 0.007696, 0.001936, -0.001229, -0.007474, -0.004071, -0.004585, -0.004381, -0.014232, -0.001326, -0.01104, 0.003955, -0.007101, -0.011034, -0.00415, -0.01197, -0.01336, -0.010088, -0.013282, -0.00555, -0.01534, 0.001752, -0.006785, 0.0035, -0.005784, -0.007929, -0.00757, -0.006511, 0.001193, -0.008909, -0.005919, 0.009912, -1.9e-05, -0.007143, -0.01648, -0.007352, -0.012278, -0.00523, 0.000941, 0.002715, -0.002446, -0.004557, -0.001727, -0.006573, -0.010529, 0.001064, -0.007575, -0.005593, -0.001148, -0.009806, 0.004737, 0.002417, -0.007084, -0.001154, 0.003137, -0.01172, -0.012237, 0.001608, -0.006458, -0.010103, -0.006977, -0.004052, -0.012586, -0.014656, -0.009572, -0.002845, -0.010888, -0.00405, -0.013561, -0.010006, -0.005212, -0.01222, -0.004777, -0.004542, 0.002614, -0.005777, -0.004966, -0.001148, -0.007078, -0.017365, 0.001323, -0.008876, 0.00153, -0.014523, -0.001731, -0.005198, -0.010162, -0.015433, -0.008891, -0.005351, -0.009541, 0.005014, -0.001202, -0.014548, -0.002528, 0.000588, 0.002603, -0.002825, -0.011836, -0.004227, -0.00604, 0.000296, -0.014162, -0.014428, -0.005901, -0.01412, -0.017753, -0.004605, -0.01233, -0.00165, -0.009253, -0.004949, -0.003171, 0.003803, -0.005824, 0.000887, 0.000306, -0.004179, -0.002379, -0.008476, -0.012575, -0.013251, -0.010312, -0.006218, -0.000567, -0.007819, -0.017152, -0.005879, -0.006253, -0.008308, -0.00171, -0.007385, -0.00801, -0.007299, -0.0042, -0.018528, -0.005322, -0.007545, -0.009477, -0.007745, -0.009069, -0.00848, 0.002887, -0.008652, -0.012732, -0.006868, -7.3e-05, 0.00207, -0.000289, -0.00139, -0.006466, -0.013998, -0.004984, -0.009423, -0.018283, -0.008807, -0.003751, -0.012931, -0.002775, -0.012117, -0.011167, -0.017764, -0.00124, -0.01087, -0.006561, -0.000416, 0.007074, -0.001657, -0.003877, -0.001692, -0.006494, 0.001315, -0.000666, -0.012383, -0.000407, -0.003373, 4.6e-05, -0.010767, -0.009957, -0.00603, -0.006695, 0.008388, -0.00062, -0.004976, -0.007232, -0.015771, -0.017215, 0.004363, 0.000663, -0.007419, -0.014939, -0.002221, -0.008964, -0.001842, 0.002399, -0.010365, -0.00861, -0.003903, -0.005317, 0.000517, -0.006399, -0.00925, -0.011284, -0.008134, -0.007317, -0.016594, -0.016161, -0.004234, -0.005883, -0.011042, -0.006601, -0.001237, -0.016281, -0.014527, -0.013618, 0.000723, -0.012025, -0.003795, -0.000812, -0.001898, -0.003397, -0.006526, -0.005319, -0.000471, -0.002018, 0.002567, -0.003879, -0.009655, 0.001571, 0.002663, -0.013268, -0.000476, -0.00506, -0.003416, -0.003797, -0.006404, -0.01918, -0.005008, -0.006306, -0.004393, -0.010251, 0.001548, 0.0081, -0.004036, -0.017332, -0.008007, -0.004384, -0.008378, 0.003308, 0.000699, 0.004781, -0.011023, -0.018384, -0.004753, -0.005761, -0.006867, -0.006102, -0.007374, -0.005273, 0.004407, 0.003671, 0.000436, -0.001928, -0.011078, -0.011645, -0.00072, 0.000567, -0.005141, -0.003668, -0.001453, -0.009159, -0.012249, -0.000381, -0.007354, -0.002816, -0.005402, 0.00271, -0.008473, -0.013626, -0.009368, -0.005964, -0.012201, 0.0054, -0.000457, -0.010647, -0.002675, -0.003751, -0.003965, -0.002168, 0.000404, -0.009696, -0.007217, -0.009371, -0.001047, -0.013418, -0.003382, 0.001837, -0.004665, -0.013212, -0.009747, -0.013779, 0.004446, -0.001136, 0.001135, -0.015508, -0.006395, -0.001898, -0.01235, -0.004545, -0.000671, -0.010387, -0.001441, -0.007653, -0.000158, -0.000673, -0.004628, -0.008367, -0.011071, -0.011865, 0.000943, -0.001609, -0.003883, -0.007708, 0.002445, -0.002627, -0.016417, 0.00246, -0.003865, -0.008151, -0.003717, -0.011208, -0.021246, 0.001793, -0.00144, 0.001157, -0.003413, 0.00412, -0.006281, 0.003048, -0.013156, -0.012205, -0.008813, 0.001851, 0.003303, -0.007285, -0.012283, -0.002458, -0.001335, -0.019185, -0.00574, -0.005005, -0.002063, -0.012185, 0.000894, -0.001597, -0.003145, -0.007871, -0.001745, -0.009808, -0.009453, 1e-06, -0.010537, -0.00651, -0.002409, 0.000234, -0.011944, -0.003228, -0.001461, -0.000605, -0.002449, -0.001102, -0.004429, 0.009575, -0.010405, -0.002502, -0.001872, -0.003623, -0.008857, 0.002597, -0.001822, -0.005769, -0.00829, -0.001395, -0.005834, -0.005965, -0.002305, -0.001479, -0.006387, -0.010963, -0.001006, 0.003068, 0.000588, -0.000574, 0.002538, -0.002215, -0.005189, -0.009723, -0.01391, 0.004314, -0.012319, -0.017737, -0.006871, -0.005466, -0.008339, -0.006705, -0.003977, -0.005899, -0.0078, -0.010176, -0.00865, -0.002835, -0.00537, 0.003122, -0.002824, -0.001614, -0.01352, -0.001342, 0.001238, -0.010113, 0.006063, -4.6e-05, -0.003457, 0.004366, -0.005628, -0.009351, -0.003863, 0.001972, -0.007676, -0.00986, -0.007052, -0.003161, 0.004623, -0.00235, -0.011452, 9.4e-05, -0.010663, -0.000386, 0.000972, -0.005464, 0.003899, -0.002766, -0.005131, 0.00092, 0.003428, -0.006334, -0.003338, -0.003802, -0.012208, -0.009357, -0.005827, -0.006067, -0.00299, -0.001781, -0.008454, -0.000754, 0.002866, -0.007146, 0.000665, -0.003526, 0.001019, -0.00988, -0.006289, -0.007721, -0.010317, -0.009647, -0.004955, -0.00783, -0.00063, 0.007876, 0.001857, -0.009998, -0.008619, -0.007988, -0.009349, -0.004505, -0.013599, 0.00721, -0.006831, -0.015792, -0.008854, -0.003397, 0.001202, -0.000793, -0.006395, 0.011051, -0.009147, -0.006865, 0.00011, -0.002317, -0.005088, -0.009004, -0.007318, -0.003621, -0.007628, -0.003565, -0.00345, 0.001071, -0.010404, -0.002593, -0.005152, -0.007037, -0.001421, 0.000144, -0.016907, -0.00257, -0.00782, 0.007906, -0.012253, -0.005063, -9.1e-05, -0.003802, -0.002763, -0.013917, -0.003387, -0.010492, -0.010215, -0.005586, -0.002326, -0.005819, 0.000175, -0.002324, -0.011132, -0.001896, -0.006423, -0.010126, -0.012311, 0.001111, -0.004598, -0.001586, -0.009051, -0.003557, -0.005143, -0.007975, -0.003528, -0.004498, -0.001913, -0.002131, -0.008947, -0.009592, -0.018292, -0.008484, -0.001147, -0.005855, -0.00108, -0.001167, -0.007697, -0.009375, 2e-05, -0.015293, -0.000962, -0.011906, -0.007761, -0.009102, -0.007881, -0.009613, -0.004747, -0.006706, -0.00823, -0.002419, -0.006116, -0.002431, -0.0097, -0.009481, -0.001273, -0.005699, -0.002942, -0.001955, -0.012377, 0.002117, -0.00359, -9.7e-05, -0.004179, -0.009861, -0.02148, -0.002187, -0.004824, -0.008395, -0.008017, -0.005364, -0.006339, 0.000652, -0.005698, -0.001312, 0.008458, -0.015583, -0.00385, -0.004444, -0.003281, -0.006638, -0.005562, -0.009787, -0.002753, -0.011835, 0.003215, -0.002192, -0.00272, -0.004975, -0.001701, 0.009147, -0.015003, -0.006995, -0.008278, -0.003232, -0.000762, -0.005539, -0.012603, -0.003327, -0.006514, 0.000139, 0.001126, 0.001883, -0.009425, -0.002391, -0.003555, -0.006622, -0.013033, -0.010824, -0.003866, -0.007913, -0.004618, -0.002437, -0.003394, -0.006364, -0.003462, -0.000178, -0.006276, -0.00468, -0.003358, -0.016316, -0.001473, -0.005788, -0.003478, 0.002821, -0.004787, 0.00137, -0.000463, 0.009948, 0.001019, -0.006959, -0.004642, -0.001381, 0.003631, -0.004681, 0.010685, -0.01263, -0.00469], ['CWNJ_C02-2F-REL01_1_FCT', '#9400D3', 'Accelerometer AVG-FS8g_ODR800HZ_Zup accel_nmDC_average_y', 0.118, -0.118, -0.01299, -0.00621, -0.01716, -0.01351, -0.01346, -0.00879, -0.00754, -0.00406, -0.01265, -0.0053, -0.01696, 0.00529, -0.00291, 0.02042, -0.01007, 0.01035, -0.00473, -0.00047, -0.01416, 0.00443, -0.01074, -0.00363, 0.00944, -0.00294, -0.01431, -0.01052, 0.01089, 0.01193, -0.01823, 0.00519, -0.00629, -0.0054, -0.01318, 0.00803, 0.00402, 0.01274, 0.00479, -0.00701, -0.0111, -0.00499, -0.00696, 0.01198, -0.00667, 0.01672, 0.00762, 0.01019, 0.00704, -0.00716, -0.01843, 0.00597, -0.00845, 0.00756, 0.00124, -0.01003, 0.00559, 0.00905, -0.01254, 0.00444, -0.01567, -0.00108, 0.00461, -0.00778, -0.01551, 0.00459, 0.00166, 0.00251, -0.01014, 0.00818, 0.00468, -0.00925, 0.01663, 0.00075, -0.01296, 0.0074, 0.01, 0.01343, 0.01196, -0.00133, 0.00934, 0.01364, -0.01073, -0.00908, 0.0108, 0.01293, -0.00151, 0.00116, -0.01224, 0.00701, 0.00167, 0.00412, -0.00077, 0.00134, -0.01423, -0.01212, 0.0048, -0.00915, -0.01051, 0.00863, -0.00582, -0.00012, -0.01431, -0.00474, -0.00154, 0.00299, -0.02589, -0.01926, -0.0133, 0.00695, -0.01206, 0.00422, -0.00675, -0.00042, -0.00017, 0.01363, -0.01388, 0.00538, -0.00318, -0.00094, -0.00057, -0.00954, -0.01007, -0.00404, -0.00783, 0.0081, -0.01236, -0.01094, -0.00779, 0.00287, -0.00361, 0.00727, 0.00969, 0.01014, -0.00783, -0.0038, -0.00488, 0.01665, -0.01496, -0.01259, -0.00689, 0.01766, -0.00946, 0.00501, -0.01066, -0.00732, -0.01233, -0.01144, 0.01166, -0.00155, -0.00095, 0.00247, -0.01513, -0.0117, 0.00566, 0.00755, -0.00598, 0.00036, 0.0, 0.00492, 0.00743, -0.00436, -0.01112, -0.01117, -0.02, -0.01568, -0.009, -0.00392, -0.01428, 0.00275, -0.01447, 0.0061, 0.00231, 0.00672, 0.00083, -0.0114, -0.00327, 0.01898, 0.00045, -0.01035, -0.01238, 0.00792, -0.01772, -0.01443, 0.00574, 0.00867, -0.00292, -0.01631, 0.01601, 0.00311, 0.01818, -0.00987, -0.00936, -0.00397, -0.0039, 0.00684, -0.00465, -0.01093, -0.00931, -0.00822, 0.00857, 0.00628, -0.00188, 0.00694, 0.01035, 0.00124, -0.01051, -0.01589, 0.00535, -0.01023, -0.00787, -0.00765, -0.01008, 0.00077, 0.00609, 0.00892, 0.01304, -0.0119, 0.00537, 0.00643, 0.00811, -0.01368, 0.00561, 0.00574, -0.00571, 0.00381, -0.00694, -0.0067, 0.01196, -0.01274, 0.00627, 0.00862, -0.00271, -0.01387, -0.00953, -0.00258, 0.00274, -0.01312, 0.00281, -0.01535, -0.00248, -0.01274, 0.01097, -0.01076, 0.00477, -0.01551, -0.01043, -0.00952, -0.00417, -0.00937, -0.009, 0.00714, 0.01364, 0.00866, 0.01344, -0.01307, 0.00343, -0.0192, -0.0054, -0.0065, -0.00458, 0.00121, 0.00441, 0.01065, -0.00331, -0.00356, 0.01435, -0.01179, -0.01244, -0.01434, 0.0073, -0.0157, 0.0047, -0.00194, 0.01669, -0.00314, 0.018, -0.01228, -0.00436, 0.00406, 0.00683, 0.00669, 0.01089, -0.00164, -0.01088, -0.01257, -0.00638, -0.01108, 0.01232, 0.00358, -0.00975, 0.00433, -0.0036, -0.01252, -0.01041, -0.01537, -0.01444, -0.01487, 0.00579, -0.00189, 0.00245, -0.01094, 0.01143, -0.00815, 0.01229, 0.00538, 0.00986, -0.00288, 0.01634, -0.00856, -0.00226, -0.00599, 0.00989, -0.00031, 0.00493, -0.00942, 0.00283, -0.00618, -0.00683, 0.00087, -0.0039, -0.01856, -0.01129, 0.0099, 0.00676, -0.00469, -0.01332, 0.00822, -0.01423, 0.00624, -0.00904, -0.00619, -0.00974, -0.0067, 0.00623, 0.00938, -0.01645, 0.00285, -0.0107, 0.00572, -0.01823, -0.01129, -0.01725, 0.00131, -0.00735, -0.0023, -0.00487, -0.00461, 0.00665, 0.01341, -0.00828, 0.00721, -0.00166, -0.00115, -0.0119, -0.00442, 0.00154, 0.00483, -0.01164, -0.01293, 0.00648, -0.01128, -0.00111, -0.00499, -0.00772, -0.00949, 0.00546, -0.01254, -0.01038, -0.01032, -0.01125, -0.01208, 0.00805, -0.00143, -0.00629, -0.01139, -0.00723, -0.01485, -0.01584, 0.00281, 0.00215, -0.00656, -0.00705, -0.00757, -0.00019, -0.00428, -0.00644, -0.00475, 0.00686, -0.01239, -0.01158, 0.00423, -0.00723, -0.00905, -0.00053, -0.01457, -0.01441, -0.01238, 0.01101, 0.00392, 0.00383, -0.00184, 0.01509, -0.00803, -0.01135, 0.00142, -0.01128, -0.01009, 0.00383, -0.01215, -0.01324, 0.00654, -0.01288, -0.00832, -0.0117, -0.01033, -0.01448, 0.00583, -0.01498, -0.0053, -0.00861, 0.00231, -0.0135, -0.01769, -0.00094, -0.00074, 0.00582, -0.0199, -0.01385, -0.01945, -0.00216, 0.01314, 0.01599, -0.01437, -0.00762, -0.01731, -0.00557, 0.0033, -0.00031, 0.00077, -0.01191, -0.01381, -0.01463, -0.01645, -0.02053, -0.01234, -0.00935, 0.00531, -0.00742, -0.00483, -0.00657, -0.01238, 0.00927, 0.00602, 0.01237, -0.0142, -0.01016, -0.00155, -0.00909, -0.01311, 0.00652, -0.00622, -0.00889, -0.0046, -0.01979, -0.00708, 0.00775, -0.01087, 0.00333, -0.00296, 0.00191, 0.00092, 0.00831, -0.02261, 0.00117, -0.01279, -0.02103, -0.01291, -0.01488, -0.00914, -0.01417, -0.0096, 0.00883, 0.009, -0.00149, -0.00565, -0.00949, -0.00801, -0.01051, 0.00401, -0.00564, -0.01268, -0.01149, -0.00978, -0.00987, -0.01099, 0.00231, -0.00416, 0.00441, -0.00755, 0.00844, 0.0009, 0.00815, 0.01255, -0.00044, -0.01181, -0.01328, -0.01369, 0.00084, 0.01021, 0.01062, 0.0008, 0.00128, -0.01253, -0.01389, 0.00368, 0.00555, -0.00029, 0.00971, -0.00106, -0.01273, -0.00794, -0.01418, -0.00889, -0.00138, -0.03984, 0.00267, -0.01093, -0.01591, 0.00113, -0.0086, 0.01066, 0.00759, -0.01848, -0.00629, 0.0043, 0.00696, 0.00188, 0.0114, 0.0013, -0.01347, -0.00846, 0.0098, 0.01054, -0.01128, -0.01122, 0.00441, -0.00468, -0.00359, -0.01263, 0.00081, -0.01166, -0.00939, -0.0103, 0.00146, 0.00463, 0.00386, -0.01111, -0.0092, 0.00495, 0.00585, -0.00524, -0.01006, -0.00733, 0.00096, -0.01156, -0.0153, -0.00901, 0.00529, 0.013, -0.00018, -0.01308, -0.01011, -0.01123, -0.00727, -0.00908, -0.00975, -0.00023, -0.01613, -0.0097, 0.00854, 0.0039, 0.00277, 0.00961, -0.01063, 0.00427, -0.01543, -0.00054, -0.00569, 0.00186, 0.00946, 0.00656, -0.00137, -0.00895, 0.00222, 0.0053, 0.00676, 0.00235, -0.00851, -0.01202, 0.00617, -0.00709, -0.00289, 0.0025, -0.00579, 0.00101, -0.0158, -0.00776, 0.00382, 0.00527, 0.00952, 0.00135, -0.01321, -0.00768, -0.01436, -0.01791, -0.00711, 0.00557, -0.00612, -0.01543, -0.00859, 0.00564, 0.00558, -0.0013, 0.00224, 0.01061, 0.0, -0.01388, 0.01211, -0.01135, -0.01037, 0.00756, -0.0118, -0.0094, 0.01114, -0.00038, 0.00804, -0.01618, 0.00481, -0.01521, -0.01603, -0.01014, -0.00851, 0.00704, 0.00369]], [['CWNJ_C02-2FAP-23_2_FCT', '#0000FF', 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_y', 0.118, -0.118, -0.0017, -0.01955, 0.00291, 0.00056, -0.01089, -0.00734, 0.01384, 0.00581, -0.01075, 0.00595, 0.00462, 0.01176, -0.01536, -0.00899, 0.01482, 0.00359, -0.01349, -0.01528, 0.01051, 0.00691, -0.01731, -0.00986, 0.00678, 0.00883, -0.01111, -0.00655, 0.00665, 0.00161, -0.0049, -0.00767, 0.0111, 0.00711, -0.0142, -0.00988, 0.01133, -0.00086, -0.02724, -0.00546, 0.01423, 0.00928, -0.01065, -0.01225, 0.00406, 0.00577, -0.01347, -0.00246, 0.01839, 0.01007, -0.01958, -0.01702, 0.00237, -0.00177, -0.0124, -0.01008, 0.01904, 0.00968, -0.00923, -0.00885, 0.00883, 0.00164, -0.00561, -0.01244, 0.0154, 0.00504, -0.01553, -0.00581, 0.00758, 0.0127, -0.01141, -0.01827, 0.01402, -0.00049, -0.00807, -0.00938, 0.00613, 0.01329, -0.01427, -0.00215, 0.01276, 0.00196, -0.01205, -0.00917, 0.00223, 0.01271, -0.00704, -0.0046, 0.00475, -0.00169, -0.01333, -0.00703, 0.0166, 0.01277, -0.01283, -0.01438, 0.02533, -0.01885, 0.02124, 0.01019, 0.02475, -0.01665, 0.01447, 0.00981, -0.0257, -0.00958, 0.01236, 0.01473, -0.02896, -0.01409, 0.01292, 0.01597, -0.01708, -0.01453, 0.01469, 0.01152, -0.01084, -0.01531, 0.01071, 0.00583, -0.01654, -0.01048, 0.00916, 0.00758, -0.02137, -0.00623, 0.01179, 0.01395, -0.00825, -0.01049, 0.01685, 0.00841, -0.01099, -0.02101, 0.02169, 0.01042, -0.01602, -0.00752, 0.00565, 0.00192, -0.01292, -0.00414, 0.01431, 0.01291, -0.0118, -0.0066, 0.02706, 0.01806, -0.01572, 0.01867, 0.02262, -0.01244, -0.01204, 0.01686, 0.00749, -0.00892, -0.00869, 0.03079, 0.00864, -0.01264, -0.00038, 0.01368, 0.02137, -0.01225, -0.00778, 0.02023, 0.00398, -0.00541, -0.01972, 0.00628, 0.01035, -0.01049, -0.01166, 0.0123, 0.01367, -0.0186, -0.00641, 0.01154, 0.00724, -0.01071, -0.00621, 0.01716, 0.01454, -0.00966, -0.02083, 0.02018, -0.00178, -0.02366, -0.00569, 0.0202, 0.01764, -0.01631, -0.0202, 0.00716, 0.00414, -0.01616, -0.00934, 0.01788, 0.00725, -0.01637, -0.01653, 0.01263, 0.00621, -0.0118, -0.01157, -0.01631, -0.01225, 0.02295, 0.01984, 0.01105, 0.02229, -0.02154, -0.00911, 0.01089, 0.00795, -0.01713, -0.01449, 0.01885, 0.00453, -0.01898, -0.00788, 0.02139, 0.01386, -0.01389, -0.01491, -0.00829, -0.01464, 0.02481, 0.00981, 0.02177, 0.00825, -0.0049, -0.017, 0.01354, 0.01159, -0.01158, -0.0155, 0.02176, 0.02241, -0.01879, -0.01494, 0.0243, 0.01254, -0.01899, -0.01555, -0.01775, -0.01303, 0.02091, 0.01687, 0.00794, -0.02397, -0.01042, 0.03153, 0.01647, -0.01229, -0.01344, -0.01819, -0.0101, 0.00905, 0.01062, 0.02095, 0.01256, -0.01232, -0.02072, 0.00686, 0.0087, -0.01891, -0.02079, 0.02433, 0.0181, -0.01573, 0.02057, 0.01986, -0.01963, -0.02079, 0.01061, 0.00862, -0.01371, -0.01292, 0.01557, 0.00886, -0.00666, -0.01103, 0.01697, 0.01729, -0.00468, -0.0127, 0.01961, 0.01604, -0.00663, -0.02168, 0.00161, 0.01556, -0.02154, -0.00865, 0.00891, 0.00506, -0.01837, -0.01141, 0.01571, 0.00126, -0.01616, -0.0213, 0.01905, 0.00912, -0.01465, -0.0098, 0.02156, 0.02164, -0.0091, -0.01052, 0.01008, 0.01472, -0.02066, -0.00596, 0.02007, 0.00526, -0.01326, -0.02099, 0.01118, 0.01481, -0.02003, -0.00936, 0.0129, 0.00797, -0.01404, -0.00945, 0.01524, 0.0137, -0.01678, -0.01682, 0.01342, -0.01428, -0.00629, 0.01125, 0.02373, -0.01438, -0.00937, 0.00927, 0.00126, -0.01402, -0.01379, -0.00379, -0.00546, 0.00829, 0.01066, 0.01419, 0.01481, -0.00812, -0.00978, -0.01523, -0.01511, 0.02133, 0.00632, 0.01457, 0.01804, -0.01679, -0.00812, -0.01538, -0.00814, 0.01166, 0.01792, 0.01802, 0.00621, -0.00769, -0.01208, 0.0052, 0.00672, -0.01624, -0.00687, -0.01971, -0.01558, 0.01143, 0.02223, 0.01019, 0.01077, -0.01739, -0.01262, -0.01263, -0.0151, -0.01559, -0.01579, -0.01139, 0.00581, 0.00471, 0.01088, 0.01093, -0.01049, -0.0085, 0.01352, 0.01102, -0.01954, -0.01076, 0.00943, 0.01293, -0.01172, -0.01046, -0.01512, -0.01254, 0.00896, 0.01372, 0.01493, 0.01147, -0.00643, -0.01101, 0.01131, 0.01087, 0.01428, -0.01646, -0.01258, -0.00636, 0.0085, 0.0228, 0.00932, 0.01075, -0.01113, -0.00969, 0.03221, 0.01446, -0.01439, 0.00944, -0.01037, 0.01696, 0.01473, -0.01271, -0.00869, -0.01372, -0.00877, 0.01282, 0.01378, 0.02352, 0.01663, -0.00765, -0.01364, 0.02203, 0.00777, -0.01661, -0.02119, 0.0091, 0.01132, -0.00848, -0.00602, -0.01484, 0.01746, -0.01387, 0.00349, -0.01298, 0.01707, -0.0165, -0.00801, 0.00337, 0.01654, -0.01243, -0.01493, -0.01804, 0.01157, 0.01286, -0.01118, -0.01805, 0.01391, 0.00644, -0.00555, -0.01024, 0.01521, 0.01092, -0.0081, -0.01883, 0.0202, 0.0054, -0.01431, -0.00854, 0.01719, 0.01367, -0.01153, -0.01027, 0.0088, 0.01021, -0.01043, -0.01344, 0.01613, 0.01043, -0.00839, -0.01989, 0.01046, 0.00771, -0.01648, -0.00625, 0.01782, 0.00767, -0.01851, -0.01186, -0.00739, 0.01082, 0.00633, -0.0171, -0.01252, -0.01257, 0.01544, 0.01583, -0.0159, -0.00989, -0.00724, 0.01899, 0.0128, 0.00124, -0.01709, -0.01184, -0.00694, 0.01851, 0.01428, 0.01367, 0.02027, -0.00749, -0.01415, -0.01528, -0.00884, -0.00608, -0.01714, 0.01133, 0.01534, 0.00488, 0.01018, 0.0209, 0.00231, -0.01323, -0.00513, 0.01684, 0.01209, -0.01962, 0.004, 0.00761, -0.02052, -0.00443, 0.02668, 0.01229, -0.01802, -0.01414, 0.0178, -0.00408, -0.02103, -0.02224, -0.01498, -0.01316, 0.01137, -0.00177, 0.00976, 0.01165, -0.01756, -0.00937, 0.01745, 0.00683, -0.01877, -0.0168, 0.02414, 0.0145, -0.01149, -0.02286, 0.02191, 0.00819, -0.01159, -0.00467, 0.01577, 0.01453, -0.01228, -0.01247, 0.02694, 0.00859, -0.00875, -0.00981, -0.01723, -0.01223, 0.01433, 0.01338, 0.01764, 0.01447, -0.01381, -0.02168, -0.01634, -0.00751, 0.01627, 0.0157, 0.01255, 0.0139, -0.01318, -0.00737, 0.01378, 0.01608, 0.00208, -0.01329, 0.01347, 0.00737, -0.00239, -0.00322, 0.01515, 0.00324, -0.00332, -0.00305, -0.00225, -0.00033, -0.01488, -0.01107, 0.00519, 0.0075, -0.00568, -0.00687, 0.00609, 0.00275, -0.00668, -0.0022, -0.00101, -0.00443, 0.0, -0.00286, 0.0088, -0.01732, -0.00455, -0.01737, -0.01773, 0.0177, 0.00722, 0.01448, 0.01066, -0.00415, -0.01805, -0.01107, -0.01893, 0.00952, 0.00693, 0.00279, 0.01168, -0.01277, -0.00665, -0.01333, -0.0074, 0.01362, 0.01631, -0.0009, 0.0157, -0.00641, -0.01472, 0.0106, 0.01295, -0.01481, -0.01326, -0.00973, -0.02479, 0.01759, 0.01562, -0.01495, 0.01412, -0.01714, -0.00957, 0.00956, 0.01273, -0.01454, -0.01396, 0.01581, 0.01662, -0.01757, -0.00915, -0.01071, -0.01552, 0.01652, 0.00875, 0.01669, 0.01392, -0.01245, -0.0131, -0.01196, -0.02116, 0.02162, 0.0165, 0.01785, 0.01763, -0.01474, -0.02265, 0.00908, 0.00469, -0.00323, -0.01791, 0.01529, 0.00392, -0.01466, -0.01257, 0.00715, 0.01009, -0.01823, -0.01169, 0.01338, 0.0126, -0.00813, -0.00855, 0.01256, 0.01409, -0.01453, -0.00744, 0.00108, 0.01047, -0.01857, -0.01342, 0.01535, 0.01513, -0.00526, -0.00771, 0.00891, 0.00869, -0.01583, -0.01557, 0.00759, 0.00753, -0.01177, -0.00908, 0.00265, 0.00767, -0.02006, -0.00496, 0.00124, -0.00804, -0.00897, 0.00714, 0.00346, -0.01156, -0.00031, 0.00194, 0.01735, -0.00267, -0.01469, 0.00203, 0.00316, -0.01688, -0.00559, 0.00069, 0.01059, -0.0147, 0.00055, 0.00308, 0.00467, 0.01267, -0.01108, -0.00751, 0.01047, 0.00482, -0.00107, 0.00327, 0.01544, 0.01175, -0.00136, -0.01328, 0.00248, 0.0071, -0.00978, -0.00717, 0.01103, 0.00626, -0.01413, -0.00523, 0.00619, 0.01007, -0.00321, -0.00154, 0.004, 0.0096, -0.0042, -0.01268, 0.00157, 0.01092, -0.00735, -0.00415, 0.00778, -0.00786, 0.00034, -0.00457, 0.00999, 0.00384, -0.01242, -0.01719, 0.01188, 0.00578, -0.00456, -0.006, 0.00803, 0.00687, -0.01262, -0.00718, -0.00027, 0.0052, -0.00068, 0.00061, 0.01561, 0.01021, -0.01847, -0.00178, 0.01113, -0.00451, -0.00407, -0.02228, 0.01422, 0.00944, -0.01435, -0.01446, -0.00027, -0.00108, -0.00917, -0.0032, -0.00457, -0.00644, 0.00594, 0.00404, -0.01683, -0.01398, 0.00858, 0.00541, -0.00408, -0.00613, 0.00556, -0.00545, -0.01103, -0.00969, 0.00189, -0.00981, -0.01676, -0.00097, -0.00752, -0.00313, -0.00084, -0.01554, -0.00769, -0.00451, -0.00286, -0.00631, -0.00903, 0.00041, -0.00715, -0.00707, -0.01049, -0.00104, -0.01338, -0.01292, -0.01095, -0.01779, -0.00416, 0.00133, -0.01941, -0.01212, -0.0018, -0.00403, -0.02301, -0.01115, -0.0083, 0.00633, -0.0114], ['CWNJ_C02-2FAP-24_1_FCT', '#FF0000', 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_y', 0.118, -0.118, -0.010068, -0.004843, -0.004042, 0.005639, -0.009609, -0.010717, 0.007167, -0.003115, -0.010679, -0.011284, 0.001923, -0.008916, -0.023735, -0.007338, 0.003154, -0.000537, -0.008813, -0.014443, -0.010224, -0.006352, -0.012431, -0.010327, 0.000214, -0.003437, -0.013305, -0.015332, -0.01124, -0.008061, -0.00372, -0.011484, -0.002802, -0.002456, -0.009067, -0.007993, -0.006528, -0.002167, -0.015184, -0.017685, -0.008271, -0.013331, -0.022377, -0.019129, 0.003236, 0.002016, -0.020507, -0.015239, 0.003298, -0.016499, -0.02686, -0.012612, -0.00249, -0.006625, -0.024198, -0.017856, -0.004741, 0.000375, -0.013125, -0.009633, 0.004931, -0.000576, -0.019589, -0.024052, 0.005959, 0.002807, -0.01581, -0.010888, -0.003417, 0.000693, -0.016875, -0.01206, 0.004702, -0.004951, -0.01083, -0.021879, -0.001761, -0.005058, -0.016435, -0.009125, 0.007233, -0.010693, -0.012661, -0.004326, -0.005208, -0.002729, -0.014726, -0.007421, -0.00977, -0.004741, -0.012149, -0.017592, 0.003081, 0.010185, -0.010766, -0.004873, 0.000277, -0.007377, -0.011264, -0.011772, -0.009462, -0.003701, -0.028408, -0.024243, -0.01236, -0.016797, -0.005337, -0.002575, -0.005834, -0.014389, -0.005532, -0.030092, -0.018631, 4.8e-05, 0.001181, -9e-05, -0.004471, -0.013974, -0.007207, 0.001708, -0.001005, -0.018032, -0.014653, -0.002265, 0.006025, -0.021801, -0.020039, 0.002288, -0.006025, -0.006557, -0.015371, -0.020781, -0.013312, -0.007885, -0.006416, -0.01957, -0.012534, 0.003095, -0.007182, -0.022397, -0.014438, -0.010878, -0.008616, -0.012402, -0.003378, -0.008334, -0.004101, -0.006084, -0.007978, -0.005385, -0.015815, 0.00268, -0.003642, -0.016103, -0.02202, -0.015917, -0.007255, 0.00381, -0.021401, -0.015014, 0.005097, -0.001528, -0.016175, -0.007929, 0.006989, -0.000191, -0.006743, 0.001845, -0.003422, -0.025307, -0.016074, -0.007187, -0.002573, -0.023896, -0.02229, -0.009628, 0.000366, -0.024384, -0.021333, -0.000914, -0.008886, -0.014409, -0.01718, -0.018383, -0.012333, -0.002924, 0.008842, -0.023403, -0.011646, -0.018544, -0.001567, 0.003159, -0.005275, -0.019179, -0.020148, 0.001737, -0.006069, -0.022241, -0.016665, -0.001172, -0.000854, -0.008266, -0.019946, -0.014165, -0.014404, -0.003547, -0.011782, -0.015556, -0.016655, -0.004624, -0.009648, -0.019042, -0.017617, -0.001953, 0.003054, -0.008955, -0.019291, -0.016533, -0.019892, -0.003692, -0.004734, -0.000765, 0.000214, -0.01478, -0.027519, 0.004272, 0.002021, -0.023417, -0.022705, -0.01471, -0.024078, 0.002158, 0.004538, -0.000502, 0.002924, -0.025947, -0.01478, -0.023994, -0.018305, -0.010097, -0.008627, -0.002592, 0.002397, -0.020639, -0.014083, -0.007329, -0.010341, -0.016464, -0.011606, -0.009204, 0.001416, 0.001704, -0.037387, -0.021943, 0.004217, 0.00618, -0.018908, -0.022304, -0.015937, -0.014575, -0.011835, -0.007036, -0.005698, -0.012819, -0.017597, -0.014848, 0.001582, -0.002875, -0.014365, -0.022714, -0.001464, 0.001494, -0.020373, -0.01227, -0.002778, 0.002309, -0.01123, -0.013437, -0.019311, -0.011268, -0.002988, 0.010927, -0.000866, -0.001541, -0.022685, -0.014868, 0.001335, 0.009516, -0.023872, -0.011464, -0.002958, 0.002036, -0.022299, -0.025166, -0.003999, -0.009116, -0.006201, -0.018793, -0.023017, -0.025404, 0.003657, -0.004165, -0.005347, -0.003695, -0.013247, -0.01413, -0.003159, -0.00185, -0.017511, -0.014575, -0.01853, -0.006654, -0.010141, -0.004212, 0.00226, -0.003015, -0.030507, -0.01288, 0.002675, 0.003245, -0.009009, -0.022285, 0.001596, 0.004047, -0.019837, -0.000805, -0.005409, -0.001201, -0.010895, -0.007924, 0.004054, 0.000871, -0.018999, -0.007324, -0.003058, -0.008808, -0.023892, -0.013457, 0.005213, -0.00333, -0.011723, -0.010387, -0.001455, 0.001694, -0.0058, -0.004571, 0.001517, 0.001488, -0.008678, -0.007084, -0.008003, -0.007391, -0.010654, -0.006225, -0.005561, 0.002949, -0.014485, -0.015092, 0.000468, 0.000971, -0.01892, -0.017573, -0.002851, -0.003437, -0.011852, -0.017697, -0.001536, 0.003417, -0.015098, -0.015683, -0.003325, -0.006586, -0.007993, -0.010029, 0.001132, -0.003911, -0.014711, -0.015964, -0.017903, -0.006284, 0.006328, 0.002618, -0.004833, -0.00666, -0.017458, -0.012758, 0.006669, -0.00708, -0.024466, -0.011611, 0.002416, -0.00415, -0.014433, -0.003476, -0.009612, -0.02479, -0.006274, 0.001113, -0.003256, -0.018857, -0.011064, 0.004033, 0.003471, -0.017611, -0.022749, 0.001402, 0.004584, -0.021494, -0.008398, -0.000148, -0.001489, -0.012781, -0.015756, -0.002766, -0.000488, -0.022817, -0.013745, -0.002374, -0.008339, -0.0181, -0.017204, -0.009306, -0.008966, -0.001929, 0.002416, -0.008664, 0.007592, -0.0056, -0.006773, 0.002993, -0.005586, -0.018291, -0.012802, -0.015053, -0.012768, -0.00351, -0.002967, 0.000122, 0.00394, -0.010804, -0.007451, -0.01144, -0.00236, -0.010287, -0.01145, -0.009243, -0.01122, -0.005678, -0.004453, -0.001596, 0.003793, -0.011513, -0.00977, 0.003666, 0.000991, -0.019946, -0.013564, 0.002597, 0.00311, -0.013442, -0.021655, -0.003725, 0.001528, 0.005854, -0.021474, -0.010938, -0.005419, -0.002565, -0.02351, -0.011948, 0.002573, -0.000341, -0.01332, -0.004131, -0.013959, -0.008139, 0.00331, -0.000205, -0.002867, -0.016245, -0.012817, -0.003369, -0.010587, -0.016274, -0.00726, -0.000932, -0.008872, -0.014238, -0.01206, -0.005776, -0.000302, -0.01184, -0.015543, -0.001752, -0.001215, -0.012119, -0.008315, -0.002573, 0.001098, -0.018857, -0.011386, -0.000655, -0.009851, -0.017978, -0.01269, -0.005478, -0.001982, -0.008374, -0.011596, -0.00482, -0.004562, -0.010052, -0.00223, 0.004692, -0.000722, -0.008984, -0.007353, -0.002822, 0.003823, -0.010971, -0.00871, 0.000219, -0.004033, -0.024443, 0.001796, -0.003312, 0.00303, -0.020795, -0.011137, -0.004379, -0.001035, -0.016692, -0.011303, 0.009199, -0.001079, -0.021342, 0.000698, -0.004711, -0.004892, -0.008349, 0.001313, -0.004538, -0.016697, -0.008295, 0.007451, 0.002939, -0.015327, -0.015263, 0.002309, -0.0019, -0.005839, -0.017207, 0.006552, -0.001254, -0.017276, -0.005029, 0.015058, 0.003574, -0.01184, -0.009355, -0.00011, 0.000254, -0.021621, -0.025078, -0.001987, -0.00059, -0.01596, -0.012426, -0.001416, -0.009311, -0.014926, -0.027416, -0.007495, -0.014174, -0.023159, -0.009443, 0.000424, 0.00456, -0.017539, -0.021542, -0.000312, -0.001665, -0.008238, -0.010861, 0.001279, -0.006572, -0.003745, -0.006762, -0.006396, -0.019837, -0.018837, -0.017956, -0.013365, -0.002954, -0.010839, -0.012729, -0.000439, -0.01034, -0.012368, -0.01371, 0.009057, -0.020129, -0.017241, 0.004204, -0.009106, -0.012422, -0.012221, 0.002792, 0.004447, -0.010579, -0.005664, -0.006285, -0.009492, -0.016161, -0.018188, -0.01337, -0.004765, -0.006616, -0.00133, 0.006718, -0.001025, -0.011728, -0.012216, 0.001801, -0.007797, -0.013696, -0.011742, -0.012377, -0.007978, 0.004433, -0.009169, 0.000312, -0.003686, -0.014848, -0.007875, 0.008378, 0.001616, -0.023686, -0.013413, 0.000385, -0.00061, -0.017089, -0.0085, 0.00477, 0.003145, -0.019125, -0.010659, -0.003295, -0.000888, -0.016972, -0.007353, 0.002456, -0.01, -0.014125, -0.015797, -0.013979, -0.012077, -0.005883, -0.004492, -0.002919, 0.005771, -0.018061, -0.019469, 0.004528, -0.003374, -0.024555, -0.010136, 0.002241, -0.006606, -0.018217, -0.016422, 0.002565, -0.002068, -0.021001, -0.012133, 0.003369, -0.007499, -0.005708, -0.01124, 0.000532, 0.000986, -0.008491, -0.009404, -0.004165, -0.014145, -0.001069, -0.010791, -0.000981, -0.0037, -0.012812, -0.01581, -0.004853, 0.001123, -0.008227, -0.014062, -0.000312, -0.000136, -0.004116, -0.006655, -0.008611, -0.007124, -0.014467, -0.015302, 0.000847, 0.002294, -0.012846, -0.009228, -0.009365, -0.005717, -0.007426, 0.002622, -0.006669, -0.004552, -0.015352, -0.010922, -0.008876, 0.001742, -0.018256, -0.011103, -0.004291, -0.000689, -0.008511, -0.015043, -0.005551, 0.006194, -0.009257, -0.011381, -0.002977, -0.003063, -0.017989, -0.006186, -0.006635, -0.000426, -0.017167, -0.006414, -0.007429, 0.00226, 1.4e-05, -0.013442, -0.016596, 0.002924, -0.009194, -0.01269, -0.009564, -0.000273, -0.002363, -0.011772, -0.006503, -0.00561, -0.004025, -0.017983, -0.009648, 0.005063, 0.001464, -0.009847, -0.01404, -0.003598, -0.002776, -0.015017, -0.005043, 0.005063, -0.017683, -0.00977, 0.001372, -0.002111, -0.021376, -0.012329, -0.006005, -0.001586, -0.001079, -0.00288, 0.003283, -0.010234, -0.01976, -0.013535, -0.015429, -0.010195, 0.004755, -0.011035, -0.001132, -0.006079, -0.005991, -0.004687, -0.002197, 0.000415, -0.020019, -0.010273, -0.002094, 0.003833, -0.003037, -0.005537, 0.002446, -0.003515, -0.007739, -0.020143, -0.00943, -0.006914, -0.000512, 0.001541, -0.019956, -0.005739, -0.006538, 0.001513, -0.012812, -0.018842, -0.010071, -0.010766, 0.000622, 0.00228, -0.001669, -0.000224, -0.011577, -0.023657, 0.000395, -0.006298, -0.016323, -0.013178, 0.002939, -0.000288, -0.014289, -0.006902, 0.005782, 0.007143, -0.008388, 0.002829, -0.003632, -0.015644, -0.011191, -0.004414, -0.003027, -0.006093, -0.009765, -0.001148, 0.009326, -0.011083, -0.005429, -0.01517, -0.00168, -0.003359, -0.001425, 0.001367, -0.002101, -0.011635, -0.003896, 0.003949, 0.000302, -0.00852, -0.001186, -0.002075, -0.016289, -0.012753, -0.003437, -0.003599, -0.015688, -0.018393, -0.006304, -0.00541, -0.016684, -0.008984, 0.006275, 0.005209, -0.013466, -0.024687, -0.002099, -0.009848, -0.013837, -0.022341, -0.001391, -0.004936, -0.022866, -0.012421, 0.008466, -0.000732, -0.011826, -0.006621, -0.001871, -0.001518], ['CWNJ_C02-2FAP-23_1_FCT', '#008000', 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_y', 0.118, -0.118, -0.01208, 0.01629, 0.00439, -0.00563, -0.00374, -0.00868, -0.00249, -0.00036, -0.01437, -0.00595, -0.00152, -0.00438, -0.0061, -0.00155, -0.00046, -0.0032, -0.00619, 0.00828, 0.00406, 0.00445, -0.00242, -0.00057, 0.00911, 0.00484, 0.0027, 0.00227, 0.00525, -0.01162, 0.0016, 0.00445, -0.00802, -0.0003, -0.00238, -0.00421, -0.00148, -0.00118, -0.00233, -0.01256, 0.00761, -0.00925, 0.00124, -0.0005, -0.00354, -0.00506, -0.00277, -0.00698, -0.00202, -0.00216, 0.0008, -0.00145, 0.00418, -0.00217, -0.00261, -0.01398, 0.00403, -0.00326, -0.0086, -0.00488, -0.00439, -0.00569, -0.00435, 0.00355, 0.0016, -0.0019, -0.01115, -0.00725, 0.00658, 0.00019, 0.00837, -0.009, -0.00431, -0.01104, 0.00252, -0.00045, 0.01141, -0.00404, -0.01001, -0.00249, -0.00256, 0.00445, -0.0005, 0.00467, -0.00498, -0.00296, 0.00241, -0.00338, 0.00327, -0.00375, 0.00767, -0.0092, -0.00665, -0.01129, 0.00666, 0.00271, -0.00211, -0.00201, -0.00416, -0.0037, -0.00033, 0.0025, -0.00691, -0.00343, -0.01024, 0.00528, 0.00584, -0.0034, 0.00198, -0.00114, -0.00769, 0.00436, -0.00082, -0.00459, 0.0019, -0.00674, -0.00073, 0.00198, 0.00113, -0.00857, 0.00348, -0.00567, -0.0057, -0.00216, 0.00738, 0.00548, 0.00574, 0.00275, -0.0068, 0.00764, 0.00745, 0.01188, -0.00498, 0.00324, 0.0019, -0.01285, 0.00054, -0.00018, 0.00439, 0.00899, 0.00026, 0.00508, -0.00544, 0.00895, 0.00256, -0.00137, -0.00452, -0.00667, -0.00497, -0.0048, -0.00474, -0.00146, 0.00642, -0.00311, -0.00083, -0.00122, 0.01127, -0.00249, -0.00713, -0.01089, 0.00358, -0.01884, -0.00503, -0.00348, -0.00185, 0.00554, 0.00464, -0.0115, -0.00808, 0.00082, -0.00056, -0.00957, -0.00656, -0.00449, 0.00563, -0.0067, 0.00743, -0.01106, -0.00354, -0.01586, -0.00038, -0.00199, 0.00084, 0.00194, 0.00033, -0.00694, -0.00057, 0.00024, 0.00876, -0.00127, 0.00477, -0.00748, 0.00138, -0.01269, -0.00904, 0.00631, -0.00123, -0.01192, 0.00465, 0.00052, -0.00661, -0.00556, -0.00354, 0.0035, 0.00253, 0.00086, 0.00858, 0.00574, -0.00482, -0.00548, -0.00406, 0.00606, -0.00118, 0.00492, 0.00975, -0.00562, 0.0022, -0.00498, -0.00774, -0.00027, -0.00178, -0.00598, -0.00281, -0.00067, -0.00539, 0.00128, 0.0, -0.01414, -0.00643, 0.00778, -0.00408, 0.00922, 0.00419, 0.00121, 0.00141, 0.00221, 0.00749, -0.00691, 0.00205, 0.00015, 0.00522, 0.00241, 0.00106, -0.01122, -0.01, 0.00042, 0.00074, -0.00863, 0.00489, -0.00523, 0.0, -0.00514, -0.00202, 0.0087, 0.00755, 0.00623, 0.00181, -0.00987, -0.0013, 0.00783, 0.0061, -0.00341, 0.0051, -0.00096, -0.00235, -0.00034, 0.00733, -0.01231, -0.00252, -0.00503, 0.00441, 0.00476, -0.00652, -0.00355, 0.00357, -0.0042, 0.00237, 0.00381, -0.01215, -0.00281, -0.01696, -0.00456, 0.00306, -0.00297, 0.00462, 0.00133, -0.00733, -0.00154, -0.00354, -0.00018, -0.00084, 0.0, 0.00195, -0.00608, 0.00538, -0.00438, -0.01031, -0.00408, -0.0017, 0.00656, -0.00086, -0.00096, 0.0054, 0.00117, -0.00645, 0.00345, -0.00372, 0.00317, 0.00337, 0.00366, -0.00827, -0.0052, 0.00142, -0.00512, -0.00344, -0.00771, 0.00166, -0.00467, -0.00296, -0.00175, -0.00181, 0.00787, -0.00041, -0.00524, 0.00171, 0.00415, 0.00442, 0.0091, 0.00169, 0.00191, 0.00164, -0.00515, -0.00532, -0.0091, -0.009, 0.00064, 0.00095, -0.00968, -0.00989, 0.00681, -0.00717, -0.01132, 0.0017, -0.00549, 0.01664, 0.00072, 0.0023, -0.00154, 0.00233, 0.00708, -0.00399, -0.00688, -0.00166, -0.00228, -0.00364, -0.00018, 0.00736, 0.00051, 0.00054, -0.01047, -0.00072, 0.00662, -0.00117, -0.00655, -0.00811, -0.00242, -0.00011, -0.00873, -0.00275, 0.00731, -0.00434, -0.00321, -0.0043, 0.00209, -0.01229, -0.00402, -0.00642, -0.00732, -0.01025, -0.00698, 0.001, 0.00473, -0.01367, 0.00272, 0.0013, -0.00344, 0.00368, -0.00279, 0.00277, 0.00368, 0.00839, 0.00503, 0.0011, -0.02952, -0.00227, 0.00298, 0.01251, -0.00931, 0.00173, -0.00471, -0.00222, 0.0073, -0.02232, -0.00338, -0.00082, -0.00846, -0.00797, -0.00283, -0.01163, 0.00124, 0.00081, -0.00348, -0.00034, 0.00433, -0.00255, -0.00177, -0.00658, -0.0021, 0.0044, -0.00614, -0.01655, 0.00028, 0.00175, 0.0116, -0.00843, -0.00519, -0.00198, 0.00236, -0.00547, -0.01008, -0.00717, -0.00412, -0.00161, 0.00615, -0.00569, -0.00419, 0.00203, -0.0003, -0.00841, -0.00138, -0.00411, -0.01103, -0.00131, -0.00106, 0.00828, 0.00847, 0.00261, 0.00162, 0.0026, 0.00303, -0.00662, 0.00505, 0.00536, 0.00743, -0.00288, -0.00264, -0.00109, -0.00333, -0.00041, -0.00316, -0.00462, 0.0041, -0.00194, 0.00041, 0.00056, 0.00508, -0.00048, -0.00266, 0.00022, -0.00284, 0.01149, 0.0032, 0.00562, 0.00442, 0.00153, -0.00018, -0.00247, -0.00339, -0.00036, -0.00624, -0.00296, -0.00334, -0.00209, -0.00399, -0.00916, -0.00709, 0.00012, 0.00596, -0.00597, 0.00505, -0.00375, -0.00355, -0.00193, -0.00505, 0.00433, 0.00189, 0.00025, -0.0019, -0.00129, -0.00462, -0.01412, 0.00079, 0.00439, -0.01506, -0.00196, 0.00758, -0.00287, 0.00713, -0.00592, -0.00106, -0.00806, -0.00529, 0.00229, -0.00107, -0.00825, -0.00164, 0.0, -0.003, -0.00772, -0.00809, -0.00405, 0.00342, 0.00733, -0.00497, 0.0, 0.00504, -0.00147, -0.00683, 0.00275, 0.00376, -0.00366, -0.00357, 0.00781, 0.01318, -0.00198, -0.00513, 0.00348, -0.00543, -0.00776, 0.00422, -0.00184, -0.00522, -0.00159, 0.00069, -0.00782, -0.00285, -0.00167, -0.00729, -0.00354, -0.00323, 0.00087, 0.00703, 0.00217, -0.00023, -0.00324, -0.0014, -0.00334, -0.00283, 0.00597, 0.00688, 0.00247, -0.00524, -0.01168, -0.00797, 0.00582, -0.00916, -0.0102, -0.00901, 0.00737, 0.00628, 0.00177, -0.006, 0.00549, -0.00614, -0.01483, -0.00034, 0.01229, -0.0051, -0.00613, 0.00544, -0.00618, -0.00297, -0.00677, -0.00362, 0.00505, 0.0, 0.00127, -0.01243, -0.00886, -0.01075, -0.01057, -0.00466, -0.00348, 0.00145, 0.00017, -0.00766, -0.00747, -0.00374, -0.00515, -0.01826, -0.01083, 0.00487, -0.00121, -0.01431, -0.00646, -0.00879, -0.00482, -0.01233, -0.01101, -0.01042, -0.00631, -0.01294, -0.0118, -0.01105, -0.00944, -0.00749, -0.00969, -0.00294, -0.00376, -0.01165, -0.0005, -0.01192, -0.00175, 0.00432, -0.00347, -0.01584, -0.00557, 0.00304, -0.00307, 0.00063, -0.00229, -0.00291, -0.006, -0.00782, 0.00365, 0.0049, -0.01185, -0.02077, -0.00415, -0.00097, -0.019, -0.01401, -0.00716, 0.00234, -0.01634, -0.01902, -0.00298, -0.00501, -0.0065, -0.02269, -0.00674, -0.00289, -0.01247, -0.01099, -0.00439, 0.0084, -0.01665, -0.01295, 0.0007, -0.00109, -0.01376, -0.01102, -0.00937, -0.01284, -0.00348, 0.00048, -0.01841, -0.02052, -0.01211, -0.01396, -0.0105, -0.00779, -0.007, -0.00686, -0.00914, -0.01175, -0.00305, -0.01004, -0.01483, -0.00501, 0.00753, -0.00645, -0.00149, 0.00746, -0.00122, -0.01047, -0.00546, 0.00353, -0.00797, -0.00433, -0.01687, -0.00664, -0.00736, -0.01477, -0.00545, -0.00351, -0.0052, -0.0146, -0.00654, -0.01326, -0.01086, -0.01035, -0.01572, -0.00703, -0.019, -0.00438, -0.01235, -0.01217, -0.00529, -0.01931, -0.01358, -0.00311, -0.00954, -0.0182, 0.0, -0.00241, -0.0066, -0.01176, -0.0141, 0.00429, -0.01271, -0.01196, -0.00686, -0.0029, -0.01433, -0.01169, -0.00337, 0.0004, -0.01598, -0.00391, -0.00696, 0.00163, -0.01377, -0.01252, -0.00917, -0.01135, -0.00224, -0.01518, -0.01783, -0.00863, -0.01388, -0.00758, -0.00608, -0.00988, -0.01289, -0.00249, -0.00261, -0.00395, -0.00738, -0.0077, -0.00575, -0.00944, 0.00245, -0.00371, -0.00855, -0.01358, 0.00636, -0.00126, -0.008, -0.02402, -0.01654, -0.0049, -0.00984, -0.00561, -0.0032, -0.00769, -0.01344, -0.01572, -0.00202, -0.01666, -0.00812, -0.00627, -0.00432, -0.00947, -0.0073, 0.00212, -0.00766, -0.0147, -0.00743, 0.00158, -0.00055, -0.00963, -0.00568, -0.01055, -0.01272, -0.01839, -0.00076, -0.01046, -0.00372, -0.01129, -0.01126, -0.01254, -0.01, -0.00392, -0.00487, -0.0098, -0.00057, -0.01797, -0.01439, -0.00423, -0.01425, -0.01374, -0.00223], ['CWNJ_C02-2FAP-24_2_FCT', '#00FFFF', 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_y', 0.118, -0.118, -0.013793, -0.000385, -0.004653, -0.002612, 0.0079, 0.00224, -0.001425, -0.007597, -0.004185, -0.004736, -0.004379, -0.014423, -0.001523, -0.010976, 0.004218, -0.00684, -0.011086, -0.004126, -0.011733, -0.013398, -0.01019, -0.013351, -0.005512, -0.015385, 0.002364, -0.006826, 0.003305, -0.006438, -0.008233, -0.007666, -0.006491, 0.001181, -0.00894, -0.005878, 0.010068, -3.9e-05, -0.007113, -0.016314, -0.007495, -0.012255, -0.004863, 0.001025, 0.002929, -0.002578, -0.004726, -0.001733, -0.006773, -0.010546, 0.000976, -0.00746, -0.005541, -0.001206, -0.009918, 0.004848, 0.002333, -0.007011, -0.001181, 0.003476, -0.011982, -0.012163, 0.00144, -0.006907, -0.010063, -0.006796, -0.004169, -0.01269, -0.014763, -0.009794, -0.003095, -0.010737, -0.003916, -0.013815, -0.009789, -0.005126, -0.012011, -0.00436, -0.004715, 0.002637, -0.005815, -0.004922, -0.001196, -0.007171, -0.017491, 0.001284, -0.009018, 0.001732, -0.014394, -0.002089, -0.005126, -0.0102, -0.015434, -0.009114, -0.00528, -0.010078, 0.005122, -0.001181, -0.014658, -0.002766, 0.000961, 0.002537, -0.003002, -0.01178, -0.00458, -0.005854, 0.000366, -0.01392, -0.014248, -0.006074, -0.013881, -0.017963, -0.004824, -0.012373, -0.001567, -0.009334, -0.005073, -0.003427, 0.003632, -0.005471, 0.001142, 0.000419, -0.004477, -0.002626, -0.008857, -0.012671, -0.013481, -0.01038, -0.005864, -0.000406, -0.008085, -0.017314, -0.005786, -0.005996, -0.008579, -0.001743, -0.007211, -0.008066, -0.00725, -0.00434, -0.018608, -0.005428, -0.0077, -0.009497, -0.007651, -0.008994, -0.008632, 0.002895, -0.008168, -0.012432, -0.006743, -0.000607, 0.001958, -0.000161, -0.001206, -0.006557, -0.01435, -0.00516, -0.00944, -0.018779, -0.008918, -0.002583, -0.011347, -0.003876, -0.011694, -0.011311, -0.0181, -0.00123, -0.010917, -0.006562, -0.000263, 0.006941, -0.001533, -0.00373, -0.001546, -0.006689, 0.001292, -0.000888, -0.012382, -0.000283, -0.0033, 0.000141, -0.010751, -0.010141, -0.006098, -0.006649, 0.008593, -0.00059, -0.004951, -0.007209, -0.015585, -0.017113, 0.004257, 0.000646, -0.007397, -0.015219, -0.002312, -0.008984, -0.002109, 0.00206, -0.010263, -0.008579, -0.004054, -0.005045, 0.000373, -0.00664, -0.009531, -0.011196, -0.008195, -0.007094, -0.016547, -0.015979, -0.003964, -0.006079, -0.010976, -0.006401, -0.001362, -0.016367, -0.014571, -0.013623, 0.0006, -0.012192, -0.003853, -0.000859, -0.001743, -0.003652, -0.006218, -0.00517, -0.00019, -0.002338, 0.002919, -0.003657, -0.009609, 0.001679, 0.002646, -0.013158, -0.000114, -0.005107, -0.003232, -0.003798, -0.00665, -0.019449, -0.005126, -0.006515, -0.004628, -0.010038, 0.001665, 0.008212, -0.004155, -0.017099, -0.007939, -0.004413, -0.008198, 0.003379, 0.001049, 0.004571, -0.011225, -0.018203, -0.005034, -0.00581, -0.007285, -0.006098, -0.007519, -0.005371, 0.0046, 0.003476, 0.000378, -0.001855, -0.01119, -0.012016, -0.000689, 0.000421, -0.004973, -0.003486, -0.001391, -0.008789, -0.012368, -0.000351, -0.007519, -0.003017, -0.005639, 0.002805, -0.008779, -0.013735, -0.009531, -0.006036, -0.012041, 0.005351, -0.000156, -0.010581, -0.002807, -0.003738, -0.00373, -0.002264, 0.000339, -0.009614, -0.007443, -0.009243, -0.001053, -0.013359, -0.003354, 0.001943, -0.004672, -0.013303, -0.009638, -0.013833, 0.004316, -0.001088, 0.00104, -0.015566, -0.006269, -0.001835, -0.012568, -0.004506, -0.000903, -0.01059, -0.00166, -0.007861, 0.000185, -0.000639, -0.004858, -0.008505, -0.010948, -0.011843, 0.000922, -0.001608, -0.003608, -0.007949, 0.002168, -0.002723, -0.016256, 0.002543, -0.003973, -0.008544, -0.003857, -0.011274, -0.021186, 0.001823, -0.001274, 0.001318, -0.00352, 0.00414, -0.006235, 0.003237, -0.013279, -0.012329, -0.008798, 0.001875, 0.003432, -0.00726, -0.012321, -0.002573, -0.001321, -0.019287, -0.005712, -0.004839, -0.002109, -0.012377, 0.001323, -0.00144, -0.003251, -0.008027, -0.00206, -0.009727, -0.00957, -0.000167, -0.010229, -0.006679, -0.002592, 0.000166, -0.012006, -0.003046, -0.001328, -0.000693, -0.002273, -0.000986, -0.004487, 0.009689, -0.010444, -0.002392, -0.001748, -0.00371, -0.008906, 0.002548, -0.001756, -0.005893, -0.008549, -0.001622, -0.005864, -0.005932, -0.002519, -0.001498, -0.00651, -0.010957, -0.001431, 0.002875, 0.000634, -0.000439, 0.002814, -0.002397, -0.005234, -0.009784, -0.013984, 0.004136, -0.012617, -0.017695, -0.006977, -0.005849, -0.008593, -0.006928, -0.004052, -0.005727, -0.008114, -0.010014, -0.008691, -0.002905, -0.005538, 0.002944, -0.003011, -0.001826, -0.013633, -0.00133, 0.001035, -0.010004, 0.006215, 0.00019, -0.003608, 0.004101, -0.00561, -0.009389, -0.004165, 0.001821, -0.007617, -0.009926, -0.007241, -0.003198, 0.004555, -0.002377, -0.011263, 0.000317, -0.010844, -0.000483, 0.001396, -0.005677, 0.004016, -0.002699, -0.005165, 0.000927, 0.00359, -0.006767, -0.003496, -0.003652, -0.01225, -0.009477, -0.0058, -0.006206, -0.002905, -0.002065, -0.008382, -0.000791, 0.003046, -0.007446, 0.000639, -0.003714, 0.001105, -0.009885, -0.006015, -0.007558, -0.010311, -0.009497, -0.004863, -0.008022, -0.000634, 0.007769, 0.001806, -0.010043, -0.009077, -0.008516, -0.008969, -0.004586, -0.013515, 0.007405, -0.007075, -0.01623, -0.008803, -0.003344, 0.001225, -0.000938, -0.006472, 0.011059, -0.009223, -0.006845, -1.4e-05, -0.002314, -0.005371, -0.009057, -0.007358, -0.003891, -0.007458, -0.003403, -0.003494, 0.000825, -0.010395, -0.002556, -0.004995, -0.00707, -0.00124, 5.3e-05, -0.016831, -0.003022, -0.007869, 0.008047, -0.012148, -0.005232, 0.000181, -0.003803, -0.002829, -0.01392, -0.003432, -0.010747, -0.010282, -0.005874, -0.001948, -0.005747, 0.000488, -0.002431, -0.011308, -0.002101, -0.00642, -0.009824, -0.012216, 0.001054, -0.004609, -0.001665, -0.009219, -0.003754, -0.005423, -0.008149, -0.003881, -0.004458, -0.002089, -0.002221, -0.008564, -0.009746, -0.018238, -0.008707, -0.001264, -0.005712, -0.001313, -0.001157, -0.007993, -0.009375, -0.000175, -0.015395, -0.00124, -0.01168, -0.007597, -0.009218, -0.007646, -0.009487, -0.004741, -0.006855, -0.008671, -0.002524, -0.006347, -0.002578, -0.009946, -0.009584, -0.001293, -0.005751, -0.003236, -0.001929, -0.012578, 0.002412, -0.003632, 5.3e-05, -0.004301, -0.009951, -0.021474, -0.002294, -0.004946, -0.008359, -0.0082, -0.005356, -0.006474, 0.000654, -0.005649, -0.001372, 0.008291, -0.015473, -0.004013, -0.004609, -0.003261, -0.006494, -0.005653, -0.010311, -0.002827, -0.012041, 0.003317, -0.002421, -0.002958, -0.005004, -0.001918, 0.008925, -0.014921, -0.007281, -0.00812, -0.003266, -0.000629, -0.00561, -0.012543, -0.003305, -0.006611, 0.000253, 0.001328, 0.001811, -0.009665, -0.00225, -0.003535, -0.00686, -0.013037, -0.010737, -0.003535, -0.007929, -0.004814, -0.002656, -0.003471, -0.006425, -0.003515, -0.00017, -0.006235, -0.004547, -0.003291, -0.016161, -0.001464, -0.006049, -0.003422, 0.00289, -0.004799, 0.00143, -0.000214, 0.009975, 0.000766, -0.006904, -0.00456, -0.001289, 0.003857, -0.004902, 0.010336, -0.012709, -0.004833], ['CWNJ_C02-2F-REL01_1_FCT', '#9400D3', 'Accelerometer AVG-FS8g_ODR100HZ_Zup accel_only_average_y', 0.118, -0.118, -0.0128, -0.00623, -0.01704, -0.01349, -0.01247, -0.00883, -0.00771, -0.00404, -0.01261, -0.00529, -0.01706, 0.00519, -0.00298, 0.02047, -0.01036, 0.01062, -0.00476, -0.00049, -0.0142, 0.00431, -0.01226, -0.00427, 0.00933, -0.00331, -0.0145, -0.00933, 0.01094, 0.01213, -0.01819, 0.00512, -0.00716, -0.00588, -0.01296, 0.00798, 0.00368, 0.01265, 0.00513, -0.00724, -0.01116, -0.00504, -0.00695, 0.01174, -0.00573, 0.01672, 0.00754, 0.00948, 0.00725, -0.0072, -0.01835, 0.00584, -0.00811, 0.00772, 0.00133, -0.01023, 0.00561, 0.00893, -0.01231, 0.00444, -0.01554, -0.00111, 0.00478, -0.00783, -0.01557, 0.00502, 0.00161, 0.00255, -0.01042, 0.00812, 0.00552, -0.00934, 0.0165, 0.00071, -0.01304, 0.00768, 0.00991, 0.01338, 0.0117, -0.0014, 0.01098, 0.01338, -0.01056, -0.00927, 0.01071, 0.01315, -0.0015, 0.00108, -0.01229, 0.00695, 0.00155, 0.00396, -0.00082, 0.00117, -0.015, -0.01232, 0.00396, -0.00912, -0.01107, 0.00876, -0.0057, 0.0, -0.01414, -0.00461, -0.00164, 0.00302, -0.02617, -0.01916, -0.01395, 0.0072, -0.01186, 0.00432, -0.00687, -0.00046, -0.00021, 0.01353, -0.01379, 0.00571, -0.00318, -0.00079, -0.00057, -0.00952, -0.01015, -0.00376, -0.00767, 0.00812, -0.01252, -0.01104, -0.00774, 0.00293, -0.00362, 0.00722, 0.00975, 0.01004, -0.00792, -0.00401, -0.00455, 0.01658, -0.01508, -0.01256, -0.00733, 0.01782, -0.0095, 0.00489, -0.01049, -0.00714, -0.01217, -0.01153, 0.01158, -0.00144, -0.0009, 0.0026, -0.01518, -0.01193, 0.00575, 0.00808, -0.00606, 0.0, -0.00042, 0.00476, 0.00726, -0.00438, -0.01125, -0.01127, -0.02003, -0.01584, -0.00914, -0.00377, -0.0142, 0.00266, -0.0144, 0.00602, 0.00204, 0.0068, 0.00103, -0.01132, -0.00438, 0.01863, 0.00033, -0.01041, -0.01244, 0.00781, -0.01785, -0.01449, 0.00563, 0.00838, -0.00349, -0.01639, 0.01594, 0.00305, 0.01875, -0.0097, -0.00941, -0.00385, -0.00401, 0.00678, -0.00492, -0.01185, -0.0088, -0.00818, 0.00856, 0.00662, -0.00181, 0.0068, 0.0113, 0.00115, -0.01072, -0.01587, 0.00551, -0.01029, -0.00784, -0.00781, -0.01017, 0.0006, 0.00604, 0.00884, 0.01268, -0.01208, 0.00527, 0.00588, 0.00804, -0.01359, 0.00562, 0.00589, -0.00586, 0.00382, -0.0069, -0.00698, 0.01182, -0.01279, 0.00595, 0.0086, -0.0027, -0.01367, -0.00947, -0.00263, 0.00249, -0.01309, 0.00277, -0.01518, -0.00251, -0.01234, 0.01074, -0.01082, 0.00488, -0.01546, -0.01041, -0.00975, -0.00444, -0.00933, -0.00942, 0.00716, 0.01354, 0.00854, 0.01346, -0.01293, 0.00355, -0.01931, -0.00557, -0.00648, -0.00457, 0.00097, 0.00456, 0.01033, -0.0033, -0.00355, 0.01416, -0.01178, -0.01249, -0.01433, 0.00736, -0.01563, 0.00472, -0.00189, 0.01659, -0.00308, 0.01798, -0.0123, -0.00457, 0.00401, 0.00688, 0.00697, 0.01055, -0.00192, -0.01079, -0.01273, -0.00621, -0.01114, 0.01227, 0.00349, -0.00955, 0.0042, -0.00313, -0.01222, -0.01082, -0.01539, -0.01456, -0.01518, 0.00604, -0.00169, 0.00228, -0.01093, 0.01115, -0.00835, 0.01268, 0.00557, 0.00982, -0.00279, 0.01643, -0.00817, -0.00229, -0.00606, 0.00982, -0.00139, 0.00472, -0.00966, 0.00315, -0.00635, -0.00677, 0.00113, -0.00388, -0.01861, -0.01122, 0.00983, 0.00692, -0.00489, -0.0133, 0.00817, -0.01433, 0.00618, -0.00902, -0.00621, -0.00961, -0.00677, 0.00615, 0.00922, -0.01663, 0.00297, -0.01081, 0.00569, -0.0184, -0.01136, -0.0175, 0.0013, -0.00743, -0.00225, -0.00476, -0.00487, 0.00659, 0.01373, -0.00826, 0.00774, -0.0018, -0.00099, -0.01198, -0.0046, 0.00135, 0.00476, -0.01167, -0.01305, 0.00643, -0.01133, -0.00133, -0.00504, -0.00788, -0.00927, 0.0054, -0.01251, -0.01032, -0.01023, -0.01076, -0.0121, 0.00892, -0.00162, -0.00636, -0.01148, -0.00744, -0.01505, -0.01577, 0.0027, 0.00216, -0.00661, -0.00705, -0.00738, -0.0006, -0.00403, -0.00648, -0.00503, 0.00701, -0.01266, -0.01172, 0.00442, -0.00723, -0.00914, -0.00072, -0.01448, -0.01414, -0.01238, 0.01096, 0.00349, 0.00377, -0.00193, 0.01497, -0.00782, -0.01155, 0.00181, -0.01159, -0.01017, 0.00387, -0.01218, -0.01328, 0.0064, -0.01331, -0.00834, -0.01159, -0.01053, -0.01453, 0.00593, -0.01542, -0.0057, -0.0086, 0.00233, -0.01363, -0.01779, -0.00094, -0.00075, 0.00578, -0.01994, -0.01385, -0.01964, -0.00207, 0.01314, 0.01568, -0.01423, -0.00746, -0.01732, -0.00559, 0.00359, -0.0001, 0.00052, -0.01208, -0.01365, -0.01464, -0.01655, -0.02025, -0.01238, -0.00948, 0.00521, -0.00757, -0.0047, -0.00652, -0.01231, 0.00925, 0.0062, 0.01217, -0.01429, -0.01042, -0.00145, -0.00879, -0.01361, 0.00651, -0.00613, -0.00892, -0.00451, -0.01997, -0.00742, 0.00773, -0.01103, 0.00328, -0.00295, 0.00211, 0.00084, 0.00822, -0.02271, 0.00168, -0.01284, -0.02105, -0.0127, -0.01466, -0.00913, -0.01466, -0.00968, 0.00866, 0.00894, -0.00106, -0.00589, -0.00978, -0.0082, -0.01072, 0.00416, -0.00564, -0.01294, -0.01143, -0.00992, -0.00991, -0.01152, 0.00233, -0.00428, 0.00434, -0.00756, 0.00846, 0.00071, 0.00831, 0.0121, -0.00039, -0.01177, -0.01331, -0.01348, 0.00077, 0.01022, 0.01055, 0.00053, 0.00127, -0.01282, -0.01394, 0.00352, 0.00562, -0.0003, 0.00972, -0.00115, -0.01284, -0.00805, -0.01432, -0.00932, -0.00153, -0.03988, 0.00274, -0.01153, -0.016, 0.00113, -0.00851, 0.01053, 0.00764, -0.01857, -0.00617, 0.0044, 0.00664, 0.00205, 0.01155, 0.00136, -0.01342, -0.00851, 0.00981, 0.01046, -0.01151, -0.0111, 0.0045, -0.00481, -0.00366, -0.0129, 0.00075, -0.0114, -0.00944, -0.0103, 0.00135, 0.00451, 0.00403, -0.01112, -0.00925, 0.00475, 0.00582, -0.00562, -0.0102, -0.00742, 0.0009, -0.01134, -0.01551, -0.0089, 0.00545, 0.01304, -0.00014, -0.013, -0.01018, -0.01126, -0.00707, -0.00915, -0.00948, -0.00023, -0.01616, -0.00985, 0.00852, 0.00367, 0.0028, 0.00933, -0.0108, 0.00436, -0.01551, -0.00096, -0.00583, 0.00188, 0.0096, 0.0069, -0.00168, -0.00899, 0.00235, 0.0053, 0.00688, 0.00264, -0.00853, -0.01204, 0.00616, -0.00717, -0.00297, 0.00231, -0.00592, 0.00092, -0.01574, -0.00786, 0.00368, 0.00449, 0.00936, 0.00133, -0.01327, -0.00776, -0.01451, -0.01802, -0.00735, 0.00544, -0.00624, -0.01543, -0.00873, 0.00549, 0.00541, -0.00109, 0.00219, 0.01081, 0.00082, -0.01423, 0.01225, -0.01077, -0.01086, 0.00734, -0.01147, -0.00931, 0.01072, -0.00044, 0.00813, -0.01622, 0.00481, -0.01533, -0.01575, -0.00994, -0.00854, 0.00713, 0.00367]]]
    # table_category_data = []#it's empty when color_by == 'Off'

    new_y_lsl = ''
    new_y_usl = ''
    new_x_lsl = ''
    new_x_usl = ''
    start_time_first = '2020/3/11 14:16'
    start_time_last = '2020/3/26 9:58'
    pic_path = '/Users/rex/Desktop/CPK_Log/'
    correlation_plot(table_data,table_category_data,pic_path,start_time_first,start_time_last,new_y_lsl,new_y_usl,new_x_lsl,new_x_usl)

    #------------------------ debug correlation_plot end --------------------------------


