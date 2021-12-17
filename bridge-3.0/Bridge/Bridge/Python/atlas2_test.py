#! /usr/bin/env python3
# --*-- coding: utf-8 ---*---


import copy
import json
errorMsgs = ''
try:
    import sys,os,time,math,re
except Exception as e:
    print('e---->',e)
    errorMsgs = str(e)+'\r\n'

BASE_DIR=os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0,BASE_DIR+'/site-packages/')


try:
    import datetime
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

try:
    import pytz
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

try:       
    from pytz import timezone
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'
try:
    import time
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'
try:    
    import threading
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'



try:
    import csv
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ---->matplotlib')
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> matplotlib.colors')
try:
    import matplotlib.colors as colors
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> FontProperties')
try:
    from matplotlib.font_manager import FontProperties
except Exception as e:
    print('e---->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'


# print('python import ----> numpy')
try:
    import numpy as np
except Exception as e:
    print('e--->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> pandas')
try:
    import pandas as pd
except Exception as e:
    print('e--->',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> openpyxl')
try:
    import openpyxl
except Exception as e:
    print('import openpyxl error:',e) 
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> xlsxwriter')
try:
    import xlsxwriter
except Exception as e:
    print('import xlsxwriter error:',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> diptest')
try:
    import diptest
except Exception as e:
    print('import diptest error:',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> zmg')    
try:
    import zmq
except Exception as e:
    print('import zmq error:',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

# print('python import ----> redis')
try:
    import redis
except Exception as e:
    print('import redis error:',e)
    errorMsgs = errorMsgs + str(e) +'\r\n'

print('check for error Msgs module:',errorMsgs)
userdocuments = os.path.join(os.path.expanduser("~"), 'Documents')
with open(userdocuments + '/.errormodule.txt', 'w') as file_obj:
    file_obj.write(errorMsgs)

print(sys.getdefaultencoding())


redisClient = redis.Redis(host='localhost', port=6379, db=0)

context = zmq.Context()
socket = context.socket(zmq.REP)
socket.setsockopt(zmq.LINGER,0)
socket.bind("tcp://127.0.0.1:3310")

filelogname = '/tmp/CPK_Log/temp/.logcpk.txt'


from post import *

import operator

g_my_datas = {}

g_box_datas = None

def get_redis_data(zmqMsg):
    tb = redisClient.get(zmqMsg)
    tb_data=[]
    if tb:
        tb=tb.decode('utf-8')
        tb=tb.split("\n")
        tb=(tb[1:-1])   #去掉数据库首尾元素
        for i in tb:
            k=re.sub('\"','',i)  #去掉数据库引号
            h=re.sub(',','',k)   #去掉数据库逗号
            m=h.strip()          #去掉数据库首尾空白
            if is_number(m):
                tb_data.append(eval(m))   #去掉数字的引号
            else:
                tb_data.append(m)
    else:
        tb_data.append('')
    return tb_data

# ------------------------------------------------------------------------------------------------------------------
# 公共方法
# 毫秒级时间转为时间戳，方便时间做运算
def time_data(time_str):
    # 定义传入时间的格式
    if '-' in time_str:
        datetime_obj=datetime.datetime.strptime(time_str,'%Y-%m-%d %H:%M:%S.%f')
    elif '/' in time_str:
        datetime_obj=datetime.datetime.strptime(time_str,'%m/%d/%Y %H:%M:%S.%f')
    time_stamp=int(time.mktime(datetime_obj.timetuple())*1000.0+datetime_obj.microsecond/1000.0)
    return time_stamp
#删除特殊字符，以防乱码的存在
def remove_special_characters(strfile):
    word = ''
    for i in strfile:
        if ord(i) >= 32 and ord(i) <= 127 or i == '\n' or i == '\r':
            word = word + i
    return word

tm=re.compile(r'\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\.\d{3}')
tech=re.compile(r'\w+?(?=[\r\n])')
cov=re.compile(r'(?<===SubTest: ).*?(?=[\r\n])')
sub=re.compile(r'(?<=SubSubTest: ).*?(?=[\r\n])')

def add_cov_2b(coverage,my_dict):
    if '_' in coverage:
            coverage_list=coverage.split('_',1)
            my_dict['Coverage_Tag']=coverage_list[0]
            my_dict['Column_2b']='_'+coverage_list[1]
    else:
        my_dict['Coverage_Tag']=coverage
        my_dict['Column_2b']=""

def re_flow(n,flow_str,time_list,cycle_time,flow_list):
    my_dict={}
    cyctime_list=[]
    my_dict['num']=n
    my_time=tm.findall(flow_str)
    my_dict["Technology_Tag"]=tech.search(flow_str).group()
    coverage=cov.search(flow_str).group()
    if '@'in coverage:
        coverage=coverage.split('@',1)[0]
        add_cov_2b(coverage,my_dict)
    elif '-'in coverage:
        coverage=coverage.split('-',1)[0]
        add_cov_2b(coverage,my_dict)
    else:
        add_cov_2b(coverage,my_dict)
    my_dict["SubSubTest"]=sub.search(flow_str).group()
    my_dict["StartTime"]=time_list[-1][-1]
    my_dict["EndTime"]=my_time[-1]
    # 计算耗时
    my_dict['SpendTime']=time_data(my_dict['EndTime'])-time_data(my_dict['StartTime'])

    tag_name=my_dict['Technology_Tag']
    cyctime_list.append(n)
    cyctime_list.append(tag_name)
    cyctime_list.append(my_dict['SpendTime'])

    cycle_time.append(cyctime_list)
    time_list.append(my_time)
    flow_list.append(my_dict)
    return cycle_time,time_list,flow_list

    
def create_flow_log_list(path):
    flow_list=[]
    time_list=[]
    cycle_time=[]
    data=[]
    n=0
    with open(path,'r')as f:
        my_str=f.read()
    my_str=remove_special_characters(my_str)
    spl_list=my_str.split('==Test: ')
    # 读取spl列表中第一个元素里的时间，存入时间列表
    spl_time_1=re.findall(r'\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\.\d{3}',spl_list[0])
    time_list.append(spl_time_1)

    # 去掉spl_list中比较特殊的头和尾，得到一个新列表
    new_list=spl_list[1:-1]

 
    for i in new_list:
        n=n+1
        cycle_time,time_list,flow_list=re_flow(n,i,time_list,cycle_time,flow_list)
    cycle_time,time_list,flow_list=re_flow(n+1,spl_list[-1],time_list,cycle_time,flow_list)
    # 取出要生成表格所需要的数
    for i in flow_list:
        # msg_list=[i['Technology_Tag'],i['Coverage_Tag'],None,i['SubSubTest'],i['SpendTime']]
        msg_list=[i['Technology_Tag'],i['Coverage_Tag'],i['Column_2b'],i['SubSubTest'],i['SpendTime']]
        data.append(msg_list)
    # 将表格先按列表中第一个元素排序，再按第二个元素排序
    # data.sort(key=operator.itemgetter(0,1),reverse=False)

    return data
def create_xlsx(data,path):
    info=['Tech Tag','Coverage Tag','_Column_2b','SUBSUBTESTNAME','Item Cycle Time (ms)','Coverage Tag Cycle Time (ms)','Tech Tag Cycle Time (ms)']
    # data列表添加表头
    data.insert(0,info)
    # 实例化workbook对象
    workbook=openpyxl.Workbook()
    # 激活一个sheet
    sheet=workbook.active
    # 为sheet设置title
    sheet.title='Item_CycleTime'
    # 列表，将数据写入表格
    for row_index,row_item in enumerate(data):
        for col_index,col_item in enumerate(row_item):
            sheet.cell(row=row_index+1,column=col_index+1,value=col_item)
    # 表格插入第一列
    sheet.insert_cols(idx=1)
    # 遍历表格的每一行
    st_row1=2
    st_row2=2
    my_sum1=0
    my_sum2=0
    my_sum3=0
    for i in range(1,sheet.max_row+1):
        if i ==1:
            for i in range(1,sheet.max_column+1):
                # 给第一行的表格上色，第一行字体加粗
                sheet.cell(row=1,column=i).fill=PatternFill('solid',fgColor='FFBB02')
                sheet.cell(row=1,column=i).font=Font(bold=True)
        else:
            my_sum3=my_sum3+sheet.cell(row=i,column=6).value
            # 给第一列每个单元格赋值
            sheet.cell(row=i,column=1).value=i-1
            # 根据第3列合并单元格
            if sheet.cell(row=i,column=3).value==sheet.cell(row=i+1,column=3).value and sheet.cell(row=i,column=2).value==sheet.cell(row=i+1,column=2).value:
                my_sum1=my_sum1+sheet.cell(row=i,column=6).value
            else:
                my_sum1=my_sum1+sheet.cell(row=i,column=6).value
                sheet.cell(row=st_row1,column=7).value=my_sum1
                sheet.cell(row=st_row1,column=7).alignment=Alignment(horizontal='center',vertical='center')
                sheet.merge_cells(start_column=7,end_column=7,start_row=st_row1,end_row=i)
                st_row1=i+1
                my_sum1=0
            # 根据第二列合并单元格
            if sheet.cell(row=i,column=2).value==sheet.cell(row=i+1,column=2).value:
                my_sum2=my_sum2+sheet.cell(row=i,column=6).value
            else:
                my_sum2=my_sum2+sheet.cell(row=i,column=6).value
                sheet.cell(row=st_row2,column=8).value=my_sum2
                sheet.cell(row=st_row2,column=8).alignment=Alignment(horizontal='center',vertical='center')
                sheet.merge_cells(start_column=8,end_column=8,start_row=st_row2,end_row=i)
                st_row2=i+1
                my_sum2=0
    # 给表格增加一行，表示总耗时
    sheet.cell(row=sheet.max_row,column=1).value=sheet.max_row-1
    sheet.cell(row=sheet.max_row,column=2).value='Total Time(ms)'
    sheet.cell(row=sheet.max_row,column=sheet.max_column-2).value=my_sum3
    sheet.cell(row=sheet.max_row,column=6).alignment=Alignment(horizontal='center',vertical='center')
    sheet.merge_cells(start_column=2,end_column=5,start_row=sheet.max_row,end_row=sheet.max_row)
    sheet.merge_cells(start_column=6,end_column=sheet.max_column,start_row=sheet.max_row,end_row=sheet.max_row)
    # 写入excel文件
    timer_log = path
    workbook.save(timer_log)
#给param1，param2添加规则
def modification_rules(TestActions):
    dic = {'relaySwitch':{'param1':'netname','param2':'state'},'checkVBUS':{'param1':'netname','param2':'reference'},'readVoltage':{'param1':'netname','param2':'gain'},
    'parseLDCMString':{'param1':'pattern','param2':'pattern_limit'},'sendCmdAndParse':{'param1':'pattern','param2':'mark'},'sendCmdAndParseWithDC':{'param1':'pattern','param2':'attribute'},
    'parseWithRegexString':{'param1':'pattern','param2':'attribute'},'readAndCheckOTP':{'param1':'pattern','param2':'attribute'},'readGPIOState':{'param1':'netname','param2':'reference'},
    'checkScannedSNAndMLBSN':{'param1':'pattern','param2':'attribute'},'setAmplification':{'param1':'netname','param2':'netname2'},'orionTest':{'param1':'param1','param2':'param2'},
    'powerSupply':{'param1':'powertype','param2':'start'},'ampMeasBBTL':{'param1':'pattern','param2':'param1'}}

    param1 = ''
    param2 = ''
    allkeys = dic.keys()  
    if TestActions in allkeys:
        param1 = dic[TestActions]['param1']    
        param2 = dic[TestActions]['param2']        
    return param1,param2
#填充字典       
def design_dic(Technology,TestName,subsubtestname,param1,param2,timeout,TestActions,description,fail_count,key,val,nyquist_group,nyquist_proposed_rate):
    dic_mas = {'unit':"",'low':"",'high':""}
    dic_mas['technology'] = Technology
    dic_mas['testname'] = TestName
    dic_mas['subsubtestname'] = subsubtestname
    dic_mas['timeout'] = timeout
    dic_mas['testactions'] = TestActions
    dic_mas['description'] = description
    dic_mas['fail_count'] = fail_count
    dic_mas['key'] = key
    dic_mas['val'] = val
    dic_mas['param1'] = param1
    dic_mas['param2'] = param2
    dic_mas['nyquist_group'] = nyquist_group
    dic_mas['nyquist_proposed_rate'] = nyquist_proposed_rate
    return dic_mas
#去掉'"',' ','}'
def remove_symbol(mas_param):
    while 1:
        if mas_param[:1] != '"' and mas_param[:1] != ' ':
             break
        mas_param = mas_param[1:]
    while 1:
        if mas_param[-1:] != '"' and mas_param[-1:] != '}':
            break
        mas_param = mas_param[:-1]
    return mas_param

def create_testplan(path):
    mas_all = []
    path_limit=path+'/Limits.csv'
    path_main=path+'/Main.csv'
    path_Sampling=path+'/Sampling.csv'
    # 打开Limits.csv，去掉表头
    with open(path_limit)as a:
        limit_list = list(csv.reader(a))
        limit_list.pop(0)
    # 打开Main.csv，去掉表头
    with open(path_main, 'r') as f:
        reader_main = list(csv.reader(f))
        reader_main.pop(0)
    # 打开Sampling.csv，去掉表头
    with open(path_Sampling, 'r') as f:
        reader_Sampling = list(csv.reader(f))
        reader_Sampling.pop(0)
    #根据Main.csv中的Technology来打开对应的csv文件
    for mas in reader_main:
        path_csv = path+'/Tech/' + mas[1] + '.csv'
        with open(path_csv, 'r') as f:
            reader_Tech = list(csv.reader(f))
            reader_Tech.pop(0)
        #一行行遍历csv的内容，并取得所需内容
        num_Tech = 0
        while num_Tech < len(reader_Tech):
            if reader_Tech[num_Tech][0] == '':
                reader_Tech[num_Tech][0] = reader_Tech[num_Tech-1][0]
            if reader_Tech[num_Tech][0] == mas[0]:
                fail_count = ''
                nyquist_group = ''
                nyquist_proposed_rate = ''

                nyquist_group = mas[8]
                if nyquist_group != '':
                    for x in reader_Sampling:
                        if nyquist_group == x[0]:
                            nyquist_proposed_rate = x[1]
                if mas[9] == 'Y':
                        fail_count = 1
                #去掉"subsubtestname","timeout","unit","record"
                mas_list_param1 = reader_Tech[num_Tech][7].split('",')
                num = 0
                a = len(mas_list_param1)
                while num < a:
                    if '"subsubtestname"' in mas_list_param1[num].lower() or '"timeout"' in mas_list_param1[num].lower() or '"unit"' in mas_list_param1[num].lower() or '"record"' in mas_list_param1[num].lower():
                        mas_list_param1.remove(mas_list_param1[num])
                        a -= 1
                        num -= 1
                    num += 1
                mas_list = re.findall(r'(?<=")(.*?)(?=")',reader_Tech[num_Tech][7])
                #AdditionalParameters为空的情况
                if len(mas_list)<1:
                    param1 = ''
                    param2 = ''
                    subsubtestname = ''
                    timeout = ''
                    key = ''
                    val = ''
                    TestActions = ''
                    
                    list_TestActions = reader_Tech[num_Tech][1].split(':')
                    if len(list_TestActions) > 1:
                        TestActions = list_TestActions[1]
                    else:
                        TestActions = reader_Tech[num_Tech][1]
                    if reader_Tech[num_Tech][5] != '':
                        timeout = reader_Tech[num_Tech][5]
                    if '==' in reader_Tech[num_Tech][12]:
                        conditon = reader_Tech[num_Tech][12].split('==')
                        key = "{{" + str(conditon[0]) + "}}"
                        val = conditon[1]
                    param2 = reader_Tech[num_Tech][10]
                    mas_all.append(design_dic(mas[1],mas[0],subsubtestname,param1,param2,timeout,TestActions,mas[11],fail_count,key,val,nyquist_group,nyquist_proposed_rate)) 
                num_list = 0    
                while num_list < len(mas_list):
                    param1 = ''
                    param2 = ''
                    subsubtestname = ''
                    timeout = ''
                    key = ''
                    val = ''
                    TestActions = ''

                    list_TestActions = reader_Tech[num_Tech][1].split(':')
                    if len(list_TestActions) > 1:
                        TestActions = list_TestActions[1]
                    else:
                        TestActions = reader_Tech[num_Tech][1]
                    param1_rule,param2_rule = modification_rules(TestActions)
                    if len(param1_rule) > 0 or len(param2_rule) > 0:
                        for x in mas_list_param1:
                            if param1_rule in x:
                                param1 = x.split(':')[1]
                                param1 = remove_symbol(param1)
                            if param2_rule in x:
                                param2 = x.split(':')[1]
                                param2 = remove_symbol(param2)
                    else:
                        num_param = 0      
                        while num_param < len(mas_list_param1):
                            mas_param = mas_list_param1[num_param].split('":')[1]
                            mas_param = remove_symbol(mas_param)
                            param1 = param1 + mas_param + ';'
                            num_param += 1
                        if param1 == ';':
                            param1 = ''
                        else:
                            param1 = param1[:-1]
                    if reader_Tech[num_Tech][10]:
                        param1 = reader_Tech[num_Tech][10]
                    if 'timeout' == mas_list[num_list].lower():
                        timeout = mas_list[num_list+2]
                    if reader_Tech[num_Tech][5] != '':
                        timeout = reader_Tech[num_Tech][5]
                    if '==' in reader_Tech[num_Tech][12]:
                        conditon = reader_Tech[num_Tech][12].split('==')
                        key = "{{" + str(conditon[0]) + "}}"
                        val = conditon[1]       
                    if 'subsubtestname' == mas_list[num_list]:
                        subsubtestname = mas_list[num_list+2]
                        mas_all.append(design_dic(mas[1],mas[0],subsubtestname,param1,param2,timeout,TestActions,mas[11],fail_count,key,val,nyquist_group,nyquist_proposed_rate)) 
                    num_list += 1
            num_Tech += 1
    for i in mas_all:
        for x in limit_list:
            if i['testname']==x[0] and i['subsubtestname']==x[1]:
                i['unit']=x[2]
                i['high']=x[3]
                i['low']=x[4]
    # 制作最后要生成的testplan
    # 表头
    data=[['TESTNAME','SUBTESTNAME','SUBSUBTESTNAME','DESCRIPTION','FUNCTION','TIMEOUT','PARAM1','PARAM2','UNIT','LOW','HIGH','KEY','VAL','FAIL_COUNT','NYQUIST_GROUP','NYQUIST_PROPOSED_RATE']]
    for i in mas_all:
        if '@'in i['testname']:
            i['testname']=re.findall(r'.*?(?=@)',i['testname'])[0]
        elif '-'in i['testname']:
            i['testname']=re.findall(r'.*?(?=-)',i['testname'])[0]
        else:
            i['testname']=i['testname']
        my_list=[i['technology'],i['testname'],i['subsubtestname'],i['description'],i['testactions'],i['timeout'],i['param1'],i['param2'],i['unit'],i['low'],i['high'],i['key'],i['val'],i['fail_count'],i['nyquist_group'],i['nyquist_proposed_rate']]
        data.append(my_list)
    # with open ('./testplan.csv','w')as f:
    #     for x in data:
    #         csv.writer(f).writerow(x)
    return data

def get_pst_time():
    date_format=  '%Y-%m-%d_%H-%M-%S' #'%m-%d%Y_%H_%M_%S_%Z'
    date = datetime.datetime.now(tz=pytz.utc)
    date = date.astimezone(timezone('US/Pacific'))
    pstDateTime=date.strftime(date_format)
    return pstDateTime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Font, Side

DIAG_KEY = [
    "[31337-Send:]",  # key of PRM
    "[7603-Send:]",  # key of PRM
    "cmd-send: ",  # key of SC
    "[0016099E:3E6B001E] :-) ",  # key of other vendor
    "[000931D0:1423001E] :-) ",  # key of other vendor
    "[001A7109:02EA001E] :-) ",  # key of other vendor
    "[dut-send:]"
]

smokey_key = "          :-) "
SMOKEY_PAT = re.compile("          :\-\) (.*)")
DIAG_PAT = re.compile("({})(.*)".format(re.escape("!".join(DIAG_KEY)).replace('!', '|')))

class ParseFlowLog(object):
    def __init__(self):
        super(ParseFlowLog, self).__init__()
        self.selectedStation = dict()
    #将选中的文件路径保持到selectedStation
    def select_flow_file(self,path_list):
        path_list = path_list.split(';')
        path_list.pop(0)
        _validateFlogLogPath = list()
        _error = []
        # select all files ends with "_flow.log", string before _flow.log is will be mark as station name
        for filePath in path_list:
            _fileName = os.path.basename(filePath)
            if _fileName.endswith("flow.log"):
                _station = _fileName
                self.selectedStation[_station] = filePath
                _validateFlogLogPath.append(filePath)
            else:
                _error.append(_fileName)
        if _error:
            print("Invaild flow logs:\n{}".format(_error))

    def generate(self,choose):
        #判断是对加载的文件进行全选还是单选
        _selectedStation = choose
        if not _selectedStation:
            return
        elif _selectedStation.upper() == 'ALL':
            _fileNameToParse = list(self.selectedStation.keys())
        else:
            _fileNameToParse = [_selectedStation]
        #表头
        headerName = ["Tech Tag", "Test Item", "Diags Commands", "Station", "Is\nSmokey\n/Selftest",
                      "Commands Qty by Tech Tags"]
        #编辑xlsx文件名
        diag_name_list = ["DiagsCmd"]
        if self.parseType == 0:
            diag_name_list.append("ParsebyTestItems")
        elif self.parseType == 1:
            diag_name_list.append("ParsebyTechTag")
        if self.isSmokey:
            diag_name_list.append("Somkey")
        diag_name_list.append(datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S'))
        self.diag_file_name = "_".join(diag_name_list) + ".xlsx"
        #编辑xlsx文档
        _workBook = Workbook()
        self._workSheet = _workBook.active
        for index, value in enumerate(headerName):
            cell_position = str(self._workSheet.cell(1, index + 1)).split('.')[1].split('>')[0]
            self._workSheet[cell_position] = value
            self._workSheet[cell_position].font = Font(bold=True)
            self._workSheet[cell_position].alignment = Alignment(horizontal='center', vertical='center')
        for col, width in zip('ABCDEF', [15, 80, 80, 30, 25, 30]):
            self._workSheet.column_dimensions[col].width = width
        # 判断制作test item还是tech tag
        self._current_row = 2  
        if self.parseType == 0:
            for _name in _fileNameToParse:
                self.parse_by_test_item(_name)
        elif self.parseType == 1:
            self.parse_by_tech_tag(_fileNameToParse)
        else:
            return
        self.merge_test_item_column()
        #写入xlsx文档
        _workBook.save(os.path.join(self.path_xlsx, self.diag_file_name))
        _workBook.close()

    def parse_by_test_item(self, station_name):
        file_path = self.selectedStation[station_name]
        with open(file_path, "r", encoding='gb18030', errors='ignore') as f:
            data = f.read()
            items = re.split("lower:.*; upper:.*; value:.*", data)
            _lastTest = ""
            _qty = 0
            _testCount = 1
            _testColumnStartRow = self._current_row
            _end = False
            for item in items:
                try:
                    _test = re.findall("==Test:(.*)", item)[0].strip()
                    _subtest = re.findall("==SubTest:(.*)", item)[0].strip()
                    _subsubtest = re.findall("==SubSubTest:(.*)", item)[0].strip()
                except Exception as e:
                    _end = True
                # write Commands Qty by Tech Tags and clear buff
                if (_lastTest and _test != _lastTest) or _end:
                    _line = ["", "", "", "", "", _qty]
                    merge_cell_position = "A{}:A{}".format(_testColumnStartRow, _testColumnStartRow + _testCount - 1)
                    self._workSheet.merge_cells(merge_cell_position)
                    _testCount = 1
                    _qty = 0
                    self._workSheet.append(_line)
                    self._workSheet.cell(self._current_row, 6).fill = PatternFill(fill_type="solid",
                                                                                  fgColor="F0F005")
                    self._workSheet.cell(self._current_row, 6).alignment = Alignment(horizontal='left',
                                                                                     vertical='center')
                    self._current_row += 1
                    _testColumnStartRow = self._current_row
                    if _end:
                        break
                _lastTest = _test
                _commands = DIAG_PAT.findall(item)
                _testItem = ' '.join([_test, _subtest, _subsubtest])
                if not _commands:
                    _line = [_test, _testItem, 'NA', station_name, "", ""]
                    self._workSheet.append(_line)
                    self._workSheet.cell(self._current_row, 1).alignment = Alignment(horizontal='left',
                                                                                     vertical='center')
                    self._workSheet.cell(self._current_row, 2).alignment = Alignment(horizontal='left',
                                                                                     vertical='center')
                    self._current_row += 1
                    _testCount += 1
                else:
                    for val in _commands:
                        if len(val) == 2 and val[1]:
                            _qty += 1
                            cmd = val[1]
                            _isSmokey = 'Y' if cmd.startswith("smokey") or 'selftest' in " ".join(
                                [_testItem, cmd]).lower() else "N"

                            _line = [_test, _testItem, cmd, station_name, _isSmokey, ""]
                            self._workSheet.append(_line)
                            self._workSheet.cell(self._current_row, 1).alignment = Alignment(horizontal='left',
                                                                                             vertical='center')
                            self._workSheet.cell(self._current_row, 2).alignment = Alignment(horizontal='left',
                                                                                             vertical='center')
                            self._current_row += 1
                            _testCount += 1
                if self.isSmokey:
                    smokeyItems = SMOKEY_PAT.findall(item)
                    for cmd in smokeyItems:
                        if val:
                            _qty += 1
                            _isSmokey = 'Y'
                            _line = [_test, _testItem, cmd, station_name, _isSmokey, ""]
                            self._workSheet.append(_line)
                            self._workSheet.cell(self._current_row, 1).alignment = Alignment(horizontal='left',
                                                                                             vertical='center')
                            self._workSheet.cell(self._current_row, 2).alignment = Alignment(horizontal='left',
                                                                                             vertical='center')
                            self._current_row += 1
                            _testCount += 1

    def parse_by_tech_tag(self, files):
        _dictByTech = dict()
        # collect all files item by tech
        for stationName in files:
            file_path = self.selectedStation[stationName]
            with open(file_path, "r", encoding='gb18030', errors='ignore') as f:
                data = f.read()
                items = re.split("lower:.*; upper:.*; value:.*", data)
                for item in items:
                    try:
                        _test = re.findall("==Test:(.*)", item)[0].strip()
                        if _test in _dictByTech.keys():
                            _dictByTech[_test].append({"item": item, "station": stationName})
                        else:
                            _dictByTech[_test] = [{"item": item, "station": stationName}]
                    except:
                        continue
        _qty = 0
        _testCount = 1
        self._current_row = 2
        _testColumnStartRow = self._current_row
        for key, items in _dictByTech.items():
            _isRecord = False
            _currentTagCmds = set()
            for data in items:
                item = data['item']
                station_name = data['station']
                _test = re.findall("==Test:(.*)", item)[0].strip()
                _subtest = re.findall("==SubTest:(.*)", item)[0].strip()
                _subsubtest = re.findall("==SubSubTest:(.*)", item)[0].strip()
                _testItem = ' '.join([_test, _subtest, _subsubtest])
                _commands = DIAG_PAT.findall(item)

                for val in _commands:
                    if len(val) != 2:
                        continue
                    cmd = val[1].strip()
                    if not cmd or cmd in _currentTagCmds:  # if repeat command, skip
                        continue
                    _currentTagCmds.add(cmd)
                    if len(val) == 2 and val[1]:
                        _isRecord = True
                        _qty += 1
                        cmd = val[1]
                        _isSmokey = 'Y' if cmd.startswith("smokey") or 'selftest' in " ".join(
                            [_testItem, cmd]).lower() else "N"

                        _line = [_test, _testItem, cmd, station_name, _isSmokey, ""]
                        self._workSheet.append(_line)
                        self._workSheet.cell(self._current_row, 1).alignment = Alignment(horizontal='left',
                                                                                         vertical='center')
                        self._workSheet.cell(self._current_row, 2).alignment = Alignment(horizontal='left',
                                                                                         vertical='center')
                        self._current_row += 1
                        _testCount += 1

                if self.isSmokey:
                    _smokeyCommands = SMOKEY_PAT.findall(item)
                    for cmd in _smokeyCommands:
                        if not cmd.strip() or cmd.strip() in _currentTagCmds:
                            continue
                        _currentTagCmds.add(cmd)
                        _isRecord = True
                        _qty += 1
                        _isSmokey = 'Y'
                        _line = [_test, _testItem, cmd, station_name, _isSmokey, ""]
                        self._workSheet.append(_line)
                        self._workSheet.cell(self._current_row, 1).alignment = Alignment(horizontal='left',
                                                                                         vertical='center')
                        self._workSheet.cell(self._current_row, 2).alignment = Alignment(horizontal='left',
                                                                                         vertical='center')
                        self._current_row += 1
                        _testCount += 1
                    # if line startswith ":-) ", it's a inner smokey command
                    for line in item.split('\n'):
                        _match = re.match("^:\-\) (.+)", line)
                        if _match:
                            cmd = _match.group(1)
                            if not cmd.strip() or cmd.strip() in _currentTagCmds:
                                continue
                            _currentTagCmds.add(cmd)

                            _line = [_test, _testItem, cmd, station_name, "Y", ""]
                            self._workSheet.append(_line)
                            self._workSheet.cell(self._current_row, 1).alignment = Alignment(horizontal='left',
                                                                                             vertical='center')
                            self._workSheet.cell(self._current_row, 2).alignment = Alignment(horizontal='left',
                                                                                             vertical='center')
                            self._current_row += 1
                            _qty += 1
                            _isRecord = True
                            _testCount += 1
            if not _isRecord:
                continue
            _line = ["", "", "", "", "", _qty]
            merge_cell_position = "A{}:A{}".format(_testColumnStartRow, _testColumnStartRow + _testCount - 1)
            self._workSheet.merge_cells(merge_cell_position)
            self._workSheet.append(_line)
            self._workSheet.cell(self._current_row, 6).fill = PatternFill(fill_type="solid",
                                                                          fgColor="F0F005")
            self._workSheet.cell(self._current_row, 6).alignment = Alignment(horizontal='left',
                                                                             vertical='center')
            self._current_row += 1
            _testCount = 1
            _qty = 0
            _testColumnStartRow = self._current_row

    def merge_test_item_column(self):
        _lastItem = ""
        _itemCount = 0
        _startRow = 1
        for i in range(1, self._workSheet.max_row):
            _testItem = self._workSheet.cell(i, 2).value
            if (_lastItem and _lastItem != _testItem) or i == self._workSheet.max_row:
                if _itemCount > 1:
                    self._workSheet.merge_cells("B{}:B{}".format(_startRow, _startRow + _itemCount - 1))
                _startRow = i
                _itemCount = 0
            if not _lastItem:  # if last item is empty, update start row
                _startRow = i
            _lastItem = _testItem
            if _lastItem:
                _itemCount += 1

    def dorun(self,path,path_xlsx):
        self.path_xlsx = path_xlsx 
        self.parseType = int(path[2])
        self.isSmokey = int(path[1])
        self.select_flow_file(path[4])
        self.generate(path[3]) 
        return self.diag_file_name
 
def send_result(key,path):
    global g_my_datas
    
    if key=='atlas2_timer':
        data =  create_flow_log_list(path)
        PostJsonInfo( "atlas2_timer^&^{}".format(json.dumps(data)))
        
        g_my_datas["atlas2_timer"] = data

    elif key=='atlas2_timer_export':
        cpklog = os.path.expanduser("~/Desktop/CPK_Log/")
        if not os.path.exists(cpklog):
            os.makedirs(cpklog)

        if "atlas2_timer" in g_my_datas.keys():
            path =  os.path.expanduser("~/Desktop/CPK_Log/")+ "Timer" + str(get_pst_time()) + ".xlsx"
            create_xlsx(copy.deepcopy(g_my_datas["atlas2_timer"]),path)
            PostJsonInfo( "finish^&^{}".format(path))
        else:
            PostJsonInfo( "exception^&^Please Load First !! ")

    elif key=='atlas2_profile':
        data = create_testplan(path)

        g_my_datas["atlas2_profile"] = data
        data_dump = copy.deepcopy(data)
        del(data_dump[0])
        PostJsonInfo( "atlas2_profile^&^{}".format(json.dumps(data_dump)))
        
    elif key=='atlas2_profile_load':
        # convert atls2 profile to legacy & return save path as new_path
        data = create_testplan(path)

        cpklog = os.path.expanduser("~/Desktop/CPK_Log/")
        if not os.path.exists(cpklog):
            os.makedirs(cpklog)

        cpklog = os.path.expanduser("~/Desktop/CPK_Log/")+ "LegacyTestPlan" + "__" + str(get_pst_time()) + ".csv"

        with open (cpklog,'w')as f:
            for x in data:
                #print("atlas2_ write line {}".format(x))
                csv.writer(f).writerow(x)

        # new_path = "/usr/lib" + cpklog

        PostJsonInfo( "atlas2_profile_load^&^{}".format(json.dumps({"path":cpklog})))

    elif key=='atlas2_profile_export':

        cpklog = os.path.expanduser("~/Desktop/CPK_Log/")
        if not os.path.exists(cpklog):
            os.makedirs(cpklog)

        cpklog = os.path.expanduser("~/Desktop/CPK_Log/")+ "LegacyTestPlan" + "__" + str(get_pst_time()) + ".csv"

        if "atlas2_profile" in g_my_datas.keys():
            with open (cpklog,'w')as f:
                for x in g_my_datas["atlas2_profile"]:
                    #print("atlas2_ write line {}".format(x))
                    csv.writer(f).writerow(x)
            PostJsonInfo( "finish^&^{}".format(cpklog))
        else:
            PostJsonInfo( "exception^&^Please Load First !! ")
     
    elif key=='atlas2_diags_extrator':
        cpklog = os.path.expanduser("~/Desktop/CPK_Log/")
        if not os.path.exists(cpklog):
            os.makedirs(cpklog)
        app = ParseFlowLog()
        filepath = app.dorun(path,cpklog)
        g_my_datas["atlas2_diags_extrator"] = filepath

        #[self sendAtlas2ZmqMsg:[NSString stringWithFormat:@"atlas2_diags_extrator@%d@%d@%@@%@",isSmokey,parseOption,fileopetion,ret]];
        
        if "atlas2_diags_extrator" in g_my_datas.keys():
            path = os.path.expanduser("~/Desktop/CPK_Log/") + g_my_datas["atlas2_diags_extrator"]
            PostJsonInfo( "finish^&^{}".format(path))
        else:
            PostJsonInfo( "exception^&^Please Load First !! ")

def run(n):
    while True:
        try:
            print("wait for atlas2 client ...")
            zmqMsg = socket.recv()
            socket.send(b'atlas2.com')
            if len(zmqMsg)>0:

                key = zmqMsg.decode('utf-8')

                print(key)
                keys = key.split("@")

                if len(keys) ==2:
            
                    send_result(keys[0],keys[1])
                elif len(keys) >2:
                    send_result(keys[0],keys)
            else:
                time.sleep(0.05)
        except Exception as e:
           PostJsonInfo( "exception^&^" + "\n" + str(e) + "\n [In Atlas test Server]") 

#def Test(data ,label=[],title="Test",select_items = ["A@1","B@2","C@3","D@4"]):
    
    plt.title(title,fontsize=20)
if __name__ == '__main__':
    # t1 = threading.Thread(target=run, args=("<<cpk1>>",))
    # t1.start()
    run(0)
    #Test()
    
    




